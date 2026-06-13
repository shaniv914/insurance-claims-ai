"""
Standard append script for Storm Case Database — Schema v1 (21 columns).

Usage
-----
1. Read the source PDF(s) and extract the fields listed in NEW_CASES below.
2. Add one dict per case to NEW_CASES following the extraction rules.
3. Run from the repo root:
       py scripts/append_storm_v1.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Storm_Case_Database.xlsx

===========================================================================
FIELD EXTRACTION RULES
===========================================================================

Case ID         : Format STORM-NNN (zero-padded to 3 digits)
FOS Decision ID : DRN-XXXXXXX or DRNXXXXXXX as printed in the PDF
Insurer Name    : Formal registered name from the FOS decision
FOS Decision Date : DD Mon YYYY — accept-or-reject deadline in final paragraph
Claim Type      : Physical incident and nature of dispute in one sentence
Leak Source     : Physical storm damage mechanism / source of water ingress
                  e.g. "Wind-lifted roof tiles allowing water ingress"
                       "Storm debris impact on roof structure"
                       "Wind-driven rain through failed pointing"
Property Type   : "Residential home" / "Unoccupied residential property" /
                  "Leasehold flat" / "Commercial" / etc.
Dispute Type    : Controlled vocab (7 values)
Coverage Decision : Controlled vocab (5 values)
Rejection Reason  : Insurer's stated reason for declining
Evidence Dispute  : What evidence each party relied on
Outcome Category  : Controlled vocab (4 values)
Outcome           : Full FOS remedy instructions
Compensation Awarded (£) : Integer — D&I only; 0 if none
Is Core Case      : Controlled vocab (5 values)
Key Policy Clause : Policy wording or FOS/FCA principle applied
Missing Evidence  : Evidence that was absent and affected the outcome
Ombudsman Reasoning : How the ombudsman weighed the evidence
Workflow Insight  : Operational rule for the claims workflow
AI Rule Candidate : Machine-evaluable rule for the rules engine
Source PDF        : Filename only (e.g. DRN-1207086.pdf)
===========================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — populate before running
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "STORM-031",
        "FOS Decision ID": "DRN3019884",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "26 Oct 2015",
        "Claim Type": "Storm damage to roof of residential home declined; Mr and Mrs M claimed storm damage; two independent experts appointed by RSA found no storm damage — second expert found no missing or dislodged slates or tiles, lead flashing intact, and roof deteriorating over years; policyholders' contractor claimed storm damage but could not displace two concordant expert opinions; prejudice from delayed claim handling noted but not determinative",
        "Leak Source": "Roof — two expert inspections found no missing or dislodged slates or tiles, all lead flashing in good repair; damage attributed to multi-year deterioration not a storm event; policyholders' contractor claimed storm damage but provided no independent expert support",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Two independent experts appointed by RSA found damage inconsistent with storm damage; second expert explicitly found no signs of a one-off insurable event, no missing or dislodged slates or tiles, and confirmed the roof had been deteriorating for years; policyholders' contractor's opinion of storm damage not supported by independent expert findings",
        "Evidence Dispute": "RSA: two concordant expert reports both finding no storm damage; first expert noted damage had occurred a long time before and RSA's position was prejudiced as damage worsened over time; second expert found no missing or dislodged slates or tiles, lead flashing in good repair, no signs of a one-off insured event, roof deteriorating for years. Policyholders: their contractor said storm damage; argued claim was made months earlier but no forms were sent by RSA; neighbour's claim was settled by their insurer; internal damage also raised but excluded from complaint at policyholders' request",
        "Outcome Category": "Not Upheld",
        "Outcome": "FOS did not uphold; RSA's decline confirmed as fair; Q2 and Q3 answered no; contractor's assertion of storm damage insufficient to displace two concordant expert findings; prejudice from delayed claim handling noted but not determinative as storm damage itself not established; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Q2 and Q3 failed — two concordant expert findings of no storm damage, including explicit finding of no signs of a one-off insurable event and documented multi-year roof deterioration, are sufficient to decline the claim; a policyholders' contractor's assertion of storm damage does not displace two independent expert opinions finding no storm damage; prejudice from delayed claim handling is relevant only if the underlying storm damage claim would have succeeded — where Q2 fails it does not need to be resolved",
        "Missing Evidence": "Expert evidence supporting storm causation (policyholders had only their contractor's opinion, not an independent expert report); evidence that damage was the result of a one-off insured event rather than progressive deterioration",
        "Ombudsman Reasoning": "Q1 — storms occurred in the area (no specific date pinpointed but accepted generally on overall records). Q2 — policyholders' contractor said storm damage; two RSA experts disagreed; second expert found no missing or dislodged slates or tiles, lead flashing in good repair, no signs of a one-off insurable event, roof deteriorating for years; answer to Q2 no. Q3 not considered once Q2 failed. RSA acted reasonably. Prejudice from delayed claim noted but not determinative once storm damage not established. Neighbour's claim with different insurer irrelevant. Internal damage excluded from complaint by policyholders.",
        "Workflow Insight": "Two concordant expert findings of no storm damage are sufficient to decline Q2; a contractor's storm assertion does not displace two independent expert reports where both find no signs of a one-off insurable event; where the underlying storm damage claim fails at Q2, prejudice arguments (delayed forms, late inspection) need not be resolved; confirm at first handling whether internal damage is also claimed and assess it separately to avoid creating a separate complaint head",
        "AI Rule Candidate": "IF two_independent_experts_find_no_storm_damage AND no_missing_or_dislodged_slates AND roof_deteriorating_years THEN storm_q2 = no AND decline_reasonable; IF policyholder_counter_evidence_is_contractor_assertion_only THEN insufficient_to_displace_concordant_expert_findings; IF storm_q2_fails THEN delay_prejudice_argument_need_not_be_resolved",
        "Source PDF": "DRN3019884.pdf",
    },
    {
        "Case ID": "STORM-032",
        "FOS Decision ID": "DRN5013915",
        "Insurer Name": "Ageas Insurance Limited",
        "FOS Decision Date": "12 Sep 2019",
        "Claim Type": "Storm damage to flat roof of residential home declined; January 2019 claim; felt and boards reported ripped off during high winds; Ageas declined based on a self-imposed 55mph storm threshold not in policy terms; FOS rejected threshold but found Q3 failed — 9+ year old unmaintained flat roof with no pre-storm condition evidence",
        "Leak Source": "Flat roof — felt and boards underneath reported ripped off during high winds; policyholders had contractors carry out permanent repairs before any inspection could be arranged; repair invoice listed work done but did not establish storm causation",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Ageas declined on basis of self-imposed 55mph storm threshold (not in policy terms — rejected by FOS); actual Q3 failure confirmed by FOS: flat roof had been in place for over 9 years without maintenance, approaching end of typical 10-15 year lifespan; no pre-storm condition report; repair invoice listed repairs carried out but not that storm was the main cause; immediate permanent repairs prevented any inspection to validate storm causation",
        "Evidence Dispute": "Insurer: weather records showing 48mph winds; self-imposed 55mph policy storm threshold (not in policy — FOS rejected); no inspection carried out as repairs completed before Ageas could inspect. Policyholders: Met Office postcode-level data showing 58mph (closer to property); reported felt and boards ripped off; repair invoice; acknowledged flat roof had been in place 9+ years without any maintenance",
        "Outcome Category": "Not Upheld",
        "Outcome": "FOS did not uphold; Ageas 55mph threshold rejected as absent from policy terms; Q1 accepted (48-58mph can cause structural damage = storm); Q2 accepted (felt and boards ripped off consistent with storm); Q3 failed — flat roof 9+ years without maintenance, no pre-storm condition evidence, repair invoice shows repairs not causation; storm not established as main cause; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "An insurer cannot apply a numeric storm threshold (here 55mph) that does not appear in the policy wording; Q3 — a flat roof in place for over 9 years without maintenance, approaching the end of its typical 10-15 year lifespan, with no evidence of good repair prior to the storm, fails Q3 as it is more likely that the weather highlighted pre-existing deterioration rather than caused the damage; a repair invoice listing work done does not establish storm causation; the burden of proof is on the policyholder to show storm was the main cause — absent pre-storm condition evidence this cannot be met",
        "Missing Evidence": "Pre-storm condition report showing flat roof was in good repair; specialist storm causation report establishing that the 9+ year old unmaintained flat roof was damaged by the specific storm rather than pre-existing deterioration; expert inspection of damage (prevented by immediate permanent repairs)",
        "Ombudsman Reasoning": "Q1 — weather records showed minimum 48mph; Met Office postcode data showed 58mph; FOS accepted structural damage could arise at these speeds; Ageas 55mph threshold not in policy terms and potentially unfair; Q1 yes. Q2 — felt and boards ripped off is consistent with storm damage; Q2 yes. Q3 — onus on policyholders to show storm was main cause; flat roof in place 9+ years without any maintenance; flat roofs typically last 10-15 years; damage could have been storm-caused but more likely weather highlighted pre-existing issues; no pre-storm condition report; invoice lists repairs only not causation; immediate repairs prevented inspection; Q3 no.",
        "Workflow Insight": "A self-imposed policy storm threshold (55mph) not in policy wording will be rejected by FOS; where a flat roof is nearing the end of its typical lifespan with no maintenance history, Q3 is likely to fail unless the policyholder can evidence good pre-storm condition; document any spontaneous policyholder admissions about roof age and maintenance history at FNOL as these are key Q3 evidence; a repair invoice alone does not establish storm causation",
        "AI Rule Candidate": "IF insurer_storm_threshold_not_in_policy THEN threshold_cannot_be_applied AND q1_assessed_on_standard_structural_damage_criteria; IF flat_roof AND roof_age_approaching_end_of_lifespan AND no_maintenance_history AND no_pre_storm_condition_report THEN storm_q3 = no; IF repair_invoice_lists_repairs_only_not_causation THEN insufficient_to_establish_q3; IF immediate_repairs_prevent_inspection THEN does_not_relieve_policyholder_of_q3_burden",
        "Source PDF": "DRN5013915.pdf",
    },
    {
        "Case ID": "STORM-033",
        "FOS Decision ID": "DRN5397298",
        "Insurer Name": "U K Insurance Limited",
        "FOS Decision Date": "28 Aug 2015",
        "Claim Type": "Storm damage to garden wall of residential home; brick boundary wall with trellis collapsed; UKI declined on wear and tear grounds; loss adjuster report given reduced weight due to factual errors including wrong insurer and inapplicable exclusion cited; FOS found storm was most likely dominant cause and upheld the complaint",
        "Leak Source": "Garden wall — boundary brick wall with small trellis on top blown over and collapsed in one piece; no roof water ingress; structural storm damage to boundary structure",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "UKI's surveyor found mould growth to mortar joints, nearby vegetation reducing structural strength, and no similar storm damage to surrounding fences facing the same direction; claim declined on wear and tear grounds",
        "Evidence Dispute": "UKI: loss adjuster report citing mould growth to joints, nearby vegetation, and no similar damage to surrounding fences. Policyholders: no wear and tear exclusion in their policy (only a maintenance condition); photographs of wall before and after collapse; incurred telephone and postage costs chasing UKI for response over two months. FOS: reduced reliance on loss adjuster report due to factual errors — wrong insurer cited, wear and tear exclusion cited that did not exist in the policy",
        "Outcome Category": "Upheld",
        "Outcome": "FOS upheld complaint; UKI required to: settle the storm claim for wall collapse with 8% simple interest on cash settlement; reimburse cost of replacement fence erected due to declined claim with 8% simple interest; pay £30.51 for telephone calls and postage costs; pay £100 compensation for distress and inconvenience",
        "Compensation Awarded (£)": 100,
        "Is Core Case": "Yes",
        "Key Policy Clause": "A loss adjuster's report carries reduced weight where it contains factual errors including citing the wrong insurer and applying a wear and tear exclusion that does not exist in the policy; photographs showing a wall fell in one piece and broke on impact are consistent with Q2 (blown over by storm-force wind); where there is no wear and tear exclusion in the policy (only a maintenance condition) the insurer cannot decline on that basis; neighbouring structures withstanding a storm is not conclusive against Q3 — shielding effects may explain the difference; a policyholder who cannot rebuild a declined structure is entitled to reimbursement of reasonable temporary remediation costs incurred as a direct consequence",
        "Missing Evidence": "Not applicable — FOS found storm causation established on balance of photographic evidence and discredited loss adjuster report",
        "Ombudsman Reasoning": "Q1 not disputed — storm on the day confirmed. Q2 — photographs showed wall fell in one piece and broke on impact with ground, consistent with being blown over by storm-force wind; yes. Q3 — UKI's loss adjuster cited mould growth to mortar joints, vegetation, and unaffected neighbouring fences; FOS placed reduced reliance on this report due to errors (wrong insurer referenced, wear and tear exclusion cited that did not exist in the policy — policy had a maintenance condition instead); photographs of wall before and after showed only minor weathering and limited moss, not enough to cause collapse without storm; small vegetation across bottom not shown to grow into wall affecting structural integrity; neighbouring fences may have been shielded by the wall; on balance storm was most likely dominant cause and wall would probably have remained intact but for storm.",
        "Workflow Insight": "Accuracy of decline documentation is critical — a loss adjuster report citing the wrong insurer or an inapplicable exclusion will carry reduced weight at FOS; always verify the exact policy exclusions before issuing a decline and ensure the surveyor has checked the actual policy wording; where a policyholder incurs reasonable remediation costs as a direct consequence of a wrongly declined claim, reimbursement will be required; a neighbour's structure withstanding a storm is not conclusive — consider shielding effects before relying on this as Q3 evidence",
        "AI Rule Candidate": "IF loss_adjuster_report_contains_factual_errors_including_wrong_exclusion_cited THEN report_weight_reduced; IF wall_fell_in_one_piece_and_broke_on_impact THEN storm_q2 = yes; IF no_wear_and_tear_exclusion_in_policy THEN cannot_decline_on_wear_and_tear_grounds; IF policyholder_incurs_reasonable_remediation_costs_due_to_wrongly_declined_claim THEN reimbursement_required",
        "Source PDF": "DRN5397298.pdf",
    },
    {
        "Case ID": "STORM-034",
        "FOS Decision ID": "DRN7021460",
        "Insurer Name": "Liverpool Victoria Insurance Company Limited",
        "FOS Decision Date": "21 Jul 2017",
        "Claim Type": "Storm damage to roof of residential home in exposed location; LV paid initial £5,000 for 2016 storm damage but declined additional £21,000 quote; policyholder had history of multiple prior claims (2012-2015) and could not produce invoices proving prior settled repairs were carried out; LV inspector also found non-storm causes",
        "Leak Source": "Roof — LV inspector found some storm damage but also inferior valley lead, tile damage consistent with workmanship not storm, and crumbling ridge mortar from gradual deterioration; policyholder's roofer reported sagging from water ingress following storm tile dislodgement; engineer's report addressed structural integrity only, not specific leaking areas or repair scope",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Partial",
        "Rejection Reason": "Policyholder could not provide invoices showing prior settled claims (2013, 2014, 2015) had been paid for and repairs carried out; LV inspector found non-storm causes including workman damage to tiles, inferior valley lead quality, and crumbling ridge mortar; engineer's report established general structural soundness but did not identify which areas were leaking, how long problems had existed, or what repairs were needed for the specific disputed scope",
        "Evidence Dispute": "LV: inspector report finding some storm damage plus workman tile damage, inferior valley lead, crumbling mortar; absence of invoices for 2013, 2014 and 2015 claim repairs; concerns about 2012 invoice. Policyholder: roofer's report saying sagging caused by storm-driven water ingress; engineer's report confirming good construction, quality award, correct purlin spacing, leaks attributable to storm not structural deficiencies; claimed to have sent proof of repairs to LV but could not provide evidence to FOS; offer to obtain structural engineer's report not followed through",
        "Outcome Category": "Not Upheld",
        "Outcome": "FOS did not uphold; LV's initial settlement of ~£5,000 not disturbed; additional ~£21,000 quote not payable — policyholder failed to show damage was attributable to 2016 storm rather than cumulative unrepaired prior storm damage; LV's non-renewal decision also confirmed as LV's commercial prerogative; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Where a policyholder has multiple prior settled storm claims paid in cash, the burden is on the policyholder to demonstrate those repairs were actually carried out before claiming for what may be cumulative deterioration; an engineer's report establishing general structural soundness and attributing leaks generally to storm does not establish Q3 for a specific additional claim scope where it does not identify which areas are leaking, the duration of problems, or the specific repair scope required; evidence offered but not provided to FOS cannot be considered; an insurer's non-renewal decision is a commercial decision FOS will not override",
        "Missing Evidence": "Invoices or receipts demonstrating that prior claim cash settlements for 2013, 2014 and 2015 were used to carry out the agreed repairs; an expert report identifying which specific areas of the roof were leaking, how long problems had persisted, and the scope of repairs attributable to the 2016 storm as distinct from cumulative unrepaired prior damage",
        "Ombudsman Reasoning": "Mr M provided evidence of storm damage to roof generally but not that current damage was attributable to the 2016 storm specifically; LV had paid cash settlements for multiple prior claims (2012-2015); LV required invoices confirming repairs were carried out; only 2012 invoice provided and that raised concerns; FOS found it more likely than not that prior invoices were not sent to LV; if prior repairs were not carried out and damage accumulated, LV not required to pay for that cumulative deterioration; engineer's report confirmed structural soundness and attributed leaks generally to storm but did not identify specific leaking areas, duration, or repair scope — insufficient to establish Q3 for the additional claim; LV's initial £5,000 settlement not disturbed; non-renewal is LV's commercial decision.",
        "Workflow Insight": "For policyholders with a history of multiple storm claims paid in cash, require proof of completed repairs before paying further claims — the prior repair requirement is FOS-endorsed; where a policyholder submits a general expert opinion attributing leaks to storm damage but the report does not address which specific areas are leaking or the current repair scope, it is insufficient to establish Q3 for the disputed additional amount; document why prior invoices are inadequate or absent in decline reasoning",
        "AI Rule Candidate": "IF multiple_prior_settled_storm_claims AND policyholder_cannot_evidence_prior_repairs_completed THEN current_claim_may_be_declined_as_cumulative_unrepaired_damage; IF expert_report_attributes_leaks_generally_to_storm BUT does_not_identify_specific_leaking_areas_or_repair_scope THEN insufficient_to_establish_q3_for_additional_claim; IF prior_claims_settled_in_cash AND no_repair_invoices THEN require_proof_of_completed_repairs_before_paying_further_claims",
        "Source PDF": "DRN7021460.pdf",
    },
    {
        "Case ID": "STORM-035",
        "FOS Decision ID": "DRN7244667",
        "Insurer Name": "Liverpool Victoria Insurance Company Limited",
        "FOS Decision Date": "2 Aug 2020",
        "Claim Type": "Storm damage to tiled roof and internal ceiling declined; winds reached 41mph; LV's trusted provider inspection found damage not consistent with a one-off event and pre-existing; policyholder's independent report suggested tiles moved by high winds; FOS found Q1 failed — 41mph winds and low hourly rainfall do not constitute storm conditions",
        "Leak Source": "Tiled roof — tiles reportedly moved by high winds allowing water ingress causing internal ceiling collapse; LV's trusted provider and review of independent report concluded damage was not consistent with a one-off storm event and appeared pre-existing and gradual",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Weather records showed maximum winds of 41mph — below storm threshold; hourly rainfall not exceptional; LV's inspector found damage not consistent with a one-off event (pre-existing, gradual); independent report submitted by policyholder attributed damage to high winds but high winds at 41mph do not constitute storm conditions",
        "Evidence Dispute": "LV: weather records showing 41mph winds; trusted provider inspection finding damage not consistent with one-off event and appearing pre-existing; LV's review of independent report finding damage not consistent with storm damage. Policyholder (represented by Mr A): independent report suggesting high winds moved tiles and rain entered; roofer said damage caused by weather conditions; workmen could not access roof during first visit due to wind; alleged further damage caused while awaiting repair",
        "Outcome Category": "Not Upheld",
        "Outcome": "FOS did not uphold; Q1 failed — winds 41mph below storm threshold and hourly rainfall not exceptional; storm conditions not established; damage more likely from gradual wear and tear and multiple causes; LV's decline confirmed as fair; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "41mph winds without exceptional accompanying rainfall do not constitute storm conditions; storm conditions can exist without very strong winds only if accompanied by exceptional rainfall, hail, or snowfall — both wind and precipitation must be assessed together; workmen being unable to access a roof for safety reasons does not constitute evidence of storm conditions; strong winds are not the same as a storm; a roof in poor condition that cannot withstand normal weather does not give rise to a storm claim where storm conditions are not established",
        "Missing Evidence": "Evidence of exceptional rainfall, hail, or snowfall concurrent with the 41mph winds; expert evidence establishing storm causation rather than pre-existing gradual deterioration",
        "Ombudsman Reasoning": "Q1 — weather records showed 41mph winds; FOS considers storm winds to be stronger; storm conditions can exist without very strong winds only if rainfall, hail, or snowfall is exceptional; hourly rainfall was not high; no evidence of storm conditions around the date of claim; Q1 answered no. Q2 and Q3 not considered. LV's inspector found damage not consistent with a one-off event (pre-existing, gradual); LV reviewed independent report and found damage not consistent with storm; roofer's opinion that tiles moved by high winds does not establish storm conditions; delay in repair did not alter the pre-existing damage finding.",
        "Workflow Insight": "When declining on sub-storm wind speeds, also document the precipitation levels — if both wind and rainfall are below storm thresholds Q1 is clearly answered no; being unable to access a roof due to wind is a workman safety consideration, not storm conditions evidence; where a policyholder's independent report attributes tile movement to high winds but does not establish storm conditions, the insurer can rely on its own inspector's pre-existing damage finding; once Q1 fails Q2 and Q3 need not be addressed",
        "AI Rule Candidate": "IF winds_41mph AND rainfall_not_exceptional AND no_exceptional_hail_or_snow THEN storm_q1 = no; IF storm_conditions_not_established THEN q2_and_q3_not_considered AND claim_fails; IF workmen_cannot_access_roof_due_to_wind THEN not_evidence_of_storm_conditions; IF independent_report_attributes_damage_to_high_winds_only AND high_winds_not_storm THEN report_does_not_establish_storm_q1",
        "Source PDF": "DRN7244667.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Storm_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
