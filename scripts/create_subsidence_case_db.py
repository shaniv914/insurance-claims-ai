"""
Creates an empty Subsidence_Case_Database.xlsx with the 21-column v1 schema.
Column 6 is "Movement Cause" (replaces "Leak Source" used in Flood/EOW/Storm databases).

Run ONCE from the repo root:
    py scripts/create_subsidence_case_db.py

Do NOT re-run — it will overwrite the existing database.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Movement Cause",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

COLUMN_WIDTHS = {
    "Case ID": 12,
    "FOS Decision ID": 18,
    "Insurer Name": 32,
    "FOS Decision Date": 18,
    "Claim Type": 50,
    "Movement Cause": 40,
    "Property Type": 28,
    "Dispute Type": 28,
    "Coverage Decision": 28,
    "Rejection Reason": 50,
    "Evidence Dispute": 50,
    "Outcome Category": 20,
    "Outcome": 50,
    "Compensation Awarded (£)": 22,
    "Is Core Case": 22,
    "Key Policy Clause": 60,
    "Missing Evidence": 50,
    "Ombudsman Reasoning": 60,
    "Workflow Insight": 60,
    "AI Rule Candidate": 60,
    "Source PDF": 28,
}


def main() -> None:
    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    out_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Subsidence_Case_Database.xlsx"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subsidence Cases"

    header_font   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill   = PatternFill(start_color="6B4226", end_color="6B4226", fill_type="solid")
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side          = Side(style="thin", color="AAAAAA")
    header_border = Border(left=side, right=side, top=side, bottom=side)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = header_border
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_name, 20)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Created: {out_path}")
    print(f"Columns : {len(COLUMNS)}")
    print("Ready for append_subsidence_v1.py")


if __name__ == "__main__":
    main()
