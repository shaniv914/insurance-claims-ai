"""
Standard append script for Roof Damage Case Database — Schema v1 (21 columns).
Column 6 is "Damage Cause" (the physical mechanism causing the roof/water damage).

Usage
-----
1. Read the source PDF(s) and extract the fields listed in NEW_CASES below.
2. Add one dict per case to NEW_CASES following the extraction rules.
3. Run from the repo root:
       py scripts/append_roof_damage_v1.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Roof_Damage_Case_Database.xlsx

===========================================================================
FIELD EXTRACTION RULES
===========================================================================

Case ID             : Format ROOF-NNN (zero-padded to 3 digits)
FOS Decision ID     : DRN-XXXXXXX or DRNXXXXXXX as printed in the PDF
Insurer Name        : Formal registered name from the FOS decision
FOS Decision Date   : DD Mon YYYY — accept-or-reject deadline in final paragraph;
                      use "Not stated in document" if deadline not printed
Claim Type          : Policy type, physical incident and nature of dispute in one sentence
Damage Cause        : Physical mechanism of the roof/water damage, e.g.
                      "Roof ridge pointing deteriorating — gradually operating cause (wear and tear)"
                      "Front guttering overwhelmed by intense rainfall"
                      "Split lead flashing (laid too long) and slipped tiles — poor workmanship"
Property Type       : "Residential home" / "Commercial (...)" / "Residential home (landlord let)" etc.
Dispute Type        : Controlled vocab (7 values)
Coverage Decision   : Controlled vocab (5 values)
Rejection Reason    : Insurer's stated reason for declining or disputing
Evidence Dispute    : What evidence each party relied on
Outcome Category    : Controlled vocab (4 values)
Outcome             : Full FOS remedy instructions
Compensation Awarded (£) : Integer — D&I only; 0 if none
Is Core Case        : Controlled vocab (5 values)
                      Use "No — Commercial" for commercial/all-risks policies outside
                      standard residential home/buildings insurance scope
Key Policy Clause   : Policy wording or FOS/FCA principle applied
Missing Evidence    : Evidence that was absent and affected the outcome
Ombudsman Reasoning : How the ombudsman weighed the evidence
Workflow Insight     : Operational rule for the claims workflow
AI Rule Candidate   : Machine-evaluable rule for the rules engine
Source PDF          : Filename only (e.g. DRN0228808.pdf)
===========================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Damage Cause",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 6 (FINAL BATCH): ROOF-051 to ROOF-059 (9 cases)
# Note: ROOF-055 turned out NOT to be roof damage at all — it was initially
# investigated as a possible roof claim (policyholder suspected stone-throwing
# damage) but three separate roof assessors found nothing, and the insurer's
# own records plus a plumber's report established the true cause was a
# leaking radiator/heating pipe. Flagged but kept Is Core Case = Yes since it
# is a legitimate escape-of-water precedent, just not a roof peril one.
# ROOF-056 is a MOTOR insurance claim (car roof bodywork damage from a
# motorway accident), not home/building roof damage — flagged and marked
# Is Core Case = No — Commercial.
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "ROOF-051",
        "FOS Decision ID": "DRN7600763",
        "Insurer Name": "Millennium Insurance Company Limited",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": "Home insurance — claim for internal damage (ceiling in spare room, water damage to floor) following heavy rainfall; insurer initially declined citing no storm evidence, but the claim was for accidental damage, not solely storm",
        "Damage Cause": "Internal ceiling/floor damage from water ingress via the roof following heavy rainfall; the loss adjuster's own report confirmed the internal damage was recent and the insured was unaware of any roof damage beforehand — consistent with a single unforeseen event covered by accidental damage cover, not requiring proof of storm conditions",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Loss adjusters found no evidence of storm conditions on or around the time of the damage and declined on the basis no insured event (storm) had occurred; Millennium separately argued if the roof damage was not storm-related it must be a maintenance issue",
        "Evidence Dispute": "Mr and Mrs Z were clear in their original complaint letter that their claim was specifically for internal damage to the ceiling and floor, not primarily about the roof; the insurer's own loss adjuster report stated the internal event 'would seem to be recent' and that the insured was unaware of the roof damage, suggesting accidental damage cover should be considered. Millennium argued that if damage was due to a storm it must have occurred before the policy started, and if not storm-related it was a maintenance issue — but did not address the accidental damage cover point directly",
        "Outcome Category": "Upheld",
        "Outcome": "Millennium to deal with the claim for internal damage under accidental damage cover, add 8% simple interest to any cash payment from date of loss to settlement, and pay £200 compensation for distress and inconvenience",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Where a policyholder's claim, as clearly and specifically framed in their complaint, concerns internal damage (not the roof itself), and the policy includes accidental damage cover without a specific definition of 'accidental damage', the ombudsman will apply the ordinary meaning (unforeseen and unintentional) rather than requiring the claim to satisfy the storm-specific three-part test; an insurer's own loss adjuster report confirming the damage was recent and that the policyholder was unaware of any pre-existing roof problem is itself evidence supporting a finding of a single unforeseen event covered under accidental damage; an insurer cannot defeat an accidental damage claim merely by arguing that if the damage isn't storm-related it must be 'a maintenance issue' without evidence establishing gradual/maintenance-related causation specific to the claimed internal damage",
        "Missing Evidence": "Any evidence from Millennium establishing the internal damage was itself gradual/maintenance-related, as opposed to a single recent event — the loss adjuster's own report supported the opposite conclusion",
        "Ombudsman Reasoning": "The policyholders' original complaint letter framed their claim specifically as being for internal ceiling and floor damage, not the roof — this framing determined the analytical focus; the loss adjuster's own report confirmed the internal damage was recent and that the policyholder was unaware of it beforehand, meaning it likely occurred during the policy period and was unforeseen; because the policy's accidental damage section had no specific definition and no exclusion applicable to these facts, and the loss adjuster's own words described a 'single unforeseen one-off event', the internal damage was covered; Millennium's argument that a non-storm cause must be a 'maintenance issue' was unsupported since it hadn't shown the internal damage specifically resulted from gradual deterioration",
        "Workflow Insight": "When a policyholder's claim is specifically and narrowly framed (e.g., in their letter of complaint) as being about internal water damage rather than the roof itself, assess that internal damage claim on its own terms under accidental damage cover, rather than treating it as automatically dependent on establishing storm causation for the roof; where a policy's accidental damage section lacks a specific definition, the ordinary meaning (a single unforeseen and unintentional event) governs, and the insurer's own loss adjuster's language can itself support the claim; an insurer asserting a 'maintenance issue' alternative cause bears the burden of evidencing that specific conclusion, it cannot rely on process of elimination alone",
        "AI Rule Candidate": "IF policyholder_complaint_specifically_and_narrowly_frames_the_claim_as_being_for_internal_water_damage_rather_than_the_roof_itself THEN assess_the_internal_damage_claim_under_accidental_damage_cover_independently_of_whether_storm_causation_for_the_roof_is_established; IF policy_accidental_damage_section_lacks_a_specific_definition AND insurer_own_loss_adjuster_report_confirms_the_damage_was_recent_and_the_policyholder_was_unaware_of_it_beforehand THEN this_supports_a_finding_of_a_single_unforeseen_unintentional_event_covered_under_accidental_damage; insurer_asserting_damage_must_be_a_maintenance_issue_because_it_was_not_storm_related_bears_the_burden_of_evidencing_that_specific_conclusion_and_cannot_rely_on_process_of_elimination_alone",
        "Source PDF": "DRN7600763.pdf",
    },
    {
        "Case ID": "ROOF-052",
        "FOS Decision ID": "DRN8186649",
        "Insurer Name": "Elite Insurance Company Limited",
        "FOS Decision Date": "28 Oct 2016",
        "Claim Type": "Home insurance — storm claim for external roof damage and internal bathroom damage; insurer accepted a storm occurred but declined both roof and internal damage, saying its policy only covers wind damage, not rain",
        "Damage Cause": "Roof — missing mortar, found not to be typical storm damage, with the roof generally not in a great state of repair; the fact no leak recurred since repair confirmed the repair was effective but not what originally caused the need for it. Bathroom — internal water damage caused by rain entering through the storm-damaged roof (a wind-caused defect allowed rain ingress), not a gradual/pre-existing leak",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Partial",
        "Rejection Reason": "Elite said the roof wasn't damaged by the storm (attributed to missing mortar/poor repair, not storm) and that the bathroom damage, even if caused by rain, wasn't covered because its policy only covers damage from storm winds, not rain",
        "Evidence Dispute": "Mrs M argued there had been no roof leak since the repair work was done, so both the original roof and internal damage must have been storm-caused. Elite's report and commentary on the roof damage found missing mortar (not typically storm-caused) and a roof not in great repair; on the bathroom, Elite argued its cover was limited to wind damage, excluding rain. The ombudsman found the successful post-repair outcome only proved the fix worked, not what caused the original need for it, and found no evidence rain had been entering the property during non-storm conditions, meaning the wind (storm) had created the pathway that allowed the otherwise-excluded rain to cause the internal damage",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Elite to settle the claim for internal (bathroom) damage in line with the remaining policy terms and conditions; roof damage decline upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Where a policy's storm cover is worded to cover damage from storm winds specifically (excluding rain), this does not mean that internal damage caused by rain entering the property because storm winds created or exploited an opening is excluded — 'but for' the storm winds, the rain would not have caused the damage, so the wind remains the operative/proximate cause even though rain was the immediate mechanism of the internal damage; missing mortar is not typically indicative of storm-caused roof damage, and a successful post-repair outcome (no further leaks) is evidence the repair was effective, not evidence of what originally caused the damage requiring repair; internal damage that is not shown to be gradual (no evidence rain had been entering during non-storm conditions) supports a finding the damage was a direct, one-off consequence of the storm event",
        "Missing Evidence": "Any evidence that rain had been entering Mrs M's property during non-storm conditions (which would have supported a gradual/pre-existing leak theory) — none was found, supporting the internal damage being a direct storm consequence",
        "Ombudsman Reasoning": "The roof damage (missing mortar, general poor condition) was not typical storm damage, and the successful post-repair outcome only demonstrated the repair was done well, not that a storm had caused the original problem — so the roof decline was fair; but for the bathroom, the policy's coverage for storm wind damage extends to rain damage that occurs specifically because storm winds allowed rain to enter that would not otherwise have entered — the wind is the 'but for' cause even though rain caused the visible damage; because there was no evidence of a pre-existing, gradual rain-ingress problem and the bathroom damage did not appear gradual, Elite could not fairly deny liability for the internal damage",
        "Workflow Insight": "When a storm policy's wording distinguishes between wind and rain damage, do not treat this as automatically excluding rain-caused internal damage where storm winds created the pathway (e.g., a roof opening) for that rain — apply a 'but for' causal chain analysis: would the rain have caused the damage absent the storm winds' role in creating the opening; a successful, leak-free outcome following a repair is evidence the repair fixed the problem, not evidence of what originally caused the problem, and should not be used by either party to prove or disprove storm causation of the underlying damage; check for evidence of rain ingress during non-storm periods before concluding internal water damage was gradual rather than a one-off storm consequence",
        "AI Rule Candidate": "IF policy_storm_cover_wording_is_limited_to_damage_from_storm_winds_and_excludes_rain_but_storm_winds_created_or_exploited_a_roof_opening_that_allowed_rain_to_enter_and_cause_internal_damage THEN the_wind_remains_the_operative_but_for_cause_and_the_internal_rain_damage_is_covered_notwithstanding_the_rain_only_exclusion; a_leak_free_outcome_following_a_roof_repair_is_evidence_the_repair_was_effective_not_evidence_of_the_original_causation_of_the_damage_requiring_repair; absence_of_evidence_of_rain_ingress_during_non_storm_conditions_supports_treating_internal_water_damage_as_a_direct_one_off_storm_consequence_rather_than_gradual",
        "Source PDF": "DRN8186649.pdf",
    },
    {
        "Case ID": "ROOF-053",
        "FOS Decision ID": "DRN8381346",
        "Insurer Name": "Royal Sun Alliance Insurance Plc",
        "FOS Decision Date": "4 Mar 2020",
        "Claim Type": "Home insurance — storm claim for damage to the roof of a bay window",
        "Damage Cause": "Bay window roof tiles lay under a slurry render that was breaking down and covered in moss — surveyor found this to be gradual deterioration (wear and tear) of the render, not storm damage; the main roof of the house (undamaged by the storm) appeared in good condition, supporting the inference that the bay window roof's deteriorated condition (not the storm) explained why it alone was damaged",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "RSA's surveyor found the roof tiles lay under a deteriorating, moss-covered slurry render, concluding the damage resulted from gradual material breakdown (wear and tear) rather than storm, applying the standard three-question storm framework",
        "Evidence Dispute": "Mr M provided an alternative builder's opinion attributing the damage to the storm; RSA's surveyor's detailed report and photographs showed accumulated moss and deteriorating slurry render, and noted the main roof of the house (unaffected by the storm) was in good condition, suggesting that if the bay window roof had been in similarly good condition it too would likely have survived the storm undamaged. FOS found the builder's report insufficiently detailed/persuasive to outweigh the surveyor's photographic and technical evidence",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — RSA's decline of the claim upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Storm conditions (here, 61mph winds) and roof-tile-consistent damage pattern (missing tiling) can satisfy the first two parts of the standard three-part storm test, but the claim can still fail on the third (main cause) part where a surveyor's detailed report identifies specific, visible evidence of gradual material deterioration (e.g., breaking-down slurry render, accumulated moss) underlying the damaged tiles; comparing the condition of an undamaged part of the same roof against the damaged section is a valid comparative technique — if the undamaged section was in good condition and withstood the storm, this supports an inference that the damaged section's pre-existing poor condition (not the storm itself) explains why only that section failed; a bare alternative builder's opinion, without matching technical/photographic detail, will generally not outweigh a detailed surveyor's report addressing the same evidence",
        "Missing Evidence": "Technical detail or photographic evidence in the builder's report specifically addressing or countering the surveyor's findings on the deteriorating slurry render and moss accumulation — the builder's opinion alone lacked comparable evidential weight",
        "Ombudsman Reasoning": "Storm conditions (61mph) were undisputed and roof tile damage (missing tiling) was consistent with typical storm impact, satisfying the first two questions; but the surveyor's detailed report and photographs showed the bay window roof tiles sat under a deteriorating, moss-covered slurry render, indicating gradual material breakdown; critically, the main roof of the house — unaffected by the same storm — was shown to be in good condition, supporting the inference that had the bay window roof also been well-maintained, it too would likely have withstood the storm; this comparative evidence, combined with the detailed surveyor findings, outweighed the more general builder's opinion that the damage was storm-caused",
        "Workflow Insight": "When assessing whether storm conditions were the main cause of damage to one specific roof section (while a broader roof largely survived undamaged), compare the condition and outcome of the surviving section(s) against the damaged section — if the surviving parts were undamaged and in reasonable condition, this comparative evidence supports attributing the damaged section's failure to its own pre-existing poor condition rather than the storm generally; when a policyholder's alternative expert report simply asserts storm causation without technical detail matching the level of the insurer's surveyor report, it is unlikely to be found persuasive enough to displace the insurer's more detailed evidence",
        "AI Rule Candidate": "IF surveyor_report_identifies_specific_gradual_deterioration_evidence_(e.g.,_breaking_down_slurry_render_moss_accumulation)_under_the_damaged_roof_tiles AND an_undamaged_section_of_the_same_roof_survived_the_same_storm_in_good_condition THEN this_comparative_evidence_supports_a_wear_and_tear_finding_for_the_damaged_section_over_storm_causation; policyholder_alternative_expert_opinion_asserting_storm_causation_without_technical_detail_or_photographic_evidence_matching_the_insurers_surveyor_report_is_unlikely_to_outweigh_the_insurers_more_detailed_findings",
        "Source PDF": "DRN8381346.pdf",
    },
    {
        "Case ID": "ROOF-054",
        "FOS Decision ID": "DRN8624847",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": "Home insurance — storm claim for internal water damage (bedroom) caused by water entering through the chimney (lead flashing or brickwork); roof/chimney element and internal damage element both disputed",
        "Damage Cause": "AXA's expert found water entering through the chimney's lead flashing or brickwork was caused by 'gradual process loss (wear and tear)' both to the roof/chimney area and internally; the adjudicator found the extreme rainfall levels constituted storm conditions and that the internal bedroom damage specifically was caused suddenly and directly by that storm-level rainfall, distinct from the chimney/roof element's gradual causation",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Partial",
        "Rejection Reason": "AXA's expert report found the water ingress through the chimney flashing/brickwork was due to gradual wear and tear, not storm; AXA maintained storm conditions had not prevailed and, without accidental damage cover, the internal damage was also excluded",
        "Evidence Dispute": "Weather records showed extreme rainfall levels supporting storm conditions; the roof/chimney element of the damage was found consistent with gradual wear and tear; the internal (bedroom) damage was found to have occurred at the time of the heavy rainfall and was accepted by both parties to have happened contemporaneously with it — supporting a finding that the storm-level rainfall was the dominant/effective cause of the internal damage specifically, distinguishable from the roof/chimney's own gradual deterioration",
        "Outcome Category": "Upheld in Part",
        "Outcome": "AXA to settle the internal (bedroom) damage claim subject to the remaining policy terms and conditions; roof/chimney element decline upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Extreme rainfall levels (assessed against weather records) can independently satisfy the storm-conditions threshold even without reference to wind speed; where a roof/chimney entry point (e.g., lead flashing or brickwork) is found to have gradually deteriorated (wear and tear) allowing water ingress, this gradual-causation finding for the entry point does not automatically extend to the internal damage that results from a specific rainfall event coinciding with storm conditions — the internal damage element should be separately assessed against the three-part storm test, and can be found storm-caused (sudden, direct result of the storm) even where the underlying building defect that let the water in was itself a gradual deterioration",
        "Missing Evidence": "None decisive — the parties' apparent agreement that the internal damage occurred at the time of the heavy rainfall was itself sufficient to establish the storm-conditions timing link for that specific element",
        "Ombudsman Reasoning": "Weather records showing extreme rainfall levels were sufficient to establish storm conditions prevailed; the roof/chimney damage (water entry via lead flashing/brickwork) was reasonably found to result from gradual wear and tear, so AXA's decline of that element was not unreasonable; but the internal bedroom damage was accepted as having occurred at the time of the heavy rainfall, was sudden and unexpected, and was a direct result of the storm-level rainfall — supporting a finding that storm conditions were the dominant/effective cause of that specific damage, independent of the chimney's own gradual deterioration being the entry mechanism",
        "Workflow Insight": "When a building defect (e.g., deteriorated chimney flashing) that gradually developed over time serves as the entry point for water, and a specific storm/heavy-rainfall event then causes internal damage on a particular date, assess the internal damage separately under the three-part storm test — a gradual-deterioration finding for the roof/chimney defect itself does not necessarily defeat a storm claim for the internal damage that resulted from a specific, dominant rainfall event; extreme rainfall levels recorded in weather data can independently establish storm conditions without needing to separately assess wind speed",
        "AI Rule Candidate": "IF building_entry_point_(e.g.,_chimney_flashing_or_brickwork)_is_found_to_have_gradually_deteriorated_over_time_allowing_water_ingress AND internal_damage_is_shown_to_have_occurred_specifically_at_the_time_of_a_storm_level_rainfall_event THEN the_internal_damage_may_still_be_found_storm_caused_under_the_three_part_test_independent_of_the_gradual_deterioration_of_the_entry_point_itself; extreme_recorded_rainfall_levels_can_independently_satisfy_the_storm_conditions_threshold_without_reference_to_wind_speed",
        "Source PDF": "DRN8624847.pdf",
    },
    {
        "Case ID": "ROOF-055",
        "FOS Decision ID": "DRN8751154",
        "Insurer Name": "Royal and Sun Alliance Insurance Plc",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": "FLAGGED — NOT ROOF DAMAGE: home insurance — claim for water damage to a living room ceiling; the policyholder initially suspected roof damage (possibly from children throwing stones), and RSA sent three separate assessors to inspect the roof, but none found any roof damage; RSA's own internal records and a plumber's report instead pointed to the true cause being a leaking radiator/heating pipe, which the ombudsman ultimately found determinative — this is a plumbing/heating escape-of-water case initially (and unsuccessfully) investigated as a possible roof claim, not an actual roof damage case",
        "Damage Cause": "FLAGGED — no roof damage was ever found by any of RSA's three roof assessors; RSA's own claim notes progressively identified the likely cause as heating pipes, then a leak in pipework from the bedroom floor, then specifically a plumbing problem; a plumber's report (obtained by the policyholder) confirmed the actual cause was a leaking radiator, which the plumber capped off and disconnected — the water damage to the ceiling resulted from this radiator/pipe leak, not from any roof defect or storm",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "RSA requested a surveyor's report to determine the cause of damage despite its own internal records already suggesting a plumbing/heating pipe leak was the likely cause, rather than accepting the claim under the escape-of-water section covering water/oil escaping from fixed water or heating systems",
        "Evidence Dispute": "Mrs A initially believed the damage was roof-related (possibly from children throwing stones), which led RSA's roof assessors to find nothing, since they were looking in the wrong place; RSA's own internal claim notes across April-August progressively and consistently pointed to a plumbing or heating pipe leak as the likely cause; a plumber's report and invoice obtained by Mrs A confirmed a leaking radiator was found and capped off/disconnected. RSA disputed the credibility and dating of the plumber's report, but the ombudsman found the plumber's account corroborated by telephone discussions between the investigator and the plumber, and consistent with RSA's own internal suspicions",
        "Outcome Category": "Upheld",
        "Outcome": "RSA to reconsider the claim in line with policy terms and conditions (under the escape-of-water/fixed heating system cover), add 8% simple interest to any cash settlement from date of loss to settlement, and pay £250 compensation for distress and inconvenience",
        "Compensation Awarded (£)": 250,
        "Is Core Case": "Yes",
        "Key Policy Clause": "FLAGGED — NOT A ROOF DAMAGE PRECEDENT: where a policyholder incorrectly suspects and reports a cause (e.g., roof damage from thrown objects) that leads an insurer's assessors to investigate and find nothing, but the insurer's OWN internal claim records independently and consistently point to a different specific cause (e.g., a plumbing/heating pipe leak) covered under a different section of the policy (e.g., escape of water from fixed heating systems), the insurer should pursue and properly investigate that alternative cause identified in its own records rather than requesting further generic surveys that don't address the specific plumbing/heating theory its own staff had already identified; a policyholder-obtained plumber's report and invoice, corroborated by follow-up telephone discussion with the tradesperson, can be persuasive evidence of causation even where the insurer disputes the report's precise dating or completeness",
        "Missing Evidence": "None decisive against the policyholder — RSA's own internal records consistently supported the plumbing/heating leak theory, and its request for a general surveyor's report (rather than pursuing the plumbing angle it had already identified) was found unreasonable",
        "Ombudsman Reasoning": "Despite the claim initially being investigated as roof damage (based on the policyholder's own mistaken belief), RSA's own internal notes from April through August consistently and progressively identified a plumbing or heating pipe leak as the likely cause, culminating in a request for 'a plumber's report confirming the cause of damage' — this was itself an acknowledgment that RSA suspected a plumbing cause; the plumber's report and invoice, further corroborated by direct telephone conversations between the investigator and the plumber, was persuasive and was not adequately countered by anything in RSA's file; RSA had therefore acted unreasonably in not accepting the plumber's evidence and continuing to request a general roof surveyor's report instead of pursuing its own identified plumbing theory",
        "Workflow Insight": "NOT DIRECTLY A ROOF DAMAGE PLAYBOOK PRECEDENT — this is a claim-investigation-direction principle: when a policyholder's own initial theory of causation (e.g., roof damage) proves unsupported by repeated expert inspection, but the insurer's own internal claim notes independently and consistently identify a different specific, covered cause (e.g., a heating/plumbing leak), the insurer should pursue and properly evaluate evidence relevant to that alternative cause rather than continuing to request generic reports unconnected to its own identified theory; retained for reference on escape-of-water claim-handling standards, not for building roof peril causation rules",
        "AI Rule Candidate": "NOT A ROOF DAMAGE PERIL RULE — retained for reference on escape-of-water claim handling: IF policyholder_initial_theory_of_damage_cause_(e.g.,_roof_damage)_is_repeatedly_unsupported_by_expert_inspection_finding_no_roof_damage AND insurers_own_internal_claim_records_independently_and_consistently_identify_a_different_specific_covered_cause_(e.g.,_a_plumbing_or_heating_pipe_leak) THEN insurer_should_pursue_and_properly_evaluate_evidence_relevant_to_that_alternative_identified_cause_rather_than_requesting_further_generic_reports_unconnected_to_its_own_identified_theory; policyholder_obtained_tradesperson_report_corroborated_by_follow_up_verbal_confirmation_can_be_persuasive_causation_evidence_notwithstanding_insurer_disputes_about_the_reports_precise_dating_or_completeness",
        "Source PDF": "DRN8751154.pdf",
    },
    {
        "Case ID": "ROOF-056",
        "FOS Decision ID": "DRN8937043",
        "Insurer Name": "esure Insurance Limited",
        "FOS Decision Date": "27 Nov 2017",
        "Claim Type": "FLAGGED — NOT HOME/BUILDING ROOF DAMAGE: motor (car) insurance — after a motorway collision where another vehicle lost control and both cars drove onto the grass verge, the policyholder's car sustained roof damage (dents, creases, ripples); esure's engineers (two internal, one independent) all concluded the roof damage was not caused by the accident, so esure repaired the roof under a separate 'fault' claim rather than including it in the accident claim",
        "Damage Cause": "FLAGGED — vehicle bodywork damage, not a building/home roof peril: multiple dents, creases and ripples of varying severity across the entire car roof area, found by an independent engineer to show no evidence consistent with impact from tree branches/debris as the policyholder alleged, and inconsistent overall with the low-speed grass-verge accident circumstances",
        "Property Type": "Motor vehicle (personal car insurance) — NOT a home/property risk",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Two esure engineers and a further independent engineer all concluded the roof damage (dents, creases, ripples across the entire roof) was inconsistent with having been caused by the motorway accident, so esure processed the roof repairs as a second, separate 'fault' claim rather than as part of the (non-fault) accident claim",
        "Evidence Dispute": "Mr W said over twenty witnesses could confirm his roof wasn't damaged before the accident, that his passenger heard debris hitting the roof during the collision, that the other driver's vehicle broke tree branches which then struck his roof, and that a car wash attendant confirmed the roof was undamaged that morning; the independent engineer found no evidence of damage consistent with tree branch contact, and instead found multiple dents/creases/ripples across the entire roof of varying severity, inconsistent with a single impact event; the independent engineer also incorrectly reported that Mr W had chosen not to attend the inspection, which esure accepted was an error (he was never given the choice)",
        "Outcome Category": "Compensation Only",
        "Outcome": "esure to pay Mr W £75 compensation for the lost opportunity to attend the independent engineer's inspection; the decision to process the roof damage as a separate claim (rather than part of the accident claim) upheld as fair and reasonable",
        "Compensation Awarded (£)": 75,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "FLAGGED FOR PLAYBOOK EXCLUSION — this is a motor accident-causation and claim-recording dispute about a car's bodywork roof, not a home roof peril case: where multiple independent engineers (including one specifically instructed to resolve a dispute about causation) all conclude that specific damage was not caused by the accident being claimed for, an insurer is entitled to process that damage as a separate claim rather than including it in the primary (often more favourable, e.g., non-fault) accident claim; a procedural error by an inspecting engineer (e.g., wrongly reporting the policyholder chose not to attend an inspection) that does not affect the substantive outcome still warrants modest compensation for the lost opportunity/expectation, separate from the underlying causation finding",
        "Missing Evidence": "Not applicable to the causation finding — the independent engineer's inspection was the decisive technical evidence and was not successfully countered by the policyholder's lay/circumstantial evidence (witness statements, car wash confirmation, passenger's account of noises)",
        "Ombudsman Reasoning": "Two esure engineers and a further independent engineer (instructed specifically because Mr W disputed the first two engineers' findings) all concluded the roof damage was not accident-related; the independent engineer's finding of multiple dents, creases and ripples of varying severity across the entire roof area was inconsistent with a single tree-branch-impact event as alleged, and esure was entitled to reasonably rely on this consistent, independently corroborated technical evidence over Mr W's lay and circumstantial evidence; however, the independent engineer incorrectly reported Mr W had chosen not to attend the inspection when this wasn't true, and esure accepted this was an error — while this didn't change the substantive outcome, it warranted modest compensation for the lost opportunity to be present",
        "Workflow Insight": "NOT APPLICABLE TO THE HOME ROOF DAMAGE (BUILDING) PLAYBOOK — this is a motor accident-causation and claim-classification principle; retained only as a general reference that consistent findings from multiple independent engineers can be reasonably relied upon by an insurer to exclude specific damage from a primary claim, and that procedural errors not affecting the substantive outcome still warrant modest standalone compensation — not applicable to building roof peril rules",
        "AI Rule Candidate": "NOT APPLICABLE TO HOME ROOF DAMAGE PERIL — flagged as an out-of-scope motor claim; IF multiple_independent_engineers_including_one_specifically_instructed_to_resolve_a_disputed_causation_question_all_consistently_conclude_specific_vehicle_damage_was_not_caused_by_the_claimed_accident THEN insurer_may_reasonably_process_that_damage_as_a_separate_claim_rather_than_including_it_in_the_primary_accident_claim; procedural_error_by_an_inspecting_engineer_that_does_not_alter_the_substantive_technical_outcome_still_warrants_modest_standalone_DI_compensation_for_the_lost_opportunity",
        "Source PDF": "DRN8937043.pdf",
    },
    {
        "Case ID": "ROOF-057",
        "FOS Decision ID": "DRN9411289",
        "Insurer Name": "UK Insurance Limited",
        "FOS Decision Date": "27 Dec 2018",
        "Claim Type": "Home insurance — roof leak under heavy snow; claim for external roof damage (declined) and internal damage (accepted under accidental damage), plus service complaints about delay, lack of emergency roof repairs, and alternative accommodation handling",
        "Damage Cause": "Loss adjuster found the roof damage was mainly due to the condition and age of the (flat) roof — wear and tear — with the report noting that while the weather (heavy snowfall) highlighted the problem, a roof in good condition would have withstood the snowfall",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Partial",
        "Rejection Reason": "Loss adjuster concluded the roof damage wasn't due to an insured peril (storm or flood) but mainly the condition and age of the roof; UKI also said it didn't provide emergency roof repairs because the policyholder didn't hold Home Emergency cover",
        "Evidence Dispute": "Mrs M felt UKI should have provided emergency repairs given the danger to her property, and disputed the roof decline; UKI's loss adjuster (the only expert evidence available, as Mrs M was invited to but didn't provide a report from a local builder) found the flat roof had probably reached the end of its life and was suffering from wear and tear; the weather conditions were found not to approach the criteria for a snow storm; on service, UKI accepted causing delays (25 days to send a loss adjuster) and failing to properly communicate about extensions to Mrs M's alternative accommodation, causing her to believe she had nowhere to stay",
        "Outcome Category": "Upheld in Part",
        "Outcome": "UKI's decline of the external roof damage claim upheld (wear and tear, no storm); UKI to pay a total of £500 compensation (including £200 already paid) for its poor service and the distress caused; internal repairs already accepted and completed under accidental damage cover",
        "Compensation Awarded (£)": 500,
        "Is Core Case": "Yes",
        "Key Policy Clause": "For a buildings insurance claim to succeed, the loss must fall within a specifically insured peril (e.g., storm or escape of water) — weather conditions that don't approach the policy's storm criteria (here, snow that did not amount to a 'snow storm') cannot support a storm-based roof claim regardless of the practical severity of the resulting leak; where the only available expert evidence (an insurer's loss adjuster) attributes roof damage to age/wear and tear, and the policyholder is invited to but does not provide a competing expert report, the insurer's expert evidence stands and supports a fair decline; lack of Home Emergency cover means an insurer is not obliged to provide emergency temporary roof repairs even where a declined-claim roof is causing an active leak, though this doesn't excuse poor handling of accepted claim elements or communication about alternative accommodation",
        "Missing Evidence": "A competing report from a local builder or other independent expert on the cause of the roof damage — UKI specifically invited Mrs M to provide one and she did not",
        "Ombudsman Reasoning": "The weather conditions (heavy snowfall) did not meet the criteria for a snow storm under the policy, and the only expert evidence (UKI's loss adjuster) attributed the roof damage to its age and wear and tear, unrebutted by any competing expert report despite an invitation to provide one; this made the roof decline fair and reasonable; separately, UKI's own extensive and repeated service failures caused significant additional distress beyond the disappointment of the roof decline itself, warranting substantial compensation (£500 total) even though the underlying coverage decision on the roof was correct",
        "Workflow Insight": "Snow-related roof damage claims must be assessed against the policy's specific storm definition — heavy snowfall causing visible damage does not automatically qualify as 'storm' if it doesn't meet the wind speed/precipitation thresholds defined; when a policyholder is specifically invited to provide a competing expert report to challenge an insurer's loss adjuster findings and does not do so, the insurer's unrebutted expert evidence should generally be accepted; service failures around communication (e.g., failing to inform a policyholder that alternative accommodation has been extended) can independently warrant substantial compensation, separate from and in addition to any compensation for delay, even where the underlying coverage decision is upheld as correct",
        "AI Rule Candidate": "IF weather_event_is_snow_related_and_does_not_meet_the_policys_specific_storm_definition_thresholds_(e.g.,_wind_speed_or_precipitation_rate) THEN claim_cannot_succeed_under_the_storm_peril_regardless_of_the_severity_of_resulting_leak_damage; IF insurer_specifically_invites_policyholder_to_provide_a_competing_expert_report_on_roof_damage_causation_and_policyholder_does_not_provide_one THEN insurers_unrebutted_loss_adjuster_findings_should_generally_be_accepted; insurer_failure_to_communicate_extensions_to_alternative_accommodation_causing_policyholder_to_believe_they_have_nowhere_to_stay_independently_warrants_substantial_DI_compensation_separate_from_delay_related_compensation",
        "Source PDF": "DRN9411289.pdf",
    },
    {
        "Case ID": "ROOF-058",
        "FOS Decision ID": "DRN9726991",
        "Insurer Name": "Liverpool Victoria Insurance Company Limited",
        "FOS Decision Date": "13 Dec 2019",
        "Claim Type": "Home insurance — storm claim for roof damage: a single dislodged tile (accepted by LV as storm-caused, but not pursued by the policyholders due to the excess) and separately disputed damage to mortar between roof tiles",
        "Damage Cause": "Single tile — accepted by LV as dislodged by the storm (not pursued due to excess). Mortar — damage (dislodged, broken, crumbling in various places) found to be located centrally on the roof (not in exposed positions like ridge/edge tiles or chimney pots typically vulnerable to wind), with the surrounding tiles themselves not out of position, supporting a finding of gradual deterioration rather than storm causation; the mortar's function (resisting wind, keeping tiles secure) also made storm-caused mortar failure without corresponding tile displacement unusual",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "LV accepted the single tile was dislodged by the storm, but concluded the separate mortar damage was due to wear and tear, not storm, because it was located in a less wind-exposed central roof position, unaccompanied by tile displacement, and mortar's structural purpose is specifically to resist wind",
        "Evidence Dispute": "Mr F argued the storm's high winds had caused the tiles to rattle and lift, which then caused the mortar to break, and that other equally-old areas of the roof still had secure tiles; the ombudsman accepted this was theoretically possible but found no supporting evidence for it, and specifically noted the tiles in the mortar-damaged area did not themselves appear out of position, undermining the 'tiles rattled and lifted' theory",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — LV's decline of the mortar damage claim upheld; the tile damage was not part of this decision as the policyholders chose not to pursue it given the excess",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Mortar between roof tiles serves specifically to resist wind and keep tiles secured — its failure without accompanying tile displacement is atypical of storm damage; storm damage typically manifests at more exposed roof positions (ridge tiles, edge tiles, chimney pots) rather than centrally located mortar; where a policyholder offers a plausible theoretical mechanism for storm causation but provides no supporting evidence, and the physical evidence (tiles remaining in position) is inconsistent with that theory, the theoretical possibility alone is insufficient to establish storm as the main cause; 'accidental damage' cover requiring sudden causation cannot apply to damage found to be the result of gradual deterioration",
        "Missing Evidence": "Evidence supporting Mr F's theory that wind-induced tile rattling/lifting (without visible displacement) caused the mortar to dislodge — none was provided, and the tiles' undisturbed position was itself evidence against this theory",
        "Ombudsman Reasoning": "Storm conditions (60mph+ winds) were undisputed, and the single dislodged tile was accepted by LV as storm-caused but not pursued by the policyholders; for the mortar, its position (central, not at exposed edges/ridges/chimneys) and its specific wind-resisting function meant storm-caused failure would be atypical, especially absent any corresponding tile displacement in that area; Mr F's rattling/lifting theory was plausible in principle but unsupported by evidence and contradicted by the tiles' undisturbed position; accidental damage cover (requiring sudden causation) also could not apply given the gradual deterioration finding",
        "Workflow Insight": "When assessing storm damage to a specific building element (e.g., roof mortar), consider that element's structural position (exposed vs. sheltered/central) and specific function (e.g., wind resistance) against the typical pattern of storm damage — damage to elements specifically designed to resist the claimed causal force, located in less exposed positions, and unaccompanied by displacement of adjacent/dependent elements, is atypical of storm causation and supports a gradual deterioration finding; a policyholder's theoretical causal mechanism should be checked against the physical evidence before being given weight",
        "AI Rule Candidate": "IF damaged_roof_element_(e.g.,_mortar)_is_located_in_a_less_wind_exposed_central_position_rather_than_typically_vulnerable_exposed_positions_(ridge_edge_tiles_chimney_pots) AND the_elements_specific_function_is_to_resist_the_claimed_causal_force_(e.g.,_wind) AND no_corresponding_displacement_of_adjacent_dependent_elements_(e.g.,_tiles)_is_present THEN this_pattern_is_atypical_of_storm_damage_and_supports_a_gradual_deterioration_finding; policyholder_theoretical_causal_mechanism_for_storm_damage_unsupported_by_evidence_and_contradicted_by_physical_evidence_is_insufficient_to_establish_storm_as_main_cause",
        "Source PDF": "DRN9726991.pdf",
    },
    {
        "Case ID": "ROOF-059",
        "FOS Decision ID": "DRN9797551",
        "Insurer Name": "Ageas Insurance Limited",
        "FOS Decision Date": "22 Nov 2019",
        "Claim Type": "Home insurance — flat roof storm claim; water entered and damaged a kitchen; roof damage and internal (kitchen) damage both disputed",
        "Damage Cause": "Ageas found the flat roof was worn in places, had cracked edges, and had likely reached the end of its natural life — the storm merely highlighted the pre-existing state of the roof rather than causing the damage; the internal kitchen damage (wet ceiling, walls, and floor) showed no sign of being long-standing or of rain having been getting in over a prolonged period, supporting a finding that it was a direct, one-off result of the storm",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Partial",
        "Rejection Reason": "Ageas's inspection found the roof worn in places and likely at the end of its natural life, concluding the storm had merely highlighted its condition rather than caused the damage",
        "Evidence Dispute": "Mrs S argued all roofs have some deterioration and it wasn't fair to decline on that basis, and said there'd been no sign of any problem before the storm, adding that her own roofer said the damage was wind-caused (though no report or supporting detail from the roofer was ever provided); Ageas's report and photos showed cracked edges and wear consistent with age-related failure, with no opposing expert opinion to counter it; Ageas separately argued it shouldn't have to cover the internal damage because if the roof had been in good condition the internal damage wouldn't have occurred",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Ageas to accept and consider Mrs S's claim for internal (kitchen) damage caused by the storm in line with the remaining policy terms and conditions; roof damage decline upheld; no compensation awarded as Mrs S's primary concern (the roof) was correctly declined",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Even where a roof is found to be suffering from wear and tear (correctly excluding the roof damage claim itself), if resulting internal damage occurs as a direct, one-off consequence of the same storm event — with no sign that the internal damage itself has been ongoing or gradual — the insurer will still be required to cover that internal damage; an insurer's argument that 'if the roof had been in good condition the internal damage wouldn't have occurred' does not defeat this principle, since the wear-and-tear roof condition is the mechanism that let the water in, while the storm remains the operative trigger for the specific, sudden internal damage event; a bare assertion that 'my roofer said the damage was caused by winds', without any actual report, technical detail, or explanation from the roofer, carries no persuasive evidential weight and does not displace an insurer's detailed surveyor findings",
        "Missing Evidence": "Any actual report or technical explanation from Mrs S's roofer supporting her assertion that wind caused the damage — she only relayed a verbal claim without documentation",
        "Ombudsman Reasoning": "Ageas's inspection (cracked edges, general wear, likely end of natural life) was unrebutted by any actual expert report from Mrs S's side (only an unsupported verbal claim from her roofer), so the roof decline was fair and reasonable; separately, following this service's standard approach, even where a roof itself is found to be worn/deteriorated, resulting internal damage caused directly by a specific storm event should still be covered — Ageas's argument that a good-condition roof wouldn't have let water in doesn't override this, since the wear-and-tear finding explains the entry mechanism, not whether the specific damage event itself was sudden and storm-caused; because Mrs S's real concern was always the roof (correctly declined), no compensation was awarded despite Ageas being wrong about the internal damage",
        "Workflow Insight": "Apply the 'even a worn/wear-and-tear roof's resulting internal damage can be storm-covered' principle consistently — a roof being correctly excluded as wear and tear does not automatically extend to exclude internal damage from a specific storm event, provided the internal damage itself shows no signs of being long-standing/gradual; insurers should specifically check internal damage for gradual indicators separately from the roof's own wear-and-tear condition before declining the internal element; an unsupported second-hand claim about a roofer's opinion, without any actual report or documentation, should be given no persuasive weight against a detailed surveyor report",
        "AI Rule Candidate": "IF roof_is_correctly_found_to_be_worn_or_at_the_end_of_its_natural_life_(wear_and_tear_exclusion_properly_applied) AND resulting_internal_damage_shows_no_sign_of_being_long_standing_or_gradual_and_occurred_at_the_time_of_a_specific_storm_event THEN the_internal_damage_should_still_be_covered_under_the_storm_peril_notwithstanding_the_roofs_own_wear_and_tear_condition_being_the_entry_mechanism; insurer_argument_that_a_good_condition_roof_would_not_have_allowed_water_entry_does_not_defeat_coverage_for_internal_damage_that_is_itself_sudden_and_storm_caused; unsupported_second_hand_claim_about_a_tradespersons_opinion_without_an_actual_report_or_documentation_carries_no_persuasive_weight_against_a_detailed_surveyor_report",
        "Source PDF": "DRN9797551.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "F5E6E6" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Roof_Damage_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
