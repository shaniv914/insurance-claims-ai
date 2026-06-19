"""
Standard append script for Flood Case Database — Schema v1 (21 columns).
Batch 5: FLOOD-041 to FLOOD-050

Usage
-----
Run from the repo root:
    py scripts/append_flood_v5.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Flood_Case_Database.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 5: FLOOD-041 to FLOOD-050
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "FLOOD-041",
        "FOS Decision ID": "DRN-5267073",
        "Insurer Name": "The National Farmers' Union Mutual Insurance Society Limited",
        "FOS Decision Date": "Mar 2025",
        "Claim Type": "Residential buildings and contents insurance — January 2024 flood accepted by NFU; dispute about whether insurer must fund comprehensive flood defences or rebuild/lift property to prevent future flooding; NFU offered £5,000 flood resilience benefit per policy terms but declined to fund rebuild or lifting as betterment",
        "Leak Source": "External flood water — January 2024 flood event; NFU accepted the flood claim; groundwater flood risk from Chalk aquifer identified in expert flood report (moderate to high 1.3% annual probability)",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "NFU: proposed repairs are effective and lasting per independent expert flood risk report (1.3% annual probability); any further payment beyond the policy £5,000 flood resilience benefit would constitute betterment; policy does not oblige insurer to fund rebuild or lifting of property",
        "Evidence Dispute": "NFU provided independent expert flood risk report: moderate to high 1.3% annual probability of groundwater levels rising above surface level; not necessarily indicating property will flood. Policyholders argued local mismanagement (not weather) causes higher flood risk not captured by the report. No independent expert evidence submitted by policyholders to counter NFU's report. FOS found NFU's expert report persuasive in absence of rebuttal.",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld; NFU's repair settlement (including £5,000 flood resilience benefit) is effective and lasting per expert flood report; insufficient evidence that property will definitely re-flood post-repair; NFU not required to fund rebuild or lift property",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood resilience benefit clause: up to £5,000 for flood resilience measures where repair cost >£10,000 with prior insurer consent; 'effective and lasting repair' principle — FOS requires evidence that repair will not be effective and lasting before ordering betterment; betterment principle limits insurer liability to like-for-like repair not structural enhancements",
        "Missing Evidence": "Independent expert evidence on local flood risk factors (drainage/watercourse mismanagement) to counter insurer's flood risk report; expert evidence that property will definitely flood again post-repair",
        "Ombudsman Reasoning": "FOS principle: repairs must be effective and lasting. To require insurer to fund rebuild/lift, consumer must show property will definitely flood again post-repair. NFU's independent expert flood risk report (1.3% annual probability, moderate risk) was uncontested by independent expert counter-evidence. Consumer's assertion of local mismanagement as a higher-risk factor was not evidenced. Without rebuttal of insurer's expert report, FOS cannot order betterment beyond policy benefit cap.",
        "Workflow Insight": "Where flood claim is accepted and dispute concerns flood defences/betterment, insurer must obtain and rely on independent expert flood risk assessment. Consumer must provide specific expert evidence to rebut that assessment — general assertions about local conditions are insufficient. The policy flood resilience benefit is capped and conditional; further flood prevention beyond that cap requires proof of definite re-flooding risk. FOS cannot direct insurer to fund betterment on basis of possibility alone.",
        "AI Rule Candidate": "IF flood_claim_accepted AND dispute_about_flood_defences AND insurer_has_independent_flood_risk_report AND consumer_provides_no_expert_rebuttal THEN decline_flood_prevention_beyond_policy_benefit_cap; repair_deemed_effective_and_lasting_unless_consumer_demonstrates_definite_reflooding_risk",
        "Source PDF": "DRN-5267073.pdf",
    },
    {
        "Case ID": "FLOOD-042",
        "FOS Decision ID": "DRN-5285327",
        "Insurer Name": "U K Insurance Limited",
        "FOS Decision Date": "Feb 2025",
        "Claim Type": "Residential buildings insurance — water pooling under floorboards from rising water table following heavy rainfall; Churchill declined claim as water table changes constitute neither storm nor flood; all parties agreed rising water table was the cause",
        "Leak Source": "Groundwater — rising water table; water pooled under floorboards as water table increased beneath property; all three parties (Churchill's loss adjuster, Churchill's drying specialist, consumer's builder) independently agreed water table rise was cause",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Churchill: rainfall did not meet policy storm threshold; gradual water table fluctuation is not a flood (no sudden/violent event; loss adjuster definition required event such as river bursting its banks or coastal flooding); policy includes gradual damage exclusion; no sudden release of water identifiable",
        "Evidence Dispute": "All three parties — Churchill's loss adjuster, Churchill's drying specialist, and consumer's own builder — independently agreed rising water table caused water to pool under property. No factual dispute on cause. Dispute was definitional: Churchill argued flood requires sudden/violent external event and applied extra-contractual narrow definition; policy had no flood definition. FOS: no policy flood definition exists; FOS broad definition applies; gradual water table rise causing pooling = flood.",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld; Churchill must reconsider claim under flood peril; gradual water table rise causing water to pool under property constitutes flood where policy has no flood definition; Churchill must then assess whether the flood caused the specific damage claimed (damage to solom/damp proof course)",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood peril (undefined in policy) — where no flood definition exists FOS applies broad definition: flood can occur gradually through slow and steady build-up of water including rising water table; insurer cannot apply extra-contractual narrow flood definition (rivers/coastal only) not set out in policy terms; all-party consensus on cause of water entry is decisive on flood peril question",
        "Missing Evidence": "Evidence of whether the flood (water table rise) specifically caused the claimed damage to the solom/damp proof course — Churchill had not assessed this when declining",
        "Ombudsman Reasoning": "All parties agreed water table rise caused water to pool under property — no factual dispute. Churchill's policy has no flood definition. Churchill cannot apply an extra-contractual narrow definition requiring sudden/violent event or specific external source (rivers, coastal) when that definition is not in the policy. FOS broad definition: flood can occur gradually through slow and steady water build-up; water table rise causing pooling = flood. Churchill must now assess whether the flood caused the specific claimed damage.",
        "Workflow Insight": "If policy has no flood definition and insurer declines on grounds water table rise is not a flood, FOS will apply its broad flood definition and require reconsideration. Loss adjuster conceding water table rise as cause while arguing it is not a flood is inconsistent where no policy definition supports that distinction. All-party agreement on cause of water entry is decisive on flood peril question. Reconsideration does not automatically mean payment — insurer must still assess causation of specific damage.",
        "AI Rule Candidate": "IF policy_has_no_flood_definition AND water_table_rise_causes_water_to_pool_under_property AND all_parties_agree_on_cause THEN classify_as_flood_and_reconsider; IF insurer_applies_extra_contractual_narrow_flood_definition THEN FOS_overrides_with_broad_definition; gradual_water_table_rise_can_constitute_flood_where_water_enters_or_pools_in_property",
        "Source PDF": "DRN-5285327.pdf",
    },
    {
        "Case ID": "FLOOD-043",
        "FOS Decision ID": "DRN-5349922",
        "Insurer Name": "Covea Insurance plc",
        "FOS Decision Date": "Mar 2025",
        "Claim Type": "Flood insurance premium repricing dispute — no flood claim; Covea increased buildings and contents premium from £3,521 to £5,220 at March 2024 renewal citing increased flood risk assessment; Covea also sent renewal documents late (after renewal date), denying consumer opportunity to shop around",
        "Leak Source": "N/A — no flood occurred; dispute concerns flood risk assessment methodology used for underwriting and late communication of renewal premium",
        "Property Type": "Residential home",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": "N/A — no claim made; Covea's premium increase driven by Flood Zone 2 classification within 25m boundary radius per proprietary risk algorithm exceeding Covea's underwriting threshold, triggering Flood Re referral and removal of discounts",
        "Evidence Dispute": "Covea: proprietary risk data shows Flood Zone 2 within 25m boundary; algorithm produces overall flood score above underwriting threshold; data applied consistently to all customers. Consumer: Environment Agency data shows 'Very Low' flood risk (<0.1%); property built 1m higher than neighbours; never flooded in 41 years; 25m radius methodology arbitrary. FOS: FOS cannot override insurer's consistent proprietary pricing methodology; government data is one input among many; Covea acted consistently and without discrimination.",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Premium pricing upheld as fair and consistently applied; £200 compensation awarded for late renewal documents (not received until 8 days after renewal date), which denied consumer opportunity to compare alternatives before cover auto-renewed at significantly increased premium",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "No — Administrative",
        "Key Policy Clause": "FCA ICOBS consumer fairness principles — insurer must provide timely renewal documents to allow consumer to make informed decision before renewal; FOS cannot direct insurer on premium pricing methodology if applied consistently to all customers in same risk profile without discrimination or data error; Flood Re referral is risk transfer mechanism not a coverage change; insurer's flood risk assessment can consider area within defined radius of property boundary",
        "Missing Evidence": "N/A — pricing methodology dispute; FOS cannot compel use of alternative risk data source",
        "Ombudsman Reasoning": "FOS has no power to override insurer's pricing methodology where applied consistently to all customers and no discrimination or data error demonstrated. Covea's 25m boundary radius methodology applied uniformly; their data legitimately showed Flood Zone 2 proximity. Not a data error. However, renewal documents not sent until 8 days after renewal date — consumer missed ability to shop around before policy auto-renewed at higher premium. Cover had already started when documents arrived. £200 compensation reflects loss of opportunity to compare alternatives at significant premium uplift.",
        "Workflow Insight": "FOS cannot interfere with insurer's flood risk underwriting methodology if applied consistently without data error — consumers cannot compel insurers to price risk using government data over proprietary models. However, late renewal documentation for significant premium increases is a separate administrative failure: consumer must have timely pre-renewal notice to exercise right to shop around. Flood Re referral means insurer can no longer offer discounts on the flood element of the premium.",
        "AI Rule Candidate": "IF insurer_applies_consistent_proprietary_flood_risk_methodology AND no_discrimination_or_data_error THEN premium_challenge_fails; IF renewal_documents_sent_after_renewal_date AND premium_significantly_increased THEN compensation_for_loss_of_ability_to_shop_around; FOS_cannot_direct_insurer_to_use_alternative_flood_risk_data_source",
        "Source PDF": "DRN-5349922.pdf",
    },
    {
        "Case ID": "FLOOD-044",
        "FOS Decision ID": "DRN-5387216",
        "Insurer Name": "Covea Insurance plc",
        "FOS Decision Date": "Apr 2025",
        "Claim Type": "Residential buildings and contents insurance — March 2024 cellar flood from rising water table through cracks in walls and floor; Covea accepted some damage but declined flood prevention measures (tanking, larger sump pump) as betterment; Covea characterised flood as one-off; property flooded again in early 2025",
        "Leak Source": "Groundwater — temporary rise in water table caused water ingress through cracks in cellar walls and floor; sump pump insufficient to manage ingress; Covea's own expert acknowledged no guarantee standard repairs would be sufficient under high water pressure",
        "Property Type": "Residential home (with cellar/basement)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "Covea: cracks in cellar walls and floor were due to long-standing wear and tear/deterioration; flood was a one-off incident with no foreseeable risk of recurrence; flood prevention measures (tanking) constitute betterment as consumer did not have tanking previously; government flood risk data showed no significant flood risk",
        "Evidence Dispute": "Covea's expert: water entered from 3 areas — repairs possible but 'no guarantee measures will be sufficient should water pressures be overbearing'; 'as basement is not tanked, it is anticipated there will be some moisture/dampness; should water table have risen, further works likely necessary'. Property flooded again in early 2025 in similar manner. FOS: Covea's own expert report undermined the 'one-off' characterisation; consecutive years of flooding disproves it definitively.",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld; Covea must arrange and pay for suitable flood prevention measures (agreed between parties with their experts) to make the repair effective and lasting; betterment argument overridden by consecutive floods in 2024 and 2025 and Covea's own expert acknowledgement of ongoing risk; £250 D&I compensation",
        "Compensation Awarded (£)": 250,
        "Is Core Case": "Yes",
        "Key Policy Clause": "'Effective and lasting repair' principle — where property at significant risk of re-flooding, flood prevention measures are required to make repair effective and lasting rather than a repair that will need redoing after a short period; betterment argument fails where insurer's own expert acknowledges repair may be insufficient under certain conditions AND property subsequently re-floods; betterment principle cannot prevent insurer from funding measures necessary for durable repair",
        "Missing Evidence": "Covea's final response contradicted by own expert report which qualified repair effectiveness; second flood in 2025 provided conclusive real-world evidence against one-off characterisation",
        "Ombudsman Reasoning": "Covea's final response claimed one-off event with no foreseeable recurrence risk — directly contradicted by Covea's own expert who acknowledged 'no guarantee' of sufficiency under high water pressure and anticipated further works if water table risen. Property then flooded again in early 2025 in same manner, definitively disproving one-off characterisation. FOS principle: repairs must be effective and lasting; no point completing repair work requiring redoing after short period. With floods in 2024 and 2025, flood prevention measures are logically necessary. Betterment argument cannot override effective-and-lasting repair obligation.",
        "Workflow Insight": "When insurer's own expert qualifies effectiveness of proposed repair AND property subsequently re-floods, betterment argument against flood prevention measures fails. The 'effective and lasting repair' principle requires insurer to fund prevention measures if short-term repair would need to be redone. A second flood within 12 months is decisive evidence against the 'one-off' defence. Where parties cannot agree on specific works, FOS can order parties to agree with their experts rather than specifying exact measures in the decision.",
        "AI Rule Candidate": "IF insurer_own_expert_qualifies_repair_sufficiency AND property_re_floods_within_12_months THEN flood_prevention_measures_required_for_effective_lasting_repair; IF betterment_argument AND prevention_measures_necessary_for_durability THEN betterment_argument_fails; IF cellar_or_basement AND repeated_water_table_flooding THEN tanking_or_equivalent_prevention_likely_required",
        "Source PDF": "DRN-5387216.pdf",
    },
    {
        "Case ID": "FLOOD-045",
        "FOS Decision ID": "DRN-5601561",
        "Insurer Name": "Advantage Insurance Company Limited",
        "FOS Decision Date": "Jun 2025",
        "Claim Type": "Residential home insurance — April 2024 heavy water ingress into bedroom floor of extension; consumer claimed groundwater hydrostatic pressure surge (flood); Advantage declined — consumer's own damp proofing expert (sole inspector) found structural movement evidence; Advantage's engineer confirmed hydrostatic pressure insufficient to cause observed concrete failure",
        "Leak Source": "Claimed — hydrostatic pressure from groundwater surge causing damp proof membrane failure in concrete extension floor; insurer attributed damage to structural movement (differential settlement of concrete slab, possibly tree root action or poor compaction) damaging damp proof membrane; consumer's own inspector found structural movement evidence",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Advantage: structural movement (differential settlement of concrete slab) damaged damp proof membrane — not flood; structural movement is gradual cause excluded by policy; policy defines flood as external water entering at/below ground level with substantial/abnormal volume — excludes gradual seepage; consumer's own inspector (only party to physically inspect) found structural movement evidence; Advantage's senior engineer concluded hydrostatic pressure insufficient to cause observed concrete failure pattern",
        "Evidence Dispute": "Consumer's damp proofing company (Company W, only party to physically inspect): visible structural movement caused cracks in slab and created vulnerable floor/wall junction; concluded damage caused by movement in structure which had damaged damp proof membrane. Advantage's senior engineer: concrete requires 3,500–5,000 psi of compressive strength before failure; volume of water needed would have caused far greater damage than occurred; most likely cause was differential settlement from tree root action or poor compaction. Consumer provided Met Office above-average rainfall data and Environment Agency article on groundwater flooding occurring weeks/months after rainfall. FOS: consumer's own inspector evidence and Advantage's technical analysis together favoured structural cause on balance of probabilities.",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld; consumer failed to demonstrate on balance of probabilities that flood caused the damage; structural movement (differential slab settlement) is the more likely cause per consumer's own inspector findings and Advantage's technical evidence on concrete failure thresholds; Advantage acted reasonably in declining",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Policyholder bears burden of proving insured event caused damage on balance of probabilities; policy flood definition: water from external source entering at or below ground level with substantial and abnormal volume/weight/force; excludes gradual seepage/rising damp; where consumer's own expert evidence supports alternative (non-flood) structural cause, claim properly declined; consumer's refusal of further investigation under alternative perils means insurer need not take further action",
        "Missing Evidence": "Independent structural engineer report contradicting consumer's own inspector's structural movement conclusion; consumer declined Advantage's offer to investigate further under alternative perils (e.g. subsidence); hydrostatic pressure measurements at time of event",
        "Ombudsman Reasoning": "Consumer's own damp proofing company (only party to physically inspect property) found structural movement in concrete slab — cracks, vulnerable floor/wall junction — and concluded damp proof membrane was damaged by structural movement. Advantage's senior engineer provided technical reasoning why hydrostatic pressure from groundwater would be insufficient to cause observed concrete failure (failure thresholds require volumes that would have caused far greater damage). Consumer's Met Office and EA evidence establishes heavy rainfall and groundwater possibility but not that hydrostatic pressure specifically caused this failure pattern. Balance of probabilities favours structural movement. Consumer also declined Advantage's offer of further investigation under alternative perils.",
        "Workflow Insight": "Where consumer's own expert (sole inspector) concludes structural movement caused damage, insurer's decline on that basis is appropriate even if consumer disputes that finding. Consumer cannot selectively cite inspector's positive observations (rainfall/groundwater) while ignoring inspector's structural movement conclusion. Consumer refusal of further investigation under alternative perils (subsidence) weakens their position and means insurer need not act further. Technical evidence on concrete failure thresholds (psi requirements) is persuasive in defeating hydrostatic pressure flood theory.",
        "AI Rule Candidate": "IF consumers_own_inspector_finds_structural_movement_evidence AND insurer_provides_technical_evidence_hydrostatic_pressure_insufficient_to_cause_damage_pattern THEN decline_on_balance_of_probabilities; IF consumer_declines_further_investigation_under_alternative_perils THEN insurer_need_not_take_further_action; concrete_failure_threshold_evidence_defeats_hydrostatic_pressure_flood_theory",
        "Source PDF": "DRN-5601561.pdf",
    },
    {
        "Case ID": "FLOOD-046",
        "FOS Decision ID": "DRN5640983",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "Jun 2020",
        "Claim Type": "Residential home insurance — water damage to converted basement through failed tanking; RSA declined claiming damage attributable to wear and tear or gradual causes; RSA acknowledged it had not had sufficient time to investigate flood as a potential cause before declining",
        "Leak Source": "Water ingress through failed tanking in converted basement; RSA attributed to wear and tear or gradual causes; water table rise identified as possible alternative cause but RSA claimed specifically excluded by policy; precise flood source not established before claim was declined",
        "Property Type": "Residential home (with converted basement)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "RSA: no insured peril occurred; damage attributable to wear and tear or gradual cause (tanking failure); flood requires sudden/violent event or established external cause (RSA's loss adjuster position); water table rise is specifically excluded by policy; RSA acknowledged it had insufficient time to investigate flood cause before declining",
        "Evidence Dispute": "RSA obtained loss adjuster report and drains survey — no pipe/drain issues found; attributed damage to tanking failure and lack of DPM (wear and tear). Consumer obtained report that basement needed tanking protection from groundwater; paid for rectification including tanking and waterproofing. FOS investigator: flood can occur gradually through slow and steady build-up (court-established principle); RSA failed to eliminate flood as cause before declining. RSA response: courts require established clear external cause for flood; water table is specifically excluded; investigator's broad definition is flawed. RSA acknowledged insufficient investigation time — FOS: must reassess.",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld; RSA must reassess claim against flood peril using FOS broad flood definition (water building up slowly and steadily in a property can constitute flood regardless of source); RSA not yet required to pay claim — must first obtain expert evidence on flood as cause then respond accordingly",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood peril (undefined in policy) — FOS broad definition: flood can occur when water enters or builds up in a property slowly and steadily regardless of source; insurer cannot decline for wear and tear/gradual without first completing investigation to eliminate flood as a potential cause; burden of proving exclusion applies rests with insurer; insurer's acknowledgement of incomplete investigation is significant",
        "Missing Evidence": "RSA acknowledged it did not complete investigation of flood as potential cause before declining; expert evidence on whether accumulated water in basement constituted flood under FOS broad definition",
        "Ombudsman Reasoning": "RSA declined without completing investigation to eliminate flood as potential cause — and acknowledged this. FOS approach: flood can occur gradually through slow and steady water build-up. Key question: was water building up in basement in a way that constitutes flood? RSA has not properly investigated this. RSA must reassess against flood peril (not necessarily pay at this stage) — obtain expert evidence on cause, then respond. If further dispute arises, consumers can bring new complaint.",
        "Workflow Insight": "Insurer cannot decline basement/cellar water damage for wear and tear without first properly investigating and eliminating flood as a cause. FOS applies broad flood definition to basement water ingress: slow water build-up qualifies as flood regardless of source. Acknowledgement by insurer of incomplete investigation is dispositive — cannot maintain decline without completing investigation. Correct process: investigate → obtain expert evidence → then decide. Reassessment order does not necessarily result in payment.",
        "AI Rule Candidate": "IF insurer_declines_basement_water_damage AND insurer_acknowledges_incomplete_flood_investigation THEN FOS_orders_reassessment; IF no_policy_flood_definition AND water_builds_up_slowly_in_basement THEN apply_broad_FOS_flood_definition; insurer_must_eliminate_flood_as_cause_before_relying_on_wear_and_tear_exclusion",
        "Source PDF": "DRN5640983.pdf",
    },
    {
        "Case ID": "FLOOD-047",
        "FOS Decision ID": "DRN-5827621",
        "Insurer Name": "Aviva Insurance Limited",
        "FOS Decision Date": "Sep 2025",
        "Claim Type": "Residential buildings and contents insurance — October 2023 flood accepted by Aviva; dispute about scope of flood damage coverage (kitchen, downstairs bathroom not fully covered); Aviva's contractor fitted cloakroom joists not to building regulations; Aviva's delay in upstairs drying caused consequential bedroom mould",
        "Leak Source": "External flood water — October 2023 flood event (source not specified); Aviva accepted flood entered property and damaged hallway and part of kitchen flooring; dispute about whether flood water reached full kitchen, bathroom, and whether contractor caused further damage",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "Aviva: moisture readings taken promptly showed flood water only entered first three rows of kitchen flooring; no evidence flood reached bathroom; kitchen appliances not in kitchen at time of flood (renovation in progress); contractor repairs to cloakroom joists were appropriate; upstairs bedroom damp attributed to pre-existing damp (estate agent photos)",
        "Evidence Dispute": "Aviva: contemporaneous moisture readings limited flood extent to kitchen flooring; estate agent photos from pre-purchase showed bedroom damp. Consumer: surveyor report (2 months later, restricted bathroom access) said all ground floor rooms affected; moisture readings unreliable; contractor damaged waste pipe and fitted joists not to building regulations; all damp caused by flood and Aviva's drying delay. FOS: contemporaneous moisture readings outweigh later surveyor report with restricted access; contractor's detailed expert report proved non-compliant joist fitting; drying delay caused avoidable upstairs mould.",
        "Outcome Category": "Upheld",
        "Outcome": "Aviva must pay: £200 toward cloakroom joist replacement; £100 for joist inspection report (on evidence of payment); cost of fungi wash and emulsion to upstairs bedroom walls and ceilings; £150 D&I compensation. Kitchen/bathroom additional claims not required.",
        "Compensation Awarded (£)": 150,
        "Is Core Case": "No — Handling Dispute",
        "Key Policy Clause": "Insurer liable for consequential damage caused by its contractor's substandard work (joist fitting not to building regulations); insurer's duty to take reasonable steps to dry property promptly after flood to prevent consequential mould; contemporaneous moisture readings from insurer outweigh later surveyor reports with restricted access; pre-existing damp documented in pre-purchase photographs limits insurer's liability for all subsequent bedroom damp",
        "Missing Evidence": "Consumer had no contemporaneous photographs of kitchen appliances in kitchen at time of flood (photo taken 5 days later); surveyor's bathroom inspection was restricted by stored items and taken 2 months after flood",
        "Ombudsman Reasoning": "Kitchen/bathroom: contemporaneous moisture readings (taken promptly) are more reliable than surveyor report taken 2 months later with restricted bathroom inspection. Pre-existing damp in estate agent photographs limits full insurer liability for bedroom damp. Cloakroom joists: consumer's detailed contractor report with photographs proved joists did not meet building regulations — Aviva could not contest this; £200 for remediation plus £100 for report cost (if evidenced). Bedroom mould: Aviva's delay in installing upstairs drying caused avoidable mould — cost of fungi wash and emulsion awarded. £150 D&I for drying delay and non-compliant joist discovery.",
        "Workflow Insight": "Insurer's contemporaneous moisture readings taken promptly after flood are the most reliable evidence of flood extent; later surveys with restricted access carry less weight. Where insurer's contractor fails to meet building regulations in repair work, insurer bears cost of remediation even where consumer was also carrying out private works in the same area. Delay in installing drying equipment upstairs creates liability for consequential mould that could have been avoided — limited to treatment cost (fungi wash/emulsion), not full redecoration where pre-existing damp exists.",
        "AI Rule Candidate": "IF contemporaneous_insurer_moisture_readings AND later_surveyor_report_with_restricted_access THEN moisture_readings_prevail_on_flood_extent; IF contractor_repair_not_to_building_regulations AND consumer_provides_detailed_expert_report THEN insurer_liable_for_remediation_cost; IF insurer_delays_upstairs_drying AND consequential_mould_occurs THEN insurer_liable_for_mould_treatment_not_full_redecoration_where_pre_existing_damp; pre_existing_damp_in_pre_purchase_photos_limits_insurer_liability",
        "Source PDF": "DRN-5827621.pdf",
    },
    {
        "Case ID": "FLOOD-048",
        "FOS Decision ID": "DRN-6051732",
        "Insurer Name": "Allianz Global Corporate & Specialty SE",
        "FOS Decision Date": "Jan 2026",
        "Claim Type": "Commercial property insurance (charity) — November 2024 flood claim; Allianz said flood cover was accidentally included in policy schedule by broker (acting as Allianz's agent); flood cover was not intended at renewal; Allianz declined claim; charity forced to close and suffered significant losses",
        "Leak Source": "External flood — commercial charity premises November 2024; flood caused extensive damage and forced closure; flood source not specified in decision",
        "Property Type": "Commercial (charity premises)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Allianz: flood cover was excluded at renewal but broker accidentally included it in policy schedule; mutual intention at renewal was to exclude flood (October 2024 renewal invite explicitly stated no flood cover; prior 3 years' policies excluded flood; broker instructed to renew on same terms); erroneous schedule does not create valid flood cover",
        "Evidence Dispute": "Charity (T): schedule clearly showed flood cover; Allianz's own Loss Adjuster confirmed flood cover; schedule is legally binding contract. Allianz: renewal invite (October 2024) explicitly stated no flood; T's broker asked to renew on same terms as prior year (no flood); S (broker acting for Allianz) made the schedule error. FOS: common intention at renewal was to exclude flood; Allianz was not satisfied T would have sought flood cover if schedule had been correct; Allianz's delay in confirming the error (November 2024 to January 2025) caused T inconvenience warranting £250 compensation.",
        "Outcome Category": "Compensation Only",
        "Outcome": "Flood claim not required to be paid — flood cover was not genuinely in place; £250 compensation to T for inconvenience of being led to believe claim was being considered under valid flood cover while Allianz investigated and confirmed error in January 2025",
        "Compensation Awarded (£)": 250,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "Policy schedule error by broker acting as insurer's agent — mutual intention at renewal governs true policy coverage despite schedule error; FOS applies fairness standard not strict contractual enforcement; prior policy history (3 years without flood cover) and renewal documentation (explicit flood exclusion) are evidence of mutual intention; delay in communicating coverage error causes inconvenience warranting D&I compensation even where underlying claim fails",
        "Missing Evidence": "N/A — mutual intention clear from renewal invite, prior 3 years' policies, and broker's instruction to renew on same terms; no evidence T would have sought alternative flood cover",
        "Ombudsman Reasoning": "FOS is not a contract enforcement body. On fairness: mutual intention at renewal was clearly to exclude flood (renewal invite said so; broker asked to renew on same terms; 3 prior years had no flood cover). Erroneous schedule did not reflect what was agreed. Allianz taking from November 2024 to January 2025 to identify and confirm the error meant T was led to believe claim was being considered — T could not take alternative action in that period. £250 compensation for that inconvenience. No additional material loss causally established from the period of erroneous claim consideration.",
        "Workflow Insight": "Where insurer's broker accidentally includes cover in schedule that was explicitly excluded in renewal documents, FOS looks to mutual intention (renewal terms agreed, prior policy history) rather than schedule to determine actual coverage. Insurer's time in identifying and communicating schedule error causes inconvenience even where no material loss results — warrants modest D&I compensation. Evidence of mutual intention: prior policy history without that cover, broker's renewal instruction on same terms, and renewal documentation pre-dating the erroneous schedule.",
        "AI Rule Candidate": "IF policy_schedule_includes_cover AND renewal_documents_explicitly_excluded_cover AND broker_asked_to_renew_on_same_terms AND prior_policies_excluded_cover THEN erroneous_schedule_does_not_create_valid_cover; IF insurer_delays_confirming_schedule_error AND claimant_led_to_believe_claim_being_considered THEN D_and_I_compensation_warranted; mutual_intention_at_renewal_governs_coverage_despite_schedule_errors",
        "Source PDF": "DRN-6051732.pdf",
    },
    {
        "Case ID": "FLOOD-049",
        "FOS Decision ID": "DRN6137899",
        "Insurer Name": "UK General Insurance (Ireland) Limited",
        "FOS Decision Date": "May 2015",
        "Claim Type": "Commercial shop insurance (public liability and business interruption) — sewer surge from water company caused flooding at shop (fourth flood at premises); dispute about floor replacement under public liability (insurer cannot produce policy copy), business interruption payment (insufficient evidence provided), and recovery of uninsured losses from water company",
        "Leak Source": "Sewer surge — water company's responsibility; surged sewer caused flooding at commercial shop premises; fourth flood at the same premises from same cause",
        "Property Type": "Commercial (shop/business premises)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "UKI: floor claim rejected under public liability — expert evidence suggested floor damage occurred over long period from previous floods (pre-policy inception); business interruption declined pending consumer providing daily trading statements (annual accounts insufficient); uninsured loss recovery from water company: UKI eventually withdrew assistance when consumer could not provide required evidence before limitation period",
        "Evidence Dispute": "UKI: cannot produce full policy copy (only business interruption section); floor damage pre-dates policy inception per expert evidence; Mr and Mrs P did not provide daily trading statements for BI claim. Consumer: shop closed longer than UKI thinks; supplied annual accounts (not daily statements); UKI confused responsibility for water company recovery; Mrs P's serious illness prevented responding to UKI's requests. FOS: without policy copy UKI cannot apply exclusions; public liability trigger is when liability to landlord arose (post-inception), not when damage occurred; BI claim requires specific daily trading statements not annual accounts.",
        "Outcome Category": "Upheld in Part",
        "Outcome": "UKI must reconsider floor claim under public liability without applying exclusions (cannot evidence exclusions without policy copy); UKI must reconsider business interruption claim if consumer provides daily trading statements for closure period within covered period; uninsured loss recovery handling upheld as reasonable voluntary assistance; no D&I compensation awarded",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "Insurer who cannot produce relevant policy sections cannot apply exclusions in those sections — burden of proving exclusion rests with insurer; public liability trigger is when liability to third party (landlord) arose, not when physical damage occurred (relevant for cumulative/historical damage); business interruption requires policyholder to provide evidence of losses within coverage period (daily trading statements not annual accounts); insurer's voluntary assistance with uninsured loss recovery from third party creates no ongoing obligation",
        "Missing Evidence": "Full policy document (public liability section not available); daily trading statements for business interruption claim (only annual accounts provided)",
        "Ombudsman Reasoning": "Without policy copy, UKI cannot show what exclusions or limitations applied to public liability floor claim. Public liability trigger is when liability to landlord arose (post-inception) not when floor was originally damaged — historical damage argument therefore uncertain without seeing policy terms. UKI must reconsider floor claim without applying unproven exclusions. Business interruption: shop closed only 1.5 days within covered period; consumer must provide daily trading statements (not annual accounts) to quantify losses. Water company recovery: UKI maintained active pursuit for 4+ years and contacted parties monthly; withdrew only when evidence not forthcoming and limitation period approaching — reasonable decision not attributable to UKI's fault.",
        "Workflow Insight": "Insurer who cannot produce policy document cannot apply its exclusions — burden of proving exclusion rests with insurer who must evidence the terms they seek to rely on. Public liability cover triggers when liability arises (post-inception), not when physical damage accumulated — relevant for latent or cumulative flood damage to third-party property. Voluntary assistance with third-party recovery creates no obligation — insurer can withdraw when evidence is not forthcoming without incurring liability. Consumer must provide appropriate evidence type for BI claims: daily/weekly trading statements, not annual accounts.",
        "AI Rule Candidate": "IF insurer_cannot_produce_policy_document THEN insurer_cannot_apply_exclusions; IF public_liability_cover AND physical_damage_pre_inception AND liability_to_third_party_arose_post_inception THEN claim_potentially_valid; IF insurer_voluntarily_pursues_third_party_recovery AND evidence_not_forthcoming AND limitation_approaching THEN insurer_can_withdraw_without_liability; BI_claim_requires_daily_or_weekly_trading_statements_not_annual_accounts",
        "Source PDF": "DRN6137899.pdf",
    },
    {
        "Case ID": "FLOOD-050",
        "FOS Decision ID": "DRN6619758",
        "Insurer Name": "International Insurance Company of Hannover SE",
        "FOS Decision Date": "Jan 2018",
        "Claim Type": "Residential home insurance — water damage and mould under carpet from moisture ingress into floor void through low-set air bricks; consumer claimed flood (external water ingress); insurer declined — five independent experts all agreed cause was rot from water ingress through air bricks and poor ventilation, with no evidence of standing water accumulating in void",
        "Leak Source": "Water ingress through low-set air bricks into floor void; external ground level too high relative to air bricks allowing rain water to seep into void; poor ventilation in void caused retained moisture and rot — no evidence of standing water accumulation; cause was wet ground conditions not flood",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "IICH: damage caused by rot and gradual processes (both excluded by policy); no flood demonstrated — five experts agreed water entered void through air bricks causing rot from poor ventilation and moisture; no evidence of water building up and remaining as standing water on top of ground; consumer failed to discharge burden of proving flood occurred",
        "Evidence Dispute": "Five independent experts all agreed: cause was rot in flooring from water ingress through air bricks and inadequate ventilation; no standing water in void; external ground level too high relative to air bricks. Consumer: external water ingress = flood; consensus on external water source satisfies flood definition; no standing water because ground was soil (absorbed it). FOS: water ingress that soaks into soil without accumulating as standing water = wet ground, not flood; flood requires water to build up on top of ground; consumer failed to prove flood occurred.",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld; consumer failed to demonstrate that a flood occurred; water ingress that soaks into soil without accumulating as standing water is not a flood; all five experts agreed cause was rot from gradual moisture ingress and poor ventilation, not a flood event; IICH acted fairly in declining",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood requires standing water to accumulate on top of ground — mere water ingress or seepage that soaks into soil without accumulating as standing water is not a flood ('wet ground' is not flood); policyholder bears burden of proving insured event caused damage on balance of probabilities; gradual damage exclusion and rot exclusion are secondary — only relevant once consumer first proves an insured event occurred; five-expert consensus on non-flood cause is decisive",
        "Missing Evidence": "Evidence of standing water accumulating in floor void (all experts found none — ground was soil that absorbed water); evidence of a specific identifiable flood event (as opposed to general water ingress over time)",
        "Ombudsman Reasoning": "FOS definition of flood requires water to build up and sit on top of ground as standing water — not merely for water to soak into soil leaving it damp. Consumer's own argument (no standing water because soil absorbed it) proves absence of flood: water soaking into soil without accumulating is just wet ground. Five independent experts agreed water entered through air bricks, soil became damp, rot resulted from moisture and poor ventilation — no evidence of standing water accumulation. Consumer failed to discharge burden of proving flood occurred on balance of probabilities. Exclusions for gradual damage and rot are secondary — only become relevant after insured event is first established.",
        "Workflow Insight": "Flood requires standing water to accumulate on top of ground — not just water ingress that soaks into soil. Where multiple independent experts agree cause was rot/moisture/poor ventilation (not flood) and no evidence of standing water accumulation exists, claim properly declined. Gradual damage and rot exclusions are secondary to the burden-of-proof question — they only become relevant if consumer first proves an insured event occurred. Distinguish: damp soil under floor from air brick seepage (not flood) vs water building up under floorboards or in basement as standing water (potential flood).",
        "AI Rule Candidate": "IF all_independent_experts_agree_cause_was_rot_from_gradual_water_ingress AND no_evidence_of_standing_water_accumulation THEN claim_fails_on_balance_of_probabilities; flood_requires_standing_water_accumulation_not_mere_soil_dampness_or_seepage; IF consumer_fails_to_prove_insured_event THEN exclusions_not_relevant; five_expert_consensus_on_non_flood_cause_is_decisive_where_no_counter_evidence",
        "Source PDF": "DRN6619758.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Flood_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
