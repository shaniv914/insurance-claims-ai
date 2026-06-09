"""
Standard append script for Escape of Water Case Database — Schema v2 (21 columns).

Usage
-----
1. Read the source PDF(s) and extract the fields listed in NEW_CASES below.
2. Add one dict per case to NEW_CASES, following the extraction rules in the
   FIELD EXTRACTION RULES section.
3. Run from the repo root:
       py scripts/append_eow_v2.py

The script appends NEW_CASES rows to:
    knowledge/case-databases/Escape_of_Water_Case_Database.xlsx

It automatically assigns the next available row number and applies
consistent alternating-row formatting.

===========================================================================
FIELD EXTRACTION RULES
===========================================================================

IDENTIFICATION FIELDS
─────────────────────
Case ID
    Format: EOW-NNN  (zero-padded to 3 digits, e.g. EOW-016)
    Source: assign sequentially; last used = EOW-015

FOS Decision ID
    Format: DRN-XXXXXXX or DRNXXXXXXX (match exactly as printed in the PDF)
    Source: first line of the PDF, or the "Ref:" header

Insurer Name
    Source: opening complaint paragraph, e.g.
        "Mr X complains that [INSURER NAME] declined his claim…"
    Use the formal registered name as it appears in the decision.
    For Lloyd's syndicates use: "Insurers at Lloyd's (Society of Lloyd's)"
    For brokers: use the broker's name (not the underlying insurer).

FOS Decision Date
    Format: DD Mon YYYY  (e.g. "15 Aug 2023")
    Source: final paragraph of the decision:
        "I'm required to ask [party] to accept or reject my decision
         before [DATE]."
    Use the accept-or-reject deadline date as printed.

SOURCE / PHYSICAL EVENT FIELDS
───────────────────────────────
Claim Type
    Free text. Describe the physical incident and the nature of the dispute
    in one sentence, e.g.:
        "Escape of water — burst supply pipe in kitchen causing floor damage"
    Do NOT embed dispute classification here — that belongs in Dispute Type.

Leak Source
    Free text. Describe the physical origin of the water.  Examples:
        "Supply pipe — copper pipe behind kitchen units, elbow joint failure"
        "WC cistern overflow — cistern cracked and leaking"
        "Tap left on by resident in managed flat"
        "Pre-existing — inherited damage from previous owner, source unknown"

Property Type
    Free text, but use consistent terms:
        "Residential home"
        "Residential home (kitchen)"  — if damage confined to one room
        "Unoccupied residential property"
        "Unoccupied residential property (intended for refurbishment)"
        "Residential home (recently purchased)"
        "Leasehold flat"
        "Commercial / Management Company"

DISPUTE CLASSIFICATION FIELDS
──────────────────────────────
Dispute Type
    Controlled vocabulary — use EXACTLY one of:
        "Coverage Dispute"
            Insurer declined coverage and customer disputed that decision.
        "Handling / Reinstatement Dispute"
            Insurer accepted claim but dispute arose over reinstatement
            scope, quality of work, or settlement quantum.
        "Endorsement / Exclusion Challenge"
            Insurer applied a specific endorsement or exclusion to decline;
            customer challenged its validity or applicability.
        "Pre-Inception Damage Dispute"
            Insurer declined on the basis that damage occurred before the
            policy start date; may overlap with gradual cause.
        "Peril Classification Dispute"
            Dispute is not about coverage but about which peril applies
            (affects excess level or policy section).
        "Claim Recording / Administrative Dispute"
            Complaint concerns how the claim was recorded or administered,
            not the coverage decision itself.
        "Broker Conduct Dispute"
            Complaint concerns a broker's conduct (disclosure, advice,
            renewal notification), not the insurer's claim decision.

Coverage Decision
    What the INSURER originally decided on coverage — not the FOS outcome.
    Controlled vocabulary — use EXACTLY one of:
        "Declined — Full"
            Insurer declined the entire claim.
        "Declined — Partial"
            Insurer accepted part of the claim and declined the remainder.
        "Accepted"
            Insurer accepted the claim without substantive dispute.
        "Accepted — Disputed Settlement"
            Insurer accepted the claim but the settlement amount, scope,
            or reinstatement quality is disputed.
        "Not Applicable"
            No coverage decision was made (admin / broker disputes,
            claim recording errors, etc.).

FOS OUTCOME FIELDS
──────────────────
Outcome Category
    What the FOS decided.  Controlled vocabulary — use EXACTLY one of:
        "Upheld"
            Complaint fully upheld; insurer required to accept or extend
            the claim.
        "Upheld in Part"
            Some elements upheld, others not.  Coverage may be partially
            extended or a combination of coverage and compensation awarded.
        "Not Upheld"
            Insurer's position maintained throughout.
        "Compensation Only"
            Insurer's coverage decline was upheld as correct, but
            compensation was awarded for a separate handling failure
            (e.g. avoidable delays in processing).

Outcome
    Free text.  Full description of what the FOS required the insurer to do.
    Include: settlement instructions, compensation amount, interest
    obligations.  Match detail level of existing rows.

Compensation Awarded (£)
    Numeric (integer).  The total compensation awarded by the FOS for
    distress and inconvenience.
    - Do NOT include claim settlement amounts (these are in Outcome).
    - Do NOT include interest.
    - If no compensation: 0
    - Example: £150 compensation → 150

Is Core Case
    Whether the case should drive rules in the core residential claims
    assessment playbook.
    Controlled vocabulary — use EXACTLY one of:
        "Yes"
            Standard residential claim — drives core playbook rules.
        "No — Administrative"
            Claim recording or administrative dispute; no coverage
            analysis; retain as reference only.
        "No — Handling Dispute"
            Claim was accepted; dispute entirely about reinstatement
            quality or settlement quantum; no coverage principle.
        "No — Commercial"
            Commercial or all-risks policy; classification rules may
            not apply to standard residential policies.
        "No — Broker Dispute"
            Broker conduct / renewal disclosure complaint; no claim
            assessment principle applies.

ANALYSIS FIELDS  (free text — follow existing depth/style)
────────────────────────────────────────────────────────────
Key Policy Clause   — specific contractual wording or FOS/FCA principle applied
Missing Evidence    — evidence that was absent and affected the outcome
Ombudsman Reasoning — how the ombudsman weighed the evidence
Workflow Insight    — operational rule for the claims workflow
AI Rule Candidate   — machine-evaluable rule for the rules engine

Source PDF
    Format: DRNXXXXXXX.pdf  (match filename in knowledge/raw-cases/escape-of-water/)
===========================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Column definition — must match live workbook exactly
# ---------------------------------------------------------------------------
COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

# Controlled-vocabulary fields — validated on write
CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

# Columns that get centred alignment (not wrapped)
CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — add one dict per case following the extraction rules above.
# Leave empty until you are ready to process the next batch of PDFs.
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID":                  "EOW-036",
        "FOS Decision ID":          "DRN-5088221",
        "Insurer Name":             "Accelerant Insurance Europe SA/NV UK Branch",
        "FOS Decision Date":        "15 Nov 2024",
        "Claim Type":               "Escape of water causing ceiling damage in a social housing property; insurer declined citing prior workmanship as cause rather than EOW; £200 compensation directed for unexplained 3-month handling delay only",
        "Leak Source":              "Disputed — surveyor attributed ceiling cracking to prior defective workmanship rather than EOW; freeholder unable to confirm a water event occurred",
        "Property Type":            "Residential home (rented to local council — social housing)",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "Surveyor attributed ceiling cracking to prior defective workmanship; freeholder could not confirm an EOW event had occurred; insurer declined on causation grounds",
        "Evidence Dispute":         "Surveyor's report attributed ceiling damage to prior workmanship; freeholder's inability to confirm EOW event; 3-month unexplained handling delay not disputed",
        "Outcome Category":         "Upheld in Part",
        "Outcome":                  "FOS upheld only in relation to a 3-month unexplained handling delay; directed Accelerant to pay £200 compensation for that delay; coverage decline upheld",
        "Compensation Awarded (£)": 200,
        "Is Core Case":             "No — Commercial",
        "Key Policy Clause":        "Where a surveyor attributes damage to prior workmanship rather than an escape of water, and the freeholder cannot confirm the water event, the insurer is entitled to decline on causation grounds; compensation remains available for unexplained handling delays even where coverage is not established",
        "Missing Evidence":         "Confirmation from the freeholder or maintenance records of an EOW event occurring at the property; independent assessment of whether ceiling cracking pattern was consistent with EOW or prior workmanship",
        "Ombudsman Reasoning":      "Surveyor found no evidence linking ceiling cracking to an EOW event; freeholder's inability to confirm the event further weakened the causation claim; however, a 3-month unexplained handling delay warranted £200 compensation",
        "Workflow Insight":         "Where a surveyor attributes damage to prior workmanship and the freeholder cannot confirm an EOW event, the coverage decline is sustainable; always investigate and compensate for unexplained handling delays even where the coverage decision is upheld",
        "AI Rule Candidate":        "IF surveyor_attributes_damage_to_workmanship AND freeholder_cannot_confirm_eow THEN coverage_decline = defensible; IF unexplained_handling_delay_3_months_plus THEN compensation = payable_regardless_of_coverage_outcome",
        "Source PDF":               "DRN-5088221.pdf",
    },
    {
        "Case ID":                  "EOW-037",
        "FOS Decision ID":          "DRN-5193042",
        "Insurer Name":             "Allied World Assurance Company (Europe) dac",
        "FOS Decision Date":        "21 Jan 2025",
        "Claim Type":               "Escape of water from failed grouting and sealant around bath and shower at a tenanted property; insurer declined under gradually operating cause; £200 compensation directed for handling delay on a separate plasterwork element",
        "Leak Source":              "Failed grouting and sealant around bath and shower enclosure — gradual deterioration over approximately 3 months",
        "Property Type":            "Residential home (tenanted landlord policy)",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "Failed grouting and sealant establishes gradual deterioration over approximately 3 months with landlord and tenant awareness; gradually operating cause exclusion properly applied",
        "Evidence Dispute":         "Insurer's inspection identified failed grouting and sealant with evidence of ongoing moisture damage over months; tenant and landlord were or ought to have been aware of the deteriorating condition; £200 compensation directed for a separate handling delay on the plasterwork element",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold the EOW coverage dispute — gradually operating cause exclusion properly applied; directed Allied World to pay £200 compensation for a handling delay on a separate plasterwork element",
        "Compensation Awarded (£)": 200,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Failed grouting and sealant around bathroom fixtures establishes a gradually operating cause; where damage has been occurring over approximately 3 months and the landlord and tenant were or ought to have been aware, the gradual cause exclusion is properly applied; compensation is available for handling failures on separate claim elements even where the main coverage decision is upheld",
        "Missing Evidence":         "Independent assessment of precisely when the grouting failure began and whether it was reasonably discoverable by the tenant or landlord; evidence of the landlord's inspection schedule and whether the property was checked during the deterioration period",
        "Ombudsman Reasoning":      "Failed grouting and sealant is a classic gradual deterioration scenario; approximately 3 months of deterioration with parties who were or ought to have been aware means the exclusion applies; handling delay on the plasterwork element warranted £200 compensation",
        "Workflow Insight":         "Failed grouting and sealant damage should always be assessed against the gradual cause exclusion — the duration of deterioration and whether the policyholder or tenant was aware are the key questions; compensation may still be awarded for handling failures even where the coverage decision is upheld",
        "AI Rule Candidate":        "IF eow_source_is_failed_grouting_or_sealant AND deterioration_duration_3_months_plus AND landlord_or_tenant_ought_to_have_known THEN gradual_cause_exclusion = applicable; IF handling_delay_on_separate_claim_element THEN compensation_available_regardless_of_coverage_decision",
        "Source PDF":               "DRN-5193042.pdf",
    },
    {
        "Case ID":                  "EOW-038",
        "FOS Decision ID":          "DRN-5198749",
        "Insurer Name":             "Accredited Insurance (Europe) Ltd",
        "FOS Decision Date":        "4 Feb 2025",
        "Claim Type":               "Burst pipe caused by freezing discovered after policyholder switched insurers; dispute over which insurer's policy period covers the loss",
        "Leak Source":              "Burst pipe — frozen pipe failure; exact date of burst disputed between prior insurer (Accredited) and new insurer",
        "Property Type":            "Residential home",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Accepted",
        "Rejection Reason":         "Dispute was between insurers about which policy period the loss fell within — Accredited (on risk during the likely freeze period) sought to decline; new insurer argued the damage occurred after their coverage began",
        "Evidence Dispute":         "Weather data showing sub-zero temperatures during Accredited's policy period; electricity tripping as corroborating evidence of the likely freeze event timing; Accredited's weather data analysis vs new insurer's position on discovery date",
        "Outcome Category":         "Upheld",
        "Outcome":                  "FOS upheld — Accredited Insurance (Europe) Ltd to accept the claim as the insurer on risk during the likely freeze period; directed to pay £500 compensation for distress and inconvenience; may seek contribution from the successor insurer",
        "Compensation Awarded (£)": 500,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "The relevant date for a frozen pipe loss is the date of the freeze event, not the date of discovery; weather data (sub-zero temperatures) and corroborating physical evidence (electricity tripping) are sufficient to establish which policy period a frozen pipe loss falls within; the insurer on risk during the likely freeze period takes the lead and can seek contribution from the successor insurer",
        "Missing Evidence":         "Precise date the pipe burst confirmed by independent expert; specific temperature readings at the property location on the relevant dates",
        "Ombudsman Reasoning":      "Weather data established sub-zero temperatures during Accredited's policy period; electricity tripping corroborated the timing of the freeze event; date of discovery is not determinative for a freeze-pipe loss — the relevant date is when conditions causing the burst occurred; Accredited was on risk at that time",
        "Workflow Insight":         "For frozen pipe losses, always establish the likely date of the freeze event using weather data and corroborating evidence (electricity trips, witness accounts), not just the date of discovery; the insurer on risk at the time of the freeze event bears primary liability regardless of when the damage became visible",
        "AI Rule Candidate":        "IF frozen_pipe_loss AND weather_data_shows_sub_zero_during_prior_insurer_period AND corroborating_evidence_supports_freeze_timing THEN prior_insurer = liable; IF policyholder_switches_insurer AND loss_date_disputed THEN use_freeze_event_date_not_discovery_date",
        "Source PDF":               "DRN-5198749.pdf",
    },
    {
        "Case ID":                  "EOW-039",
        "FOS Decision ID":          "DRN-5199107",
        "Insurer Name":             "Admiral Insurance (Gibraltar) Limited",
        "FOS Decision Date":        "11 Feb 2025",
        "Claim Type":               "Accepted EOW claim at a property being marketed for sale; dispute about occupancy enquiries, market value reduction, and adequacy of cash settlement",
        "Leak Source":              "Not specified — EOW claim accepted by Admiral; dispute concerns handling and settlement, not the EOW source",
        "Property Type":            "Residential home (being marketed for sale at time of claim)",
        "Dispute Type":             "Handling / Reinstatement Dispute",
        "Coverage Decision":        "Accepted",
        "Rejection Reason":         "N/A — claim accepted; Admiral disputed certain elements of the extended settlement process including additional losses alleged to result from property being sold at auction under financial pressure",
        "Evidence Dispute":         "Policyholder alleged Admiral's occupancy enquiries caused delay leading to forced auction sale at undervalue; Admiral's extended enquiries based on minimal furniture and property being marketed for sale; FOS found occupancy enquiries reasonable; market value reduction from repair not a covered loss; £525 D&I compensation paid by Admiral confirmed as fair",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — Admiral's occupancy enquiries were reasonable; insurer not responsible for policyholder's decision to sell at auction; market value reduction from repair not a covered loss; cash settlement (£43,243 plus interest) confirmed as correct; Admiral's existing £525 compensation payment confirmed as adequate",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "No — Handling Dispute",
        "Key Policy Clause":        "Where a property shows minimal furniture and is being marketed for sale, an insurer is entitled to make extended occupancy enquiries before progressing an EOW claim; market value reduction resulting from disclosed repair history is not a consequential loss covered under the policy; an insurer is not responsible for a policyholder's voluntary decision to sell at auction under financial pressure",
        "Missing Evidence":         "Evidence linking Admiral's handling delay directly to the decision to sell at auction; independent valuation confirming the market value reduction specifically attributable to the EOW repair work",
        "Ombudsman Reasoning":      "Admiral's enquiries were reasonable given the property was sparsely furnished and being marketed for sale; the decision to sell at auction under financial pressure was the policyholder's own; market value reduction from disclosed repair history is consequential loss and not covered; £525 compensation already paid by Admiral was fair for any handling inconvenience",
        "Workflow Insight":         "Extended occupancy enquiries are justified where a property is minimally furnished and being marketed for sale; the insurer is not responsible for consequential financial decisions the policyholder makes during the claims process; always confirm whether D&I compensation has already been paid before making an additional award",
        "AI Rule Candidate":        "IF property_minimally_furnished AND being_marketed_for_sale THEN occupancy_enquiries = justified; IF market_value_reduction_from_repair_history THEN consequential_loss = NOT covered; IF insurer_already_paid_di_compensation AND fos_confirms_adequate THEN no_additional_award",
        "Source PDF":               "DRN-5199107.pdf",
    },
    {
        "Case ID":                  "EOW-040",
        "FOS Decision ID":          "DRN-5396824",
        "Insurer Name":             "Aviva Insurance Limited",
        "FOS Decision Date":        "16 Apr 2025",
        "Claim Type":               "Escape of water from under-floor waste pipe; claim accepted but dispute about whether kitchen unit damage was caused by the EOW and whether compensation for poor expectation management was adequate",
        "Leak Source":              "Under-floor waste pipe (kitchen area)",
        "Property Type":            "Residential home",
        "Dispute Type":             "Handling / Reinstatement Dispute",
        "Coverage Decision":        "Accepted — Disputed Settlement",
        "Rejection Reason":         "Aviva declined to include kitchen unit damage in scope — expert evidence showed wastewater from the under-floor waste pipe did not reach joist height; kitchen units not demonstrably damaged by the EOW",
        "Evidence Dispute":         "Policyholder argued kitchen units were damaged by the EOW; Aviva's expert evidence showed wastewater was confined to the floor void and did not reach joist height, meaning kitchen units were not in the flood path; inconsistent expert reports did not establish the link between kitchen unit damage and this specific EOW",
        "Outcome Category":         "Upheld in Part",
        "Outcome":                  "FOS upheld in part — Aviva directed to increase compensation from £200 to £500 for poor expectation management during the claims process; kitchen unit coverage dispute not upheld — policyholder failed to establish the link between kitchen unit damage and this specific EOW event",
        "Compensation Awarded (£)": 500,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Policyholder bears the burden of linking each claimed item to the specific EOW event; where expert evidence shows the escape was localised to the floor void and did not reach the height of kitchen units, there is no evidential basis for including unit damage in the scope; an insurer may be required to pay increased compensation for poor expectation management even where the coverage decision is upheld",
        "Missing Evidence":         "Independent expert evidence establishing that wastewater from the under-floor waste pipe reached the height of kitchen base units; photographic evidence of unit damage consistent with waterline from a floor void leak",
        "Ombudsman Reasoning":      "Expert evidence showed wastewater confined to floor void below joist level — kitchen units not in the water path; inconsistent expert reports did not resolve the causation question in the policyholder's favour; however, Aviva's poor expectation management during the claim was separately established and warranted increased compensation from £200 to £500",
        "Workflow Insight":         "Always obtain expert confirmation of the water path and maximum flood level before including items in a scope of works; poor expectation management during a claim can attract compensation even where the underlying coverage decision is correct; kitchen unit damage from a floor void pipe requires specific evidence the water reached unit level",
        "AI Rule Candidate":        "IF eow_from_floor_void AND expert_confirms_water_did_not_reach_unit_level THEN unit_damage = NOT established; IF poor_expectation_management_during_claim THEN compensation = payable_even_if_coverage_decision_upheld",
        "Source PDF":               "DRN-5396824.pdf",
    },
    {
        "Case ID":                  "EOW-041",
        "FOS Decision ID":          "DRN5611706",
        "Insurer Name":             "CIS General Insurance Limited",
        "FOS Decision Date":        "",
        "Claim Type":               "Escape of water while property was unoccupied; insurer declined under EOW-when-unoccupied exclusion; adviser had explained the theft exclusion but not the separate EOW unoccupied exclusion",
        "Leak Source":              "Escape of water during unoccupied period (specific source not detailed in this older decision)",
        "Property Type":            "Residential home (unoccupied at time of loss)",
        "Dispute Type":             "Endorsement / Exclusion Challenge",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "CIS declined under the EOW-when-unoccupied exclusion; policy defined unoccupied as insufficiently furnished for full habitation OR not lived in for 60 or more consecutive days — dual test",
        "Evidence Dispute":         "Policyholder stated the property was fully furnished; CIS's adviser at inception had explained the theft exclusion for unoccupied properties but had not explained the separate EOW exclusion; FOS found the failure to explain the EOW exclusion meant CIS could not properly rely on it",
        "Outcome Category":         "Upheld",
        "Outcome":                  "FOS upheld — CIS General Insurance Limited directed to accept and settle the claim; £200 compensation for distress and inconvenience; CIS must confirm whether the property met the unoccupied definition before relying on the exclusion",
        "Compensation Awarded (£)": 200,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Where an insurer's adviser explains one exclusion (theft when unoccupied) but not another (EOW when unoccupied), the insurer cannot subsequently rely on the unexplained exclusion; where a policyholder states the property was fully furnished, the insurer must also confirm the property satisfied the unoccupied definition (dual test: insufficiently furnished for full habitation OR not lived in for 60+ consecutive days) before invoking the exclusion",
        "Missing Evidence":         "Precise details of what the CIS adviser explained at inception; independent confirmation of the property's furnishing status at the time of the EOW; evidence of whether the 60-day non-habitation limb of the unoccupied definition was met",
        "Ombudsman Reasoning":      "CIS's adviser explained the theft exclusion for unoccupied properties but failed to draw attention to the separate EOW exclusion; this failure means CIS cannot fairly rely on the EOW exclusion; additionally, the unoccupied definition involves a dual test and CIS had not confirmed both limbs were met given the policyholder's claim to be fully furnished",
        "Workflow Insight":         "When explaining policy exclusions at inception or renewal, all relevant exclusions — not just some — must be drawn to the policyholder's attention; where an exclusion has a dual test (insufficiently furnished OR not lived in for 60+ days), both limbs must be assessed before the exclusion is invoked; failure to explain an exclusion prevents reliance on it",
        "AI Rule Candidate":        "IF adviser_explained_one_unoccupied_exclusion AND did_not_explain_eow_unoccupied_exclusion THEN eow_unoccupied_exclusion = cannot_be_relied_upon; IF unoccupied_definition_is_dual_test AND policyholder_claims_furnished THEN both_limbs_must_be_assessed_before_exclusion_invoked",
        "Source PDF":               "DRN5611706.pdf",
    },
    {
        "Case ID":                  "EOW-042",
        "FOS Decision ID":          "DRN-5649220",
        "Insurer Name":             "Aviva Insurance Limited",
        "FOS Decision Date":        "16 Dec 2025",
        "Claim Type":               "Escape of water at a tenanted rental property — kitchen floor damage accepted; dispute about whether wall and ceiling damage beyond the kitchen floor was also caused by the same EOW",
        "Leak Source":              "Escape of water causing kitchen floor damage (source of broader wall and ceiling damage disputed)",
        "Property Type":            "Residential home (tenanted rental property)",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Accepted — Disputed Settlement",
        "Rejection Reason":         "Aviva accepted kitchen floor damage but declined to extend coverage to walls and ceilings; damp specialist report was internally inconsistent — body cited a persistent water leak as cause but the associated quote described penetrating damp remediation; salt contamination in walls indicated external water ingress rather than internal EOW",
        "Evidence Dispute":         "Policyholder's damp specialist report attributed damage in body text to a persistent leak but provided a quote for penetrating damp remediation; salt contamination in walls was consistent with external water ingress, not EOW; FOS found the report's internal inconsistency and the salt contamination evidence undermined the claim that wall damage was caused by the EOW",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — Aviva's decision to limit coverage to the confirmed kitchen floor area was fair; the damp specialist report's internal inconsistency and salt contamination evidence meant wall damage could not be reliably attributed to the EOW; no further award directed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "An internally inconsistent specialist report (body citing one cause; associated quote describing a different remediation type) cannot establish EOW causation for damage beyond the confirmed area; salt contamination in walls is physical evidence of external water ingress inconsistent with internal EOW causation; policyholders must provide reliable, internally consistent expert evidence to extend an EOW claim beyond the confirmed damage area",
        "Missing Evidence":         "A follow-up expert report resolving the inconsistency between the damp specialist's attributed cause and the remediation description; independent assessment of whether the salt contamination pattern was consistent with EOW rather than external water ingress",
        "Ombudsman Reasoning":      "The damp specialist report cited a persistent water leak in the body but described penetrating damp remediation in the quote — an internal inconsistency that undermined its reliability; salt contamination in the walls is physical evidence pointing to external water ingress rather than internal EOW; Aviva was entitled to limit coverage to the area where EOW causation was reliably established",
        "Workflow Insight":         "When relying on a specialist report to extend an EOW claim, ensure the report is internally consistent between its attributed cause and its proposed remediation; salt contamination in walls is a key indicator of external water ingress that can defeat an EOW claim for that area; always obtain a follow-up report to resolve any inconsistency before submitting to FOS",
        "AI Rule Candidate":        "IF specialist_report_body_cites_eow AND remediation_quote_describes_penetrating_damp THEN report = internally_inconsistent AND eow_causation_not_established; IF salt_contamination_in_walls THEN external_water_ingress_indicated AND eow_causation = weakened",
        "Source PDF":               "DRN-5649220.pdf",
    },
    {
        "Case ID":                  "EOW-043",
        "FOS Decision ID":          "DRN5670903",
        "Insurer Name":             "Insurers at Lloyd's (Society of Lloyd's)",
        "FOS Decision Date":        "1 Mar 2020",
        "Claim Type":               "Escape of water at an unfurnished residential property during a 5-day unfurnished period; insurer declined under a separate unfurnished property EOW exclusion distinct from the standard 30-day unoccupied exclusion",
        "Leak Source":              "Escape of water during a period when the property was unfurnished (specific physical source not central to this decision)",
        "Property Type":            "Residential home (unfurnished at time of EOW — only 5 days)",
        "Dispute Type":             "Endorsement / Exclusion Challenge",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "Policy contained an endorsement excluding EOW cover where the property was unfurnished; property was unfurnished for 5 days at the time of the EOW; the unfurnished exclusion applied from day 1, before the 30-day unoccupied exclusion would have been triggered; insurer provided adequate justification for the higher risk",
        "Evidence Dispute":         "Policyholder argued 5 days unfurnished was unreasonably short to trigger the exclusion; insurer provided justification for the higher risk; FOS confirmed the unfurnished exclusion is distinct from and independent of the unoccupied exclusion",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — the unfurnished endorsement applied from day 1; insurer provided adequate underwriting justification (higher risk when property is vacant and unfurnished, especially in winter); no award directed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "An unfurnished exclusion for EOW is legally distinct from an unoccupied exclusion; it applies from day 1 regardless of duration and before the 30-day unoccupied trigger; once an insurer provides adequate underwriting justification for the exclusion, FOS will not override it even if the period is very short; policyholders should note that unfurnished and unoccupied are separate policy conditions",
        "Missing Evidence":         "Specific policy wording of both the unfurnished endorsement and the 30-day unoccupied exclusion for comparison; evidence of whether the property had heating or other protective measures during the 5-day unfurnished period",
        "Ombudsman Reasoning":      "The unfurnished endorsement is a separate and independent policy condition from the 30-day unoccupied exclusion; it applied from the moment the property became unfurnished, with no minimum period; the insurer provided adequate justification for the higher risk (vacant property in winter); FOS will not override a clearly worded and adequately justified underwriting decision",
        "Workflow Insight":         "Always check for unfurnished endorsements specifically, as they operate independently of and are triggered earlier than standard unoccupied exclusions; where a property moves between furnished and unfurnished states, even briefly, the unfurnished endorsement may apply; document the property's furnishing status throughout the year if an unfurnished endorsement is in place",
        "AI Rule Candidate":        "IF property_is_unfurnished AND policy_has_unfurnished_eow_exclusion THEN eow_exclusion = applies_from_day_1_regardless_of_duration; IF unfurnished_exclusion_present AND unoccupied_exclusion_present THEN they_are_independent_conditions",
        "Source PDF":               "DRN5670903.pdf",
    },
    {
        "Case ID":                  "EOW-044",
        "FOS Decision ID":          "DRN-5805040",
        "Insurer Name":             "Saga Services Limited",
        "FOS Decision Date":        "8 Oct 2025",
        "Claim Type":               "Escape of water at an estate property; Saga failed to disclose an endorsement change at renewal; insurer independently declined for breach of heating and drainage condition; counterfactual analysis showed alternative cover would also have excluded the loss",
        "Leak Source":              "Escape of water during a period where heating and drainage condition was not met (specific physical source not the focus of this decision)",
        "Property Type":            "Residential home (property of deceased — executors)",
        "Dispute Type":             "Endorsement / Exclusion Challenge",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "Saga's primary defence was a heating and drainage endorsement that was not met at the time of the loss; independently, the endorsement had been changed at renewal without adequate disclosure; however, counterfactual analysis showed any alternative cover the policyholder would have purchased would also have contained equivalent heating and drainage requirements that were not met",
        "Evidence Dispute":         "Executors challenged the non-disclosure of the endorsement change at renewal; Saga relied on energy bill evidence to establish the heating condition was not met; executors argued smart meter data could challenge the energy bill evidence but did not obtain it; FOS found the counterfactual argument decisive",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — even if the endorsement change should have been disclosed at renewal, alternative cover available to the policyholder would also have contained heating and drainage requirements that were not met; Saga's separate £1,250 goodwill offer was outside FOS jurisdiction; no FOS-directed award; FOS cannot award D&I compensation to executors or estate representatives",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Failure to disclose an endorsement change at renewal does not automatically result in coverage where counterfactual analysis shows the policyholder would not have obtained materially better cover elsewhere; FOS cannot award distress and inconvenience compensation to executors or estate representatives; a policyholder must produce positive evidence (e.g. smart meter data confirming billing errors) to challenge an insurer's documentary evidence of condition breach",
        "Missing Evidence":         "Smart meter data confirming an error in the energy bills used to establish the heating condition breach; evidence that alternative cover available at renewal would not have included an equivalent heating and drainage condition",
        "Ombudsman Reasoning":      "The endorsement change non-disclosure was a valid complaint ground; however, even without the endorsement change, any equivalent cover the policyholder would have purchased would also have required compliance with heating and drainage conditions that were not met; the counterfactual analysis defeats the endorsement non-disclosure argument; energy bill evidence of condition breach was not successfully challenged; FOS cannot award D&I compensation to estate representatives",
        "Workflow Insight":         "Always conduct a counterfactual analysis when responding to non-disclosure of endorsement changes — if alternative cover would also have excluded the same loss, the non-disclosure argument fails; executors pursuing claims for deceased policyholders cannot receive D&I compensation from FOS; challenge energy bill evidence of condition breaches with smart meter data at the earliest opportunity",
        "AI Rule Candidate":        "IF insurer_failed_to_disclose_endorsement_change_at_renewal AND alternative_cover_would_also_exclude_same_loss THEN non_disclosure_argument = fails; IF complainant_is_executor_or_estate_representative THEN fos_di_compensation = NOT available",
        "Source PDF":               "DRN-5805040.pdf",
    },
    {
        "Case ID":                  "EOW-045",
        "FOS Decision ID":          "DRN5927839",
        "Insurer Name":             "Admiral Insurance (Gibraltar) Limited",
        "FOS Decision Date":        "18 May 2019",
        "Claim Type":               "Escape of water in 2015 at a property already affected by dry rot; insurer declined citing gradual cause — dry rot requiring both wet timber and poor ventilation pre-existed the EOW; dominant cause (poor ventilation) not an insured peril",
        "Leak Source":              "Escape of water (2015 event); ongoing damp; blocked air bricks; inadequate subfloor ventilation — multiple contributing factors",
        "Property Type":            "Residential home",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "Dry rot requires both wet timber AND poor ventilation — both conditions pre-existed the claimed EOW; policyholders were notified of ongoing damp issues in 2015; the dominant cause of the dry rot outbreak was poor ventilation (blocked air bricks), not the EOW",
        "Evidence Dispute":         "Policyholders' expert argued a 2015 EOW caused the dry rot; Admiral's evidence showed both pre-conditions (wet timber and poor ventilation) pre-existed; expert who wrote the 2018 report gave contradictory views (written report vs subsequent email); FOS preferred the written report and found the gradual cause exclusion applied",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — dry rot's dual preconditions (wet timber AND poor ventilation) both pre-existed; where the dominant cause (poor ventilation from blocked air bricks) is not an insured peril, Admiral's decline is sustainable; no award directed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Dry rot requires both wet timber and poor ventilation; where both conditions pre-existed the claimed EOW event and the policyholder was notified of the conditions, the gradual cause exclusion applies even if the EOW contributed to one of those conditions; where the dominant cause of damage is not an insured peril (e.g. poor ventilation from blocked air bricks), the insurer's decline is sustainable; policyholders bear the burden of demonstrating the insured peril was the dominant cause",
        "Missing Evidence":         "Expert evidence establishing the precise chronology of when the dry rot developed relative to the EOW events; independent assessment definitively ruling out pre-existing dry rot before the 2015 EOW; consistent expert opinion",
        "Ombudsman Reasoning":      "Both dry rot preconditions pre-existed the EOW; policyholders were notified of damp issues in 2015, establishing constructive knowledge; the expert's written report was preferred over the contradictory email; dominant cause analysis determined poor ventilation (blocked air bricks) was the primary cause — an uninsured peril; EOW's contribution to timber wetness does not make it the dominant cause",
        "Workflow Insight":         "Dry rot claims involving EOW require analysis of all preconditions for dry rot (wet timber AND poor ventilation); where both pre-existed and the policyholder had notice of damp, the gradual cause exclusion is strong; dominant cause analysis is key — establish which condition was the primary driver of the rot; written expert reports are given more weight than subsequent conflicting oral or email communications",
        "AI Rule Candidate":        "IF dry_rot_claimed AND both_preconditions_preexisted AND policyholder_notified_of_damp THEN gradual_cause_exclusion = applicable; IF dominant_cause_is_uninsured_peril THEN eow_contribution = insufficient_for_coverage",
        "Source PDF":               "DRN5927839.pdf",
    },
    {
        "Case ID":                  "EOW-046",
        "FOS Decision ID":          "DRN-5979060",
        "Insurer Name":             "Ecclesiastical Insurance Office Plc",
        "FOS Decision Date":        "19 Jan 2026",
        "Claim Type":               "Second FOS complaint about reinstatement scope following a March 2024 FOS direction; dispute about whether undersized joists causing floor deflection were damaged by EOW or were a pre-existing structural issue",
        "Leak Source":              "Escape of water (specific source addressed in prior FOS decision — this decision concerns follow-on expert assessment of specific structural components)",
        "Property Type":            "Residential home (holiday home)",
        "Dispute Type":             "Handling / Reinstatement Dispute",
        "Coverage Decision":        "Accepted — Disputed Settlement",
        "Rejection Reason":         "Following the March 2024 FOS direction to reconsider, Ecclesiastical appointed a suitably qualified expert who found the floor deflection (undersized joists) was a structural issue independent of the EOW; insurer declined to include joist replacement in the scope",
        "Evidence Dispute":         "Policyholder relied on photographs to argue the joists were damaged by EOW; Ecclesiastical's qualified expert (appointed pursuant to the prior FOS direction) found undersized joists causing floor deflection is a pre-existing structural issue independent of the EOW event; FOS preferred the expert's structured physical assessment over the policyholder's photographs",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — Ecclesiastical had complied with the prior FOS direction by appointing a suitably qualified expert; that expert's findings that the joist issue was structural and pre-existing were reliable and reasonable; policyholder photographs alone were insufficient to override a structured expert report; no further award directed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Where an insurer appoints a suitably qualified expert pursuant to a prior FOS direction and that expert finds a specific component's problem is structural and pre-existing and independent of the EOW, the insurer has discharged its obligation under the FOS direction; policyholder photographs alone are insufficient to override the findings of a structured expert physical assessment; compliance with a prior FOS direction is assessed by reference to the quality of the process, not whether the outcome favours the policyholder",
        "Missing Evidence":         "Independent structural engineer assessment of the joist condition both before and after the EOW event; photographic evidence taken closer to the time of the EOW showing the joist condition at that point",
        "Ombudsman Reasoning":      "Ecclesiastical complied with the prior FOS direction by appointing a qualified expert; that expert conducted a physical assessment and found the joist deflection was caused by undersized joists — a structural issue pre-existing the EOW; the policyholder's photographs did not constitute expert evidence capable of overriding a structured assessment; FOS would not require Ecclesiastical to take further action on a properly investigated structural finding",
        "Workflow Insight":         "Compliance with a FOS direction to reconsider requires appointing a suitably qualified expert and properly documenting their findings; a structured expert physical assessment will take precedence over policyholder photographs; where a prior FOS complaint resulted in a direction, document the expert appointment and methodology carefully to demonstrate compliance",
        "AI Rule Candidate":        "IF fos_direction_issued_to_reconsider AND insurer_appoints_qualified_expert AND expert_finds_no_eow_damage THEN insurer_complied_with_direction; IF policyholder_relies_only_on_photographs AND insurer_has_structured_expert_report THEN expert_report = more_persuasive",
        "Source PDF":               "DRN-5979060.pdf",
    },
    {
        "Case ID":                  "EOW-047",
        "FOS Decision ID":          "DRN-5982848",
        "Insurer Name":             "Protector Insurance UK",
        "FOS Decision Date":        "24 Mar 2026",
        "Claim Type":               "Escape of water at a leasehold flat in a council-managed block; kitchen EOW damage accepted but bathroom and hallway damage and pre-EOW electrical works declined; dispute about causation of bathroom and hallway damage",
        "Leak Source":              "Escape of water causing kitchen damage; bathroom damage attributed to steam and insufficient ventilation; hallway mould attributed to condensation; pre-EOW electrical works confirmed by rewiring quotation obtained before the EOW",
        "Property Type":            "Leasehold flat (block policy with council as freeholder)",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Accepted — Disputed Settlement",
        "Rejection Reason":         "Inspection report identified bathroom ceiling paint peeling consistent with steam and insufficient ventilation (not EOW); hallway mould consistent with condensation; pre-EOW rewiring quotation confirmed electrical works were pre-existing; loss adjuster's report considered valid basis for partial decline",
        "Evidence Dispute":         "Policyholder argued bathroom and hallway damage was caused by the EOW; inspection report showed evidence of steam damage and condensation rather than water ingress from EOW; pre-EOW rewiring quotation was contemporaneous evidence of pre-existing electrical issue; loss adjuster conducted a physical site inspection — considered as valid as a surveyor's report",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold the disputed elements — bathroom damage consistent with steam and insufficient ventilation, not EOW; hallway mould consistent with condensation, not EOW; pre-EOW rewiring quotation established electrical works were pre-existing; kitchen EOW damage accepted as confirmed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "No — Commercial",
        "Key Policy Clause":        "Where an inspection report identifies damage patterns consistent with steam or condensation rather than EOW, insurer may decline those specific areas while accepting confirmed EOW damage elsewhere; a pre-EOW rewiring quotation is strong contemporaneous evidence that electrical works were pre-existing; a loss adjuster's physical site inspection report carries equal evidential weight to a surveyor's report",
        "Missing Evidence":         "Independent assessment distinguishing between EOW-caused moisture and steam and condensation damage in the bathroom; evidence of the ventilation conditions in the bathroom and hallway at the time of the EOW",
        "Ombudsman Reasoning":      "The inspection report's observations (peeling paint consistent with steam; mould consistent with condensation) supported the partial decline; the pre-EOW rewiring quotation was contemporaneous and compelling evidence that the electrical damage preceded the EOW; a loss adjuster with physical site access is qualified to make causation assessments; kitchen EOW damage was separately confirmed and accepted",
        "Workflow Insight":         "Always check whether bathroom damage patterns (peeling ceiling paint, mould) are more consistent with steam or condensation than EOW — these require different causation evidence; a pre-EOW contractor quotation is strong evidence of pre-existing works; loss adjuster reports carry equal weight to surveyor reports where a physical inspection was performed",
        "AI Rule Candidate":        "IF bathroom_ceiling_paint_peeling AND consistent_with_steam_not_eow THEN bathroom_damage = NOT eow_caused; IF pre_eow_contractor_quotation_obtained THEN works_described = pre_existing; IF loss_adjuster_conducted_physical_inspection THEN report = as_valid_as_surveyor_report",
        "Source PDF":               "DRN-5982848.pdf",
    },
    {
        "Case ID":                  "EOW-048",
        "FOS Decision ID":          "DRN-6004362",
        "Insurer Name":             "Aviva Insurance Limited",
        "FOS Decision Date":        "14 Jan 2026",
        "Claim Type":               "Escape of water from a burst water pipe; claim accepted December 2022 but not progressed for over 2 years; dispute about handling delays, settlement quantum, and contractor options",
        "Leak Source":              "Burst water pipe (December 2022)",
        "Property Type":            "Residential home",
        "Dispute Type":             "Handling / Reinstatement Dispute",
        "Coverage Decision":        "Accepted — Disputed Settlement",
        "Rejection Reason":         "N/A — claim accepted; Aviva failed to arrange an inspection for over 2 years; the eventual cash settlement (~£5,000 based on expert scope) was significantly lower than Mr R's own estimate; Mr R also disputed whether Aviva had offered contractor repairs",
        "Evidence Dispute":         "Aviva relied on independent engineer's inspection report and costed scope of works for the cash settlement; Mr R had no independent expert report contradicting those findings; additional correspondence produced by Aviva demonstrated the contractor option was offered on multiple occasions (January 2024, February 2025, June 2025, December 2025); Mr R's assertion of no contractor offer was not supported by the evidence",
        "Outcome Category":         "Upheld in Part",
        "Outcome":                  "FOS upheld in part — Aviva directed to pay £750 total compensation (£300 already offered plus £450 additional) for failure to progress the claim proactively; Aviva to offer Mr R the option of their own contractors completing repairs; if Mr R declines contractor option, cash settlement per the expert scope to be paid; if additional EOW-related damage found during repairs, Aviva to consider it under the policy",
        "Compensation Awarded (£)": 750,
        "Is Core Case":             "No — Handling Dispute",
        "Key Policy Clause":        "The onus is on the insurer to progress an accepted claim proactively even where the policyholder prefers to self-source contractors; an insurer is entitled to rely on an independent expert's scope and costing as the basis for settlement where no contradictory expert opinion is produced; where a policyholder declined a contractor option and chose self-sourcing, the insurer's responsibility for resulting delays is reduced; FOS will not re-underwrite the quantum of a settlement based on an unchallenged expert report",
        "Missing Evidence":         "No independent expert report from Mr R contradicting the scope and costings; clear documented evidence of when the contractor option was offered (eventually produced and accepted as evidence of three separate offers)",
        "Ombudsman Reasoning":      "Aviva accepted the claim in December 2022 but did not arrange inspection until April 2025 — over 2 years; despite this, Mr R was offered the contractor option multiple times and chose to self-source; shared responsibility for delays reduced the compensation from an initial £1,300 to £750; settlement based on the unchallenged expert report was fair; contractor option must be formally re-offered",
        "Workflow Insight":         "Accepted EOW claims must be progressed proactively by the insurer even where a policyholder is self-sourcing contractors; offer the contractor repair option in writing and retain documentary evidence of each offer; where no contradictory expert report is produced, FOS will uphold a settlement based on the insurer's own expert scope; shared responsibility for delays significantly reduces compensation",
        "AI Rule Candidate":        "IF accepted_claim_not_progressed_2_years AND insurer_did_offer_contractors AND policyholder_chose_self_sourcing THEN shared_responsibility = true AND compensation = reduced; IF no_contradictory_expert_report THEN insurer_settlement_based_on_own_expert = upheld",
        "Source PDF":               "DRN-6004362.pdf",
    },
    {
        "Case ID":                  "EOW-049",
        "FOS Decision ID":          "DRN-6046762",
        "Insurer Name":             "INTACT INSURANCE UK LIMITED",
        "FOS Decision Date":        "27 Jan 2026",
        "Claim Type":               "Escape of water at a residential property; claim accepted subject to average (underinsurance); dispute about five specific settlement points — drying costs subject to average, survey reimbursement, quote inflation, chimney inspection, Christmas food spoilage",
        "Leak Source":              "Escape of water (specific source not central to this decision — dispute concerns settlement terms on an accepted claim)",
        "Property Type":            "Residential home",
        "Dispute Type":             "Handling / Reinstatement Dispute",
        "Coverage Decision":        "Accepted — Disputed Settlement",
        "Rejection Reason":         "Claim settled subject to average (underinsurance); RSA applied average to all claim-related costs including drying invoice; declined full survey reimbursement; attributed quote inflation to policyholder's delay; declined chimney specialist inspection as no evidence of EOW damage; declined food spoilage as kitchen undamaged and property not evacuated",
        "Evidence Dispute":         "(1) Drying invoice — RSA applied average to all costs including monitoring and certificates; FOS confirmed these are claim-related and correctly subject to average; (2) Survey costs — RSA offered £90 of £360; FOS could not find on this as underinsurance position excluded from scope; (3) Quote inflation — 1.76% attributed to Miss J's delay in providing own survey; FOS found RSA's approach fair; (4) Chimney — loss adjuster found no evidence of EOW damage; (5) Christmas food £50 — kitchen undamaged and no evacuation needed",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold any of the five disputed points — all aspects of RSA's settlement approach were reasonable and fair; no further payment directed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "No — Handling Dispute",
        "Key Policy Clause":        "All claim-related costs on an underinsured claim (including drying survey fees, certificates and monitoring) are properly subject to average; a policyholder who delays providing their own survey bears responsibility for resulting quote inflation; a loss adjuster conducting a physical inspection is qualified to assess whether a specific component was damaged by EOW; food spoilage is not covered where the kitchen is undamaged and no evacuation was required",
        "Missing Evidence":         "Evidence of whether the survey delay was attributable to RSA or the policyholder; independent chimney specialist report linking chimney damage to EOW; compelling evidence linking food disposal directly to the EOW rather than the claimant's own decision",
        "Ombudsman Reasoning":      "Drying costs are claim costs subject to average regardless of their specific nature; survey cost finding was precluded as the underinsurance position was outside scope; quote inflation arose from Miss J's own delay in commissioning and providing the survey; loss adjuster found no EOW chimney damage and is qualified to make that assessment; no compelling reason why the EOW caused food to be thrown away where the kitchen was undamaged and the property was not evacuated",
        "Workflow Insight":         "On underinsured claims, all claim-related costs — including specialist monitoring and certification fees — are subject to average; policyholders must be clearly advised at the outset to provide required documentation promptly to avoid bearing quote inflation risk; where a loss adjuster conducts a physical inspection and finds no EOW damage, no specialist inspection obligation arises; ancillary loss claims (food, minor spoilage) require a clear causal link to the EOW",
        "AI Rule Candidate":        "IF claim_subject_to_average AND drying_invoice_includes_monitoring_fees THEN monitoring_fees = subject_to_average; IF policyholder_delays_own_survey AND quote_inflation_results THEN policyholder_bears_inflation_cost; IF loss_adjuster_physical_inspection_finds_no_eow_damage THEN no_specialist_inspection_required",
        "Source PDF":               "DRN-6046762.pdf",
    },
    {
        "Case ID":                  "EOW-050",
        "FOS Decision ID":          "DRN6187392",
        "Insurer Name":             "St Andrew's Insurance Plc",
        "FOS Decision Date":        "",
        "Claim Type":               "Floor cracking attributed by insurer to sulphate attack from water exposure; initially presented as heave and subsidence claim; FOS directed insurer to investigate under EOW cover based on its own causation explanation",
        "Leak Source":              "Disputed — St Andrew's attributed floor cracking to sulphate attack caused by water exposure (potentially from damp ground, fill material, or EOW); origin of the water not determined",
        "Property Type":            "Residential home",
        "Dispute Type":             "Peril Classification Dispute",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "St Andrew's declined under heave and subsidence — no heave of the site (only floor movement); also argued exclusion for solid floor movement without concurrent foundation damage; both grounds found partially deficient by FOS",
        "Evidence Dispute":         "St Andrew's relied on its expert's opinion that no site heave occurred and the floor movement exclusion applied; Miss F's representative (retired engineer) argued site investigations were needed; FOS found the sulphate attack explanation itself created a sufficient factual basis for an EOW investigation; FOS also found St Andrew's had not adequately justified the foundations exclusion",
        "Outcome Category":         "Upheld in Part",
        "Outcome":                  "FOS upheld in part — St Andrew's Insurance Plc directed to consider the damage under EOW cover in accordance with policy terms and FOS guidance; if heave of the site is later established, St Andrew's to consider under that cover without reference to the foundations exclusion; no monetary award",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Where an insurer's own explanation for floor damage (sulphate attack caused by water exposure) is consistent with an EOW cause, that explanation creates sufficient grounds for an EOW investigation; but-for causation — if EOW caused the sulphate attack, the insurer cannot dismiss liability by pointing to the sulphate attack as the proximate cause; an insurer seeking to rely on a foundations exclusion must provide specific reasoning for why foundations are unaffected, not merely assert it",
        "Missing Evidence":         "Investigation results confirming or ruling out an EOW below the floor; evidence that the foundation external walls were or were not also affected by the same process; clear photographic record of the crack to external render cited by Mr J",
        "Ombudsman Reasoning":      "St Andrew's own explanation (sulphate attack from water) is consistent with an EOW origin — this shifts the burden to St Andrew's to demonstrate no EOW occurred or a valid exclusion applies; St Andrew's did not acknowledge the crack in external render cited by Mr J, let alone explain why it was irrelevant to the foundations exclusion; FOS directed EOW investigation as the appropriate next step",
        "Workflow Insight":         "When an insurer's own damage explanation involves water exposure that is consistent with an EOW cause, the insurer must investigate under EOW cover; but-for causation applies to the EOW pathway — do not simply pass liability to the secondary mechanism (sulphate attack); when invoking a foundations exclusion, provide specific reasons why foundations are unaffected with reference to all physical evidence including any cracks noted by the policyholder",
        "AI Rule Candidate":        "IF insurer_attributes_damage_to_sulphate_attack_from_water AND eow_is_plausible_water_source THEN insurer_must_investigate_under_eow_cover; IF insurer_relies_on_foundations_exclusion AND has_not_addressed_visible_cracks_near_foundations THEN exclusion_justification = insufficient",
        "Source PDF":               "DRN6187392.pdf",
    },
    {
        "Case ID":                  "EOW-051",
        "FOS Decision ID":          "DRN-6263959",
        "Insurer Name":             "Tesco Underwriting Limited",
        "FOS Decision Date":        "28 Apr 2026",
        "Claim Type":               "Kitchen ceiling leak identified by home emergency provider; ceiling damage claimed under home insurance trace and access section; insurer applied the higher EOW section excess rather than the lower trace and access excess",
        "Leak Source":              "Leak in ceiling above kitchen; identified and fixed by home emergency contractor under separate home emergency policy",
        "Property Type":            "Residential home",
        "Dispute Type":             "Handling / Reinstatement Dispute",
        "Coverage Decision":        "Accepted — Disputed Settlement",
        "Rejection Reason":         "Tesco applied the EOW section excess (£350 compulsory plus £200 voluntary = £550) rather than the trace and access section excess, on the basis that some water damage existed before the trace and access work; FOS found the damage was confined to the trace and access area and T&A classification was fairer",
        "Evidence Dispute":         "Tesco relied on pre-existing staining and cracking in photographs to support an EOW classification; FOS examined photographs taken after the ceiling was cut for trace and access work and found damage was confined to the same small area; ceiling was in good condition with no widespread water damage consistent with EOW; repair work would be identical regardless of which section applied",
        "Outcome Category":         "Upheld",
        "Outcome":                  "Tesco Underwriting Limited directed to pay Mr H £250 to refund the excess differential (or pay a net settlement of £530 if no settlement yet made); claim to be processed under trace and access section at the lower excess",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Where damage is confined to the trace and access work area and photographs show no widespread water damage, it is fairer and more pragmatic to settle under the trace and access section even if an EOW event technically occurred; an insurer cannot use EOW peril classification to impose a materially higher excess where both sections would produce identical repair costs; policyholders are entitled to claim under the policy section most beneficial to them where multiple sections could apply to the same facts",
        "Missing Evidence":         "Evidence of widespread pre-existing water damage (tidemarks, widespread staining) beyond the trace and access work area; photographs taken before the trace and access work showing the ceiling condition",
        "Ombudsman Reasoning":      "Photographs taken after the T&A access cut showed damage confined to a small area consistent with T&A work — no frayed edges, no widespread water damage patterns; applying the T&A section would produce identical repair work and costs; Tesco should not use EOW classification to impose a higher excess where damage was localised; £250 excess differential to be refunded or credited",
        "Workflow Insight":         "When settling a claim involving both T&A and potential EOW elements, check photographs for evidence of widespread water damage beyond the T&A area; if damage is confined to the T&A area, the T&A section should be used; always apply the policy section most beneficial to the policyholder where both apply to the same facts",
        "AI Rule Candidate":        "IF damage_confined_to_ta_access_area AND no_widespread_water_damage_in_photographs THEN ta_section_classification = appropriate AND lower_excess_applies; IF two_policy_sections_applicable AND repair_costs_identical THEN use_section_most_beneficial_to_policyholder",
        "Source PDF":               "DRN-6263959.pdf",
    },
    {
        "Case ID":                  "EOW-052",
        "FOS Decision ID":          "DRN6739737",
        "Insurer Name":             "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date":        "6 Jan 2020",
        "Claim Type":               "Cracking floor tiles at residential property; policyholder suspected EOW; RSA's surveyor found no evidence of EOW and attributed cracking to poorly fitted subfloor boards; RSA declined to carry out trace and access investigation",
        "Leak Source":              "Disputed — no EOW confirmed; RSA's surveyor attributed cracking to chipboard and plywood subfloor boards not securely fixed, expanding when the house heats up",
        "Property Type":            "Residential home",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "No evidence of an escape of water or any other insurable event; RSA's surveyor attributed floor tile cracking to poor workmanship (inadequately fixed subfloor boards); without an established insured event, trace and access cover was not triggered under the policy wording",
        "Evidence Dispute":         "Miss M relied on a civil engineer's recommendation for further intrusive investigations and claimed the RSA inspector told her trace and access should apply; RSA relied on its surveyor's report finding no EOW and an alternative workmanship explanation; policy wording required evidence of an EOW before T&A cover applied; FOS read the surveyor's own notes as directing Miss M to carry out T&A if she chose, not as an insurer obligation",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — trace and access cover was not triggered without confirmed evidence of an EOW insured event; RSA's surveyor's alternative explanation was reasonable; if Miss M later funds exploratory works and an EOW is confirmed, RSA stated it would consider a claim in line with policy terms; no further action required",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "Trace and access cover is triggered by confirmed evidence of an escape of water, not by a suspected or possible EOW; where an insurer's surveyor provides an alternative non-EOW explanation for damage with expert support, the insurer is not required to fund intrusive investigations; the policyholder may self-fund exploratory works and submit a claim if an EOW is subsequently confirmed; surveyor's notes recommending T&A as an option for the policyholder are not the same as an insurer commitment to fund T&A",
        "Missing Evidence":         "Intrusive investigation results confirming whether an EOW had occurred beneath the floor tiles; independent expert refuting RSA's surveyor's alternative workmanship explanation",
        "Ombudsman Reasoning":      "Policy wording required evidence of an EOW for T&A cover to apply; RSA's surveyor provided a credible non-EOW explanation; civil engineer's recommendation for further investigation did not establish that an EOW had occurred; surveyor's paper note directing Miss M to carry out T&A was addressed to Miss M as an option, not as an insurer commitment; T&A is available once an EOW is confirmed",
        "Workflow Insight":         "Trace and access cover requires established evidence of an EOW — suspected or possible EOW is not sufficient to trigger T&A; where an insurer's surveyor identifies a plausible non-EOW cause for floor damage, the insurer may decline without further investigation; advise policyholders clearly that they may self-fund exploratory works and claim if an EOW is found",
        "AI Rule Candidate":        "IF no_evidence_of_eow AND alternative_non_eow_explanation_established THEN ta_cover = NOT triggered; IF policyholder_suspects_eow AND insurer_surveyor_finds_no_eow THEN insurer_not_required_to_fund_ta_investigations",
        "Source PDF":               "DRN6739737.pdf",
    },
    {
        "Case ID":                  "EOW-053",
        "FOS Decision ID":          "DRN7090448",
        "Insurer Name":             "UK Insurance Limited",
        "FOS Decision Date":        "19 Jan 2015",
        "Claim Type":               "Escape of water from neighbour's bath overflow entering a rental property between tenancies; insurer declined under vacancy exclusion (30+ days unoccupied); ceiling collapsed on day 34 suggesting progressive ingress began before the 30-day period expired",
        "Leak Source":              "Neighbour's bath overflow (overflow fault requiring a plumber to fix — not a tap left running); water entered through ceiling from above property",
        "Property Type":            "Residential home (rental property, tenants vacated 33 days before discovery)",
        "Dispute Type":             "Endorsement / Exclusion Challenge",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "UKI invoked the vacancy exclusion — property unoccupied, untenanted or not actively used for more than 30 consecutive days; tenants had vacated 33 days before the water was noticed; UKI concluded the escape of water occurred after the 30-day period expired",
        "Evidence Dispute":         "UKI relied on the 33-day vacancy period and stated the escape of water was due to a bath overflow discovered on the date of occurrence; Mr K argued the ceiling collapse on day 34 was evidence that water ingress had been occurring over several days — inconsistent with a single-day event; the neighbour called a plumber to fix the overflow, suggesting a pipe or overflow fault rather than a tap left running",
        "Outcome Category":         "Upheld",
        "Outcome":                  "UK Insurance Limited directed to reconsider Mr K's claim in line with the remaining policy terms and conditions; pay £200 compensation for inconvenience caused by the handling of the claim; reissue expired £24 cheque previously paid for initial customer service failure",
        "Compensation Awarded (£)": 200,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "An insurer invoking a vacancy exclusion must demonstrate on the balance of probabilities that the escape of water actually occurred after the vacancy trigger period expired — the date damage was noticed is not determinative; ceiling collapse on the day after discovery suggests progressive water ingress over multiple days starting before the 30-day mark; where a neighbour calls a plumber to fix an overflow, this is consistent with a pipe fault that may have been leaking for days rather than a single tap-on event",
        "Missing Evidence":         "Plumber's report confirming when the overflow fault began and how long it may have been present; inspection of the neighbour's bathroom at the time to establish the nature of the overflow fault",
        "Ombudsman Reasoning":      "If the escape of water was simply a tap left running, no plumber would be needed to fix it — the fact a plumber was called suggests an overflow fault that had been developing; ceiling collapse on day 34 is inconsistent with a single-day water ingress event and more consistent with progressive ingress over days; UKI did not demonstrate that the escape of water started after the 30-day vacancy period expired; EOW more likely began before the 30-day period ended",
        "Workflow Insight":         "When invoking a vacancy exclusion for EOW, analyse the physical evidence of water ingress duration (e.g. degree of ceiling collapse, saturation levels) to determine whether the EOW more likely started before or after the vacancy trigger; a neighbour requiring a plumber to fix an overflow suggests a pipe fault rather than a single negligent act, increasing the likelihood of progressive ingress starting earlier",
        "AI Rule Candidate":        "IF vacancy_exclusion_invoked AND ceiling_collapsed_day_after_discovery THEN progressive_ingress_likely_started_before_vacancy_period AND exclusion = NOT established; IF neighbour_called_plumber_to_fix_overflow THEN overflow_fault_not_single_event AND earlier_start_date = more_probable",
        "Source PDF":               "DRN7090448.pdf",
    },
    {
        "Case ID":                  "EOW-054",
        "FOS Decision ID":          "DRN7112734",
        "Insurer Name":             "Hiscox Insurance Company Limited",
        "FOS Decision Date":        "",
        "Claim Type":               "EOW from drains caused sub-soil washout and building subsidence; claim settled and paid by Hiscox; dispute about whether the claim should be recorded as EOW or subsidence for future renewal and premium purposes",
        "Leak Source":              "Escape of water from drains — caused sub-soil to be washed away, leading to downward movement of the site and cracking of building walls",
        "Property Type":            "Residential home",
        "Dispute Type":             "Claim Recording / Administrative Dispute",
        "Coverage Decision":        "Accepted",
        "Rejection Reason":         "N/A — claim was settled and paid by Hiscox under the EOW excess; post-settlement, Hiscox recorded the claim as subsidence for renewal and database purposes; Miss A disputed this recording",
        "Evidence Dispute":         "Miss A relied on the settlement form headed Escape of Water and the use of an EOW excess to argue the claim was recorded as EOW; Hiscox argued escape of water on the settlement form referred to the applicable excess rate, not the claim category; engineer's own report stated damage consistent with a recent episode of subsidence and recommended applying the EOW excess; FOS agreed with Hiscox's characterisation of the settlement form",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — the physical damage (downward movement of site and cracking) was correctly characterised as subsidence regardless of the water trigger mechanism; Hiscox's use of escape of water on the settlement form related only to the applicable excess rate; no award directed; Miss A not entitled to have the record changed or to premium refund",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "No — Administrative",
        "Key Policy Clause":        "The category of a claim is determined by the nature of the physical damage to the insured property, not by the mechanism that triggered that damage; where EOW from drains caused sub-soil washout leading to site movement and building cracking, the physical damage is subsidence in nature; where a settlement form uses escape of water in the context of applying a lower policy excess, this refers to the excess applicable — not to the claim category for recording purposes",
        "Missing Evidence":         "Evidence from the loss adjuster or engineer establishing exactly how the settlement form's use of escape of water was intended; premium comparison data showing whether Miss A paid higher premiums due to the subsidence record",
        "Ombudsman Reasoning":      "Engineer stated damage was consistent with a recent episode of subsidence — the EOW from drains was the mechanism but subsidence was the resulting physical damage type; escape of water on the settlement form referred to applying the lower EOW excess rather than categorising the claim as EOW; throughout correspondence, subsidence was the consistent terminology used; the neighbour's insurer categorising the same event as EOW does not compel Hiscox to do the same",
        "Workflow Insight":         "Where an EOW is the cause of a chain reaction (EOW leading to soil washout leading to subsidence), record the claim under the nature of the physical damage (subsidence) not the trigger mechanism; when applying a lower excess as a goodwill gesture (e.g. EOW excess rather than subsidence excess), make explicitly clear in writing that the excess reference does not alter the claim category for recording purposes",
        "AI Rule Candidate":        "IF eow_causes_soil_washout AND soil_washout_causes_building_movement THEN physical_damage_type = subsidence AND claim_category = subsidence; IF settlement_form_references_eow_only_for_excess_purposes THEN eow_reference_does_not_change_claim_recording_category",
        "Source PDF":               "DRN7112734.pdf",
    },
    {
        "Case ID":                  "EOW-055",
        "FOS Decision ID":          "DRN7147115",
        "Insurer Name":             "Legal & General Insurance Limited",
        "FOS Decision Date":        "30 Dec 2018",
        "Claim Type":               "Escape of water from dishwasher discovered November 2017; claim made January 2018; insurer declined citing gradually operating cause (alleged continued dishwasher use) and failure to prevent further damage",
        "Leak Source":              "Dishwasher — water leaking onto the dishwasher's electrical wiring, tripping the electricity supply; Mr C isolated the dishwasher after identifying the source",
        "Property Type":            "Residential home",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "Legal and General alleged Mr and Mrs C continued to use the dishwasher after discovering the leak, making the cause gradual (November 2017 to January 2018); relied on loss adjustor's observation of dirty dishes in the dishwasher in March 2018 as evidence of continued use; also alleged failure to take reasonable steps to prevent further damage after discovering the problem",
        "Evidence Dispute":         "L&G relied on loss adjustor's observation of dirty dishes in dishwasher in March 2018 (no photographs taken); also argued nature of damage (kitchen flooring tiles and unit doors) was consistent with prolonged gradual escape; Mr and Mrs C denied dirty dishes in dishwasher; argued water on electrical wiring would make continued use an electrocution risk — a compelling reason to stop immediately; Mr C was a plumber and isolated the dishwasher's plumbing",
        "Outcome Category":         "Upheld",
        "Outcome":                  "Legal and General Insurance Limited directed to reconsider the claim in line with remaining policy terms; not to decline on gradually operating cause or failure to prevent further damage grounds; no D&I compensation awarded",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "The insurer bears the burden of proving a gradually operating cause exclusion applies — it is not on the policyholder to disprove it; where the policyholder has a compelling rational reason for immediately stopping use of an appliance (water on live electrical wiring creates electrocution risk), continued use is implausible; undocumented observations of an inspector are weak evidence against a specific rational explanation for stopping use; the 2-month gap between discovering the leak and making a claim notification is not itself gradual damage",
        "Missing Evidence":         "Photographs of the dirty dishes in the dishwasher at the March 2018 inspection; independent report confirming whether the nature of the damage was consistent with prolonged gradual escape or a shorter period of standing water under flooring",
        "Ombudsman Reasoning":      "Mr and Mrs C had a compelling rational reason not to continue using the dishwasher — water on electrical wiring creates an electrocution risk; L&G's loss adjustor saw dirty dishes but took no photographs; it is implausible that Mr and Mrs C would knowingly cause continuing water damage and risk electrocution; L&G failed to demonstrate that continued use occurred; the 2-month notification delay did not constitute gradual damage; L&G did not discharge its burden of proving the exclusion applied",
        "Workflow Insight":         "When declining on gradually operating cause for an appliance leak, the insurer must produce positive evidence of continued use — circumstantial observations without photographs are insufficient where the policyholder provides a rational explanation for stopping; consider whether the policyholder had a rational safety reason to stop (e.g. electrical hazard) — this substantially undermines the gradual use theory; notification delay alone does not establish gradual damage",
        "AI Rule Candidate":        "IF gradual_cause_exclusion_invoked AND policyholder_had_safety_reason_to_stop AND insurer_evidence_is_undocumented_observation THEN insurer_fails_burden_of_proof; IF notification_delay_2_months AND leak_isolated_at_discovery THEN gradual_cause_exclusion = NOT established",
        "Source PDF":               "DRN7147115.pdf",
    },
    {
        "Case ID":                  "EOW-056",
        "FOS Decision ID":          "DRN7411842",
        "Insurer Name":             "Ageas Insurance Limited",
        "FOS Decision Date":        "",
        "Claim Type":               "Escape of water at a landlord's rental property between tenancies; insurer declined under unoccupied and untenanted exclusion; owners visited regularly for decoration works but this was not considered occupancy under the landlord's policy definition",
        "Leak Source":              "Escape of water during the period between tenancies in January 2013 (specific physical source not identified — dispute focuses on the occupancy exclusion)",
        "Property Type":            "Residential home (landlord's policy — between tenancies)",
        "Dispute Type":             "Endorsement / Exclusion Challenge",
        "Coverage Decision":        "Declined — Full",
        "Rejection Reason":         "Ageas declined under the EOW unoccupied exclusion — property defined as not lived in by a Tenant; no tenant since August 2012; owners' weekend visits for decoration did not constitute occupancy; additionally, policy required heating or drainage in the November to March period if tenants away 14+ days — this condition would also exclude cover",
        "Evidence Dispute":         "Mr and Mrs K argued their weekend visits (including overnight stays with beds and cooking equipment) constituted occupancy; Ageas countered that a landlord's policy is designed for tenanted properties and owner visits do not satisfy the tenancy-based occupancy definition; FOS confirmed that occupancy on a landlord's policy requires everyday activities on a regular basis such that a property is considered in the ordinary sense to be lived in — decoration visits do not meet this test; additionally, Ageas confirmed it would have restricted cover had it known the property was untenanted",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — the unoccupied exclusion was properly applied; on a landlord's policy, occupancy requires tenancy by a tenant; owner's decoration visits do not constitute occupancy; the heating and drainage condition (separate November to March condition) would also have excluded cover even if visits were treated as occupancy; no award directed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "On a landlord's insurance policy, occupied means tenanted — the occupancy definition is tailored to the purpose of the policy (cover for a tenanted residential property); owner's periodic visits for decoration works do not constitute occupancy under a landlord's policy; occupancy requires everyday activities on a regular basis as a primary residence; a landlord's policy November to March condition requiring heating and drainage when tenants are away 14+ days is a separate exclusion that may independently apply",
        "Missing Evidence":         "Evidence of whether the November to March heating and drainage condition was met during the period in question; evidence of whether the property had been presented as tenanted or untenanted at the November 2012 renewal",
        "Ombudsman Reasoning":      "A landlord's policy is designed for tenanted premises — the definition of occupancy as tenancy is not unusual or unreasonable; owner visits for decoration are not everyday residential activities; even if visits counted as occupancy, the November to March heating and drainage condition would independently have excluded cover; Ageas would have restricted cover had it known the property was untenanted — this further supports the decline; broker conduct at point of sale is outside the scope of the insurer's complaint",
        "Workflow Insight":         "On landlord's policies, always check whether the property is currently tenanted before accepting an EOW claim — the occupancy definition is tied to tenancy; notify insurers promptly whenever a property becomes untenanted between lettings; check for seasonal heating and drainage conditions that apply between November and March independently of the vacancy exclusion",
        "AI Rule Candidate":        "IF landlords_policy AND property_untenanted_30_days_plus THEN unoccupied_exclusion = applicable_regardless_of_owner_visits; IF november_to_march AND tenants_absent_14_days_plus AND heating_not_maintained THEN seasonal_exclusion = also_applicable",
        "Source PDF":               "DRN7411842.pdf",
    },
    {
        "Case ID":                  "EOW-057",
        "FOS Decision ID":          "DRN9891691",
        "Insurer Name":             "UK Insurance Limited",
        "FOS Decision Date":        "7 Sep 2015",
        "Claim Type":               "Escape of water from bathroom waste pipe; contents claim — electrical items accepted; soft furnishings (sofa, rug, carpet) declined as damage not consistent with EOW; upholsterer's report cost also disputed",
        "Leak Source":              "Bathroom waste pipe escape of water",
        "Property Type":            "Residential home",
        "Dispute Type":             "Coverage Dispute",
        "Coverage Decision":        "Declined — Partial",
        "Rejection Reason":         "UKI accepted the claim for electrical items; declined sofa, rug and carpet on the basis that damage was not consistent with EOW — loss assessor found carpet stain was pre-existing (no ceiling staining above it), rug damage (curling, bald patches) was not attributable to EOW, and upholsterer's report for sofa stated EOW causation was only a possibility; damage more consistent with wear and tear and general usage",
        "Evidence Dispute":         "Loss assessor found no ceiling staining above the carpet stain (indicating pre-existing damage, not EOW); rug showed wear-consistent damage (curling edges, bald patches); upholsterer's report stated the sofa damage was only a possibility from EOW and noted oil and dirt staining; UKI policy required written estimates to be provided at the policyholder's expense; Ms H claimed she was verbally told the items would be covered but no record of this conversation existed on the insurer's file",
        "Outcome Category":         "Not Upheld",
        "Outcome":                  "FOS did not uphold — damage to soft furnishings was not established as caused by the EOW; absence of ceiling staining above the carpet stain supported pre-existing damage; upholsterer's only a possibility was insufficient; policy required estimates at policyholder's expense; no record of verbal commitment found; no award directed",
        "Compensation Awarded (£)": 0,
        "Is Core Case":             "Yes",
        "Key Policy Clause":        "A policyholder bears the burden of demonstrating that each claimed item was damaged by the EOW; absence of ceiling staining above a carpet stain is strong physical evidence that the stain pre-dates the EOW; an expert or specialist report stating EOW causation is only a possibility is insufficient to discharge the policyholder's burden; where a policy requires written estimates to be provided at the policyholder's expense, the insurer is not liable to reimburse the cost of reports or assessments commissioned by the policyholder; a claimed verbal commitment must be evidenced — absence from the insurer's records makes it implausible",
        "Missing Evidence":         "Contemporaneous photographs of carpet, rug and sofa taken immediately after the EOW before cleaning was attempted; evidence that ceiling above the carpet area was affected by the EOW; upholsterer's report stating a definitive rather than qualified opinion on causation",
        "Ombudsman Reasoning":      "No ceiling staining above the carpet orange stain — physical evidence contradicting EOW causation; rug damage (curling edges, bald patches) consistent with wear and tear rather than water damage; upholsterer's report noted oil and dirt staining and gave only a qualified possibility opinion — insufficient burden discharge; Mr H attempted cleaning before inspection, reducing available physical evidence; no record of verbal commitment on insurer file; the policyholder did not act to her detriment on the alleged verbal commitment",
        "Workflow Insight":         "For EOW contents claims, photograph all claimed items before any cleaning or removal — post-EOW cleaning significantly weakens causation evidence; absence of ceiling staining above soft furnishing damage is a key indicator of pre-existing damage; expert or specialist reports must give a definitive opinion to discharge the policyholder's burden — a possibility is insufficient; verbal commitments by insurers cannot be relied upon without documentary evidence",
        "AI Rule Candidate":        "IF carpet_stain_alleged_from_eow AND no_ceiling_staining_above_stain THEN carpet_stain = pre_existing_not_eow; IF expert_report_states_eow_causation_only_a_possibility THEN policyholder_burden = NOT discharged; IF verbal_commitment_claimed AND no_record_on_insurer_file AND no_detrimental_reliance THEN verbal_commitment_claim = fails",
        "Source PDF":               "DRN9891691.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    # validate controlled-vocab fields before touching the file
    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases",
        "Escape_of_Water_Case_Database.xlsx",
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # verify live schema matches COLUMNS
    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [(i+1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
                    for i in range(max(len(live_headers), len(COLUMNS)))
                    if i >= len(live_headers) or i >= len(COLUMNS)
                    or live_headers[i] != COLUMNS[i]]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n" +
            "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row   = ws.max_row
    last_case  = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
