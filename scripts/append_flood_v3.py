"""
Standard append script for Flood Case Database — Schema v1 (21 columns).
Batch 3: FLOOD-021 to FLOOD-030

Usage
-----
Run from the repo root:
    py scripts/append_flood_v3.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Flood_Case_Database.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 3: FLOOD-021 to FLOOD-030
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "FLOOD-021",
        "FOS Decision ID": "DRN-2922192",
        "Insurer Name": "St Andrew's Insurance Plc",
        "FOS Decision Date": "Jul 2021",
        "Claim Type": "Flood risk premium loading dispute — Flood Re cession; no flood damage claim; complaint about flood risk rating methodology and large annual premium increases (20% per year, 2018–2020)",
        "Leak Source": "N/A — no flood event; complaint concerns flood risk premium pricing and address-point vs postcode rating methodology",
        "Property Type": "Residential home",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": "N/A — no claim declined; complaint about flood risk premium pricing; insurer applied address-point flood risk rating from 2015, ceded policy to Flood Re from 2016; premium increases attributed to general pricing policy changes alongside flood risk component",
        "Evidence Dispute": "Mr and Mrs W: property protected by earth flood barrier built 2009; elevated position; Flood Re website shows low-to-very-low flood risk for address; no prior flooding in 41 years. St Andrew's: uses multiple data sources including third-party flood risk models, EA data, and local authority flood maps; address-point rating from 2015 increased their flood band rating; cession to Flood Re capped flood risk element but overall premium also includes non-flood components",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — insurer entitled to use postcode or address-point flood risk rating consistently without individual property assessments; Flood Re cession correctly applied (capping flood risk element); overall premium increases since 2016 in line with general pricing policy; Mr and Mrs W were not singled out; no unfair or unreasonable treatment found",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Administrative",
        "Key Policy Clause": "Flood Re scheme cession — flood risk premium element capped; address-point rating methodology; insurer entitled to set flood risk premiums based on its own consistent model data; FOS will not interfere with insurer's commercial pricing judgements unless clearly unfair",
        "Missing Evidence": "Individual property-level flood survey distinguishing this address from higher-risk neighbours in the same postcode; breakdown of non-flood vs flood premium components at each renewal",
        "Ombudsman Reasoning": "Insurer entitled to use consistent rating methodology (postcode or address-point) without individual assessment for each property — cost-prohibitive; Flood Re scheme working as designed (capping flood premium element); premium increases since 2016 reflect general pricing adjustments alongside Flood Re, not scheme misapplication; Mr and Mrs W not singled out; no right or wrong way to assess flood risk provided consistent methodology applied",
        "Workflow Insight": "Flood Re cession caps only the flood risk component of the premium — the retained premium for other risks can still increase under general pricing policy; customers may experience overall premium increases even after Flood Re cession if non-flood pricing elements rise; FOS will not challenge insurer's flood risk rating model absent clear inconsistency or unfair singling out",
        "AI Rule Candidate": "IF flood_re_cession = True AND customer_disputes_premium_increase THEN insurer_methodology_upheld IF consistent_with_general_pricing_policy AND no_evidence_of_unfair_singling_out; Flood_Re_cession DOES NOT guarantee reduced overall premium if non-flood retained components increase",
        "Source PDF": "DRN-2922192.pdf",
    },
    {
        "Case ID": "FLOOD-022",
        "FOS Decision ID": "DRN-2928961",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "Sep 2021",
        "Claim Type": "Commercial landlord buildings insurance — claimed flash flood to shop floor on 13 August 2020; insurer declined for failure to prove insured event occurred; tenant alleged water found on shop floor during storm conditions",
        "Leak Source": "External — claimed flash flood through front of shop; insurer's surveyor found no moisture or storm-consistent damage; weather records showed only 1.6mm/hr maximum rainfall on date of incident (not storm conditions)",
        "Property Type": "Commercial (shop / retail)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Policyholder failed to discharge burden of proving an insured event occurred — insurer's surveyor (Company B, 1 month post-incident) found no moisture in floor, no wall/skirting damage, no stock damage; weather records confirmed no storm or flood conditions on the date; no photos of flooding taken; no media reports or evidence of wider area flooding; damage attributed to long-term wear and tear and material breakdown",
        "Evidence Dispute": "Company B (insurer, 1 month post-incident): no moisture, no skirting/wall damage, floor damage = wear and tear/material breakdown. Company C (policyholder, 3 months post-incident): floor saturated, concluded flash flood cause. Weather records: 1.6mm/hr maximum — not storm conditions. FOS placed greater weight on Company B's earlier inspection and independent weather data",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — AXA's decline of the claim was fair and reasonable; policyholder did not discharge burden of proving an insured event; no contemporaneous evidence of flooding; Company B's earlier, more proximate inspection findings given greater weight than Company C's later inspection",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood defined as 'escape of water from the normal confines of any natural or artificial water course, lake, reservoir, canal, drain or dam; inundation from the sea; rain induced run-off; whether resulting from storm or not'; policyholder bears burden of proving insured event occurred; earlier expert inspection given greater evidential weight",
        "Missing Evidence": "Contemporaneous photographs of flooding by tenant; weather data supporting storm or heavy rainfall on 13 August 2020; evidence of wider area flooding (media reports, neighbouring premises claims); moisture readings taken at time of incident rather than 1 or 3 months later",
        "Ombudsman Reasoning": "Onus on policyholder to prove insured event — not discharged here; Company B's earlier inspection (1 month post-incident) found no moisture, no skirting/wall damage consistent with flooding, and stock undamaged; weather records confirmed only 1.6mm/hr rainfall (not storm conditions); no corroborating evidence of flash flooding in area; Company C's saturated floor reading 3 months later cannot be reliably attributed to August 2020 incident; wear and tear excluded under accidental damage cover",
        "Workflow Insight": "In flood claims on commercial properties, the timing of expert inspections is critical — the longer the gap between incident and inspection, the weaker the evidential link; absence of contemporaneous photos, weather data, and corroborating area evidence makes proof of flash flood very difficult; tenant reporting water 'during storm conditions' is insufficient without corroboration where weather data contradicts storm conditions",
        "AI Rule Candidate": "IF flood_claim AND no_contemporaneous_evidence AND insurer_expert_inspected_earlier_than_policyholder_expert THEN weight_earlier_expert_report_more_heavily AND consider_declining_if_burden_not_discharged; absence_of_stock_damage_and_skirting_damage = negative_indicator_for_flooding",
        "Source PDF": "DRN-2928961.pdf",
    },
    {
        "Case ID": "FLOOD-023",
        "FOS Decision ID": "DRN2955063",
        "Insurer Name": "Millennium Insurance Company Limited",
        "FOS Decision Date": "Unknown",
        "Claim Type": "Rented residential property — flooded by agricultural field surface water runoff during thunderstorm; insurer voided policy and declined flood claim citing non-disclosure of proximity to watercourse (130m brook) and flood risk",
        "Leak Source": "External — agricultural field surface water runoff caused by severe thunderstorm exceptional rainfall; water ran off adjacent field into street flooding numerous properties; not river overbank flooding",
        "Property Type": "Residential (rented)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Insurer voided policy from inception citing non-disclosure — Mr A answered 'No' to questions about property being in a flood risk area and near a watercourse; a brook was located approximately 130m away; insurer concluded answers were inaccurate",
        "Evidence Dispute": "Insurer: brook 130m away = watercourse 'near' property; property in flood risk area. Mr A: questions were ambiguous ('near', 'area subject to flooding' undefined); property never previously flooded in 40+ years; neighbours confirmed no prior flooding on the road; EA flood risk data showed likelihood was 'low'; flooding was exceptional one-off event from agricultural runoff not typical flood risk. FOS agreed questions were ambiguous",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld — insurer's voidance and claim decline were unreasonable; non-disclosure questions were ambiguous and not sufficiently specific to fix Mr A with liability; 'near' undefined, no specific distance given; EA data indicated 'low' flood risk; Mr A took reasonable care in answering; insurer must reinstate policy, process claim, pay £200 D&I compensation, remove voidance from all databases",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Insurer must ask specific and clear non-disclosure questions; undefined terms ('near', 'area subject to flooding') cannot support policy voidance; if proximity threshold is material to underwriting (e.g. within 250m), question must state that threshold; pre-CIDRA reasonable care test — insurer cannot apply strict liability for ambiguous questions; one-off exceptional rainfall causing agricultural runoff is different from inherent flood risk from nearby watercourse",
        "Missing Evidence": "Specific distance threshold in questions; insurer's own underwriting acceptance criteria referencing proximity to watercourses (criteria did not reference rivers/watercourses at all); EA flood risk data for the specific property address",
        "Ombudsman Reasoning": "FOS principle: onus on insurer to ask specific, clear questions; 'near' in ordinary meaning is subjective — insurer itself suggested 250m as critical but did not ask that; acceptance criteria did not reference rivers/watercourses, undermining the materiality claim; flooding caused by exceptional rainfall runoff from agricultural field — different from typical watercourse-flooding risk scenario; EA data showed 'low' flood likelihood; property never previously flooded; Mr A's answers reflected reasonable care; voidance of policy caused significant distress — £200 D&I appropriate",
        "Workflow Insight": "Insurers cannot void policies based on vague proximity questions; if specific distance thresholds are material underwriting criteria (e.g. 'within 250 metres of a watercourse'), those thresholds must be stated explicitly in the question; one-off exceptional rainfall events causing surface water flooding may be categorised differently from inherent flood risk posed by adjacent watercourses; acceptance criteria should be internally consistent with questions asked",
        "AI Rule Candidate": "IF policy_voided_for_non_disclosure AND question_text_lacks_specific_distance_threshold AND insurer_underwriting_criteria_not_referenced_in_question THEN voidance_likely_unfair; IF flooding_from_exceptional_rainfall_runoff AND property_not_in_flood_risk_area AND no_prior_flooding THEN consider_one_off_exceptional_event_vs_inherent_risk",
        "Source PDF": "DRN2955063.pdf",
    },
    {
        "Case ID": "FLOOD-024",
        "FOS Decision ID": "DRN-2965648",
        "Insurer Name": "Ocaso SA, Compania de Seguros y Reaseguros",
        "FOS Decision Date": "Aug 2021",
        "Claim Type": "Residential buildings and contents — house flooded by surface water/drain overwhelm during wet storm; claim declined under CIDRA careless misrepresentation — Mr S stated home was >400m from river at inception (actually within 80m); insurer retrospectively applied flood exclusion endorsement",
        "Leak Source": "External — surface water flooding from drains overwhelmed by torrential rainfall during storm; river was not the source of flooding; Mr S's property within 80m of river but river did not cause the flood",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "CIDRA careless qualifying misrepresentation — Mr S declared home was >400m (and >200m on statement of fact) from river; property actually within 80m; had correct distance been declared, insurer would have assessed flood risk and applied Flood Exclusion Endorsement 011 from inception; endorsement also excludes 'flood resulting from storm' blocking the storm route; 'but for' test — storm caused drain overwhelm = flood from storm = excluded",
        "Evidence Dispute": "Mr S argued river was not the cause of the flood (surface water/drain overwhelm) so proximity misrepresentation was irrelevant; insurer argued correct proximity declaration would have triggered flood exclusion regardless of flood source; FOS agreed flood exclusion applied to all flood including from storm and was not limited to river flooding",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — CIDRA careless qualifying misrepresentation established; Mr S responsible for accuracy of proximity declarations; Ocaso entitled to apply retrospective flood exclusion under CIDRA remedy; flood exclusion covered 'flood resulting from storm' so storm route equally blocked; 'escape of water' exception in exclusion applies only to internal fixed apparatus not external public drains",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "CIDRA s.4(1) careless misrepresentation remedy — insurer may treat policy as if entered on terms it would have applied without the misrepresentation; Endorsement 011 Flood Exclusion Clause: (a) watercourse escape; (b) sea inundation; (c) flood resulting from storm or tempest or any other peril except escape of water from fixed water tanks, apparatus or pipes; 'escape of water' exception covers only internal apparatus flooding, not failure of external public drainage to absorb rainfall",
        "Missing Evidence": "Accurate proximity declaration at inception; whether insurer's underwriting criteria would have applied a narrower river-specific exclusion (rather than a broad flood exclusion) had correct distance been disclosed",
        "Ombudsman Reasoning": "CIDRA careless misrepresentation established — Mr S declared property >400m from river; actually within 80m; Mr S responsible for taking reasonable care regardless of whether river was visible; qualifying misrepresentation because correct declaration would have triggered flood exclusion; 'but for' test: storm caused drains to overwhelm = surface water flooding from storm = excluded under endorsement; 'escape of water' exception requires internal apparatus (tanks/pipes) not external public drains overwhelmed by rainfall volume",
        "Workflow Insight": "Broadly drafted flood exclusions covering 'flood resulting from storm' can simultaneously block both the flood and storm routes; 'escape of water' exception in flood exclusion clauses is narrow — covers only internal fixed water apparatus (pipes, tanks), not public drainage infrastructure; CIDRA misrepresentation allows retrospective application of terms the insurer would have used at inception even where actual flood cause was unrelated to the misrepresented fact",
        "AI Rule Candidate": "IF flood_exclusion_includes_flood_from_storm AND cause_is_storm_surface_water_drain_overwhelm THEN storm_route_also_blocked; IF cidra_careless_misrepresentation AND qualifying THEN insurer_may_retrospectively_apply_exclusion_terms_from_inception; escape_of_water_exception_in_flood_exclusion_does_not_cover_public_drainage_failure",
        "Source PDF": "DRN-2965648.pdf",
    },
    {
        "Case ID": "FLOOD-025",
        "FOS Decision ID": "DRN-3121807",
        "Insurer Name": "Ocaso SA, Compania de Seguros y Reaseguros",
        "FOS Decision Date": "Nov 2021",
        "Claim Type": "Residential home — August 2020 flood from torrential rainfall; local drainage overwhelmed, local dam and river burst; ~3 feet of water entered property via front door and conservatory; claim declined under comprehensive flood exclusion endorsement; policyholders argued some damage from burst pipes",
        "Leak Source": "External — dam/river burst and drainage overwhelm from torrential overnight rainfall; water entered via front door and conservatory at ground level; road outside flooded (20 houses on road confirmed flooded by news reports)",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Comprehensive flood exclusion endorsement (Endorsement 011) on policy schedule excludes all flood from storm, river/dam, or watercourse escape; no evidence of burst pipes separate from flood causation; assessor attributed all damage to 3 feet of flood water entering via front door and conservatory; bursting-pipes allegation unsupported",
        "Evidence Dispute": "Mr H argued some damage caused by burst pipes (separate covered peril); insurance assessor's report and local news confirmed all 20 homes on road flooded from dam/river burst; assessor attributed damage to flood water rising to 3 feet within ground floor; no independent evidence of burst pipes distinct from flood ingress; FOS found no separate cause established",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Endorsement 011 flood exclusion clause clearly applies to dam/storm/river flood events; assessor's report, local news, and Mr H's own account all confirm the damage was caused by flood water; no evidence of burst pipes independent of the flood; endorsement 011 covers flooding from storm, dam, or watercourse making the exclusion comprehensive",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Endorsement 011 — Flood Exclusion Clause: (A) escape from normal confines of artificial watercourse/lake/reservoir/canal/dam; (B) sea inundation; (C) flood from storm/tempest or other peril except escape from fixed water tanks/apparatus/pipes; assessor confirmation of 3 feet of water entering property at ground level via doors = classic flood event within exclusion",
        "Missing Evidence": "Independent evidence of burst pipes within property distinct from flood water ingress; plumber's report or inspection confirming internal pipe failure separate from external flooding",
        "Ombudsman Reasoning": "Endorsement 011 clearly listed on policy schedule and defined at page 6; torrential overnight rainfall causing drainage overflow and dam/river burst = flood event within exclusion; 20 houses on road flooded confirming scale of event; assessor confirmed water entered via front door and conservatory to 3 feet depth; this is flood ingress — not consistent with burst pipe pattern; policyholder's burst pipe allegation unsupported by any evidence; even if pipes contributed, insurer entitled to attribute damage to flood given overwhelming evidence; endorsement covers all three flood types (watercourse, sea, storm-flood)",
        "Workflow Insight": "Where comprehensive flood exclusion (covering all flood types) applies, policyholders cannot escape it by alleging a concurrent covered cause (burst pipes) without independent evidence; assessor's report attributing all damage to flood water is determinative; when 20 neighbouring properties confirmed flooded simultaneously, the flood cause is established beyond doubt for the purposes of applying the exclusion",
        "AI Rule Candidate": "IF comprehensive_flood_exclusion AND policyholder_alleges_concurrent_covered_cause THEN require_independent_expert_evidence_of_concurrent_cause; IF assessor_confirms_flood_ingress_to_significant_depth AND local_news_confirms_area_flood THEN flood_exclusion_applies_notwithstanding_alternative_cause_allegation",
        "Source PDF": "DRN-3121807.pdf",
    },
    {
        "Case ID": "FLOOD-026",
        "FOS Decision ID": "DRN-3219788",
        "Insurer Name": "QIC Europe Limited",
        "FOS Decision Date": "May 2022",
        "Claim Type": "Residential buildings insurance — two flood damage claims declined: (1) February 2021 basement flood — attributed to rising water table and failed internal waterproofing system; (2) August 2021 — road outside flooded but no evidence of water entering property at ground level; both excluded under rising water table and gradual damage policy exclusions",
        "Leak Source": "Groundwater — rising water table causing increased pressure on internal basement waterproofing system (both claims); August claim: confirmed road flooding but basement damage attributed to water table rise, not surface water ingress through ground floor",
        "Property Type": "Residential home (with basement)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "February claim: independent weather data showed minimal rainfall (max 5.6mm/hr one week before incident, not flood conditions); expert reports (Company T — waterproofing specialist; Company A — water supplier) concluded rising water table caused increased ground pressure leading to waterproofing system failure (gradual); both exclusions applied (rising water table + gradual occurrence). August claim: road flood confirmed but surveyor found no evidence of surface water entering property at ground level; no skirting/door/carpet/laminate damage consistent with surface flooding; basement damage again attributed to rising water table",
        "Evidence Dispute": "February: Policyholder's experts (Companies A and Q) attributed to heavy rainfall / surface water. Company T (QIC's waterproofing specialist) and Company A concluded rising water table most likely cause. Independent weather data: minimal February rainfall — contradicts heavy rainfall claims. August: Mr M provided photos, fire service report, newspaper article confirming road flooding; surveyor found no ground-floor damage pattern consistent with surface flooding; water damage in basement attributed to water table not surface water",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — both claims fairly declined; February: rising water table + gradual waterproofing failure = two independent exclusions; independent weather data contradicts heavy rainfall claim; QIC's expert reports more detailed and reliable than policyholder's brief, late expert reports. August: road flood confirmed but insufficient evidence that surface water entered property; surveyor found no flood-entry damage pattern; water table again most likely cause of basement damage",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood definition: 'substantial and abnormal build-up of water from an external source'; Flood exclusions: (d) loss or damage caused by rising water table levels; (e) anything that happens gradually; failed waterproofing system due to groundwater pressure = gradual occurrence + rising water table = dual exclusion; policyholder bears burden of proving flood as cause",
        "Missing Evidence": "Independent hydrological report distinguishing groundwater from surface water flood for February claim; contemporaneous moisture readings or photos of February flooding; evidence of ground-floor surface water damage pattern (skirting, doors, carpets) in August claim; weather data supporting heavy rainfall in February 2021 (independent data contradicted this)",
        "Ombudsman Reasoning": "February: detailed expert reports from QIC's specialists (Company T, Company S, Company C) given greater weight than brief, late reports from policyholder's experts; weather data confirmed minimal February 2021 rainfall; Company A (independent water supplier) referenced high groundwater levels (water table), not surface water; Company T concluded waterproofing failure from groundwater pressure; both rising water table and gradual occurrence exclusions applied. August: road flood undisputed; but surveyor found no damage to skirting boards, doors, carpet, or laminate — inconsistent with surface water entry at ground level; water damage in basement pattern consistent with rising water table not overhead flooding; policyholder must prove covered cause, not merely that an external flood occurred nearby",
        "Workflow Insight": "Basement flood claims require two-stage analysis: (1) did surface water enter at ground level (no skirting/door/carpet damage = negative indicator) or (2) did groundwater rise through basement structure (excluded); rising water table + gradual waterproofing failure = dual exclusion; expert timing matters — detailed early reports outweigh brief late reports; road flooding does not prove property flooding — separate evidence of entry required",
        "AI Rule Candidate": "IF basement_flood AND no_skirting_board_door_carpet_damage AND no_surface_water_entry_evidence THEN consider_rising_water_table_exclusion; IF waterproofing_failure AND groundwater_pressure THEN dual_exclusion_applies (rising_water_table + gradual_occurrence); confirm_external_flood_does_not_automatically_prove_property_flood_ingress",
        "Source PDF": "DRN-3219788.pdf",
    },
    {
        "Case ID": "FLOOD-027",
        "FOS Decision ID": "DRN3290959",
        "Insurer Name": "Ageas Insurance Limited",
        "FOS Decision Date": "Sep 2019",
        "Claim Type": "Residential home — June 2016 excessive rain overwhelmed drains; water backed up and flooded home via submerged air bricks at ground floor; broad flood exclusion endorsement on policy; claimants argued damage was caused by storm or escape of water from fixed apparatus rather than flood",
        "Leak Source": "External — drain overwhelm from excessive rainfall; water entered property via submerged ventilation bricks (air bricks) at ground floor level; river and/or drains overflowed; several inches of standing water inside property",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Broad flood exclusion endorsement covers all flood damage except from fixed water apparatus; drain overflow from excessive rain = flood (not storm, not escape of water from apparatus); water entering via submerged air bricks = flood ingress, not apparatus escape; storm route inapplicable — damage pattern (several inches of standing water) is flood damage not typical storm damage; customer agreed to endorsement at inception; no evidence water over-spilled from bath/sink (apparatus escape route not established)",
        "Evidence Dispute": "Mr C and Mrs T: event was storm-driven rain not river flood; storm should be covered; also damage from waste/soil pipe escape indistinguishable from river water. Ageas: water coming into property from outside = flood regardless of source (river or drains); flood endorsement covers all flood. Photos: no evidence of bath/sink overflows; air bricks at ground floor level submerged by external flood water. FOS: confirmed water entering via submerged air bricks = flood not apparatus escape",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — flood endorsement is clearly worded and broadly applies to all flood; water entering via submerged air bricks from external floodwater = flood, not escape from fixed apparatus; storm damage requires storm-type evidence (tiles, staining) not several inches of standing water; emergency strip-out by insurer's agents simultaneously with cover warnings did not constitute acceptance of cover; endorsement fairly applied",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Broad flood exclusion endorsement: 'does not cover loss or damage caused by flood other than directly resulting from escape of water from fixed water tanks, apparatus or pipes'; flood = body or mass of water appearing where it should not be, regardless of cause (river or drain overflow); storm damage = tiles missing, staining/dampness — not several inches of standing water; emergency insurer response ≠ acceptance of cover; endorsement agreed at inception",
        "Missing Evidence": "Evidence that waste/soil pipes within property overflowed independently of the external flood (photos from during flood showed no bath/sink overflows); loss adjuster's July 2017 report identified air bricks as flood entry point",
        "Ombudsman Reasoning": "Flood endorsement is clear — Mr C accepted it, even if he misunderstood its scope; endorsement covers all flood except from fixed water apparatus (escape of water exception); water entering via submerged air bricks is external flood ingress, not apparatus escape; storm cover requires damage typical of storms (tiles, water staining) — not several inches of standing water which is flood-type damage; insurer's emergency strip-out response was appropriate emergency action not cover acceptance, given simultaneous warnings about cover doubts and formal decline within a fortnight; no evidence waste/soil pipes independently overflowed — photos show no overflowing fixtures",
        "Workflow Insight": "When insurer deploys emergency contractors before confirming liability, always issue simultaneous written warnings that emergency response does not constitute cover acceptance; broad flood exclusion endorsement overrides both storm and escape-of-water routes if damage pattern is flooding (standing water); water entering via submerged air/ventilation bricks is flood ingress, not escape of water from apparatus",
        "AI Rule Candidate": "IF flood_exclusion_endorsement_agreed_at_inception AND water_entered_via_submerged_air_bricks THEN escape_of_water_from_apparatus_exception_does_not_apply; IF insurer_deploys_emergency_contractors AND exclusion_flagged_simultaneously THEN emergency_response_not_cover_acceptance; storm_damage_requires_storm_type_evidence_not_standing_water_pattern",
        "Source PDF": "DRN3290959.pdf",
    },
    {
        "Case ID": "FLOOD-028",
        "FOS Decision ID": "DRN-3295916",
        "Insurer Name": "QIC Europe Ltd",
        "FOS Decision Date": "Apr 2022",
        "Claim Type": "Residential home — August 2020 flood claim; QIC voided policy back to 2019 alleging failure to declare 2018 'flood' at renewal; 2018 incident was public drain overflow onto pavement/gravel at front of property — no damage to property, water soaked into gravel and disappeared; CIDRA reasonable consumer test",
        "Leak Source": "External — August 2020 actual flood at property; 2018 alleged prior flood was a transient public drain overflow during heavy rain: water flowed across public footpath, entered gravel at front of property, soaked in, disappeared — no entry to property, no property damage",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "QIC: 2018 drain overflow technically meets policy definition of flood ('substantial and abnormal build-up of water from an external source'); Mr H should have declared it at 2019 renewal (statement of fact said property not affected by flood in last 10 years); QIC would not have renewed if known; careless CIDRA misrepresentation; avoided policy back to 2019 and declined 2020 claim",
        "Evidence Dispute": "QIC: any external water overflow onto property = declarable flood regardless of scale. Mr H: described 2018 incident honestly — water overflowed public drain onto pavement, soaked into front garden gravel, no property entry, no damage, brief and fleeting; normal consumers would not see this as a flood requiring declaration. FOS: CIDRA test is reasonable consumer standard, not strict policy definition; brief drain overflow with no property damage not declarable by reasonable consumer",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld — CIDRA reasonable consumer standard applies, not strict policy definition; fleeting drain overflow on public footpath with no property entry or damage is not what a reasonable consumer would consider a declarable flood; QIC's policy avoidance was unfair and outside CIDRA entitlement; QIC to: pay 2020 flood claim in full (less excess and any premium refunds) including alternative accommodation, disturbance allowance, travel expenses, utility bills, and out-of-pocket expenses; 8% simple interest on settlement from claim date to offer date; £1,000 D&I compensation; remove void record from QIC files and CUE database",
        "Compensation Awarded (£)": 1000,
        "Is Core Case": "Yes",
        "Key Policy Clause": "CIDRA — standard of care is that of a reasonable consumer; QIC renewal statement of fact: 'property or any property within 100m has NOT been affected by flood within the last 10 years'; test is what a reasonable consumer would deem declarable, not the strict policy definition of flood; brief transient public drain overflow with no property damage not declarable by reasonable consumer standard",
        "Missing Evidence": "QIC had no evidence the 2018 incident was known as a flood event locally or caused damage; Mr H's consistent honest account (including knowing QIC would categorise it as a flood) supported credibility; no evidence Mr H was downplaying the 2018 incident",
        "Ombudsman Reasoning": "CIDRA test = reasonable consumer standard, not policy definition; Mr H honestly described 2018 incident — water flowed across public footpath into gravel, soaked in, no property entry, no damage, disappeared quickly during 30-minute heavy rain; this is the kind of transient drain overflow commonly experienced during sudden heavy rain across UK streets — reasonable consumer would not consider this a flood requiring insurance declaration; Mr H's honesty in acknowledging the incident while contesting its characterisation weighs in his favour; QIC's strict definitional approach misapplied CIDRA; substantial financial and personal impact on family (COVID key worker; two sets of living costs; stress diagnosis) justified £1,000 D&I",
        "Workflow Insight": "Insurers cannot apply strict policy definitions of 'flood' to defeat CIDRA's reasonable consumer standard in non-disclosure avoidance cases; a transient drain overflow with no property damage is not a declarable flood by reasonable consumer standards; asking courts to characterise minor, fleeting water incidents as prior floods for avoidance purposes is unlikely to succeed under CIDRA; policy avoidance has severe collateral consequences (CUE record, claim decline, two households to fund) that FOS weighs in setting D&I quantum",
        "AI Rule Candidate": "IF cidra_non_disclosure AND prior_incident_is_fleeting_drain_overflow AND no_property_damage AND no_property_entry THEN reasonable_consumer_would_not_declare; apply_cidra_reasonable_consumer_test_not_strict_policy_definition_to_non_disclosure_avoidance; IF policy_voided_AND_severe_personal_financial_impact THEN higher_d_and_i_quantum_warranted",
        "Source PDF": "DRN-3295916.pdf",
    },
    {
        "Case ID": "FLOOD-029",
        "FOS Decision ID": "DRN3348419",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "Jul 2016",
        "Claim Type": "Commercial business property — secondary latent damage claim; RSA accepted and settled 2007 flood claim; in 2015 Mr W claimed further damage (pool table flooring, cellar, car park, paving/tree, cellar stairs floor) allegedly caused by the same 2007 flood; RSA declined the supplementary claim",
        "Leak Source": "External — original 2007 flood (insured event accepted); secondary claim for latent moisture damage discovered years later including adhesive failure in pool table flooring area and potential cellar/car park damage",
        "Property Type": "Commercial (public house / business premises)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "RSA disputed that most secondary damage areas (cellar, car park, paving/tree, cellar stairs) were caused by the 2007 flood; insufficient expert evidence linking these areas to the flood; only pool table flooring area had sufficient evidence of flood causation; cellar damage negated by damp specialist; car park cause unclear per structural engineer; no expert opinion on paving/tree or cellar stairs",
        "Evidence Dispute": "Structural engineer: car park cause unclear, cannot link to flood; cellar may have been affected; pool table floor adhesive failed due to moisture. Damp specialist: cellar damage unlikely caused by flood. Flooring specialist: water got between vinyl and floor around pool table = adhesive degraded. No expert opinion on paving/tree or cellar stairs floor. Adjudicator and FOS agreed pool table area proven; cellar stairs to be inspected; all other areas insufficient evidence",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Complaint upheld in part — RSA to pay £5,000 for pool table area floor and ceiling repair (evidentially linked to 2007 flood via structural engineer and flooring specialist); RSA to arrange inspection of cellar stairs floor and carry out repairs if flood-caused; all other areas (cellar, car park, paving/tree) not proven as flood-caused; RSA's £5,000 cash offer reasonable given re-carpeting cost ~£4,000",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Policyholder bears burden of proof for each claimed area of damage; secondary flood damage claims years after original accepted claim require area-specific expert evidence linking damage to the original insured event; flood causation assessed separately for each area of claim; FOS will disaggregate multi-area claims and assess evidential basis for each element individually",
        "Missing Evidence": "Expert opinion on paving/tree and cellar stairs floor (no expert report for these areas); independent survey of cellar (damp specialist later negated claim after adjudicator recommendation); stronger pre-loss condition evidence for car park area",
        "Ombudsman Reasoning": "Each area of claimed damage assessed against expert evidence available: pool table flooring confirmed via structural engineer (moisture causing adhesive failure) and flooring specialist (water between vinyl and floor); cellar: damp specialist concluded unlikely to be flood-caused; car park: structural engineer could not determine flood causation; paving/tree and cellar stairs: no expert opinion at all; policyholder's burden not discharged for unsubstantiated areas; £5,000 RSA offer reasonable vs re-carpeting cost of ~£4,000; cellar stairs floor to be inspected as outstanding area of uncertainty",
        "Workflow Insight": "Secondary flood damage claims made years after the original event require separate expert reports for each claimed area — global claims without area-specific causation evidence will be disaggregated and rejected for unsubstantiated areas; FOS will assess each area independently; the original claim acceptance does not extend to subsequently discovered damage without individual expert linkage to the original flood event",
        "AI Rule Candidate": "IF secondary_flood_damage_claim AND years_after_original_accepted_event THEN require_separate_expert_causation_evidence_per_claimed_area; IF claim_partially_proven THEN partial_settlement_ordered_for_proven_areas_only; original_claim_acceptance_does_not_automatically_cover_latent_damage_discovered_later",
        "Source PDF": "DRN3348419.pdf",
    },
    {
        "Case ID": "FLOOD-030",
        "FOS Decision ID": "DRN-3370157",
        "Insurer Name": "Fairmead Insurance Limited",
        "FOS Decision Date": "Apr 2022",
        "Claim Type": "No flood claim made — complaint about policy ambiguity regarding scope of flood cover; Mr and Mrs T live in flood zone and sought confirmation that policy would cover a future flood; alleged mis-sale; Fairmead failed to respond to written queries",
        "Leak Source": "N/A — no flood event; complaint concerns policy wording clarity and insurer's obligation to pre-confirm flood cover",
        "Property Type": "Residential home",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": "N/A — no flood claim declined; complaint about policy wording ambiguity; Fairmead confirmed flood is an insured peril but could not give blanket pre-confirmation that all future flood claims would be covered (exclusions must be applied on a claim-by-claim basis)",
        "Evidence Dispute": "Mr and Mrs T: policy says covered for 'flood' but also excludes 'water entering the building' under Accidental Damage and 'underground water' under flood cover — contradictory and misleading; also verbal assurance of flood cover at inception. Fairmead: policy has flood cover; 'water entering from external source' exclusion is in Accidental Damage section only, not main flood cover section; 'underground water/rising groundwater' exclusion is standard and clearly worded. FOS: no contradiction — exclusions are section-specific",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — policy does have flood cover; 'water entering from external source' exclusion applies only under the Accidental Damage section, not the main flood cover; 'underground water' exclusion is a standard and clearly worded flood cover exclusion; no insurer can give pre-confirmation that all future claims under any peril will be accepted (exclusions apply on a claim-by-claim basis); mild failing in Fairmead not responding to written queries promptly but no substantive wrongdoing; mis-sale allegation not established",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Administrative",
        "Key Policy Clause": "Policy flood cover: buildings and contents covered against flood as standard peril; Accidental Damage exclusion for 'damage caused by water entering the buildings from an external source' — applies to AD section only, not main flood cover; flood cover exclusion for 'underground water/rising groundwater' — standard exclusion; no insurer can pre-confirm coverage for all future claims without knowing the circumstances",
        "Missing Evidence": "N/A — complaint was about policy interpretation, not a live flood claim with disputed evidence",
        "Ombudsman Reasoning": "Policy has flood cover as an insured peril — this was correctly communicated at inception; 'water entering from external source' exclusion is limited to the Accidental Damage section of the policy, not the main flood cover — the two sections are different and independent; rising groundwater exclusion under flood cover is standard and clearly worded; no insurer can give blanket pre-confirmation about future claims as exclusions must be assessed on the facts of each claim; initial verbal assurance of 'flood cover' was accurate (policy does have flood cover); Fairmead's failure to respond promptly to written queries is a minor administrative failing but not substantive unfairness",
        "Workflow Insight": "Policy documents with multiple sections (main peril cover; accidental damage; exclusions) may appear contradictory to consumers who read them linearly — advisers should proactively explain that AD exclusions operate in a separate context from main peril cover; no insurer can pre-confirm all future flood claims will be paid without knowing the claim circumstances; rising groundwater exclusion under the flood cover section is distinct from the main flood peril cover",
        "AI Rule Candidate": "IF accidental_damage_exclusion_for_water_entry AND separate_main_flood_cover_exists THEN exclusion_is_section_specific_not_global; IF consumer_queries_policy_ambiguity_about_flood THEN explain_difference_between_ad_section_and_main_flood_cover; rising_groundwater_exclusion_under_flood_section_is_standard_and_enforceable",
        "Source PDF": "DRN-3370157.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Flood_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
