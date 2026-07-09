"""
Creates the empty Unoccupied_Property_Case_Database.xlsx with header row only.
Run once. Re-running will overwrite existing data.

Schema v1 — 21 columns (same as EOW v2 / Storm v1 / Flood v1 / Subsidence v1 / Theft v1).
Column 6 renamed to "Unoccupied Period / Circumstance".
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Unoccupied Period / Circumstance",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

HEADER_FILL   = PatternFill(start_color="4F1C2E", end_color="4F1C2E", fill_type="solid")
HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

COL_WIDTHS = {
    "Case ID": 14,
    "FOS Decision ID": 18,
    "Insurer Name": 36,
    "FOS Decision Date": 20,
    "Claim Type": 55,
    "Unoccupied Period / Circumstance": 45,
    "Property Type": 28,
    "Dispute Type": 36,
    "Coverage Decision": 28,
    "Rejection Reason": 55,
    "Evidence Dispute": 55,
    "Outcome Category": 22,
    "Outcome": 55,
    "Compensation Awarded (£)": 26,
    "Is Core Case": 24,
    "Key Policy Clause": 65,
    "Missing Evidence": 55,
    "Ombudsman Reasoning": 65,
    "Workflow Insight": 65,
    "AI Rule Candidate": 65,
    "Source PDF": 26,
}


def main() -> None:
    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases",
        "Unoccupied_Property_Case_Database.xlsx"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cases"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = COL_WIDTHS.get(col_name, 30)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"Created: {xlsx_path}")
    print(f"Columns : {len(COLUMNS)}")


if __name__ == "__main__":
    main()
