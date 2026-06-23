"""
Standard append script for Subsidence Case Database — Schema v1 (21 columns).
Column 6 is "Movement Cause" (physical cause of ground movement).

Usage
-----
1. Read the source PDF(s) and extract the fields listed in NEW_CASES below.
2. Add one dict per case to NEW_CASES following the extraction rules.
3. Run from the repo root:
       py scripts/append_subsidence_v1.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Subsidence_Case_Database.xlsx

===========================================================================
FIELD EXTRACTION RULES
===========================================================================

Case ID           : Format SUBS-NNN (zero-padded to 3 digits)
FOS Decision ID   : DRN-XXXXXXX or DRNXXXXXXX as printed in the PDF
Insurer Name      : Formal registered name from the FOS decision
FOS Decision Date : DD Mon YYYY — accept-or-reject deadline in final paragraph
Claim Type        : Physical incident and nature of dispute in one sentence
Movement Cause    : Cause of ground movement — physical mechanism
                    e.g. "Tree roots causing clay soil shrinkage"
                         "Escape of water softening ground — water-induced subsidence"
                         "Excessive groundwater moisture from poor surface drainage"
Property Type     : "Residential home" / "Leasehold flats" / "Commercial" / etc.
Dispute Type      : Controlled vocab (7 values)
Coverage Decision : Controlled vocab (5 values)
Rejection Reason  : Insurer's stated reason for declining or disputing
Evidence Dispute  : What evidence each party relied on
Outcome Category  : Controlled vocab (4 values)
Outcome           : Full FOS remedy instructions
Compensation Awarded (£) : Integer — D&I only; 0 if none
Is Core Case      : Controlled vocab (5 values)
Key Policy Clause : Policy wording or FOS/FCA principle applied
Missing Evidence  : Evidence that was absent and affected the outcome
Ombudsman Reasoning : How the ombudsman weighed the evidence
Workflow Insight  : Operational rule for the claims workflow
AI Rule Candidate : Machine-evaluable rule for the rules engine
Source PDF        : Filename only (e.g. DRN0001741.pdf)
===========================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Movement Cause",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 3: SUBS-021 to SUBS-030
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "SUBS-021",
        "FOS Decision ID": "DRN-4950435",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "15 Apr 2025",
        "Claim Type": "Home insurance — outbuilding crack damage recorded and cash settled by AXA as subsidence without physical inspection (loss adjuster based decision solely on consumer's photos/videos); consumer's stonemason and chartered structural engineer both concluded damage was not subsidence; AXA agreed to reclassify record from subsidence to accidental damage but disputed reimbursing engineer's report cost and recalculating premiums",
        "Movement Cause": "Not subsidence — outbuilding cracking attributed to accidental damage; initial subsidence classification made without physical inspection or adequate investigation; chartered structural engineer confirmed no subsidence following detailed assessment",
        "Property Type": "Residential home",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "AXA: claim was properly recorded as subsidence based on photos/videos; loss adjuster's video call failed but assessment was adequate; consumer received benefit of settlement (repairs made from cash); engineer report cost not reimbursable as consumer had already benefited; premiums correctly calculated at time of renewal",
        "Evidence Dispute": "AXA loss adjuster C: assessed via consumer's photos/videos only (video call failed; no physical inspection); brief summary report with no detailed rationale; recorded damage as subsidence. Miss B's stonemason: did not think damage caused by subsidence. Miss B's chartered structural engineer S (£696 report): physical inspection not possible (repairs done); detailed analysis of photos and other information; concluded damage was not subsidence and not covered by any policy peril. FOS: subsidence claims require physical inspection by appropriately qualified person — remote assessment from consumer photos only is inadequate; S's analysis was more persuasive and compelling than C's brief summary; weight of evidence firmly against subsidence as cause of damage",
        "Outcome Category": "Upheld",
        "Outcome": "Reclassify claim record from subsidence to accidental damage internally and externally; recalculate 2023 and 2024 premiums with accidental damage classification and refund any overpayment plus 8% simple interest; reimburse £696 for S's engineering advice; pay £400 compensation for distress and inconvenience",
        "Compensation Awarded (£)": 400,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Subsidence claims require physical inspection by an appropriately qualified person (chartered structural engineer, surveyor) — remote assessment based solely on consumer-provided photos/videos without physical inspection is inadequate and challengeable; incorrect subsidence recording causes lasting harm to premium costs and insurability that is independently compensable; where insurer's inadequate investigation causes consumer to commission independent engineering evidence, insurer must reimburse that cost; insurer that cash settles a non-covered claim as subsidence without adequate investigation cannot reclaim the settlement if consumer has spent funds on repairs; all premiums charged at a higher-risk (subsidence) rate must be reviewed and any overpayment refunded when the classification is corrected",
        "Missing Evidence": "C's detailed investigation rationale (only a brief summary was produced); physical inspection of damage before repairs were completed (not possible as repairs had already been carried out by the time claim was reconsidered)",
        "Ombudsman Reasoning": "Subsidence claims require a physically qualified inspection — C's assessment was based only on consumer's photos after a failed video call; C's report was brief with little explanation; S is a chartered structural engineer who provided detailed reasoning from available information; weight of evidence firmly against subsidence; AXA settled a non-covered claim without adequate investigation; Miss B had no realistic choice but to commission S's advice given AXA's inadequate investigation and failure to act on stonemason's concerns; AXA itself compounded its mistake by suggesting Miss B obtain engineering advice and then disputing the cost; no need to recover settlement as that was AXA's mistake and consumer used funds on outbuilding repairs; premiums must be recalculated",
        "Workflow Insight": "Initial subsidence classification made without physical inspection creates a risk that the claim is misrecorded with lasting consequences for the consumer (premium increases, difficulty obtaining insurance); once a consumer challenges a subsidence classification with professional evidence, insurer must genuinely review — failing to do so compels the consumer to incur further professional costs that become the insurer's liability; where a claim is wrongly recorded as subsidence without adequate investigation, insurer must correct the classification in all records and refund all consequential premium overpayments, but need not recover the original settlement payment if the consumer spent the funds on repairs",
        "AI Rule Candidate": "IF subsidence_classification_based_solely_on_consumer_photos_without_physical_inspection THEN investigation_is_inadequate_and_classification_is_challengeable; IF insurer_fails_to_act_on_consumer_challenge_to_subsidence_classification AND consumer_commissions_engineer_report THEN insurer_must_reimburse_engineer_costs; incorrect_subsidence_recording_must_be_corrected_internally_and_externally_in_all_systems; insurer_that_settles_uncovered_claim_as_subsidence_cannot_reclaim_settlement_if_consumer_spent_funds_on_repairs; IF claim_classification_corrected_from_subsidence_to_other_peril THEN all_premiums_at_higher_risk_rate_must_be_reviewed_and_overpayment_refunded",
        "Source PDF": "DRN-4950435.pdf",
    },
    {
        "Case ID": "SUBS-022",
        "FOS Decision ID": "DRN5217766",
        "Insurer Name": "St Andrew's Insurance Plc",
        "FOS Decision Date": "2014 (est.)",
        "Claim Type": "Home insurance — conservatory subsidence claim (June 2011); two St Andrew's surveyors confirmed subsidence since 2003 but estimated repair cost below £1,100 excess; initial cause identified as clay shrinkage from neighbour's hedge; soil sample taken only after FOS complaint showed hedge roots were dead and soil not desiccated — initial cause hypothesis disproven; dispute about repair scope and adequacy of investigation",
        "Movement Cause": "Subsidence to conservatory on shallow foundations; initial cause (clay shrinkage from neighbour's hedge) disproven by soil analysis (roots dead, soil not desiccated); actual current cause undetermined at time of decision — requires proper investigation",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "St Andrew's: both surveyors confirmed subsidence since 2003; proposed limited repair scope (flashing, wall section, mastic fill) is adequate; estimated cost below £1,100 excess; historic settlement predates cover and St Andrew's not liable for demolish/rebuild; betterment principle limits liability for deeper foundations",
        "Evidence Dispute": "St Andrew's surveyors: subsidence from hedge since 2003; limited repair scope adequate and cost below excess; deeper foundations would be betterment. St Andrew's surveyor (later, after soil sample): roots dead, soil not desiccated; cause may be shrinkage of made and deeper soils aggravated by vegetation — cause uncertain. Mrs A: conservatory continues to move; gap between conservatory and main house widening; contractors would not quote for proposed limited works as they said it would not solve the problem; 2000 valuation report showed no subsidence at time of purchase. FOS: soil sample arranged during FOS investigation (not proactively by St Andrew's) showed hedge was not the cause; St Andrew's surveyor acknowledged cause is now uncertain; without identifying cause, St Andrew's cannot confirm what scope of works is adequate or that repair cost is below the excess; insurer must carry out a proper investigation before asserting works are sufficient",
        "Outcome Category": "Upheld",
        "Outcome": "St Andrew's Insurance Plc to carry out appropriate investigation to determine the cause of the current conservatory subsidence; then reassess what works need to be carried out to ensure an effective and lasting repair; pay Mrs A £500 compensation for distress and inconvenience from St Andrew's failure to properly investigate",
        "Compensation Awarded (£)": 500,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Insurer must identify the cause of subsidence before determining the scope of repair works — without knowing the cause, it cannot establish what constitutes a lasting and effective repair or confirm repair cost is below the policy excess; where initial cause hypothesis is disproven by investigation, insurer must reinvestigate rather than defaulting to the original repair plan; obligation to carry out a lasting and effective repair may require demolish and rebuild if investigation shows it is necessary — betterment argument only limits what insurer funds for improvements beyond what is needed for an effective repair; argument that repair cost is below excess cannot be used to avoid investigation if the effectiveness of proposed repairs is uncertain; delay in investigating ongoing movement that consumer has repeatedly reported causes compensable distress",
        "Missing Evidence": "Root identification and soil desiccation data at time of 2011 inspection (soil sample only taken years later during FOS investigation, not arranged proactively); monitoring data to confirm whether conservatory movement is ongoing and its current rate",
        "Ombudsman Reasoning": "Both St Andrew's surveyors confirmed subsidence since 2003; dispute was about scope of repair and cause; soil sample (arranged only after FOS complaint) showed hedge roots were dead and soil not desiccated — hedge was not the cause; St Andrew's own surveyor acknowledged this and conceded cause is uncertain; without knowing cause, St Andrew's cannot determine extent of liability or confirm proposed works are adequate; investigation could and should have been arranged years earlier; Mrs A witnessed continuing movement throughout; £500 appropriate for significant distress from failure to properly investigate over an extended period",
        "Workflow Insight": "Insurer investigating conservatory or structure subsidence must confirm the cause with adequate investigation before making any claim about repair scope or cost — assertions that repair cost is below excess are premature if cause is unknown; where cause investigation disproves initial field hypothesis (e.g. soil sample shows hedge roots dead and soil not desiccated), insurer must arrange further investigation proactively, not only when forced to do so by FOS complaint; consumer evidence of continuing movement must be taken seriously — if movement is ongoing and cause is unconfirmed, the insurer's obligation to carry out a lasting and effective repair is engaged and the excess argument cannot be relied upon",
        "AI Rule Candidate": "IF cause_of_subsidence_not_identified THEN insurer_cannot_determine_scope_of_repair_or_confirm_cost_is_below_excess; IF initial_cause_hypothesis_disproven_by_investigation THEN further_investigation_required_before_repair_scope_finalised; IF consumer_reports_ongoing_movement AND insurer_has_not_monitored THEN insurer_must_arrange_investigation_and_monitoring; delay_in_investigating_ongoing_movement_reported_by_consumer_causes_compensable_distress; excess_argument_cannot_be_relied_upon_when_cause_unknown_and_repair_effectiveness_is_uncertain",
        "Source PDF": "DRN5217766.pdf",
    },
    {
        "Case ID": "SUBS-023",
        "FOS Decision ID": "DRN-5220010",
        "Insurer Name": "U K Insurance Limited",
        "FOS Decision Date": "15 Jan 2025",
        "Claim Type": "Landlord property insurance — garden/patio landslip caused by removal of adjacent retaining land by commercial construction; UKI's expert confirmed no subsidence to main building and identified cause as landslip; garden/patio damage excluded as yard; at 2023 renewal UKI removed subsidence cover citing underwriting criteria for properties showing signs of or suffering subsidence; consumer unable to obtain subsidence cover elsewhere",
        "Movement Cause": "Landslip and ground movement caused by removal of retaining land along shared boundary during commercial multi-storey building development; no subsidence to main building foundations identified; movement is downward slope from loss of lateral support, not clay shrinkage subsidence",
        "Property Type": "Residential let property (landlord insurance — Churchill/UKI)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "UKI: (1) claim — main building undamaged; garden/patio damage excluded as yard under subsidence section of policy. (2) Cover removal — risk address showed signs of subsidence or had suffered subsidence per underwriting criteria; underwriting team confirmed subsidence cover could not be offered",
        "Evidence Dispute": "UKI's own expert: movement in garden is landslip from removal of retaining land — not subsidence to main building; recommended low-level retaining wall. Consumer: cannot obtain subsidence cover elsewhere; UKI's expert identified the exact problem. FOS investigator: ABI continuation guidance applied; UKI should stand by its own expert's findings. FOS ombudsman: ABI guidance not strictly relevant (claim declined; no repairs undertaken); key issue is whether UKI correctly applied its underwriting criteria — underwriting guide specifies 'subsidence' not 'landslip'; UKI's own expert found landslip not subsidence; policy separately defines subsidence, ground heave and landslip as distinct events; UKI has not demonstrated it applied its criteria correctly to the identified peril",
        "Outcome Category": "Upheld",
        "Outcome": "UKI to reinstate and backdate subsidence cover for the 2023–2024 policy period, subject to Mrs and Mr A paying any relevant additional premium; pay £150 compensation for unnecessary inconvenience",
        "Compensation Awarded (£)": 150,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Insurer's underwriting criteria must be correctly applied to each identified peril — where criteria specifically address 'subsidence' and not the broader SHL perils (subsidence, heave, landslip), insurer cannot rely on those criteria to remove subsidence cover when its own expert identified the peril as landslip; subsidence and landslip are distinct perils (both in fact and under policy terms that list them separately) and must be treated accordingly in underwriting decisions; insurer whose own investigation identified the cause as landslip is bound by that finding and cannot invoke subsidence criteria to remove subsidence cover; the plain text of an underwriting guide is the primary indicator of what action was authorised — insurer's assertion of a broader interpretation is insufficient",
        "Missing Evidence": "Whether UKI's underwriting guide's reference to 'subsidence' was intended to encompass all three SHL perils or only subsidence specifically (UKI asserted broader scope; FOS found plain text did not support this)",
        "Ombudsman Reasoning": "UKI's own expert confirmed landslip not subsidence; UKI's underwriting guide says subsidence cover declined for signs of or suffering subsidence — not for landslip; policy terms list three distinct events (subsidence, ground heave, landslip) confirming they are separate; UKI's argument that 'subsidence' in the guide meant all three perils is not supported by plain text; UKI has not demonstrated it applied its criteria correctly; subsidence cover must be reinstated; £150 for inconvenience of being unable to obtain subsidence cover elsewhere",
        "Workflow Insight": "When an insurer's own expert investigation identifies the damage as landslip rather than subsidence, that finding binds the insurer's subsequent underwriting decisions about each named peril — the insurer cannot then invoke subsidence criteria to remove subsidence cover; underwriting guides that use the word 'subsidence' do not automatically extend to landslip or heave unless the guide explicitly says so; each of the three SHL perils must be assessed independently in underwriting and claims decisions; the ABI continuation guidance may not apply when a claim has been declined and no repairs undertaken, but the insurer must still correctly apply its own criteria to each specific peril",
        "AI Rule Candidate": "IF insurers_expert_identifies_peril_as_landslip_not_subsidence THEN insurer_cannot_invoke_subsidence_underwriting_criteria_to_remove_subsidence_cover; subsidence_and_landslip_are_distinct_perils_that_must_be_assessed_independently_in_underwriting_and_claims; IF underwriting_guide_refers_to_subsidence_only THEN it_does_not_authorise_action_based_on_landslip_or_heave_unless_expressly_stated; insurer_is_bound_by_the_plain_text_of_its_underwriting_guide_broader_interpretation_requires_explicit_textual_support",
        "Source PDF": "DRN-5220010.pdf",
    },
    {
        "Case ID": "SUBS-024",
        "FOS Decision ID": "DRN-5315100",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "9 Jul 2025",
        "Claim Type": "Home insurance — subsidence claim accepted February 2021 (tree root clay soil shrinkage; vegetation removed); multi-year dispute about ongoing movement monitoring, drain repair coverage, AXA removing subsidence cover at renewal (admitted breach of ABI guidance), cash settlement management fee deduction, and delay and poor communication over approximately three years; claim still not at repair stage at date of decision",
        "Movement Cause": "Tree root-influenced clay soil shrinkage (vegetation removed 2021 and further tree identified January 2025 and removed); drain damage also identified but AXA disputes connection to subsidence; movement ongoing during multi-year claim period",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "AXA: property stabilised post-vegetation removal; damaged drains not caused by or contributing to subsidence (cyclical not progressive movement — not drain-related subsidence); management fee deduction from cash settlement in line with policy terms (payment capped at what AXA would pay preferred supplier, excluding management coordination costs); subsidence cover removal situation under investigation",
        "Evidence Dispute": "Mrs R: consumer-commissioned monitoring shows continued movement; surveyor says drains could be a cause; chip board flooring damaged by subsidence. AXA: cyclical not progressive movement — not drain-related; purchase survey showed flooring distortion predating subsidence; management fee is a non-repair cost excluded from cash settlement. FOS (provisional): further monitoring needed; drain reimbursement requires specific expert evidence not yet available; £1,500 excess management fee deduction is fair per policy terms; £1,500 compensation for multi-year delay. FOS (final): AXA accepted breach of ABI continuation guidance; Jan 2025 neighbour's tree removed; ongoing monitoring continuing; floor damage not conclusively subsidence-related; drain expert evidence not provided — AXA's position (cyclical movement) accepted in absence of contradicting evidence; three years continuing cover from final COSA date directed",
        "Outcome Category": "Upheld",
        "Outcome": "AXA to arrange subsidence cover on reasonable terms (as part of full buildings insurance policy) for a minimum of three years from the date of issue of a final Certificate of Structural Adequacy (COSA); pay Mrs R £1,500 compensation for distress and inconvenience from multi-year delay",
        "Compensation Awarded (£)": 1500,
        "Is Core Case": "Yes",
        "Key Policy Clause": "ABI continuation guidance on subsidence cover extends beyond claim conclusion and the repair period — the obligation to provide continuing cover applies after a final COSA is issued (post-repair), not merely until the claim is administratively closed; minimum three years cover from final COSA provides assurance to the housing market and to the consumer without creating an unlimited obligation; insurer removing cover at claim conclusion acts inconsistently with the guidance's purpose; cash settlement for subsidence repairs is capped at what insurer would pay its own preferred supplier — management/coordination costs are non-repair costs and are legitimately excluded; expert evidence specific to the dispute is required before FOS will direct reimbursement of drain costs — consumer's surveyor suggestion is insufficient; insurer dispute about repair method (underpinning vs resin injection) is not within FOS competence to direct — insurer's chosen method is acceptable if effective",
        "Missing Evidence": "Expert evidence specifically linking drain damage to the subsidence movement (not provided — FOS could not direct drain cost reimbursement); confirmation of whether chip board flooring distortion predated subsidence or was caused by it (purchase survey noted pre-existing distortion); AXA's underwriter explanation for subsidence cover removal (never provided — AXA ultimately accepted breach)",
        "Ombudsman Reasoning": "AXA accepted approximately seven months of attributable delay in the period reviewed; £1,500 compensation appropriate given Mrs R's health circumstances and her son's health, and the extended claim life; ABI continuation guidance 'after the repair is effected' language means cover continues post-repair, not that cover ceases at claim closure; three years from final COSA is proportionate and not unreasonably open-ended; management fee: policy says payment will not exceed amount AXA would pay preferred supplier — management fee is excluded as a non-repair cost, and £1,500 deduction is fair; floor distortion: purchase survey noted pre-existing unevenness — insufficient evidence subsidence worsened beyond that; drains: AXA's position (cyclical not progressive movement) accepted in absence of contradicting expert evidence from consumer",
        "Workflow Insight": "ABI continuation guidance must be read as extending beyond claim closure to the post-repair period — insurer removing subsidence cover when a claim concludes acts against the guidance's purpose; minimum cover period of three years from issue of final COSA is a proportionate industry-consistent obligation; cash settlement for subsidence repairs may legitimately exclude insurer management and coordination fees as non-repair costs within the policy's cash settlement cap; consumer must produce specific expert evidence linking a disputed damage element (drains, flooring) to subsidence before FOS will direct reimbursement — a surveyor's unsubstantiated suggestion is not sufficient",
        "AI Rule Candidate": "ABI_continuation_guidance_extends_beyond_claim_closure_to_post_repair_period; minimum_three_years_cover_from_final_COSA_is_proportionate_continuing_cover_obligation; cash_settlement_for_subsidence_repairs_may_exclude_insurer_management_fees_as_non_repair_costs; IF consumer_claims_drain_or_other_element_is_subsidence_related THEN specific_expert_evidence_required_before_FOS_will_direct_reimbursement; IF insurer_removes_subsidence_cover_at_claim_conclusion THEN removal_is_inconsistent_with_ABI_continuation_guidance",
        "Source PDF": "DRN-5315100.pdf",
    },
    {
        "Case ID": "SUBS-025",
        "FOS Decision ID": "DRN-5375880",
        "Insurer Name": "Highway Insurance Company Limited",
        "FOS Decision Date": "23 Apr 2025",
        "Claim Type": "Buildings insurance — new cracking August 2022; Highway avoided policy citing careless misrepresentation on 'Is the property and surrounding area free from subsidence?' answered 'Yes'; property had a prior 2003 subsidence claim (tree roots; same beech tree implicated in both incidents) but repairs carried out then and no further signs for nearly 20 years; beech tree is local authority property and cannot be removed",
        "Movement Cause": "Tree root-induced subsidence (same local authority beech tree identified as cause of both 2003 and 2022 incidents); 2003 repairs included root barrier, drain replacement, and wall stitching; no further movement observed for nearly 20 years between the two incidents",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Highway: property not 'free from subsidence' because the tree identified as the cause of both incidents was present and unremoved at inception; 2003 stitch repairs stabilised cracking but did not remove the cause of the subsidence; consumer should have answered 'No' to the question; careless misrepresentation at inception entitles avoidance under CIDRA 2012",
        "Evidence Dispute": "Mr and Mrs D: 2003 repairs were substantial (root barrier, drain replacement, wall stitching); no further subsidence signs for nearly 20 years; reasonably believed subsidence had been resolved; question did not ask about previous subsidence claims; tree is local authority property — cannot be removed. Highway: stitch repairs only stabilised not resolved the cause; tree still present means property is not free from subsidence; should have answered 'No.' FOS: question does not define 'property' or 'surrounding area'; no timeframe or geographic scope given; no guidance on how to answer if property was previously affected; question did not ask about prior subsidence claims specifically; consumer who carried out substantial repairs in 2003 and experienced no further movement for nearly 20 years reasonably believed the subsidence was resolved; presence of an unremovable local authority tree does not automatically mean the property is currently 'affected by subsidence'; CIDRA reasonable care test met",
        "Outcome Category": "Upheld",
        "Outcome": "Highway to reinstate policy and deal with the subsidence claim in line with remaining policy terms and conditions; consumer may need to repay any premiums that were previously refunded following avoidance",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "CIDRA 2012 — consumer must take reasonable care not to misrepresent; where a question about subsidence does not define the geographic scope of 'property and surrounding area', provide a timeframe, or give guidance for consumers who previously experienced subsidence, the insurer cannot establish that reasonable care was not taken; consumer who carried out substantial post-subsidence repairs and experienced no further movement for nearly 20 years reasonably believes the subsidence is resolved; presence of an unremovable third-party tree does not automatically mean the property is currently affected by subsidence; insurer cannot extend the meaning of its vague question beyond what a reasonable consumer would understand it to mean; statement of fact that does not include a prior claims question cannot be relied upon to show the consumer should have disclosed prior subsidence history",
        "Missing Evidence": "Whether the statement of fact specifically asked about prior subsidence claims (FOS found it did not); precise intended scope of 'property and surrounding area free from subsidence' (insurer failed to define this at time of application)",
        "Ombudsman Reasoning": "2003 repairs were substantial — root barrier, drains replaced, walls stitched; no further damage for nearly 20 years; consumers reasonably believed subsidence was resolved; statement of fact did not ask about prior subsidence claims; question 'free from subsidence' has no defined scope, timeframe or guidance for previously affected properties; Highway did not prove the question was asked in a way that meant the consumer should have answered 'No'; presence of unremovable local authority tree does not automatically mean property is currently affected; CIDRA reasonable care test met; avoidance is not available to Highway",
        "Workflow Insight": "Insurer relying on CIDRA to avoid a policy for non-disclosure of prior subsidence must show that the question asked was clear enough that a reasonable consumer would know the correct answer — a question with no geographic scope, no timeframe and no guidance for previously affected properties is too vague to support an avoidance; consumer who completed substantial post-subsidence repairs and experienced no further movement for nearly 20 years has reasonable grounds to believe the subsidence was resolved; insurer cannot treat the continued presence of an unremovable third-party tree as conclusive evidence that the property is 'affected by subsidence' for disclosure purposes",
        "AI Rule Candidate": "CIDRA_reasonable_care_test_requires_question_to_be_clear_enough_that_consumer_would_know_correct_answer; IF question_about_subsidence_has_no_defined_scope_or_timeframe THEN insurer_cannot_establish_consumer_failed_reasonable_care_standard; IF substantial_post_subsidence_repairs_carried_out AND no_further_movement_for_twenty_years THEN consumer_reasonably_believes_subsidence_is_resolved; presence_of_unremovable_third_party_tree_does_not_automatically_mean_property_is_currently_affected_by_subsidence; statement_of_fact_without_prior_claims_question_cannot_support_non_disclosure_of_prior_subsidence_history",
        "Source PDF": "DRN-5375880.pdf",
    },
    {
        "Case ID": "SUBS-026",
        "FOS Decision ID": "DRN-5643066",
        "Insurer Name": "Ageas Insurance Limited",
        "FOS Decision Date": "12 Nov 2025",
        "Claim Type": "Landlord insurance — subsidence discovered August 2022 (policy inception October 2021); Ageas initially accepted claim; reversed and declined in August 2024 on grounds damage predated inception; relied on September 2021 tenancy check-in report (uPVC doors could not fully close) and 2015 survey noting signs of settlement; window specialist confirmed door issue was unrelated to subsidence; ongoing movement confirmed at August 2023 revisit",
        "Movement Cause": "Subsidence caused by effect of nearby vegetation on clay soil (identified by loss adjuster October 2022 — moderate to severe cracking internally and externally); new damage areas observed at August 2023 revisit confirming ongoing movement during policy period",
        "Property Type": "Residential let property (landlord insurance; limited company policyholder — FOS for small businesses jurisdiction)",
        "Dispute Type": "Pre-Inception Damage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Ageas (first response): cracking appeared prior to inception; 2015 survey showed signs of settlement/subsidence not disclosed on Statement of Facts. Ageas (second response): September 2021 check-in report showed two uPVC doors could not close fully — evidence of pre-inception subsidence; claim predated start of cover",
        "Evidence Dispute": "Loss adjuster (October 2022): moderate to severe cracking; damage arose during currency of policy; new damage areas at August 2023 revisit. Check-in report (September 2021): 43 pages — no mention of structural cracking; two uPVC doors noted as not closing fully. Window/conservatory repair specialist: uPVC doors unaffected by subsidence; frames straight and flush; door catching is weather-related and needs servicing/adjustment. Decorator's quote (September 2021): mentioned filling cracks before letting but no structural concerns raised. FOS: check-in report does not mention cracks that were later observed; uPVC door issue confirmed unrelated to subsidence by specialist; decorator did not raise structural concerns; loss adjuster confirmed insurable interest and damage during policy currency; some pre-policy cracks possible but Ageas cannot differentiate between pre- and post-inception subsidence damage",
        "Outcome Category": "Upheld",
        "Outcome": "Ageas to deal with subsidence claim in line with policy terms; assess claim for loss of rental income; reinstate policy if cancelled; pay P £800 compensation for impact on rental income and inability to let the property",
        "Compensation Awarded (£)": 800,
        "Is Core Case": "Yes",
        "Key Policy Clause": "For a subsidence claim decline to succeed entirely on pre-existing damage grounds, insurer must show that the subsidence damage both started and stopped before policy inception — if subsidence began before inception but continued during the policy period, insurer must deal with the claim; where insurer cannot differentiate between pre-inception subsidence damage and post-inception damage, FOS requires insurer to deal with all the subsidence damage; an insurer that accepts a claim at first assessment and later reverses that acceptance on the same facts must show clearly that the reversal is justified; insurer that affirms a contract by issuing a second final response focused solely on a new ground (uPVC doors) may be treated as having abandoned the earlier ground (IA 2015 fair presentation breach) and cannot later revive it",
        "Missing Evidence": "Pre-policy structural survey specifically recording cracking extent (none available — no pre-policy structural survey of crack damage existed); whether the 2015 survey settlement/subsidence findings represented active ongoing subsidence or a resolved historic issue",
        "Ombudsman Reasoning": "Check-in report (43 pages) does not mention any of the cracks later identified; uPVC door issue confirmed unrelated to subsidence by specialist — frames straight and flush; decorator quote mentioned filling cracks without raising structural concerns; loss adjuster confirmed claim arose during policy currency and moderate-to-severe cracking was found; new damage areas at 2023 revisit confirm ongoing movement during policy; some pre-policy cracks may exist but Ageas cannot differentiate — must deal with all subsidence damage; Ageas appeared to have accepted the fair presentation explanation by issuing second final response focused only on uPVC doors; £800 for impact on rental income and inability to let the property",
        "Workflow Insight": "Insurer relying on pre-inception damage to decline a subsidence claim must show not only that some damage may have existed before inception but also that the current damage is entirely attributable to pre-inception movement — if subsidence was ongoing at inception and continued during the policy, the insurer must deal with the claim; where specific evidence (check-in report, decorator's quote) does not in fact show structural cracking, it cannot support a pre-inception damage exclusion argument; a specialist confirmation that an alleged sign of subsidence (uPVC door catching) is in fact weather-related is binding on the insurer unless contradicted by other expert evidence; insurer that issues a second final response focused on a new ground may be taken to have affirmed the contract on the original ground",
        "AI Rule Candidate": "IF subsidence_began_before_inception_but_continued_during_policy THEN insurer_must_deal_with_claim; IF insurer_cannot_differentiate_pre_and_post_inception_subsidence_damage THEN insurer_must_deal_with_all_damage; alleged_sign_of_subsidence_confirmed_by_specialist_as_unrelated_to_subsidence_cannot_support_pre_inception_exclusion; IF insurer_issues_second_final_response_focused_on_new_ground THEN may_be_taken_as_affirmation_abandoning_first_ground; pre_inception_damage_exclusion_requires_evidence_damage_started_and_stopped_before_inception",
        "Source PDF": "DRN-5643066.pdf",
    },
    {
        "Case ID": "SUBS-027",
        "FOS Decision ID": "DRN-5656370",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "30 Jul 2025",
        "Claim Type": "Property protection insurance (block of flats) — subsidence claim accepted 2020; AXA increased premium by 110% at 2023 renewal and removed subsidence cover when its commercial arrangement with intermediary T ended and it applied different underwriting criteria; policyholder C unable to access open market due to ongoing claim; excess increase from £1,500 to £5,000 also proposed without supporting evidence",
        "Movement Cause": "Subsidence (cause not specified — accepted 2020; claim ongoing and not resolved at time of decision)",
        "Property Type": "Block of flats (property protection insurance; commercial policy — FOS for small businesses jurisdiction; FOS applied good industry practice principles equivalent to domestic guidance)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "AXA: 2021 and 2022 premium increases reflected general pricing factors; 2023 premium reflects new underwriting criteria applied when direct policy offered following end of arrangement with T; subsidence cover cannot be offered as claim is ongoing and property stability unconfirmed; at 2023 renewal AXA was applying different (higher risk) criteria than those used under the T arrangement",
        "Evidence Dispute": "C: some increase expected but 110% while also removing subsidence cover is excessive and unfair; trapped with AXA due to ongoing claim and unable to move to another insurer. AXA (initial): limited evidence for increases. AXA (later): 2021/22 increases due to inflation, index linking, general risk changes — not claim-driven; subsidence cover removal due to claim being ongoing. FOS: 2021/22 increases fair (market-consistent, not claim-driven, subsidence cover maintained); 2023 increase unfair — AXA switched from T-arrangement criteria to its own higher-risk criteria without considering C's inability to access the open market; AXA cannot use its own unresolved claim to justify removing cover it is handling; excess increase to £5,000 proposed without underwriting rationale or evidence — rejected",
        "Outcome Category": "Upheld",
        "Outcome": "AXA to reduce 2023 premium to £6,500 (approximately 20% increase consistent with 2021–22 approach) and refund difference plus 8% simple interest; reduce 2024 premium to £6,825 and refund difference plus 8% simple interest; reinstate subsidence cover for 2023 and 2024 without additional premium or excess increase beyond £1,500; pay £200 compensation for prolonged inconvenience",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Insurer that changes its underwriting criteria because of a change in commercial arrangements (e.g. end of intermediary agreement) must consider the impact on policyholders who cannot access the open market due to an ongoing claim; a policyholder with an open subsidence claim is effectively locked in to the incumbent insurer and cannot obtain alternative cover on normal terms — the insurer cannot exploit that lock-in by applying materially different criteria; insurer handling a subsidence claim cannot remove subsidence cover on grounds that it has not resolved its own claim — the insurer's own handling failure cannot be used to disadvantage the policyholder; excess increases proposed without supporting underwriting evidence or rationale will not be accepted by FOS",
        "Missing Evidence": "AXA's calculation of what 2023 and 2024 premiums would be under the 2021–22 approach (AXA never provided this despite multiple opportunities — FOS derived the 20% figure from the two prior year average)",
        "Ombudsman Reasoning": "2021 and 2022 increases market-consistent (inflation, index-linking, general risk changes) — not claim-driven; subsidence cover maintained; these increases were fair. 2023 increase of 110% — consequence of end of T arrangement and application of different higher-risk criteria; AXA should have considered C's inability to access open market; increase is unfair. Subsidence cover removal — AXA, as claim handler, cannot say its own unresolved claim makes the risk too great to cover; AXA itself offered reinstatement in earlier discussions. Excess increase to £5,000 proposed with no supporting evidence or underwriting rationale — rejected. £200 compensation for prolonged inconvenience",
        "Workflow Insight": "When an insurer's underwriting criteria change because of a change in commercial or distribution arrangements, it must consider the impact on policyholders who are locked in because of ongoing claims — applying substantially harsher criteria to a policyholder who has no realistic alternative is unfair; insurer handling a claim cannot use its own delay or non-resolution of the claim as justification for removing cover that it was providing — a claim that remains open is not a reason to remove the very cover that applies to the claim; excess or premium increases proposed without underwriting evidence will be disregarded by FOS regardless of whether the increase itself may be technically justified",
        "AI Rule Candidate": "IF insurer_changes_criteria_due_to_change_in_commercial_arrangements AND policyholder_cannot_access_open_market_due_to_ongoing_claim THEN applying_materially_harsher_criteria_is_unfair; IF insurer_handling_subsidence_claim_removes_subsidence_cover_citing_unresolved_claim THEN removal_is_unfair_as_insurer_cannot_use_own_handling_failure_to_disadvantage_policyholder; excess_increases_without_supporting_underwriting_evidence_or_rationale_will_be_rejected_by_FOS; premium_increases_must_be_consistent_with_market_approach_not_driven_by_change_in_commercial_distribution_arrangement",
        "Source PDF": "DRN-5656370.pdf",
    },
    {
        "Case ID": "SUBS-028",
        "FOS Decision ID": "DRN-5718419",
        "Insurer Name": "West Bay Insurance Plc",
        "FOS Decision Date": "25 Dec 2025",
        "Claim Type": "Property owners insurance (block of leasehold flats) — subsidence claim accepted 2018; West Bay withdrew its property owners policy product at 2019 renewal and subsequently withdrew from the buildings insurance market entirely; T (limited company managing the building) left without subsidence cover from 2019 to 2025 (approximately six years) until a Certificate of Structural Adequacy was issued; West Bay took no steps to arrange alternative cover or mitigate impact",
        "Movement Cause": "Subsidence (cause not specified — accepted claim 2018; repairs completed early 2023; final COSA issued 2025)",
        "Property Type": "Block of leasehold flats (property owners policy; limited company managing building — FOS considered distress of individual leaseholders within the company)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "West Bay: withdrew property owners policy product at 2019 renewal as a commercial decision; subsequently withdrew from buildings insurance market entirely; obligation to offer continuing cover applies but was not possible due to market withdrawal; extension of cover would require maintaining a product withdrawn from all policyholders of that type",
        "Evidence Dispute": "T: West Bay should have taken steps to ensure continuing subsidence cover was available; without accepted claim insurer's assistance, finding alternative cover with subsidence inclusion is extremely difficult or impossible; six years without subsidence cover caused significant concern and administrative burden. West Bay: market withdrawal is a commercial necessity applicable to all such policyholders; T was informed and had time to seek alternative cover; £1,000 compensation offered. FOS: West Bay acknowledged the general continuing cover obligation; should have explored options (similar product, maintaining cover until claim complete, arranging block transfer to another insurer); West Bay took no steps beyond not renewing; no new subsidence claims arose during gap period; financial loss from premiums unlikely (subsidence claim would have meant substantially higher premiums for any cover obtained); £1,000 compensation appropriate for six years of concern and administrative burden",
        "Outcome Category": "Upheld",
        "Outcome": "West Bay Insurance Plc to pay £1,000 compensation for six years of distress and inconvenience caused by being left without subsidence cover",
        "Compensation Awarded (£)": 1000,
        "Is Core Case": "Yes",
        "Key Policy Clause": "When an insurer withdraws its product or exits the market, it must consider the likely detriment to policyholders with ongoing or recently settled subsidence claims and take steps to mitigate it — acceptable steps include offering a similar product, maintaining cover until the claim is complete, or arranging for another insurer to step in; this obligation applies regardless of whether the withdrawal applies to all policyholders of that product type or just one; failure to explore any of these options before withdrawing leaves policyholders without cover, contrary to good industry practice; where no new subsidence claims arise during the period without cover, financial loss is unlikely; compensation for distress and administrative inconvenience is the primary remedy",
        "Missing Evidence": "What steps West Bay could have taken to arrange alternative cover or a block transfer for T (never explored by West Bay); whether a block transfer to another insurer was feasible given the market at that time",
        "Ombudsman Reasoning": "West Bay was aware of the continuing cover obligation; took no steps to mitigate the impact on T; T unable to find subsidence cover elsewhere for six years; no new subsidence claims arose during the gap — financial loss is unlikely; higher premiums would have been likely if subsidence cover had been obtained; £1,000 fair and in line with FOS approach for six years of concern and administrative burden",
        "Workflow Insight": "When an insurer withdraws a product that covers policyholders with active or recently settled subsidence claims, it must proactively explore options to protect those policyholders before the cover lapses — simply not renewing the policy is insufficient; acceptable measures include maintaining the specific policy until the claim concludes, arranging a block transfer to another insurer, or offering a comparable product; insurer must actively investigate whether alternatives are possible before concluding that market withdrawal makes continuing cover impossible; where no financial loss results from the period without cover, compensation for distress and inconvenience remains the appropriate and primary remedy",
        "AI Rule Candidate": "IF insurer_withdraws_product_or_exits_market AND policyholder_has_open_or_recent_subsidence_claim THEN insurer_must_explore_steps_to_mitigate_detriment_before_cover_lapses; acceptable_mitigation_steps_include_maintaining_policy_until_claim_complete_OR_arranging_block_transfer_OR_offering_comparable_product; simply_not_renewing_the_policy_is_not_adequate_mitigation; IF_no_new_claims_arise_during_gap_period THEN_financial_loss_is_unlikely_but_distress_and_inconvenience_compensation_remains_appropriate; obligation_applies_regardless_of_whether_withdrawal_affects_all_policyholders_of_that_product_type",
        "Source PDF": "DRN-5718419.pdf",
    },
    {
        "Case ID": "SUBS-029",
        "FOS Decision ID": "DRN-5755602",
        "Insurer Name": "Saga Services Limited",
        "FOS Decision Date": "12 Sep 2025",
        "Claim Type": "Buildings insurance — Saga (broker) moved Mrs C's policy from insurer A to insurer B at renewal without pausing to consider a subsidence issue Mrs C had reported three days before renewal; insurer A subsequently had no continuing cover obligation; insurer B removed subsidence cover when it discovered the claim at renewal three; Mrs C left without subsidence cover despite an ongoing subsidence claim with insurer A",
        "Movement Cause": "Subsidence (cause not specified — ongoing claim with insurer A; new episodes would not be covered under current policy with C)",
        "Property Type": "Residential home",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": "Saga: first record of Mrs C's subsidence issue was after renewal two (contradicted by call transcript); no internal process to contact insurer A or prevent policy moving; cannot force panel insurers to provide cover; at renewal three adequately disclosed the subsidence cover exclusion",
        "Evidence Dispute": "Mrs C: told Saga about cracks three days before renewal two (call 12 December); Saga should have paused the renewal and contacted insurer A. Saga: no record of subsidence call before renewal two; no process for involving A. FOS: call transcript confirms Mrs C clearly reported cracks on 12 December before renewal two and was referred to insurer A; Saga had a working relationship with A at renewal two and should have paused the renewal to notify A; insurer A would more likely than not have renewed given industry continuing cover guidance; Saga's inaction broke the chain of cover; at renewal three Consumer Duty required much clearer communication than a generic reference to excluded perils — Saga's own later letter acknowledged its communication at renewal three was inadequate; £2,000 compensation",
        "Outcome Category": "Upheld",
        "Outcome": "Saga Services Limited to pay Mrs C £2,000 compensation for distress and inconvenience from loss of subsidence cover and ongoing difficulty obtaining replacement cover",
        "Compensation Awarded (£)": 2000,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": "Broker who is notified by a consumer of a subsidence issue (crack report) before a renewal must take action to protect the consumer's continuing subsidence cover — the broker must not allow the policy to move to a new insurer without flagging the issue to the current insurer; broker's internal process limitations or panel structure are not an excuse for inaction that causes consumer harm; Consumer Duty (and ICOBS 8 equivalent) requires brokers to ensure consumers understand significant coverage limitations — a generic reference to 'exclusions may apply' is insufficient to notify a consumer that subsidence cover has been removed; once a broker's inaction breaks the chain of continuing cover, the consumer loses the protection that industry guidance was intended to preserve",
        "Missing Evidence": "Whether insurer A would specifically have renewed Mrs C's policy had Saga contacted them (FOS found it more likely than not A would have renewed given industry continuation guidance); full extent of financial loss from broken cover chain (could not be fully quantified given involvement of insurer A's future decisions)",
        "Ombudsman Reasoning": "Call transcript contradicts Saga's claim that no prior record existed — Mrs C clearly reported cracks on 12 December before renewal two and was referred to insurer A; Saga had a working relationship with A at that time and should have paused the renewal; A would more likely than not have renewed given industry continuing cover guidance; Saga's inaction broke the chain — A had no obligation at renewal three; B was a different insurer with no continuing cover obligation; Saga also failed at renewal three under Consumer Duty (generic exclusion reference plus no follow-up call was inadequate); Saga's own letter acknowledging inadequate communication at renewal three is an admission; £2,000 appropriate for distress and ongoing difficulty obtaining subsidence cover",
        "Workflow Insight": "Broker who learns of a subsidence issue from a consumer before renewal must pause the renewal process and contact the incumbent insurer — allowing the policy to roll to a new insurer automatically when a subsidence issue has been reported is a breach of the broker's duty to act in the consumer's best interests; the broker's internal workflow or panel limitations do not override this duty; when subsidence cover is excluded from a new policy as a consequence of the broker's action or inaction, the consumer must be given specific and prominent notification of that exclusion in plain terms, not a generic reference to possible exclusions",
        "AI Rule Candidate": "IF broker_notified_of_subsidence_issue_before_renewal THEN broker_must_pause_renewal_and_contact_incumbent_insurer; broker_internal_process_limitations_do_not_excuse_inaction_causing_consumer_harm; IF broker_inaction_breaks_chain_of_continuing_subsidence_cover THEN broker_is_liable_for_D_and_I_and_consequential_cover_loss; consumer_duty_requires_specific_prominent_notification_of_subsidence_cover_exclusion_not_generic_reference_to_excluded_perils",
        "Source PDF": "DRN-5755602.pdf",
    },
    {
        "Case ID": "SUBS-030",
        "FOS Decision ID": "DRN-6019596",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "11 Jan 2026",
        "Claim Type": "Landlord HMO buildings insurance (AXA PrimeLet) — subsidence claim January 2023; AXA wrongly told Ms L in July 2024 that damage to the front of her property was not subsidence-related; by August 2024 AXA knew the front was a separate subsidence incident from the rear (different cause) but never set up a new claim; avoidable delay of several months and poor communication throughout; consumer commissioned independent structural engineer (£600) after receiving inconsistent explanations",
        "Movement Cause": "Two separate subsidence incidents: rear of property — defective drainage causing ground movement (investigation concluded June 2024; rear claim resolved); front of property — clay shrinkage from vegetation (separate cause; separate claim required; not resolved at time of decision)",
        "Property Type": "HMO (House in Multiple Occupation) let property (AXA PrimeLet commercial landlord policy)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "AXA: front of property damage not subsidence-related (July 2024 position — later reversed in August 2024); two separate subsidence events with different causes require two separate claims and two policy excesses; communication acknowledged as poor",
        "Evidence Dispute": "Ms L: subsidence across the whole structure; AXA's surveyor gave inconsistent explanations without having visited the property; AXA should automatically set up a new claim for the front. AXA's surveyor (July 2024): front crack pattern does not indicate subsidence — outside scope. Ms L's independent structural engineer (£600): front = subsidence from clay shrinkage from nearby vegetation; rear = subsidence from damaged underground pipe; bathroom damage likely not subsidence-related. AXA (August 2024): reversed July 2024 position — front is a separate subsidence incident. FOS: AXA was wrong to tell Ms L in July 2024 that the front was not subsidence-related; once AXA knew in August 2024, it should have set up a new claim immediately — this was never done; avoidable delay of several months from July to October 2024 when rear repairs could have started but did not; two different causes = two separate claims and two excesses is fair; £1,500 total D&I compensation; £600 engineer cost refundable",
        "Outcome Category": "Upheld",
        "Outcome": "AXA to increase total compensation to £1,500 (additional £1,150 on top of £350 already offered); reimburse Ms L for independent structural engineer's report cost of £600 plus 8% simple interest from date of payment to date of settlement",
        "Compensation Awarded (£)": 1500,
        "Is Core Case": "No — Handling Dispute",
        "Key Policy Clause": "When an insurer's own investigation reversal confirms that a previously excluded area is in fact subsidence, the insurer must proactively set up a new claim immediately — it cannot wait for the consumer to initiate the new claim; where a consumer receives inconsistent information from an insurer (exclusion reversed within weeks), the consumer who commissions independent professional evidence to resolve the inconsistency is entitled to reimbursement of that cost; two subsidence incidents with different identified causes are properly treated as two separate claims with two policy excesses; avoidable delay in progressing an accepted subsidence claim (where repairs could have begun but did not due to insurer's handling failures) is compensable",
        "Missing Evidence": "Whether AXA's surveyor had visited the front of the property before giving the July 2024 advice that the front was not subsidence-related (AXA's surveyor appeared not to have done so); precise timeline of when AXA's August 2024 reversal was communicated to Ms L",
        "Ombudsman Reasoning": "AXA was wrong to tell Ms L in July 2024 that the front was not subsidence-related — AXA itself reversed this within weeks in August 2024; once AXA knew the front was a separate subsidence incident it should have set up a new claim immediately; this was never done; avoidable delay of several months for the rear repairs (could have started July 2024 but did not until October 2024); Ms L was left with inconsistent explanations and had no realistic choice but to commission an independent engineer; engineer cost of £600 reimbursable; two different causes = two separate claims and two excesses is fair and not contested further by FOS; £1,500 total D&I compensation appropriate given duration of delay and inconvenience",
        "Workflow Insight": "When an insurer's investigation reverses a prior exclusion decision (e.g. front of property initially excluded as not subsidence, then confirmed as subsidence), the insurer must immediately set up the new claim without waiting for the consumer to re-initiate it — failure to do so creates both an avoidable delay and a liability for consequential costs; where a consumer receives conflicting technical advice from the insurer over a short period and commissions independent engineering evidence to resolve the inconsistency, the insurer must reimburse that cost; two simultaneously identified subsidence incidents with different identified causes are properly treated as two separate claims with separate excesses — this position is defensible provided it is explained clearly and promptly",
        "AI Rule Candidate": "IF insurer_reverses_exclusion_and_confirms_area_is_subsidence THEN insurer_must_proactively_set_up_new_claim_without_waiting_for_consumer; IF consumer_commissions_independent_engineer_to_resolve_inconsistent_insurer_advice THEN insurer_must_reimburse_engineer_cost; two_subsidence_incidents_with_different_identified_causes_may_be_treated_as_two_separate_claims_with_two_excesses; avoidable_delay_in_progressing_accepted_subsidence_claim_due_to_insurer_handling_failure_is_compensable",
        "Source PDF": "DRN-6019596.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "EDD9C8" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Subsidence_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
