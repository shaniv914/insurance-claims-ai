"""
Appends EOW-006 through EOW-015 to Escape_of_Water_Case_Database.xlsx.
Run from the repo root: py scripts/append_eow_006_015.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID", "FOS Decision ID", "Claim Type", "Leak Source", "Property Type",
    "Rejection Reason", "Evidence Dispute", "Outcome", "Key Policy Clause",
    "Missing Evidence", "Ombudsman Reasoning", "Workflow Insight",
    "AI Rule Candidate", "Source PDF",
]

NEW_CASES = [
    {
        "Case ID": "EOW-006",
        "FOS Decision ID": "DRN-1135035",
        "Claim Type": "Escape of water — lounge ceiling collapse, alleged leaking water tank",
        "Leak Source": "Alleged leaking water tank above lounge ceiling (insurer disputed — no moisture or staining found)",
        "Property Type": "Residential home",
        "Rejection Reason": (
            "Full decline — insurer's surveyor found no moisture or staining at damaged ceiling; "
            "concluded natural breakdown of 120-year-old plaster/lath ceiling materials (wear and tear), "
            "not caused by escape of water"
        ),
        "Evidence Dispute": (
            "Plumber initially said leaking tank may have caused the collapse but provided no written report. "
            "Customer said ceiling was wet when it fell but had dried by the time surveyor attended (~3 weeks later). "
            "Surveyor conducted moisture tests and found no moisture present in the damaged area."
        ),
        "Outcome": (
            "Not upheld — insurer's decision to decline was fair. Only expert written evidence available "
            "(surveyor's report) concluded damage was not due to escape of water. Customer did not obtain an "
            "independent expert report. Covea correctly declined."
        ),
        "Key Policy Clause": (
            "Coverage requires damage to be caused by a listed insured event. Wear and tear and gradual "
            "breakdown excluded. General Exclusions clause 13: loss or damage caused by wear and tear, "
            "wet or dry rot, or anything which happens gradually."
        ),
        "Missing Evidence": (
            "No written report from initial plumber. No independent expert report obtained by customer. "
            "Fallen ceiling pieces disposed of before moisture testing."
        ),
        "Ombudsman Reasoning": (
            "Only expert written evidence available concluded damage was not due to escape of water. "
            "Natural breakdown of 120-year-old plaster and lath ceiling is consistent with gradual wear "
            "and tear. Delay before surveyor arrived was due to broker, not insurer. Customer had the "
            "opportunity to commission an independent expert report but did not do so."
        ),
        "Workflow Insight": (
            "A single plumber's unwritten verbal suggestion that a leak 'may have caused' damage is "
            "insufficient to establish escape of water. Surveyor moisture tests with no moisture detected "
            "are more persuasive than unsubstantiated claims. Customer who fails to obtain an independent "
            "expert report is at risk of losing a dispute where only the insurer's expert opinion exists."
        ),
        "AI Rule Candidate": (
            "If only insurer's expert opinion is available and customer has not obtained an independent "
            "report, treat decline as valid. Plumber's verbal suggestion without a written report is "
            "insufficient to establish escape of water. If customer claims damp dried before surveyor "
            "arrived, require contemporaneous evidence (photos, moisture readings)."
        ),
        "Source PDF": "DRN-1135035.pdf",
    },
    {
        "Case ID": "EOW-007",
        "FOS Decision ID": "DRN-1525346",
        "Claim Type": (
            "Escape of water — secondary electrical failure claim; "
            "kitchen electrics attributed to 2018 hot water pipe escape"
        ),
        "Leak Source": "Mains hot water pipe in downstairs bathroom (2018 original incident; 2019 secondary electrical claim)",
        "Property Type": "Residential home",
        "Rejection Reason": (
            "Declined — insurer's engineer found a dead short in the circuit (mechanical cable break, "
            "not water-caused); kitchen was on a separate electrical ring barely affected by the original "
            "escape; 7-month delay between original escape and electrical failure makes water causation implausible"
        ),
        "Evidence Dispute": (
            "Customer argued steam from hot water pipe spread through the property and damaged the consumer "
            "unit and kitchen electrics. Customer's own 2018 electrician report recommended work on the "
            "consumer unit. AXA's engineer tested the circuit and found a dead short (mechanical cable break, "
            "not water-related). Kitchen was on a different ring to the rest of the property and showed only "
            "minor aesthetic damage from the original escape."
        ),
        "Outcome": (
            "Not upheld — AXA's decision to decline was fair. Dead short is a mechanical cable failure, "
            "not caused by water. Kitchen was on a separate electrical ring and barely affected by the "
            "original escape. Customer's 2018 electrician report related to the consumer unit, not the "
            "circuit fault causing the 2019 failure."
        ),
        "Key Policy Clause": (
            "Coverage requires damage to have been caused by an insured peril. If electrical failure was "
            "caused by a maintenance problem (dead short — mechanical fault), there is no insured peril "
            "and no cover applies."
        ),
        "Missing Evidence": (
            "No independent electrician's report specifically addressing the 2019 fault type "
            "(dead short vs. water damage). Customer's 2018 report related to consumer unit only, "
            "not the circuit."
        ),
        "Ombudsman Reasoning": (
            "AXA's engineer found a dead short (mechanical break in cable), which cannot be caused by water. "
            "Kitchen was on a separate electrical ring from the rest of the property and barely affected by "
            "the original escape of water. Seven-month delay between original escape and electrical failure "
            "makes water causation unlikely. Customer's 2018 electrician report did not address the specific "
            "circuit fault causing the 2019 failure."
        ),
        "Workflow Insight": (
            "Delayed secondary claims for electrical failure following an escape of water require evidence "
            "that the fault was caused by water, not mechanical failure. A dead short in a circuit is a "
            "distinct mechanical fault not caused by water. Insurer's expert report on the specific fault "
            "type is more persuasive than a general report about the consumer unit. Time elapsed between "
            "escape of water and reported electrical failure is relevant to causation assessment."
        ),
        "AI Rule Candidate": (
            "For secondary electrical claims following escape of water: require independent electrical "
            "inspection report addressing the specific fault type. If insurer's expert finds fault is "
            "mechanical (dead short, cable break) and customer provides no contradicting expert report, "
            "decline on maintenance grounds. 7+ months elapsed between escape of water and electrical "
            "failure is a negative indicator for water causation."
        ),
        "Source PDF": "DRN-1525346.pdf",
    },
    {
        "Case ID": "EOW-008",
        "FOS Decision ID": "DRN-1623377",
        "Claim Type": (
            "Escape of water — attic pipe failure in unoccupied property; "
            "Structural Works Clause endorsement applied from policy inception"
        ),
        "Leak Source": "Attic pipe (elbow joint failure) in unoccupied property intended for future refurbishment",
        "Property Type": "Unoccupied residential property (intended for refurbishment)",
        "Rejection Reason": (
            "Full decline — Structural Works Clause endorsement excluded escape of water cover until all "
            "disclosed structural work was fully completed; endorsement applied from policy inception "
            "regardless of whether works had yet started"
        ),
        "Evidence Dispute": (
            "Customer argued the endorsement should not apply because structural works had not yet started. "
            "Insurer maintained the endorsement applied from inception, regardless of whether works had started. "
            "Loss adjuster's preliminary report left open whether the heating clause was complied with and "
            "identified the cause as an elbow joint failure — but insurer's final coverage decision rested "
            "on the endorsement."
        ),
        "Outcome": (
            "Not upheld — policy wording was clear. Endorsement applied at all times and from inception. "
            "Customer's own disclosure of intended structural works at inception triggered the clause. "
            "Loss adjuster's preliminary assessment did not override insurer's final coverage decision."
        ),
        "Key Policy Clause": (
            "Structural Works Clause endorsement: 'This insurance does not provide cover for Escape of Water "
            "or Trace and Access until all the structural work, as disclosed to us, has been fully completed.' "
            "Schedule stated endorsements apply 'at all times.'"
        ),
        "Missing Evidence": (
            "No written confirmation from loss adjuster that heating clause was complied with (not decisive "
            "given endorsement applied regardless). No definitive cause report from plumber on elbow joint failure."
        ),
        "Ombudsman Reasoning": (
            "Schedule of Cover stated all endorsements apply 'at all times.' The Structural Works Clause "
            "clearly excluded escape of water until structural works were fully completed. Customer's own "
            "disclosure of intended works at inception triggered the endorsement from day one. Endorsement "
            "wording was unambiguous. Loss adjuster's preliminary views were not binding on the insurer."
        ),
        "Workflow Insight": (
            "Structural Works Clause endorsements on unoccupied properties exclude escape of water from "
            "policy inception — the works do not need to have started. Insurer's policy documentation "
            "controls the coverage decision; loss adjuster's preliminary on-site views are not determinative. "
            "Customer disclosure of intended structural works at inception triggers the endorsement immediately."
        ),
        "AI Rule Candidate": (
            "On unoccupied property escape of water claims: always check for Structural Works Clause "
            "endorsement. If present and disclosed works have not been fully completed, apply endorsement "
            "to decline regardless of whether works have started. Loss adjuster's preliminary views are "
            "not binding on insurer's final coverage decision."
        ),
        "Source PDF": "DRN-1623377.pdf",
    },
    {
        "Case ID": "EOW-009",
        "FOS Decision ID": "DRN-2340952",
        "Claim Type": (
            "Escape of water — accepted claim; disputed handling: scope of reinstatement, quality of works, "
            "individual items (cooker, bed, electrics, boiler, drying), and adequacy of compensation"
        ),
        "Leak Source": "Escape of water affecting bathroom, kitchen and spare room",
        "Property Type": "Residential home",
        "Rejection Reason": (
            "Not a coverage rejection — claim accepted. Individual items declined: cooker and boiler "
            "(no causation evidence); electrical safety (pre-existing wiring issues unrelated to escape); "
            "loss of rent (no evidence of prior rental intention); bed (no evidence contractor damaged it). "
            "Total £750 compensation offered (£350 + £400)."
        ),
        "Evidence Dispute": (
            "Customer disputed: kitchen tiling gap (caused by customer-chosen layout change, not insurer); "
            "quality of repairs; cooker damage (no expert evidence linking to escape of water); bed removed "
            "by contractor in error; electrical safety post-repair; adequacy of drying (drying certificate "
            "issued); boiler damage (no evidence); loss of rental income from spare room."
        ),
        "Outcome": (
            "Not upheld — revised cash settlement (£1,086 buildings + £700 contents) was fair. Individual "
            "item disputes rejected for lack of causation evidence. £750 total compensation was fair given "
            "some delays were partly due to customer's unavailability. Insurer's revised calculation "
            "correcting an earlier error was accepted."
        ),
        "Key Policy Clause": (
            "Insurer obligated to reinstate damage caused by escape of water. Items must be shown to have "
            "been damaged by the escape of water. Policy excess applies."
        ),
        "Missing Evidence": (
            "No expert evidence linking cooker, boiler, or electrics to escape of water. No evidence of "
            "pre-rental intention for spare room before March 2020. No evidence property was not properly "
            "dried (drying certificate accepted). No photographic evidence contractor damaged the bed."
        ),
        "Ombudsman Reasoning": (
            "Cash settlement methodology was fair once insurer corrected calculation error. Scope disputes "
            "resolved in insurer's favour. Each individual item dispute required specific causation evidence "
            "which customer could not provide. Customer-caused delays (unavailability) reduced compensation "
            "entitlement. Drying certificate issued by contractor was accepted as evidence of completed drying."
        ),
        "Workflow Insight": (
            "Claims handling disputes require item-by-item causation evidence linking each disputed item "
            "to the escape of water. Insurer should correct calculation errors promptly. Customer-caused "
            "delays (unavailability, refusal of access) mitigate compensation quantum. A drying certificate "
            "from insurer's contractor is presumptive evidence of completed drying unless contradicted by "
            "independent expert evidence."
        ),
        "AI Rule Candidate": (
            "For each item in a disputed claim, require specific expert causation evidence linking it to "
            "the escape of water. Accept insurer's revised cash settlement if methodology is fair. Flag "
            "customer-caused delays as a mitigating factor in compensation assessment. Drying certificate "
            "from insurer's contractor is presumptive evidence of completed drying — rebuttable only by "
            "independent expert report."
        ),
        "Source PDF": "DRN-2340952.pdf",
    },
    {
        "Case ID": "EOW-010",
        "FOS Decision ID": "DRN2512341",
        "Claim Type": (
            "Escape of water — pre-policy wet rot damage to floor joists; "
            "customer sought to invoke secondary flooding industry approach"
        ),
        "Leak Source": (
            "Wet rot in floor joists under two bedrooms — attributed to previous owner's prior 'serious flood' "
            "(determined to be likely escape of water, not a natural flood)"
        ),
        "Property Type": "Residential home (recently purchased)",
        "Rejection Reason": (
            "Full decline — damage predated policy inception; secondary flooding industry approach does not "
            "apply because original cause was escape of water, not a natural flood; current insurer only "
            "liable for damage occurring after policy start date"
        ),
        "Evidence Dispute": (
            "Customer argued original incident was 'secondary flooding' per FOS technical note. FOS "
            "distinguished natural flood from escape of water — escape of water does not trigger secondary "
            "flooding policy. Customer later argued property was in a surface water flooding risk area and "
            "found water in the floor void during recent renovations, suggesting possible natural flood origin."
        ),
        "Outcome": (
            "Not upheld — Zurich's decline was fair. No persuasive evidence the original cause was a natural "
            "flood rather than an escape of water. Secondary flooding industry approach requires confirmed "
            "natural flood cause. Customer's later evidence (surface water flood risk area, water in void) "
            "was speculative and insufficient."
        ),
        "Key Policy Clause": (
            "Policy covers damage during the period of insurance only. Existing damage exclusion. Secondary "
            "flooding industry agreement applies only where original cause was a natural flood — escape of "
            "water, even if it causes flooding of the property, does not qualify."
        ),
        "Missing Evidence": (
            "Original claim file not available. No documentary evidence original cause was a natural flood. "
            "No evidence prior insurer treated the claim as flood rather than escape of water."
        ),
        "Ombudsman Reasoning": (
            "Secondary flooding industry approach applies only to natural flood events. An escape of water, "
            "even if it floods a property, remains an escape of water event. Customer's original "
            "characterisation (escape of water) was partly their own speculation. No persuasive evidence "
            "the original incident was a natural flood. Customer's later surface water flooding argument "
            "was speculative — nothing linked the floor void water to the original damage."
        ),
        "Workflow Insight": (
            "Secondary flooding industry approach does not apply where original cause was escape of water. "
            "Current insurer is not responsible for damage preceding policy inception. Absence of original "
            "claim file is an evidential gap against the policyholder. Customer speculation about original "
            "cause is insufficient — documentary evidence of the original claim and its cause is required."
        ),
        "AI Rule Candidate": (
            "For claims where damage predates policy inception: reject unless secondary flooding industry "
            "agreement applies (requires confirmed natural flood cause with documentary evidence). Do not "
            "apply secondary flooding approach where original cause was escape of water. Require original "
            "claim file or insurer records showing original event was treated as flood before applying "
            "secondary flooding."
        ),
        "Source PDF": "DRN2512341.pdf",
    },
    {
        "Case ID": "EOW-011",
        "FOS Decision ID": "DRN-2540997",
        "Claim Type": (
            "Escape of water — two pre-existing damage claims discovered shortly after purchase "
            "(kitchen floor and bathroom floor)"
        ),
        "Leak Source": (
            "Kitchen: escape of water from kitchen services (historic); "
            "Bathroom: pipe connecting WC to cistern (historic)"
        ),
        "Property Type": "Residential home (recently purchased)",
        "Rejection Reason": (
            "Both claims declined — insurer concluded damage occurred before policy inception and/or was "
            "caused gradually. Loss adjuster assessed kitchen only; no assessment of bathroom."
        ),
        "Evidence Dispute": (
            "Customer commissioned a full pre-purchase structural survey which found no flooring issues. "
            "Damage not discovered until weeks after moving in. Specialist leak detection found no active "
            "leaks (damage historic). Insurer's loss adjuster assessed kitchen but not bathroom. Insurer's "
            "call handler implied claims would be covered — no written record."
        ),
        "Outcome": (
            "Upheld — Fairmead must accept both claims. Pay restoration costs already incurred plus 8% "
            "simple interest. Pay trace and access costs. Settle remaining uninstructed work using insurer's "
            "contractors or equivalent cost. £400 compensation. Applying pre-existing or gradual exclusion "
            "was unfair — customer could not reasonably have known about damage; full structural survey "
            "found nothing; damage was fully concealed."
        ),
        "Key Policy Clause": (
            "Pre-existing damage exclusion; gradual damage exclusion. FOS approach: applying gradual damage "
            "exclusion is unfair if consumer could not reasonably have been aware of damage occurring. "
            "Policy covers escape of water and buildings including wooden, laminate and vinyl floor coverings."
        ),
        "Missing Evidence": (
            "No active leaks at time of survey (historic damage). No written record of call handler "
            "implying coverage. No evidence prior owners or surveyors were aware of the damage."
        ),
        "Ombudsman Reasoning": (
            "Pre-purchase structural survey found no flooring issues — no one (including prior owners or "
            "surveyors) would have known about the damage. Customer discovered problems within weeks of "
            "moving in. Applying pre-existing exclusion is unfair where customer had no reasonable means "
            "of discovering the damage. Gradual damage exclusion equally unfair — customer did not live "
            "at property when damage occurred and was prevented from noticing it. Insurer must assess "
            "all claimed areas, not just kitchen."
        ),
        "Workflow Insight": (
            "If customer commissioned a pre-purchase structural survey that identified no issues and damage "
            "was discovered shortly after moving in, it is unfair to apply pre-existing or gradual damage "
            "exclusions — customer could not reasonably have known. Insurer must assess all claimed areas; "
            "sending loss adjuster to kitchen only and ignoring bathroom is an incomplete assessment."
        ),
        "AI Rule Candidate": (
            "If pre-existing or gradual damage exclusion is applied but customer commissioned a pre-purchase "
            "structural survey finding no issues, and damage was discovered shortly after moving in: reject "
            "exclusion application as unfair — customer could not reasonably have known. Insurer must assess "
            "all claimed rooms. Pay trace and access as standalone benefit where applicable."
        ),
        "Source PDF": "DRN-2540997.pdf",
    },
    {
        "Case ID": "EOW-012",
        "FOS Decision ID": "DRN-2572816",
        "Claim Type": (
            "Escape of water — leaking drain pipe under kitchen floor causing dry rot in floor joists; "
            "insurer invoked gradually operating cause and faulty design (ventilation) exclusions"
        ),
        "Leak Source": "Leaking drain pipe under kitchen floor",
        "Property Type": "Residential home (kitchen)",
        "Rejection Reason": (
            "Declined — gradually operating cause (dry rot developed over time and not linked to escape of "
            "water); faulty design exclusion (kitchen extension allegedly lacked proper ventilation). "
            "Insurer offered £100 compensation for poor service only."
        ),
        "Evidence Dispute": (
            "Insurer's second expert concluded dry rot developed over time but drew no link between the "
            "leaking drain in the same area and the dry rot. Insurer's own restoration schedule listed cause "
            "as 'escape of water.' Phone call between insurer and contractor acknowledged the leak likely "
            "caused the dry rot. Insurer provided no evidence of ventilation requirements or building "
            "regulations to support faulty design exclusion."
        ),
        "Outcome": (
            "Upheld — Admiral failed to prove either exclusion applies. Must accept claim under escape of "
            "water peril and settle in line with policy terms. Additional £300 compensation "
            "(total £400 including £100 already paid)."
        ),
        "Key Policy Clause": (
            "Gradually operating cause exclusion; faulty design exclusion. Once insured peril established "
            "(escape of water accepted), insurer must prove exclusion applies with specific evidence. "
            "Insurer cannot merely assert exclusion."
        ),
        "Missing Evidence": (
            "No reasoning in insurer's expert report why dry rot was unrelated to the leaking drain in "
            "the same area. No building regulations or ventilation standards cited to support faulty "
            "design exclusion."
        ),
        "Ombudsman Reasoning": (
            "Insurer accepted escape of water as the insured peril but declined on exclusions. Second expert "
            "report drew no causal link between the leaking drain and the dry rot — the conclusion 'unrelated' "
            "was offered with no reasoning. Dry rot develops on damp timber; a leaking drain in the same area "
            "makes the link obvious. Insurer's own restoration schedule listed cause as escape of water. Phone "
            "call with contractor confirmed insurer's acceptance of the link. No evidence provided for the "
            "ventilation/faulty design exclusion. Concealed damage under floor — customer could not reasonably "
            "have been aware."
        ),
        "Workflow Insight": (
            "When insurer relies on gradually operating cause or faulty design exclusion, it must provide "
            "specific evidence proving the exclusion — not just assert it. Insurer's own internal records "
            "(restoration schedules, contractor call notes) acknowledging escape of water as the cause are "
            "binding admissions. Dry rot in proximity to a leaking pipe does not prove independence from "
            "the leak. Concealed pipe leak under floor: customer had no reasonable opportunity to be aware "
            "of damage developing."
        ),
        "AI Rule Candidate": (
            "If insurer relies on gradually operating cause or faulty design exclusion: require specific "
            "evidence proving the exclusion (cause analysis; building regulations for ventilation claim). "
            "If insurer's own documents reference escape of water as cause, treat as binding admissions. "
            "Where dry rot is found proximate to a confirmed leaking pipe, insurer must prove dry rot was "
            "independent of the escape of water. Flag concealed pipe leaks under floors as circumstances "
            "where gradual damage exclusion should not automatically apply."
        ),
        "Source PDF": "DRN-2572816.pdf",
    },
    {
        "Case ID": "EOW-013",
        "FOS Decision ID": "DRN-2749125",
        "Claim Type": (
            "Escape of water — perished copper pipe, slow constant leak under kitchen floor, "
            "discovered when new appliance installed"
        ),
        "Leak Source": "Perished copper pipe causing slow constant leak under kitchen floor",
        "Property Type": "Residential home (kitchen)",
        "Rejection Reason": (
            "Claim declined — insurer concluded damage predated policy inception (October 2019); "
            "insurer's buildings inspector's 28-page report with time-stamped photographs showed severely "
            "rotten floorboards and joists consistent with multi-year leaking"
        ),
        "Evidence Dispute": (
            "Customer stated they commissioned a structural survey before purchase (2014) which found no "
            "issues. Customer had a new kitchen professionally installed a few years prior with no visible "
            "leak. Customer was first alerted to damage in July 2020 when a new washing machine was being "
            "fitted. Insurer's buildings inspector produced a detailed 28-page report with time-stamped "
            "photos — customer disputed inspector spent only 5 minutes at the property."
        ),
        "Outcome": (
            "Upheld in part — coverage decline upheld (damage predated policy inception; severely rotten "
            "floorboards consistent with multi-year leaking; visible signs of structural degradation — "
            "dropping kitchen units and worktop — meant customer ought reasonably to have been aware). "
            "Covea must pay £200 compensation for avoidable delays (approximately one month from report "
            "to repudiation)."
        ),
        "Key Policy Clause": (
            "Pre-existing damage exclusion; gradually operating cause exclusion; wet or dry rot exclusion. "
            "Damage must occur after policy inception. Gradual damage exclusion applies where customer "
            "ought reasonably to have been aware of damage occurring."
        ),
        "Missing Evidence": (
            "No evidence floorboards were inspected and damage-free during kitchen refurbishment. Structural "
            "survey did not assess under-floor condition adequately. No evidence from prior insurer covering "
            "period of alleged deterioration."
        ),
        "Ombudsman Reasoning": (
            "Buildings inspector's 28-page report with time-stamped photographs was detailed and persuasive. "
            "Severely rotten floorboards and joists consistent with multi-year leaking — damage unlikely to "
            "have originated after October 2019 inception. Dropping kitchen units and worktop were visible "
            "signs of structural degradation — customer ought reasonably to have been aware earlier. "
            "This distinguishes the case from EOW-011 (DRN-2540997) where damage was fully concealed "
            "with no visible signs."
        ),
        "Workflow Insight": (
            "A detailed buildings inspector report with time-stamped photographs showing multi-year "
            "deterioration is highly persuasive evidence of pre-inception damage. Visible signs of "
            "structural degradation (dropping units, sinking worktop) represent a reasonable opportunity "
            "for the customer to have investigated earlier. Contrasts with EOW-011 (DRN-2540997) where "
            "damage was fully concealed and no visible indicators existed. Always compensate for avoidable "
            "claim processing delays even where coverage is correctly declined."
        ),
        "AI Rule Candidate": (
            "Where insurer provides a detailed inspector report (with time-stamped photos) showing damage "
            "consistent with multi-year gradual deterioration, support pre-existing decline. Where visible "
            "signs of degradation existed (dropping units, sinking surfaces), customer ought reasonably to "
            "have been aware — uphold gradual exclusion. Differentiate from cases where damage is fully "
            "concealed with no visible indicators (see EOW-011). Compensate for avoidable delays regardless "
            "of coverage outcome."
        ),
        "Source PDF": "DRN-2749125.pdf",
    },
    {
        "Case ID": "EOW-014",
        "FOS Decision ID": "DRN-2771304",
        "Claim Type": (
            "Escape of water — peril classification dispute; tap left on by resident causing water damage "
            "to managed flats; AXA applied £5,000 escape of water excess instead of £250 damage excess"
        ),
        "Leak Source": "Tap left on by a resident in a managed flat (human negligence / overflow, not a pipe failure)",
        "Property Type": "Commercial / Management Company",
        "Rejection Reason": (
            "Not a coverage rejection — AXA accepted the claim but misclassified it under the escape of "
            "water peril (£5,000 excess) rather than as accidental damage or general damage (£250 excess). "
            "Dispute was solely about peril classification and applicable excess."
        ),
        "Evidence Dispute": (
            "Insurer maintained tap-left-on = escape of water (£5,000 excess). Management company argued "
            "it was accidental damage (£250 excess). FOS reviewed multiple standard insurer policies — "
            "tap-left-on is typically classified as accidental damage, not escape of water. AXA submitted "
            "a prior adjudicator view classifying tap-left-on as escape of water, but ombudsman did not follow it."
        ),
        "Outcome": (
            "Upheld — claim should be classified as 'damage' (or accidental damage if insurer wishes to "
            "classify further), not escape of water. Lower excess to apply. £150 compensation for delays "
            "and inconvenience to flat owners and residents."
        ),
        "Key Policy Clause": (
            "All-risks policy covers 'damage' unless excluded. Escape of water excess endorsement (£5,000). "
            "Industry convention: tap-left-on = accidental damage, not escape of water. Correct peril "
            "classification is required to apply the correct excess."
        ),
        "Missing Evidence": (
            "No standard industry examples found where tap-left-on is classified as escape of water."
        ),
        "Ombudsman Reasoning": (
            "Multiple standard insurer policies cover tap-left-on under accidental damage, not escape of "
            "water. No examples found where standard policies classify tap-left-on as escape of water. "
            "All-risks policy covers damage unless excluded — no relevant exclusion applied. Prior "
            "adjudicator view on a different case was not persuasive on the individual facts here. "
            "Incorrect classification causing higher excess deduction is a service failure."
        ),
        "Workflow Insight": (
            "Tap-left-on incidents should be classified as accidental damage, not escape of water, when "
            "policy contains separate excess structures. On all-risks policies, classify as 'damage' unless "
            "a specific exclusion applies. Incorrect peril classification causing a higher excess deduction "
            "is a service failure warranting compensation. NOTE: This case involves a commercial all-risks "
            "policy for a managed residential block — classification rules derived here may not apply "
            "directly to standard residential home insurance policies."
        ),
        "AI Rule Candidate": (
            "Water damage caused by a tap being left on: classify as accidental damage (not escape of water) "
            "when policy contains separate excess structures for each peril. Do not apply escape of water "
            "excess to tap-left-on incidents. On all-risks policies, classify as 'damage' unless a specific "
            "exclusion prevents it. Note: rule derived from commercial all-risks context — verify "
            "applicability before applying to standard residential policies."
        ),
        "Source PDF": "DRN-2771304.pdf",
    },
    {
        "Case ID": "EOW-015",
        "FOS Decision ID": "DRN-2806638",
        "Claim Type": "Broker Disclosure Dispute",
        "Leak Source": "Broken toilet cistern causing floor damage",
        "Property Type": "Leasehold flat (property owners policy — lessee and management company director)",
        "Rejection Reason": (
            "Not a coverage rejection — claim was paid. Dispute was whether broker (Endsleigh) adequately "
            "disclosed that the escape of water excess had increased from £200 to £500 at 2019 renewal, "
            "resulting in policyholder receiving £300 less settlement than expected."
        ),
        "Evidence Dispute": (
            "Policyholder accepted excess increase was communicated in 2019 renewal documents but argued "
            "2020 renewal documents did not link excess levels (A/B/C) to specific perils. Broker confirmed "
            "it had notified insurer that 2020 documents lacked this mapping — but this was outside the "
            "scope of the 2019 renewal complaint before FOS."
        ),
        "Outcome": (
            "Not upheld (2019 renewal complaint) — excess increase was clearly communicated across multiple "
            "2019 renewal documents; policyholder accepted renewal and paid premium. 2020 renewal document "
            "clarity issue is a separate complaint, not within scope of this FOS referral."
        ),
        "Key Policy Clause": (
            "Broker's duty to clearly disclose material policy changes at renewal. Escape of water excess "
            "endorsement increased from £200 to £500. Policyholder acceptance of renewal after clear "
            "disclosure binds them to the new terms. FOS can only review the complaint as originally "
            "submitted to the firm."
        ),
        "Missing Evidence": (
            "2020 renewal documents failed to link each excess level (A/B/C) to the specific peril it "
            "applied to — but this was a separate complaint outside this referral scope."
        ),
        "Ombudsman Reasoning": (
            "Excess increase was communicated in multiple 2019 renewal documents (renewal invitation, "
            "statement of fact, schedule, Notice to Policyholder). Policyholder accepted renewal by paying "
            "premium — bound by new excess. Only the 2019 notification complaint was within scope. The "
            "2020 document clarity issue must first be raised as a new complaint with the broker before "
            "FOS referral."
        ),
        "Workflow Insight": (
            "[NON-CORE — Broker Disclosure Dispute] Broker must clearly disclose material policy changes "
            "at renewal, preferably across multiple documents. Policyholder who pays premium after "
            "disclosure is bound by the new terms. Excess-to-peril mapping in renewal documents must be "
            "explicit — failure to link excess levels to specific perils in subsequent renewal years is a "
            "separate documentation complaint that must be raised with the broker first. This case concerns "
            "broker regulatory conduct, not claim assessment — flag as Non-Core for playbook development."
        ),
        "AI Rule Candidate": (
            "Excess increases at renewal: verify disclosure appears explicitly in multiple renewal documents. "
            "If policyholder paid premium and renewed after clear disclosure, apply the new excess. If "
            "renewal documents fail to link excess levels to specific perils, flag as potential documentation "
            "failure — but this is a separate complaint for the renewal year in which it first arose. "
            "[NON-CORE — broker conduct rule; not applicable to standard claim assessment playbook.]"
        ),
        "Source PDF": "DRN-2806638.pdf",
    },
]


def make_row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def make_border() -> Border:
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def main():
    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(repo_root, "knowledge", "case-databases", "Escape_of_Water_Case_Database.xlsx")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    cell_font = Font(name="Calibri", size=10)
    border = make_border()
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top")
    centered_cols = {"Case ID", "FOS Decision ID", "Property Type", "Source PDF"}

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = make_row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font = cell_font
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col_name in centered_cols else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    total_data_rows = ws.max_row - 1  # subtract header row
    last_case_id = ws.cell(row=ws.max_row, column=1).value
    print(f"Saved: {xlsx_path}")
    print(f"Total data rows: {total_data_rows}")
    print(f"Last Case ID: {last_case_id}")


if __name__ == "__main__":
    main()
