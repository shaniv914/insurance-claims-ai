"""
Standard append script for Flood Case Database — Schema v1 (21 columns).
Batch 4: FLOOD-032 to FLOOD-040
(FLOOD-031 DRN-3709658 excluded — home emergency contractor liability, not flood insurance)

Usage
-----
Run from the repo root:
    py scripts/append_flood_v4.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Flood_Case_Database.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 4: FLOOD-032 to FLOOD-040
# (FLOOD-031 DRN-3709658 excluded — home emergency contractor liability, not flood insurance)
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "FLOOD-032",
        "FOS Decision ID": "DRN-3710798",
        "Insurer Name": "Society of Lloyd's",
        "FOS Decision Date": "Mar 2023",
        "Claim Type": "Flood — residential; dispute about handling quality and scope of second flood repair liability; original flood claim accepted; dispute concerns whether temporary repairs were communicated as interim only, and whether second flood flooring damage was attributable to Lloyd's failure to communicate interim repair status",
        "Leak Source": "External flood (source not specified in decision — residential flood event; property flooded twice)",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "N/A — original flood claim accepted; Lloyd's disputed liability for second flood flooring damage and whether first repairs were described as permanent or temporary",
        "Evidence Dispute": "Mr D and Ms R: Lloyd's told them initial repairs were permanent; would have taken steps to prevent second flood if advised repairs were temporary. Lloyd's: repairs were temporary pending building warranty resolution; second flood not attributable to Lloyd's. FOS: even if repairs were temporary, warranty repairs would not have been completed before second flood regardless — second flood not Lloyd's fault; but Lloyd's failed to communicate interim repair nature clearly",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Lloyd's to pay £1,500 D&I compensation for delay and poor communication about interim nature of repairs; Lloyd's to cover cost of flooring damaged by second flood if not covered by building warranty provider",
        "Compensation Awarded (£)": 1500,
        "Is Core Case": "No — Handling Dispute",
        "Key Policy Clause": "Insurer's duty to clearly communicate when flood repairs are temporary or interim; scope of liability for damage from second flood event; building warranty exclusion does not eliminate insurer liability for repair-related communication failures; insurer bears consequential liability only to extent not recoverable from another source",
        "Missing Evidence": "Evidence from policyholders confirming second flood flooring damage is not covered by building warranty",
        "Ombudsman Reasoning": "Lloyd's acknowledged delay and poor communication. Even if first repairs were temporary, building warranty repairs would not have been completed before second flood — so second flood would have happened regardless. Lloyd's not responsible for warranty provider failings. However, Lloyd's should pay for any flooring from second flood not covered by warranty. £1,500 compensation fair for delay and failure to explain temporary repair nature.",
        "Workflow Insight": "When insurer installs temporary flood repairs, it must explicitly communicate in writing that repairs are interim-only. Failure to do so creates legitimate expectation of permanence and potential liability for consequential second-event damage. Award scope is limited to damage not covered by any other route (e.g. building warranty).",
        "AI Rule Candidate": "IF temporary_flood_repairs_installed AND insurer_fails_to_communicate_interim_nature THEN partial_liability_for_second_event_damage; IF second_flood_damage AND other_coverage_available THEN insurer_liable_for_uncovered_residue_only; delay_plus_communication_failure = D_and_I_award_warranted",
        "Source PDF": "DRN-3710798.pdf",
    },
    {
        "Case ID": "FLOOD-033",
        "FOS Decision ID": "DRN4280012",
        "Insurer Name": "Liverpool Victoria Insurance Company Limited",
        "FOS Decision Date": "Jul 2017",
        "Claim Type": "Commercial buildings and business interruption insurance — severe storm caused flooding at business premises; LV declined under flood exclusion clause introduced at policy renewal in 2016; policyholder argued storm was proximate cause of damage not flood",
        "Leak Source": "External — surface water accumulation from storm; rainwater accumulated on street, drains overwhelmed, water entered premises at ground level",
        "Property Type": "Commercial (business premises)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Flood exclusion clause: Material Damage All Risks excludes damage arising from or in connection with flood. Loss adjuster confirmed rainwater accumulated on street until drains overwhelmed and water entered premises — flood was the proximate cause; storm was indirect cause only. Exclusion properly communicated at renewal (highlighted in red in renewal letter; separate broker email).",
        "Evidence Dispute": "Mr D: local drains blocked by storm; storm was proximate cause; flood exclusion added at renewal without sufficient notice to broker. LV: loss adjuster report confirms surface water accumulation = flood as proximate cause; flood exclusion highlighted in red in renewal letter; broker emailed separately; broker confirmed acceptance. FOS: no expert evidence to contradict loss adjuster; flood exclusion properly communicated.",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — flood was proximate cause; storm was indirect cause only; flood exclusion properly introduced at renewal and communicated to broker; LV entitled to decline claim",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood exclusion: Material Damage All Risks excludes damage arising from or in connection with flood; proximate cause doctrine — immediate cause of damage determines the applicable peril; insurer entitled to add flood exclusion at renewal based on updated flood mapping with proper broker notification",
        "Missing Evidence": "N/A — loss adjuster report conclusive; no expert evidence contradicting proximate cause finding",
        "Ombudsman Reasoning": "Loss adjuster report clear: rainwater accumulated on street and rose until drains overwhelmed then entered premises = flood as direct cause; storm = indirect trigger only. No expert evidence to the contrary. Flood exclusion highlighted in red on renewal letter page 1; broker separately emailed; broker confirmed acceptance on Mr D's behalf. LV not required to notify Mr D separately once broker notified and accepted.",
        "Workflow Insight": "Proximate cause of damage is the peril that determines coverage even where a second peril (storm) triggered the chain of events. Surface water that accumulates and enters premises via overwhelmed drains = flood, not storm. Flood exclusion introduced at renewal is binding where communicated to broker in writing with clear highlighting and confirmed acceptance.",
        "AI Rule Candidate": "IF loss_adjuster_confirms_surface_water_accumulation_before_property_entry THEN proximate_cause_is_flood; IF flood_exclusion_highlighted_in_renewal_letter AND broker_confirmed_acceptance THEN exclusion_is_binding; storm_triggering_drain_overwhelm_is_indirect_cause_only",
        "Source PDF": "DRN4280012.pdf",
    },
    {
        "Case ID": "FLOOD-034",
        "FOS Decision ID": "DRN4396587",
        "Insurer Name": "Legal & General Insurance Limited",
        "FOS Decision Date": "Apr 2020",
        "Claim Type": "Residential home buildings insurance — storm and flash flooding damaged rear garden wall; L&G settled as flood claim; Mr S disputed peril classification, wanted storm recorded to avoid flood claim history and future premium impact; L&G also erroneously stated property is not a flood risk",
        "Leak Source": "External — surface water runoff from flash flooding during storm; water ran under property and damaged rear garden wall at or below ground level",
        "Property Type": "Residential home",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "N/A — claim was settled; dispute only concerns peril classification (flood vs storm) and impact on future premium declarations",
        "Evidence Dispute": "Mr S: heavy rainfall during storm caused damage directly = storm claim; if recorded as flood he would have withdrawn the claim had he known. L&G: rainfall caused water to accumulate at ground level and flood = flood is proximate cause; insurer determines final peril classification, not loss adjuster. L&G error: incorrectly told Mr S property is not considered a flood risk. FOS: L&G correctly recorded as flood; but error re flood risk status warrants £100 compensation.",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Flood classification upheld as correct; L&G to pay £100 compensation for incorrectly informing Mr S his property is not considered a flood risk when it is",
        "Compensation Awarded (£)": 100,
        "Is Core Case": "Yes",
        "Key Policy Clause": "L&G flood definition: water from any external source entering a building at or below ground level with substantial or abnormal volume; flood does not have to enter the building to constitute a flood event — external structures at ground level (garden walls) included; insurers determine peril classification, not loss adjustors; garden walls are buildings",
        "Missing Evidence": "N/A — peril classification clear from facts; flood risk status established from L&G's own records",
        "Ombudsman Reasoning": "Storm rainfall caused water to accumulate and flood at ground level; accumulated flood water damaged garden wall = flood peril. If storm had caused roof damage directly, that would be storm. Peril = immediate cause of damage. Loss adjuster categorised as storm/flood preliminary; L&G as insurer made final call as flood — within its authority. L&G error: told Mr S property not a flood risk when it is; this incorrect information creates future insurer declaration problems for Mr S = £100 loss of expectation.",
        "Workflow Insight": "Peril classification is determined by the immediate cause of damage: accumulated flood water causing structural damage = flood, even where storm triggered the rainfall. Insurers must accurately communicate the flood risk status of a property after recording a flood claim, as this affects future declarations. A flood need not enter a building to be a flood event — flood claims include external structures (garden walls, boundary walls) at ground level.",
        "AI Rule Candidate": "IF storm_triggers_accumulation AND flood_water_directly_causes_damage THEN peril_is_flood_not_storm; IF insurer_records_flood_claim AND incorrectly_denies_flood_risk_status THEN compensation_for_loss_of_expectation; flood_event_does_not_require_property_entry — external_structural_damage_at_ground_level_qualifies",
        "Source PDF": "DRN4396587.pdf",
    },
    {
        "Case ID": "FLOOD-035",
        "FOS Decision ID": "DRN-4415847",
        "Insurer Name": "Wakam",
        "FOS Decision Date": "Dec 2023",
        "Claim Type": "Residential home insurance — water ingress from defective gutter on adjoining neighbour's porch canopy caused damp, bubbling plasterwork and crumbling plaster in living room; no storm; Wakam declined citing damage not caused by storm, flood, or accidental damage as defined",
        "Leak Source": "Neighbour's canopy gutter overflow — defective gutter on adjoining neighbour's porch canopy abutting Ms K's property directed water into her wall",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Wakam: no storm in area (confirmed by weather data); water from neighbour's canopy is not an overflow of external water sources as defined in policy (rivers, lakes, sea); not accidental damage (Ms K and family not the cause; damage gradual not sudden). Policy exclusion: water entering by any means other than storm or flood not covered.",
        "Evidence Dispute": "Wakam: policy flood definition 'overflow of external water sources, such as rivers, lakes, and the sea' is exhaustive and excludes canopy ingress. Ms K: 'such as' is non-exhaustive and merely illustrative; neighbour's canopy is an external water source; contra proferentem applies. FOS: agreed with Ms K — 'such as' creates non-exhaustive list; canopy is an external water source; ambiguity resolved in policyholder's favour.",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld — water ingress from neighbour's canopy falls within policy flood definition under non-exhaustive 'such as' reading and contra proferentem principle; Wakam to settle claim per remaining policy terms including any excess; £100 D&I for unfair decline",
        "Compensation Awarded (£)": 100,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood definition: 'overflow of external water sources, such as rivers, lakes, and the sea' — 'such as' renders source list illustrative and non-exhaustive; contra proferentem — ambiguous policy wording interpreted against insurer/drafter in policyholder's favour; accidental damage definition limited to damage caused by policyholder or family. See flood-source-of-water-interpretation.md for full analysis.",
        "Missing Evidence": "Weather data confirming no storm (accepted by both parties); specialist report on cause of ingress from neighbour's canopy (accepted by both parties)",
        "Ombudsman Reasoning": "No storm in area — storm section does not apply. Accidental damage requires Ms K or family to have caused it — they did not. On flood: 'such as' before list of sources means list is illustrative not exhaustive. Neighbour's canopy is clearly external, clearly water-producing, clearly a source. Even if ambiguous, contra proferentem requires interpretation in policyholder's favour where insurer drafted the policy. Damage from canopy ingress = flood under policy definition.",
        "Workflow Insight": "Policy flood definitions using 'such as' before listed water sources create non-exhaustive illustrative lists — any above-ground external water source may qualify. Water ingressing from any neighbouring property structure can constitute flood where it builds up and enters the property. Where policy definition is ambiguous (source not defined), insurer cannot rely on narrow reading — contra proferentem resolves in policyholder's favour. See FLOOD-038 for contrast: rising water table (subsoil) does not qualify under same language.",
        "AI Rule Candidate": "IF flood_definition_uses_such_as_before_source_list THEN list_is_non_exhaustive; IF water_source_is_external_and_above_ground THEN may_qualify_as_flood_source; IF policy_wording_ambiguous THEN apply_contra_proferentem_in_policyholder_favour; distinguish_above_ground_external_overflow_from_naturally_occurring_subsoil_water",
        "Source PDF": "DRN-4415847.pdf",
    },
    {
        "Case ID": "FLOOD-036",
        "FOS Decision ID": "DRN-4895575",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "Aug 2024",
        "Claim Type": "Residential buildings and contents insurance — handling dispute; original October 2019 and February 2020 flood claims accepted; AXA withheld payment for specific 2020 flood contents items (sauna, red light therapy unit, two vibro plates, hot tub, summerhouse) invoking fraud condition; fraud condition applicability decided under separate DRN reference",
        "Leak Source": "External flood — river or surface water; property flooded October 2019 and February 2020; January 2021 flood addressed separately under DRN-4901901",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "AXA invoked fraud condition to withhold payment for specific 2020 flood items; fraud condition applicability determined unfavourable to AXA under separate complaint reference; AXA had no specific fraud concerns relating to the items in this decision",
        "Evidence Dispute": "AXA: fraud condition entitles it to decline all three flood claims; items not paid due to wider fraud concerns. Ms G: items existed and were damaged; AXA's own fraud condition challenge was decided against AXA separately. FOS: fraud condition inapplicable (decided separately); no dispute items existed and were damaged; AXA must settle at current market rates due to delay within its own control.",
        "Outcome Category": "Upheld",
        "Outcome": "AXA to settle claim for sauna, red light therapy unit and collagen canopy, two vibro plates, hot tub and summerhouse damaged in 2020 flood at current market rates; no D&I compensation awarded as Ms G requested FOS not assess service or delay issues (pursuing separately)",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Handling Dispute",
        "Key Policy Clause": "Fraud condition does not apply where insurer cannot show knowing falsehood (determined separately); insurer's delay in settling accepted flood claim items means settlement at current market rates — insurer bears inflation risk for withheld payments; visitor items covered by policy but payable to visitor not policyholder",
        "Missing Evidence": "Current market rate quotes for specialist items (sauna, hot tub, summerhouse) needed to quantify settlement",
        "Ombudsman Reasoning": "No dispute that items existed and were damaged in 2020 flood. AXA provisionally accounted for costs but never paid due to fraud hold. Fraud condition decided inapplicable under separate reference. AXA had control over timing of payment and chose not to pay — must now settle at current (inflated) market rates. Ms G asked FOS not to consider D&I or service quality issues — no compensation awarded in this decision.",
        "Workflow Insight": "Insurer that withholds payment for accepted flood claim items by invoking fraud condition bears the cost of price inflation if fraud condition later found inapplicable — settlement at current market rates not original claim value. Where policyholder explicitly limits scope of FOS complaint (excluding D&I), FOS will not award compensation beyond what was requested. Related flood claims from same property and period may be split across multiple FOS references, each with separate award limits.",
        "AI Rule Candidate": "IF insurer_withholds_payment_invoking_fraud_condition AND fraud_condition_found_inapplicable THEN settle_at_current_market_rates; IF policyholder_excludes_D_and_I_from_FOS_scope THEN no_compensation_awarded; multiple_floods_same_property = separate_complaint_references_and_separate_award_limits",
        "Source PDF": "DRN-4895575.pdf",
    },
    {
        "Case ID": "FLOOD-037",
        "FOS Decision ID": "DRN-4901901",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "Aug 2024",
        "Claim Type": "Residential buildings and contents insurance — January 2021 flood (third flood at property after 2019 and 2020 events); AXA refused to pay 2021 contents claim invoking fraud condition, alleging knowingly false statement about oil contamination in floodwater and exaggerated losses",
        "Leak Source": "External flood — river or surface water; January 2021; floodwater reportedly contaminated with oil from neighbouring property heating oil leak into local storm drains",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "AXA: fraud condition — Ms G made knowingly false statement about oil contamination (specialist found no residual oil 6 days post-flood); exaggerated losses given short interval between three flood claims and volume of items claimed; Ms G irresponsibly stored items at low level knowing flood risk. FOS: honest recollection defeats knowing falsehood; oil contamination corroborated by neighbours; AXA agent generated BER list not Ms G; reasonable for Ms G to live and store items at property.",
        "Evidence Dispute": "AXA: specialist test 6 days post-flood found no oil = Ms G lied about oil; photo shows small initial loss report inconsistent with large final claim. Ms G: honest belief in oil contamination (smell and visual); neighbours corroborated oil in local storm drains; she explained items removed from multiple exits not just area shown in photo; items replaced after prior floods are normal day-to-day items. FOS: honest recollection standard applies; AXA agent's own BER decisions cannot found fraud; low-value items do not require individual receipts.",
        "Outcome Category": "Upheld",
        "Outcome": "AXA cannot rely on fraud condition to refuse 2021 contents claim; must deal with claim per remaining policy terms; simple interest at 8% per annum on payment from date of initial BER report to date of payment",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Handling Dispute",
        "Key Policy Clause": "Fraud condition: requires knowing false statement or knowingly exaggerated claim; honest contemporaneous belief defeats fraud allegation even if later evidence contradicts it; insurer's agent decisions on BER scope are insurer's responsibility not policyholder's; low-value everyday items do not require strict proof of purchase; 8% simple interest applies where insurer has wrongly withheld payment",
        "Missing Evidence": "Contemporaneous oil contamination evidence (sampling at time of flood rather than 6 days later); oil may have dissipated due to heavy rain in intervening period (neighbours described this pattern)",
        "Ombudsman Reasoning": "Fraud requires knowing falsehood — not established. Ms G honestly believed oil was present (smell and visual); neighbours corroborated oil in local drains and heating oil leak from neighbouring property; declining oil levels explained by heavy rain. AXA agent generated BER list and cross-referenced Ms G's list to save time — errors in that list are agent's responsibility. Policyholder with severely limited upstairs space must keep some items downstairs; not negligent. Multiple low-value items do not require individual receipts. No evidence of intentional exaggeration.",
        "Workflow Insight": "Flood fraud allegations must clear a high bar: insurer must show knowing false statement, not merely subsequent evidence of inaccuracy. Honest contemporaneous belief — especially where corroborated by neighbours — defeats fraud condition invocation. Agent-prepared BER lists that overstate or misclassify items are the insurer's own procedural error, not policyholder fraud. Policyholders living in partly-habitable flood-damaged properties cannot be required to store all items off-floor on the assumption of re-flooding.",
        "AI Rule Candidate": "IF fraud_condition AND alleged_false_statement AND honest_contemporaneous_belief_corroborated THEN fraud_not_established; IF agent_prepares_BER_list AND errors_in_list THEN insurer_bears_responsibility_not_policyholder; low_value_everyday_items_do_not_require_strict_receipts; IF property_partially_habitable_post_flood THEN storage_of_items_at_low_level_is_not_negligence",
        "Source PDF": "DRN-4901901.pdf",
    },
    {
        "Case ID": "FLOOD-038",
        "FOS Decision ID": "DRN-4948332",
        "Insurer Name": "Aviva Insurance Limited",
        "FOS Decision Date": "Mar 2025",
        "Claim Type": "Residential home insurance (Flood Re risk) — February 2024 discovery of standing water under reception room floor; joists rotting; damage to walls; damp specialist concluded groundwater flooding from elevated water table; Aviva declined citing flood definition excludes groundwater and damage occurred gradually",
        "Leak Source": "Groundwater — rising water table; standing water discovered under floor from elevated water table level possibly linked to proximity to river",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Aviva flood definition: storm or flood from overflow of external water sources, such as rivers, lakes and the sea — this does not cover groundwater or rising water table. Aviva also relied on general exclusion for damage occurring gradually. Water table is naturally occurring subsoil water that rises — not an overflow of a surface water source.",
        "Evidence Dispute": "Mrs and Mr S: groundwater flooding is an overflow of the water table which should be covered; policy wording not sufficiently clear; water in ground likely came from river exceeding its banks and saturating soil. Aviva: water table is subsoil water — not an external surface water source overflowing. Local authority and damp specialist confirmed elevated water table as cause. FOS: agreed with Aviva; water table physically distinct from surface water overflow; indirect river causation excluded in any event.",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — groundwater from rising water table is not an overflow of external water sources as defined; Aviva's flood definition clearly distinguishes surface water overflow from subsoil groundwater; Flood Re cession does not alter peril coverage scope; Aviva acted correctly in declining claim",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood definition: storm or flood from overflow of external water sources, such as rivers, lakes and the sea — surface overflow mechanism required; water table is naturally occurring subsoil water, not a surface overflow source; gradual damage exclusion; Flood Re scheme cession does not broaden peril coverage. See flood-source-of-water-interpretation.md for full analysis.",
        "Missing Evidence": "N/A — cause confirmed by damp specialist and local authority as groundwater or water table; no factual dispute about cause",
        "Ombudsman Reasoning": "Policy flood definition describes surface water sources that overflow at ground level (rivers, lakes, sea) — these overflow downward causing flooding. Water table is naturally occurring subsoil water that rises upward through soil — physically and mechanically distinct. Even if river flooding saturated ground and raised water table, Aviva excludes indirect loss in any event. Policy definition clearly excludes rising groundwater. Flood Re cession is a premium-capping mechanism, not a coverage expansion. FOS sympathetic but cannot override clear policy wording.",
        "Workflow Insight": "Policy flood definitions using 'such as rivers, lakes and the sea' describe above-ground surface water sources that overflow — they do not include naturally occurring subsoil groundwater that rises through soil pressure. This is a physical distinction: surface overflow = flood; rising water table = groundwater (excluded). Contrast with FLOOD-035 (DRN-4415847) where same 'such as' language covered above-ground canopy water — because canopy water is above-ground overflow, not subsoil. Flood Re is a premium-risk transfer mechanism; it does not alter peril coverage.",
        "AI Rule Candidate": "IF cause_is_rising_water_table AND flood_definition_requires_overflow_of_external_water_sources THEN claim_excluded — water_table_is_subsoil_not_surface_overflow; IF river_flooding_raises_water_table AND policy_excludes_indirect_loss THEN groundwater_damage_excluded; flood_re_cession_does_not_expand_peril_coverage; distinguish_above_ground_external_overflow_COVERED_from_subsoil_groundwater_EXCLUDED",
        "Source PDF": "DRN-4948332.pdf",
    },
    {
        "Case ID": "FLOOD-039",
        "FOS Decision ID": "DRN-5057225",
        "Insurer Name": "Royal & Sun Alliance Insurance Limited",
        "FOS Decision Date": "Dec 2024",
        "Claim Type": "Residential buildings insurance — July 2021 basement flat flood accepted and repaired by RSA; secondary flooding appeared August 2022 on walls not replastered during original repair; RSA attributed secondary damp to pre-existing condition; policyholders argue RSA failed to complete effective and lasting repair",
        "Leak Source": "External flood — basement flat July 2021 (flood source not specified; accepted claim); secondary flooding from trapped flood water percolating through walls behind waterproof render",
        "Property Type": "Residential (basement flat)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "RSA denied responsibility for secondary damp/flooding, asserting it was pre-existing condition evidenced by historic injection holes to external walls; RSA claimed property dried to pre-loss moisture levels",
        "Evidence Dispute": "RSA: historic injection holes = prior damp; moisture readings reduced from 100% to 20-40% = pre-loss baseline. Mr and Mrs B: no damp in 10 years; only walls not replastered show secondary flooding; replastered walls dry; walls still wet when painter applied second coat. FOS: no evidence property was damp at 20-40% pre-loss; injection holes alone insufficient; email chain confirmed trapped flood water percolating behind waterproof render during repair works; replastered vs non-replastered correlation confirms inadequate repair.",
        "Outcome Category": "Upheld",
        "Outcome": "RSA to undertake repair works to all walls affected by secondary flooding to achieve effective and lasting repair; RSA to pay £350 D&I compensation",
        "Compensation Awarded (£)": 350,
        "Is Core Case": "No — Handling Dispute",
        "Key Policy Clause": "When insurer opts to settle flood claim by way of repair, it is responsible for ensuring an effective and lasting repair is completed; pre-loss damp assertion requires evidential baseline not assumption from structural features; trapped flood water percolating through structure during repair works = insurer's liability to remediate",
        "Missing Evidence": "Pre-loss moisture readings or surveys establishing baseline damp levels; expert report attributing secondary flooding to causes unrelated to the original flood (RSA had none)",
        "Ombudsman Reasoning": "RSA left property at 20-40% moisture — considerably high — with no evidence this matched pre-loss levels. Internal emails during repair confirmed trapped flood water percolating behind waterproof render (surveyor: not rising damp, it is trapped flood water from capillary action). Timing of secondary flooding immediately after works completed points to causal link. Correlation between replastered walls (dry) and non-replastered walls (secondary flooding) is decisive. RSA did not discuss pre-loss damp concerns or alternatives with Mr and Mrs B before proceeding with repair it knew might fail. Injection holes establish possibility of prior damp, not actuality or extent.",
        "Workflow Insight": "Insurer settling flood claim by repair must complete an effective and lasting repair — if it discovers pre-loss structural concerns that may cause the repair to fail, it must discuss alternatives (e.g. cash settlement) with the policyholder rather than proceeding with a doomed repair. Pre-loss damp defence requires affirmative evidence of pre-loss moisture levels at the extent claimed, not just structural features that might indicate prior damp. Secondary flooding correlated with areas not properly stripped and replastered is attributable to inadequate flood repair.",
        "AI Rule Candidate": "IF insurer_settles_by_repair AND secondary_flooding_occurs_on_areas_not_fully_stripped_replastered THEN inadequate_repair_attributable_to_insurer; IF pre_loss_damp_defence AND no_pre_loss_moisture_readings THEN defence_fails; IF trapped_flood_water_identified_during_repair_works AND no_remediation THEN insurer_liable_for_subsequent_ingress; repair_obligation_requires_effective_and_lasting_outcome_not_cosmetic_completion",
        "Source PDF": "DRN-5057225.pdf",
    },
    {
        "Case ID": "FLOOD-040",
        "FOS Decision ID": "DRN-5186962",
        "Insurer Name": "Society of Lloyd's",
        "FOS Decision Date": "Jul 2025",
        "Claim Type": "Holiday home buildings and contents insurance — September 2022 water ingress from blocked drainpipe or drainage fault; Lloyd's declined claiming damage was not from a single insured event but from continuous gradual ingress from series of floods; property unlettable since September 2022; significant reinstatement required",
        "Leak Source": "External — drainage fault; blocked drainpipe allowed rainwater to enter holiday home; source unidentified until May 2023 when neighbour's drainage fault identified and remediated",
        "Property Type": "Residential (holiday home / rental property)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Lloyd's: damage not from a single insured event; damage caused by rainwater entering over prolonged period = continuous or gradual damage from series of floods rather than one-off event; policy requires single insured event; gradual damage exclusion also referenced; no AD cover for buildings.",
        "Evidence Dispute": "Mr and Mrs R: blocked drainpipe = single incident; unaware of ingress until flooding visible inside; took immediate steps to investigate and remediate. Lloyd's: series of water ingress events over months = not a single insured peril event; gradual damage exclusion applies. Lloyd's own surveyor confirmed property flooded: 'over Xmas 2022 and after a particularly heavy storm the property was flooded again on a much more severe scale.' FOS: flood can occur gradually (Rohan Investments); policyholders took reasonable steps to mitigate once aware; series of floods does not defeat coverage where policyholder acted reasonably.",
        "Outcome Category": "Upheld",
        "Outcome": "Lloyd's to accept claim under flood peril; reimburse reinstatement expenditure and replaced contents at rate paid by policyholders plus 8% simple interest; cash settle outstanding works at current market rates; pay monthly rental income losses at £991 per month from February 2024 until 3 months after outstanding works settled plus 8% interest; pay £5,000 loss of rent policy benefit plus 8% interest from January 2023; pay £1,600 total D&I compensation (including £450 already offered)",
        "Compensation Awarded (£)": 1600,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood peril (undefined in policy) = buildup of water over what should be dry land (Rohan Investments Ltd v Cunningham 1998); flood can occur gradually through slow and steady buildup; gradual damage exclusion inapplicable where policyholder takes reasonable mitigation steps once aware; insurer's unfair decline makes it liable for consequential losses beyond policy limits including loss of rental income; loss of rent policy benefit (£5,000 cap) triggered from date insurer should have accepted claim",
        "Missing Evidence": "N/A — flood causation clear from surveyor report, video evidence of water buildup, and spread of damage; policyholder mitigation steps clearly documented",
        "Ombudsman Reasoning": "Flood = buildup of water over dry land — can occur gradually (Rohan Investments 1998). Video showed water entering at a rate causing running water buildup. Surveyor confirmed flooding. Spread of damage to flooring, walls and furniture typical of flood. Whether one flood or series: policyholders took reasonable steps once aware — investigated cause, installed drying machines, arranged CCTV, stripped back sections. Source could not be identified until May 2023 despite reasonable efforts. Gradual damage exclusion cannot apply where policyholder acted reasonably. Lloyd's unfair decline means it bears consequential rental income loss from February 2024 when property would have been relet. £5,000 loss of rent benefit payable from January 2023. £1,600 D&I for significant financial distress (loans, pension drawdown, personal repair work) over extended period.",
        "Workflow Insight": "Unfair flood claim decline makes insurer liable for consequential losses beyond policy benefit limits — including rental income from date property would have been relet had claim been accepted. Series of water ingress events from unidentified latent cause = single flood claim where policyholders diligently investigate and mitigate. Gradual damage exclusion requires showing policyholder failed to act reasonably — not applicable where latent cause not identifiable despite good-faith efforts. For holiday home policies with loss of rent benefits, benefit runs from date of insured event or date insurer should have accepted claim.",
        "AI Rule Candidate": "IF insurer_unfairly_declines_flood_claim AND property_unlettable THEN insurer_liable_for_rental_income_loss_from_date_property_would_have_been_relet; IF series_of_water_ingress_events AND policyholder_takes_reasonable_mitigation THEN gradual_damage_exclusion_does_not_apply; flood_definition_undefined_in_policy = use_Rohan_Investments_buildup_of_water_definition; loss_of_rent_benefit_triggered_from_date_insurer_should_have_accepted_claim",
        "Source PDF": "DRN-5186962.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Flood_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
