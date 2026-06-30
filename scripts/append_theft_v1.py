"""
Standard append script for Theft Case Database — Schema v1 (21 columns).
Column 6 is "Entry / Theft Method" (how the theft was committed or item came to be stolen).

Usage
-----
1. Read the source PDF(s) and extract the fields listed in NEW_CASES below.
2. Add one dict per case to NEW_CASES following the extraction rules.
3. Run from the repo root:
       py scripts/append_theft_v1.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Theft_Case_Database.xlsx

===========================================================================
FIELD EXTRACTION RULES
===========================================================================

Case ID             : Format THEFT-NNN (zero-padded to 3 digits)
FOS Decision ID     : DRN-XXXXXXX or DRNXXXXXXX as printed in the PDF
Insurer Name        : Formal registered name from the FOS decision
FOS Decision Date   : DD Mon YYYY — accept-or-reject deadline in final paragraph;
                      use "Not stated in document" if deadline not printed
Claim Type          : Policy type, physical incident and nature of dispute in one sentence
Entry / Theft Method: How the theft was committed or how items came to be stolen
                      e.g. "Forced entry — front door kicked in"
                           "Theft by invited persons — items taken during house move"
                           "Snatch theft from person — bag grabbed in street"
                           "Motor vehicle theft — keys left in ignition"
Property Type       : "Residential home" / "Motor vehicle (personal)" /
                      "Commercial vehicle" / "Personal gadget / mobile phone" /
                      "Personal travel insurance" etc.
Dispute Type        : Controlled vocab (7 values)
Coverage Decision   : Controlled vocab (5 values)
Rejection Reason    : Insurer's stated reason for declining or disputing
Evidence Dispute    : What evidence each party relied on
Outcome Category    : Controlled vocab (4 values)
Outcome             : Full FOS remedy instructions
Compensation Awarded (£) : Integer — D&I only; 0 if none
Is Core Case        : Controlled vocab (5 values)
                      Use "No — Commercial" for motor, travel, gadget or commercial
                      policies that are outside home/contents insurance scope
Key Policy Clause   : Policy wording or FOS/FCA principle applied
Missing Evidence    : Evidence that was absent and affected the outcome
Ombudsman Reasoning : How the ombudsman weighed the evidence
Workflow Insight    : Operational rule for the claims workflow
AI Rule Candidate   : Machine-evaluable rule for the rules engine
Source PDF          : Filename only (e.g. DRN0600921.pdf)
===========================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Entry / Theft Method",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 4: THEFT-031 to THEFT-039
# Note: 3 of 9 cases are home/contents insurance (THEFT-032, THEFT-035,
# THEFT-036 — flagged Is Core Case = Yes). Remaining 6 are motor or
# commercial policies flagged No — Commercial or No — Administrative.
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "THEFT-031",
        "FOS Decision ID": "DRN-5901721",
        "Insurer Name": "Aviva Insurance Limited",
        "FOS Decision Date": "5 Dec 2025",
        "Claim Type": "Motor (car) insurance — car stolen September 2023; Aviva declined to process the theft claim as the policy required an Aviva-approved tracking device to be fitted, in operation, and with subscription maintained; Mr S initially denied having a tracker then claimed he did; named driver Mrs K confirmed in a September 2023 call that a tracker was purchased but was never set up or installed in the car; invoice produced did not specify a tracker; app screenshots showed a basic device not registered to the insured car; secondary complaints concern automatic policy renewal (£75 D&I awarded but not paid), being told to keep paying premiums until theft claim resolved, and whether the premium refund was correctly backdated",
        "Entry / Theft Method": "Motor vehicle theft — car stolen 16 September 2023; method of theft not described; dispute centres on whether the policy tracker condition was satisfied at the time of theft",
        "Property Type": "Motor vehicle (personal car insurance)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Policy clause required an Aviva-approved tracking device to be fitted, in operation, and with subscription maintained at the time of theft; Mr S could not demonstrate the tracker was fitted and working; named driver Mrs K confirmed in a phone call that the tracker was bought but was not set up or installed in the car",
        "Evidence Dispute": "Mr S: claimed to have a tracker; produced dealership invoice for £500 (did not specify tracker) and screenshots of a tracker app. Mrs K: confirmed in September 2023 call that tracker was bought but didn't work and wasn't set up or installed. Aviva: app screenshots showed basic device not shown to be registered to Mr S' car or any car. FOS: Mrs K's own admission is clear evidence the tracker was not fitted and working; Aviva correctly declined the theft claim; automatic renewal error acknowledged — £75 D&I not yet paid and must be paid now; £50 D&I for not responding to backdating email confirmed paid; premium refund correctly backdated to October 2023 (date of Aviva's claim decision); complaint not upheld save for outstanding £75 payment",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Aviva correctly declined Mr S' theft claim as he could not show a working Aviva-approved tracker was fitted as required by the policy; Aviva must pay the outstanding £75 compensation it had already committed to for automatically renewing the policy against Mr S' instructions",
        "Compensation Awarded (£)": 75,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "A motor policy requiring an Aviva-approved tracking device 'fitted, in operation, and with subscription maintained' is enforceable as a condition precedent to theft cover — the policyholder must demonstrate the tracker was operational at the time of theft; a dealership invoice that does not identify a tracking device and app screenshots not registered to the insured vehicle do not satisfy the tracker condition; where the named driver has confirmed in a recorded call that the tracker was purchased but not set up or installed, that admission is sufficient to uphold the insurer's decision not to pay the theft claim; an insurer should not require a policyholder to continue paying monthly instalments as a condition of processing a theft claim — remaining premium can be deducted from any claim payment; a premium refund following policy cancellation should be backdated to the date the insurer made its decision on the theft claim, not before, as the policyholder had the benefit of third-party liability cover up to that date",
        "Missing Evidence": "Evidence that the tracker was specifically fitted to and registered to Mr S' car and that its subscription was maintained at the time of theft (absent and decisive — Mrs K's own statement confirmed it was not set up or installed)",
        "Ombudsman Reasoning": "Mrs K confirmed in a September 2023 call that Mr S bought a tracker but it didn't work and wasn't set up or installed in the car; subsequent dealership invoice for £500 did not specify it was for a tracker; app screenshots showed a basic tracking device not registered to Mr S' car or any car; on balance no working tracker was fitted as required by the policy; Aviva was entitled to decline the theft claim; telling Mr S to continue paying instalments was not wrong in principle (premium is payable upfront — instalments are a concession; remaining premium would have been deducted from any claim payment); Aviva adviser wrongly said claim wouldn't be paid if premium stopped but no compensation warranted; premium refund correctly backdated to October 2023 (date of claim decision) — Mr S had the benefit of third-party liability cover until then; £50 D&I for not responding to backdating email confirmed paid; £75 D&I for auto-renewal committed but not paid — must be paid now",
        "Workflow Insight": "When a policy requires a tracker as a condition of theft cover, handlers must obtain evidence that the device was (1) specifically fitted to the insured vehicle, (2) in operation, and (3) had an active subscription at the time of theft — a generic invoice and app screenshots without vehicle registration proof are insufficient; a named driver's own admission in a call that the tracker was not set up is strong evidence the condition was not met and should be recorded and relied upon; all D&I compensation committed by the insurer must be paid promptly — failure to pay previously committed compensation compounds the complaint and requires FOS to direct payment",
        "AI Rule Candidate": "IF motor_policy_requires_tracker_fitted_in_operation_with_active_subscription AND policyholder_cannot_demonstrate_tracker_registered_to_insured_vehicle AND named_driver_confirms_tracker_not_set_up THEN tracker_condition_not_met_and_theft_claim_excluded; dealership_invoice_not_specifying_tracker_and_app_screenshots_not_registered_to_vehicle_are_insufficient_tracker_evidence; IF claim_declined_and_policy_cancelled THEN remaining_premium_should_be_deducted_from_claim_payment_not_collected_as_condition_of_processing; premium_refund_backdated_to_claim_decision_date_not_date_of_theft_as_policyholder_had_benefit_of_third_party_cover_until_then",
        "Source PDF": "DRN-5901721.pdf",
    },
    {
        "Case ID": "THEFT-032",
        "FOS Decision ID": "DRN5939533",
        "Insurer Name": "Zurich Insurance PLC",
        "FOS Decision Date": "24 Jul 2017",
        "Claim Type": "Home contents insurance — Mr and Mrs G's domestic employee stole multiple items of jewellery from their home over at least two separate occasions; one item was noticed missing in November 2011 and five further items were reported stolen in February 2012; Zurich recorded two separate theft claims and applied two excesses; Mr and Mrs G argued all thefts were committed by the same person (their employee) and should be treated as one claim",
        "Entry / Theft Method": "Theft by trusted domestic employee — employee had access to the home and stole jewellery on at least two separate occasions: one item taken no later than November 2011 and at least one item taken in February 2012; five further items reported missing on the same day as the February 2012 theft with unknown specific dates",
        "Property Type": "Residential home",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "Not applicable — both theft instances were accepted and paid by Zurich; dispute concerns whether two separate excesses were correctly applied by recording the thefts as two claims rather than one",
        "Evidence Dispute": "Mr and Mrs G: all items were taken by the same person (their employee) — should be one claim; FOS Technical Helpdesk had advised their broker that FOS would generally treat this as one claim. Zurich: items were taken at different times — at least two separate theft occasions; could reasonably have recorded six separate claims but chose to treat as two. FOS: Technical Helpdesk advice was general guidance only and not binding on this specific case; items were taken at different times (November 2011 and February 2012 are distinct occasions); same perpetrator does not collapse multiple theft occasions into one claim event; Zurich's two-claim recording was fair and actually favourable (could have recorded up to six); complaint not upheld",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Zurich correctly recorded two separate theft claims as items were taken on at least two different occasions; the same perpetrator does not make multiple theft incidents a single claim; Zurich acted generously in recording only two claims rather than up to six",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Multiple thefts by the same perpetrator do not constitute a single theft claim event if the items were taken on different occasions — each separate occasion of theft is a distinct claim to which a separate excess applies; the number of claim events is determined by the number of distinct theft occasions, not the number of items stolen or the identity of the perpetrator; the FOS Technical Helpdesk's general guidance to a broker about how FOS approaches a type of case is not binding on the ombudsman's determination of the specific complaint; where an insurer cannot pinpoint the exact dates of all thefts but can identify at least two distinct occasions, recording the minimum identifiable number of claim events is fair; same proximate cause (same person) does not determine the number of claim events — timing and occasion of each theft do",
        "Missing Evidence": "Specific dates for all six missing items (only November 2011 and February 2012 occasions were clearly established) — absence of dates supported Zurich's decision to record two rather than up to six separate claims",
        "Ombudsman Reasoning": "Items were taken at different times — November 2011 and February 2012 are at least two separate theft incidents; same perpetrator does not make it one claim event; Technical Helpdesk gave general guidance not a specific ruling on this case and had not seen all the evidence; Zurich could reasonably have recorded up to six separate claims but fairly chose to treat as two; complaint not upheld",
        "Workflow Insight": "When a policyholder reports multiple items stolen by the same person over an extended period, handlers must identify and record each distinct theft occasion as a separate claim event — the number of identified occasions (not items) is the basis for claim counting; where dates of some thefts are uncertain, recording the minimum number of clearly identifiable occasions rather than the maximum is a reasonable and favourable approach; claims handlers must not conflate 'same perpetrator' with 'same incident' — they are different concepts",
        "AI Rule Candidate": "IF multiple_items_stolen_by_same_person AND thefts_occurred_on_different_occasions THEN each_occasion_constitutes_separate_claim_event_and_separate_excess_applies; same_perpetrator_does_not_make_multiple_theft_occasions_into_one_claim; IF exact_dates_of_some_thefts_unknown THEN recording_minimum_number_of_confirmed_occasions_is_fair_and_favourable; fos_technical_helpdesk_guidance_is_general_and_not_binding_on_specific_complaint_outcome",
        "Source PDF": "DRN5939533.pdf",
    },
    {
        "Case ID": "THEFT-033",
        "FOS Decision ID": "DRN-5954307",
        "Insurer Name": "Advantage Insurance Company Limited",
        "FOS Decision Date": "7 Jan 2026",
        "Claim Type": "Motor insurance — Mr M purchased a policy in March 2025 and made a claim in September 2025; Advantage discovered he had failed to disclose a theft claim from 28 February 2025 (less than one month before policy inception) when Miss M (named driver) took out the policy and answered 'no' to a clear question asking about all motor accidents, claims or losses in the past 5 years including unsettled claims; Advantage applied a CIDRA proportionate settlement at 71%",
        "Entry / Theft Method": "Motor vehicle theft — vehicle subject to September 2025 claim; method of theft not described; a prior undisclosed motor theft claim from 28 February 2025 triggered the CIDRA proportionate settlement dispute",
        "Property Type": "Motor vehicle (personal car insurance)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "Not applicable — claim not declined; Advantage applied a CIDRA proportionate settlement at 71% based on the ratio of premium actually paid (£1,401.47) to the premium that should have been charged had the prior theft claim been disclosed (£1,975.21); careless misrepresentation classified as a qualifying misrepresentation",
        "Evidence Dispute": "Miss M: the question was ambiguous; she was going through a very difficult personal period and any error was a genuine mistake not a failure to take reasonable care; she was not aware the prior unsettled claim had to be declared. Advantage: prior theft claim from 28 February 2025 was not declared; would have charged a higher premium if disclosed. FOS: the question is clear and explicitly states it includes unsettled claims; the prior theft claim occurred less than one month before policy inception — a reasonable person would have declared it; personal difficulties acknowledged but do not override failure to take reasonable care; misrepresentation is qualifying (Advantage would have charged more); careless misrepresentation; 71% proportionate settlement is fair",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Advantage correctly identified a careless qualifying misrepresentation under CIDRA; proportionate settlement at 71% (£1,401.47 / £1,975.21) is fair and reasonable",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "Under CIDRA, an insurer may settle a claim proportionately where: (1) the consumer failed to take reasonable care not to make a misrepresentation; (2) the misrepresentation is qualifying (the insurer would have offered the policy on different terms or not at all); and (3) the misrepresentation was careless; a clear disclosure question that explicitly includes unsettled claims means that a consumer who fails to declare a motor claim made less than one month before policy inception has not taken reasonable care, regardless of personal circumstances at the time; the CIDRA proportionate settlement ratio is calculated as: (premium actually paid) divided by (premium that should have been charged), applied to the claim settlement; personal difficulties at the time of taking out a policy are relevant context but do not automatically constitute a defence to a careless misrepresentation",
        "Missing Evidence": "No material missing evidence — undisclosed theft claim confirmed by Advantage; premium differential supported by underwriting evidence; CIDRA analysis straightforward on the facts",
        "Ombudsman Reasoning": "Question was clear and explicitly included unsettled claims; prior theft claim occurred 28 February 2025 — less than one month before March 2025 policy inception; a reasonable person in this position would have declared it; Miss M's difficult personal circumstances are acknowledged but do not override the failure to take reasonable care; misrepresentation is qualifying (Advantage would have charged more — £1,975.21 vs £1,401.47 actually charged); Advantage treated the misrepresentation as careless — appropriate; CIDRA entitles Advantage to pay 71% of the claim; complaint not upheld",
        "Workflow Insight": "CIDRA proportionate settlement is the correct remedy for a careless misrepresentation where the insurer would have charged a higher premium — calculate the ratio of premium paid to premium that should have been charged and apply it to the claim settlement; a disclosure question that explicitly names unsettled claims removes any ambiguity — even a claim made days before policy inception must be declared; a consumer's difficult personal circumstances at the time of taking out a policy are relevant context but do not automatically constitute a defence to careless misrepresentation",
        "AI Rule Candidate": "IF consumer_fails_to_disclose_prior_motor_claim_when_buying_motor_policy AND question_explicitly_includes_unsettled_claims THEN failure_to_disclose_is_careless_misrepresentation; IF CIDRA_qualifying_misrepresentation_is_careless AND insurer_would_have_charged_higher_premium THEN proportionate_settlement_equals_premium_paid_divided_by_premium_that_should_have_been_charged; prior_motor_claim_less_than_one_month_before_policy_inception_must_be_declared_regardless_of_settlement_status; personal_difficulties_at_time_of_application_are_relevant_context_but_do_not_negate_duty_of_reasonable_care",
        "Source PDF": "DRN-5954307.pdf",
    },
    {
        "Case ID": "THEFT-034",
        "FOS Decision ID": "DRN6242378",
        "Insurer Name": "Covea Insurance plc",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": "Commercial motor insurance taken out through a broker — Mr M's vehicle suffered what appeared to be fire damage (found to be an electrical fault not covered under the policy); Covea voided the policy when it discovered Mr M had not disclosed a previous vehicle theft claim made two years earlier that was subsequently withdrawn after the vehicle was recovered with no damage; the broker had asked whether Mr M had any motoring incidents, accidents or claims including theft in the last 5 years whether or not a claim was made; Mr M believed a withdrawn claim with no payment did not need to be disclosed; dispute before FOS limited to the voidance (vehicle disposal complaint withdrawn by Mr M)",
        "Entry / Theft Method": "Not applicable to the current claim — the claim before Covea was for fire/electrical fault damage not theft; the prior theft referenced was a vehicle theft two years earlier that was reported to the insurer and then withdrawn when the vehicle was recovered with no damage",
        "Property Type": "Commercial (commercial motor insurance — not home/contents)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Broker asked clear question covering all motoring incidents including theft whether or not a claim was made; Mr M failed to disclose a previous vehicle theft claim made two years earlier (withdrawn after vehicle recovery with no damage); Covea voided the policy on non-disclosure grounds; fire/electrical fault damage was separately not covered under the policy terms regardless of voidance",
        "Evidence Dispute": "Mr M: genuinely believed a withdrawn claim with no payment did not need to be disclosed; didn't claim so shouldn't count. Covea: clear question asked; non-disclosure was deliberate or reckless; even notification-only disclosure would have increased the premium. Adjudicator: withdrawn claim was a notification-only incident not a settled claim; Covea would still have insured Mr M at slightly increased premium; voidance was unfair; fire/electrical fault claim correctly declined; recommended voidance removal and £200 compensation plus salvage amounts for disposal. FOS: Mr M was asked a clear question and the withdrawn claim should have been disclosed as an incident; non-disclosure was not deliberate or reckless; Covea would still have insured Mr M had it been disclosed as a notification-only incident (at slightly increased premium); full voidance was disproportionate and unfair; fire/electrical fault claim correctly declined regardless; complaint upheld in part",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Covea to remove the voidance from Mr M's record and pay £200 compensation for the distress and inconvenience caused by the unfair voidance; fire/electrical fault claim correctly declined as not a covered peril — no remedy for that element",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "An insurer cannot void a policy for non-disclosure of an incident unless: (1) the consumer was asked a clear question; AND (2) had the true facts been disclosed, the insurer would not have offered insurance on the same terms or at all; where the insurer would still have offered insurance (at a higher premium) had the withdrawn claim been disclosed as a notification-only incident, a full voidance is disproportionate — the correct remedy is a premium adjustment not voidance; a withdrawn theft claim where no payment was made is nonetheless a disclosable incident where the broker's question asks about all motoring incidents including theft regardless of whether a claim was made; failure to disclose a withdrawn claim that the policyholder genuinely believed was not disclosable is careless not deliberate or reckless",
        "Missing Evidence": "Covea's underwriting evidence as to what premium would have been charged had Mr M disclosed the withdrawn theft claim as a notification-only incident (adjudicator obtained and reviewed — confirmed Covea would still have offered cover at a slightly increased premium)",
        "Ombudsman Reasoning": "Mr M was asked a clear and unambiguous question by his broker covering all incidents including theft whether claimed or not; the withdrawn theft claim was an incident that should properly have been disclosed; Mr M's belief it was not disclosable is understandable but non-disclosure was not deliberate or reckless; Covea's underwriting criteria show it would still have insured Mr M had the incident been disclosed as notification-only at a slightly increased premium; full voidance was therefore disproportionate and unfair; fire/electrical fault damage is separately not covered under the policy regardless; Covea to remove voidance and pay £200 D&I compensation",
        "Workflow Insight": "A full policy voidance is only appropriate where the insurer demonstrates it would not have offered cover at all had the true facts been disclosed — where it would have offered cover at a higher premium, the correct remedy is a premium adjustment; handlers must check underwriting criteria before voiding for non-disclosure; a withdrawn claim where the vehicle was recovered with no payment is still a disclosable incident where the question asks about all incidents including theft regardless of whether a claim was made",
        "AI Rule Candidate": "IF insurer_would_still_have_offered_cover_at_higher_premium_had_true_facts_been_disclosed THEN full_voidance_is_disproportionate_and_only_premium_adjustment_is_permitted; IF policy_question_asks_about_all_incidents_including_theft_whether_claimed_or_not THEN withdrawn_theft_claim_is_disclosable_regardless_of_no_payment; non_disclosure_of_withdrawn_claim_policyholder_believed_undisclosable_is_careless_not_deliberate; voidance_requires_proof_insurer_would_not_have_offered_insurance_at_all_not_merely_on_different_terms",
        "Source PDF": "DRN6242378.pdf",
    },
    {
        "Case ID": "THEFT-035",
        "FOS Decision ID": "DRN6596319",
        "Insurer Name": "UK General Insurance (Ireland) Limited",
        "FOS Decision Date": "2 Oct 2015",
        "Claim Type": "Home insurance arranged by telephone May 2013 — Mr B disclosed a legal expenses claim but failed to disclose a prior home theft claim; made a claim under the policy January 2014; UK General cancelled the policy from the start date when it discovered the undisclosed theft claim, stating it would not have offered insurance had the theft claim been disclosed; Mr B argued he forgot about the theft claim and that disclosing a larger legal expenses claim should have covered his obligation",
        "Entry / Theft Method": "Home theft — specific circumstances of the prior theft claim not described in this decision; the dispute concerns non-disclosure of that prior theft claim when purchasing home insurance in May 2013; the January 2014 claim type is also not specified",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Undisclosed prior home theft claim — UK General's underwriting criteria show it would not have offered Mr B insurance had the theft claim been disclosed; Mr B failed to disclose the theft claim during the application call despite clear and unhurried questions asking about any home insurance claims and reported losses; answered the claims question by referencing only his legal expenses claim and confirmed no home insurance claims existed when a follow-up question was asked",
        "Evidence Dispute": "Mr B: forgot about the theft claim; disclosed the bigger legal expenses claim; questions were unclear; advisor did not take sufficient care; now having difficulty getting insurance. UK General: clear questions asked about all home insurance claims and losses; theft claim would have made a difference to the offer of insurance; Mr B's answers confirmed no home insurance claims. FOS: clear and unhurried questions were asked about any home insurance claims or reported losses; the word 'any' meant all claims not just the biggest; Mr B answered 'no, but' and mentioned only the legal expenses claim; a follow-up question confirmed no home insurance claims — this should have prompted disclosure of the theft claim; UK General's underwriting criteria show the theft claim would have prevented an offer of insurance; legal expenses policy was with a different underwriter so its claims did not affect UK General's home insurance underwriting criteria; Mr B acted carelessly not deliberately; UK General fairly cancelled the policy from inception",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — UK General fairly and reasonably cancelled Mr B's home insurance policy from inception; Mr B failed to take reasonable care to disclose a prior home theft claim despite clear and unhurried questions being asked during the application call",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "A policyholder asked a clear and unhurried question about 'any' home insurance claims or reported losses must disclose all relevant prior claims — disclosing one (larger) claim does not discharge the obligation to disclose all claims; a follow-up confirmation question ('as long as there were no claims on the home insurance itself or reported losses…is that right') reinforces the duty to disclose and if answered incorrectly does not assist the policyholder; forgetting a prior theft claim when answering a clear question is careless misrepresentation not an absence of misrepresentation; where the insurer demonstrates it would not have offered cover had the prior theft claim been disclosed, cancellation from inception is proportionate and fair; home insurance and legal expenses insurance arranged in the same call may be with different underwriters — the underwriting criteria for each apply independently and a claim under one does not affect underwriting eligibility under the other",
        "Missing Evidence": "Specific details of the prior theft claim (amount, circumstances, date) — absent but UK General's underwriting criteria confirmed that a theft claim of Mr B's amount would have made a difference to their decision to offer insurance",
        "Ombudsman Reasoning": "UK General asked clear and unhurried questions about any home insurance claims or losses; Mr B answered 'no, but' and described only his legal expenses claim; a follow-up question confirmed no home insurance claims — Mr B confirmed it was fine; legal expenses policy was with a different underwriter — UK General's home insurance underwriting criteria did not take legal expenses claims into account; a theft claim in Mr B's amount would have made a difference to UK General's offer of insurance; Mr B didn't act deliberately but acted carelessly; UK General fairly cancelled the policy from inception; complaint not upheld",
        "Workflow Insight": "When assessing non-disclosure in home insurance, handlers must specifically check whether the insurer's underwriting criteria treat the undisclosed claim as a difference-maker — not all undisclosed claims warrant voidance; where a consumer discloses one claim but omits another during the same call, the insurer can rely on a clear follow-up question that the consumer confirms negatively; a policyholder who mentions a claim under one product (legal expenses) arranged in the same call does not thereby discharge disclosure obligations under a separate product (home insurance) sold by a different underwriter",
        "AI Rule Candidate": "IF policyholder_asked_clear_question_about_any_home_insurance_claims AND fails_to_disclose_prior_theft_claim THEN non_disclosure_is_at_minimum_careless; IF insurer_demonstrates_undisclosed_prior_theft_claim_would_have_prevented_offer_of_cover THEN cancellation_from_inception_is_fair; disclosing_larger_claim_under_different_product_in_same_call_does_not_discharge_disclosure_of_smaller_claim_under_home_insurance; legal_expenses_claims_do_not_affect_home_insurance_underwriting_criteria_when_policies_are_with_different_underwriters",
        "Source PDF": "DRN6596319.pdf",
    },
    {
        "Case ID": "THEFT-036",
        "FOS Decision ID": "DRN6785039",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": "Home insurance commenced online March 2010 — Mr B made a theft claim December 2011 and a fire damage claim March 2012; RSA checked an external claims database and discovered Mr B had four previous claims (March 2009, August 2009, October 2009, January 2010) but had only disclosed one (October 2009) in the online application; RSA voided the policy from March 2010 and sought reimbursement of £1,486.22 it had already paid toward the December 2011 theft claim; Mr B disputed that the application was made online (claiming it was by telephone) and that he had disclosed all claims",
        "Entry / Theft Method": "Home theft — specific circumstances of the December 2011 theft claim not described; RSA partially settled the claim before discovering the non-disclosure and avoiding the policy",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Non-disclosure of three of four previous claims (March 2009, August 2009, January 2010) made within three years of policy inception — only October 2009 claim was disclosed; online application asked clearly for all claims in the previous three years; RSA demonstrated it would not have offered cover if all four previous claims had been disclosed; RSA also sought reimbursement of £1,486.22 already paid toward the December 2011 theft claim",
        "Evidence Dispute": "Mr B: application was made by telephone not online; disclosed all claims during the telephone call; RSA's online application documentation is inaccurate; Mr B does not complete application forms himself. RSA: online application exists in Mr B's name in March 2010 showing only one claim disclosed; IP address traced to Mr B's area; four prior claims found on external database at time of fire claim investigation. FOS: balance of probabilities the application was made online — online application record exists, IP address is consistent with Mr B's area, and Mr B cannot provide telephone records; question about previous claims was clear; failure to disclose three claims was inadvertent; RSA demonstrated it would not have offered cover if all four claims disclosed — voidance is reasonable; RSA chose not to check external database at inception — unreasonable to now pursue reimbursement of claims paid during the policy period; RSA accepted it will not pursue the £1,486.22; complaint upheld in part",
        "Outcome Category": "Upheld in Part",
        "Outcome": "RSA required not to pursue Mr B for reimbursement of its outlay of £1,486.22 in respect of the December 2011 theft claim; no other award — voidance stands, theft and fire claims are not paid, no compensation awarded",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Where a consumer inadvertently fails to disclose previous claims in response to a clear question on an insurance application, this service will generally expect an insurer to apply a proportionate response — where the insurer demonstrates it would not have offered cover at all, avoidance from inception is proportionate; an insurer that chooses not to check a policyholder's external claims database at inception cannot subsequently seek reimbursement of claim payments made during the policy period when it later avoids the policy — the delay in checking is the insurer's own commercial decision and the resulting cost cannot be transferred to the policyholder; where an application method (online vs telephone) is disputed and there is evidence of both an online application record and subsequent phone calls (none of which concerned prior claims disclosure), the balance of probabilities test is applied — an online application record in the consumer's name supported by an IP address consistent with their location outweighs an unverified claim of telephone disclosure not supported by any telephone records; renewal documents listing disclosed claims constitute evidence of the claims history maintained at each renewal",
        "Missing Evidence": "Evidence of a telephone call at policy inception during which Mr B disclosed all four prior claims — absent; no telephone records provided by Mr B to support the telephone disclosure claim; RSA traced four calls from Mr B's landline but all post-inception and none concerned prior claims",
        "Ombudsman Reasoning": "Balance of probabilities — online application made; online question asked clearly for all claims in previous three years; only October 2009 claim of four was disclosed; failure to disclose three other claims was inadvertent; RSA demonstrated it would not have offered cover if all four disclosed — avoidance from inception was reasonable; RSA chose not to check external database at application inception — unreasonable to pursue reimbursement of claim payments made during the policy period; RSA already accepted it would not pursue the £1,486.22; upheld in part — only remedy is that RSA must not pursue reimbursement; no compensation, voidance stands",
        "Workflow Insight": "Insurers who do not conduct external claims database checks at inception take on the risk that they have paid claims they cannot recover if they later discover non-disclosure and avoid the policy — this cost cannot be transferred to the policyholder after the fact; where a dispute arises about whether an application was made online or by telephone, the insurer's burden is to demonstrate on the balance of probabilities that the documented online application is authentic; a claimed telephone disclosure that cannot be supported by any telephone records is unlikely to displace a contemporaneous online application record in the consumer's name",
        "AI Rule Candidate": "IF insurer_chooses_not_to_check_external_claims_database_at_inception AND later_discovers_non_disclosure_and_avoids_policy THEN insurer_cannot_recover_claim_payments_made_during_policy_period; IF online_application_record_exists_in_policyholders_name_and_IP_address_matches_policyholders_area THEN balance_of_probabilities_supports_online_application_over_unverified_telephone_claim; inadvertent_non_disclosure_of_prior_claims_entitles_insurer_to_void_if_insurer_demonstrates_it_would_not_have_offered_cover_at_all; renewal_documents_listing_disclosed_claims_constitute_evidence_of_claims_history_maintained_through_policy_renewals",
        "Source PDF": "DRN6785039.pdf",
    },
    {
        "Case ID": "THEFT-037",
        "FOS Decision ID": "DRN7287812",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": "Motor insurance — Mr A's vehicle was stolen; RSA avoided the policy from inception for non-disclosure of an undisclosed 2008 theft claim (and a named driver's 2009 theft claim) and refused to meet the theft claim; RSA provided conflicting and inaccurate information throughout including conflicting accounts of how the claims were discovered and a letter from RSA's agent falsely stating Mr A had previously been found guilty of fraud; RSA's information failures caused Mr A to inadvertently remove the wrong claim from his CUE record and denied him the opportunity to seek alternative cover, leaving him uninsured when he had a separate accident in 2012",
        "Entry / Theft Method": "Motor vehicle theft — vehicle stolen; specific circumstances of the theft not described; principal dispute concerns RSA's avoidance of the policy for non-disclosure and the consequential uninsured accident loss arising from RSA's handling failures",
        "Property Type": "Motor vehicle (personal car insurance)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Undisclosed 2008 theft claim (Mr A's own) and named driver's 2009 theft claim — RSA demonstrated it would not have offered cover had the 2008 theft claim been disclosed; policy avoided from inception; theft claim not paid",
        "Evidence Dispute": "Mr A: RSA provided conflicting information about how claims were discovered (claimed CUE was checked when evidence suggests it wasn't); RSA gave inaccurate claim dates that confused Mr A into inadvertently removing the real 2010 claim from his CUE record with the previous insurer; RSA falsely stated he had been found guilty of fraud causing business harm. RSA: Mr A failed to disclose 2008 theft claim in answer to a clear 5-year claims question; avoidance was correct; previous insurer confirmed two claims. FOS: non-disclosure of 2008 theft claim is undisputed — avoidance is reasonable and voidance may remain; however RSA created significant confusion through conflicting and inaccurate information; RSA could have avoided the policy in August 2011 when it had the information — Mr A could then have sought alternative cover and been insured at the time of his 2012 accident; RSA's inaccurate information caused Mr A to remove the wrong claim; false fraud allegation caused significant distress and inconvenience — £800 compensation appropriate",
        "Outcome Category": "Upheld in Part",
        "Outcome": "RSA to pay Mr A's 2012 accident damage claim with 8% simple interest from date of loss to date of settlement; RSA to pay £800 compensation for distress and inconvenience caused by poor customer service including false fraud allegation; voidance may remain on Mr A's record; no other award",
        "Compensation Awarded (£)": 800,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "An insurer that provides conflicting and inaccurate information about the basis for a non-disclosure avoidance may be required to pay claims it would otherwise be entitled to decline — where the insurer's own information failures prevented the policyholder from taking corrective action (seeking alternative cover) and the policyholder was consequently uninsured for a subsequent loss, the insurer bears responsibility for that loss; an insurer must not state in official correspondence that a policyholder has previously been found guilty of fraud when this is false — doing so causes significant reputational and business harm and warrants substantial D&I compensation; while an insurer may legitimately avoid a policy for undisclosed theft claims, inaccurate information about how those claims were discovered can give rise to consequential liability if it prevents the policyholder from responding appropriately and seeking alternative cover",
        "Missing Evidence": "Recording of the pivotal August 2011 phone call between RSA and Mr A (not available to FOS); confirmed account from previous insurer of exactly what was communicated to RSA (disputed accounts between RSA and previous insurer)",
        "Ombudsman Reasoning": "Non-disclosure of 2008 theft claim undisputed — RSA entitled to avoid policy; but RSA provided conflicting and inaccurate information throughout (claimed CUE was checked but evidence strongly suggests it was not; gave inaccurate claim dates to Mr A); RSA's inaccurate information caused Mr A to contact previous insurer and inadvertently change the wrong claim record; RSA could have avoided the policy in August 2011 with the information it had — Mr A would then have been able to seek alternative cover and been insured for the 2012 accident; RSA should therefore pay for the accident damage; false fraud allegation caused significant distress and potential business harm — £800 compensation appropriate; voidance stands",
        "Workflow Insight": "When communicating with a policyholder about a non-disclosure avoidance, insurers must provide accurate and consistent information about how the undisclosed claims were discovered — conflicting accounts of the investigation undermine the avoidance process and may result in consequential liability for subsequent uninsured losses; a single false allegation of fraud in official correspondence can expose an insurer to substantial D&I compensation even where the underlying avoidance decision is correct; insurers should act promptly once they identify non-disclosure — delay in avoiding a policy can extend the period during which the policyholder remains uninsured and unable to seek alternative cover",
        "AI Rule Candidate": "IF insurer_provides_conflicting_inaccurate_information_about_non_disclosure_investigation AND policyholder_cannot_seek_alternative_cover_as_a_result THEN insurer_may_be_liable_for_subsequent_uninsured_losses; IF insurer_sends_official_correspondence_falsely_alleging_prior_fraud_conviction THEN substantial_DI_compensation_warranted_regardless_of_underlying_avoidance_correctness; insurer_must_provide_accurate_consistent_account_of_how_undisclosed_claims_were_discovered; IF insurer_could_have_avoided_policy_earlier_with_correct_information_and_did_not THEN insurer_bears_responsibility_for_losses_that_occurred_in_gap_during_which_policyholder_was_uninsured",
        "Source PDF": "DRN7287812.pdf",
    },
    {
        "Case ID": "THEFT-038",
        "FOS Decision ID": "DRN8409372",
        "Insurer Name": "Zurich Insurance PLC",
        "FOS Decision Date": "9 Jan 2017",
        "Claim Type": "Home contents insurance — Mr and Mrs G complained that Zurich unfairly applied an increased compulsory excess (a special endorsement of £500 for any future jewellery loss) at their 2015 renewal in response to their recent claims history including theft claims; the endorsement required disclosure to prospective insurers and they said it prevented them obtaining competitive renewal quotes elsewhere",
        "Entry / Theft Method": "Jewellery theft from home — the underlying theft claims (previous incidents recorded by Zurich) involved the theft of jewellery from the home; specific method not described; this case concerns the consequence of those recorded claims on renewal terms, not the circumstances of the theft itself",
        "Property Type": "Residential home",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": "Not applicable — no current theft claim declined; dispute concerns Zurich's decision to apply a compulsory £500 jewellery excess endorsement at renewal based on Mr and Mrs G's recent claims history (including theft claims)",
        "Evidence Dispute": "Mr and Mrs G: only one theft claim was made not two (addressed in a separate FOS complaint reference); if only one claim recorded, only one special term should have been applied; Zurich should remove the increased excess both historically and for the future. Zurich: three or more claims triggers referral to underwriters; renewal terms are offered at underwriters' discretion; increased excess fairly applied in line with internal guidelines; almost two months' notice given. FOS: Zurich's underwriting criteria show policyholders with three or more claims are referred to underwriters and renewal terms are at underwriters' discretion; Zurich demonstrated the endorsement was applied consistently with its internal guidelines; annual contract — no obligation to renew on the same terms as previous years; two months' notice to broker is reasonable; endorsement requiring disclosure to prospective insurers is how insurance works — some insurers declining to offer cover because of the endorsed excess does not make the endorsement unfair; question of whether one or two theft claims were correctly recorded is being addressed under a separate FOS complaint reference; complaint not upheld",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Zurich fairly applied an increased jewellery excess endorsement at renewal in line with its underwriting criteria and internal guidelines; annual contracts can be renewed on different terms; two months' notice was reasonable; the endorsement's required disclosure to prospective insurers is a standard insurance consequence, not an unfair practice in itself",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Administrative",
        "Key Policy Clause": "An insurer is entitled to apply a special excess endorsement or increase the policy excess at renewal based on the policyholder's claims history — there is no obligation to renew an annual home insurance contract on the same terms as previous years; a special endorsement applied at renewal must be disclosed to prospective insurers when the policyholder seeks alternative quotes — the fact that disclosure makes other insurers less willing to offer competitive cover does not make the original endorsement unfair; where an insurer's internal underwriting criteria require referral to underwriters for policyholders with three or more claims and renewal terms are at underwriters' discretion, an endorsement applied consistently with those criteria is a reasonable exercise of underwriting judgment; a dispute about whether the number of underlying claims was correctly recorded is a separate complaint that must be brought independently",
        "Missing Evidence": "Resolution of the parallel complaint about whether one or two theft claims were correctly recorded — FOS accepted Zurich's claim count of two for the purposes of this renewal terms complaint pending outcome of the separate FOS reference",
        "Ombudsman Reasoning": "Zurich's evidence shows three or more claims triggers underwriter referral and renewal terms offered at underwriter discretion; increased jewellery excess endorsed to policy consistently with internal guidelines; annual contract — no renewal on same terms obligation; almost two months' notice given to broker before renewal (reasonable); endorsement requires disclosure to prospective insurers — standard insurance practice; some prospective insurers declining due to disclosed endorsement does not make the endorsement unfair; complaint about claim count being looked at under a separate reference; complaint not upheld",
        "Workflow Insight": "When applying special excess endorsements at renewal based on claims history, insurers must document that they have followed their own underwriting criteria for referral and endorsement application — consistent application of internal guidelines is the primary defence to a challenge that the endorsement is unfair; handlers should advise policyholders at renewal that any special endorsements must be disclosed to prospective insurers when seeking alternative quotes, managing expectations about market availability",
        "AI Rule Candidate": "IF claims_history_triggers_insurer_referral_to_underwriters_per_internal_criteria AND underwriter_applies_increased_excess_endorsement THEN renewal_endorsement_is_fair_exercise_of_underwriting_discretion; special_excess_endorsement_must_be_disclosed_to_prospective_insurers_when_seeking_alternative_quotes; fact_that_prospective_insurers_decline_due_to_disclosed_endorsement_does_not_make_the_endorsement_unfair; dispute_about_number_of_underlying_theft_claims_correctly_recorded_must_be_brought_as_separate_complaint",
        "Source PDF": "DRN8409372.pdf",
    },
    {
        "Case ID": "THEFT-039",
        "FOS Decision ID": "DRN8838516",
        "Insurer Name": "Saga Services Limited",
        "FOS Decision Date": "16 Apr 2018",
        "Claim Type": "Motor (car) insurance — Mr W's car insurance premium increased by approximately 20% at renewal; a further increase was applied when Mr W disclosed the theft of a motorcycle insured under a separate motor policy; Mr W complained that a motorcycle theft claim on a different policy should not affect his car insurance premium",
        "Entry / Theft Method": "Motor vehicle theft (motorcycle) on a separate motor policy — motorcycle stolen and claimed under a different motor insurance policy; specific circumstances of theft not described; the motorcycle theft claim was disclosed at car insurance renewal and factored into the car insurance pricing by Saga's underwriters",
        "Property Type": "Motor vehicle (personal car insurance)",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": "Not applicable — no claim declined; dispute concerns Saga's decision to increase car insurance renewal premium following disclosure of a motorcycle theft claim on a separate motor policy",
        "Evidence Dispute": "Mr W: unfair that a motorcycle theft claim on a separate policy affects his car insurance premium; pricing criteria must be underhand if they cannot be disclosed; surcharge is not morally fair. Saga: motorcycle theft is a motor claim that underwriters factor into any motor insurance pricing; pricing criteria are commercially sensitive but applied consistently to all policyholders in comparable circumstances; £37 loyalty discount offered. FOS: Saga's pricing information was reviewed and found to be applied consistently to all policyholders in comparable circumstances; motorcycle theft is a motor claim relevant to all motor insurance (both motorcycles and cars are motor vehicles requiring insurance for legal road use); pricing algorithms and criteria are legitimately commercially sensitive — insurers are not required to disclose them to policyholders; no unfair treatment of Mr W",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Saga fairly increased car insurance renewal premium to reflect the motorcycle theft claim on a separate motor policy; insurers are entitled to consider any motor claim (including claims on other motor policies) when pricing motor insurance; commercially sensitive pricing criteria do not need to be disclosed to policyholders",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Commercial",
        "Key Policy Clause": "An insurer may legitimately factor in claims made under separate motor insurance policies when pricing a motor insurance renewal — motorcycles and cars are both motor vehicles requiring insurance for legal road use and claims on either are relevant risk factors; commercially sensitive pricing algorithms and criteria do not need to be disclosed to policyholders — the insurer need only demonstrate to FOS that the pricing is applied consistently to all comparable policyholders; a motorcycle theft claim on a separate policy is correctly classified as a fault/bonus-disallowed claim where the insurer cannot recover the claim payment from a third party, and it is not unreasonable for a subsequent insurer to factor this into motor insurance pricing",
        "Missing Evidence": "No material missing evidence — Saga's pricing criteria were reviewed by FOS (as commercially sensitive information) and found to be consistently applied to all comparable policyholders",
        "Ombudsman Reasoning": "Saga demonstrated its pricing was applied consistently to all comparable policyholders; motorcycle theft is a motor claim relevant to car insurance pricing — both are motor vehicles requiring legal insurance; pricing information is legitimately commercially sensitive and does not need to be disclosed to Mr W; £37 loyalty discount was offered; renewal price was given in advance and then updated when Mr W disclosed the motorcycle theft; no unfair treatment; complaint not upheld",
        "Workflow Insight": "Motor insurers may lawfully factor claims on any motor vehicle (including motorcycles or vehicles insured under different policies) into renewal pricing for car insurance — both are motor vehicles requiring legal road insurance; when a policyholder challenges a premium increase based on a claim on a separate motor policy, the insurer should demonstrate that its pricing criteria are consistently applied to comparable policyholders using commercially sensitive information reviewed by FOS if necessary",
        "AI Rule Candidate": "IF motorcycle_theft_claim_exists_on_separate_motor_policy AND car_insurance_is_being_renewed THEN motorcycle_theft_claim_is_a_relevant_motor_risk_factor_that_can_be_factored_into_car_insurance_premium; commercially_sensitive_pricing_criteria_need_not_be_disclosed_to_policyholders_if_demonstrated_to_be_consistently_applied; theft_claim_on_any_motor_vehicle_where_insurer_cannot_recover_costs_is_fault_bonus_disallowed_claim_relevant_to_all_future_motor_insurance_pricing",
        "Source PDF": "DRN8838516.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "F5E6E6" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Theft_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
