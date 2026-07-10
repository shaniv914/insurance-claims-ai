"""
Append script for Unoccupied Property batches (schema v1, 21 columns — same
schema as EOW v2 / Storm v1 / Flood v1 / Subsidence v1 / Theft v1; column 6 =
"Unoccupied Period / Circumstance").

v2 supersedes append_unoccupied_property_v1.py for Batch 4 onward (v1 is left
untouched as the historical record of Batches 1-3). Reuse this script for
each future batch: replace NEW_CASES below and run again. Appends only —
never modifies existing rows in
knowledge/case-databases/Unoccupied_Property_Case_Database.xlsx.

Current contents (this run): Batch 6 — FINAL batch (UNOC-051 to UNOC-060).
After this run the Unoccupied Property case database is complete (60/60).
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Unoccupied Period / Circumstance",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": [
        "Coverage Dispute",
        "Endorsement / Exclusion Challenge",
        "Broker Conduct Dispute",
        "Claim Recording / Administrative Dispute",
        "Claim Quantum Dispute",
        "Third-Party / Liability Dispute",
        "Causation Dispute",
    ],
    "Coverage Decision": [
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — With Deductions",
        "Not Applicable",
    ],
    "Outcome Category": [
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Not Applicable",
    ],
    "Is Core Case": [
        "Yes",
        "No — Administrative",
        "No — Broker Dispute",
        "No — Quantum Only",
        "No — Time-Barred",
    ],
}

ROW_FILL_ODD  = PatternFill(start_color="F9EDF2", end_color="F9EDF2", fill_type="solid")
ROW_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_FONT      = Font(name="Calibri", size=10)
ROW_BORDER    = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

NEW_CASES = [
    {
        "Case ID": "UNOC-051",
        "FOS Decision ID": "DRN7122779",
        "Insurer Name": "Society of Lloyd's",
        "FOS Decision Date": "24 Aug 2019",
        "Claim Type": (
            "Unoccupied residential property policy (executors of a deceased "
            "owner's estate) — escape of water from a burst pipe; insurer declined "
            "citing breach of a policy condition requiring the central heating to be "
            "set to operate continuously at not less than 12°C between 1 November and "
            "31 March"
        ),
        "Unoccupied Period / Circumstance": (
            "Policy taken out by the executors in February 2018 specifically to cover "
            "their late relative's former home, unoccupied pending sale; a buyer had "
            "already been found; burst pipe discovered 10 March 2018, about two weeks "
            "after cover started (24 February 2018)"
        ),
        "Property Type": "Residential property (deceased owner's estate, unoccupied pending sale)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Insurer's loss adjuster didn't think the executors had complied with the "
            "policy condition requiring continuous heating at a minimum of 12°C "
            "between 1 November and 31 March, based on a comparison of utility bills "
            "between the relevant period in 2018 and the same period the year before"
        ),
        "Evidence Dispute": (
            "Executors: they had every reason to keep the heating on given a buyer "
            "had already been found and there'd be little point taking out an "
            "unoccupied property policy for such a short period without intending to "
            "comply with its terms. SOL: gas usage for January-April 2018 (£72.36) was "
            "far lower than the equivalent 2017 period (£327.29, when the property was "
            "occupied by an elderly, infirm former owner and his carer), and it also "
            "implied the executors delayed the loss adjuster's visit to allow time to "
            "turn the heating back on and inflate usage. FOS: the relevant compliance "
            "window was only about two weeks (24 February to 10 March) during which a "
            "thermostatically-controlled system cycles on and off rather than running "
            "continuously, and no hot water use would register either; the loss "
            "adjuster's own April 2018 visit found the thermostat set to 15°C, above "
            "the 12°C minimum; the delay in the loss adjuster's visit was because one "
            "executor was awaiting a hospital operation, not to manipulate usage; the "
            "boiler needed repairs after the leak, giving further reason to leave "
            "heating off pending inspection; and the prior year's higher usage "
            "reflected genuine occupation by an elderly, infirm resident, which would "
            "naturally run warmer than an empty property on a minimum-temperature "
            "setting"
        ),
        "Outcome Category": "Upheld",
        "Outcome": (
            "SOL required to settle the claim (repair costs) subject to the remaining "
            "policy terms and conditions, pay 8% simple interest on any reimbursement "
            "from the date payment was made by the executors, and pay £200 "
            "compensation for trouble and upset; SOL not required to cover ongoing "
            "property-maintenance costs pending sale, as these aren't covered by the "
            "policy"
        ),
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "Where an insurer relies on a lower-than-comparable-period utility bill to "
            "infer breach of a minimum-continuous-heating condition, the burden "
            "remains on the insurer to show the shortfall isn't otherwise explained — "
            "a short compliance window (where a thermostatic system cycles on/off "
            "rather than running continuously and no hot water is used), a "
            "documented reason for delaying reconnection after a loss (e.g. awaiting "
            "boiler repair or the policyholder's ill health), and a comparator period "
            "reflecting materially different circumstances (e.g. genuine occupation by "
            "an elderly resident versus an empty property on a minimum setting) can "
            "all undermine an insurer's inference of non-compliance from bill "
            "comparison alone"
        ),
        "Missing Evidence": (
            "Direct evidence of the thermostat setting at the time of the loss (the "
            "loss adjuster's reading of 15°C was taken a month later, in April 2018) — "
            "not available, requiring the ombudsman to weigh circumstantial usage "
            "evidence instead"
        ),
        "Ombudsman Reasoning": (
            "This was a finely balanced case, but SOL hadn't done enough to show "
            "non-compliance with the heating condition given: the short (roughly "
            "two-week) compliance window before the loss, during which a "
            "thermostatically-controlled system wouldn't run continuously and no hot "
            "water would be used; a documented, credible reason (one executor's "
            "upcoming hospital operation) for delaying the loss adjuster's visit rather "
            "than any attempt to manipulate usage; a plausible reason (awaiting boiler "
            "repair, living some distance away) for not immediately turning heating "
            "back on post-loss; and the prior year's higher comparator usage reflecting "
            "genuine occupation by an elderly, infirm resident rather than the same "
            "type of unoccupied-property heating regime"
        ),
        "Workflow Insight": (
            "When assessing a suspected breach of a minimum-continuous-heating "
            "condition via utility bill comparison, check (a) how short the actual "
            "compliance window was between cover starting and the loss (a "
            "thermostatic system won't show continuous draw), (b) whether there's a "
            "credible non-fraudulent reason for any delay in the insurer's inspection "
            "or in reconnecting services post-loss, and (c) whether the comparator "
            "period reflects a materially different occupancy status (e.g. a "
            "previously occupied elderly resident versus a minimum-heated empty "
            "property) before relying on the bill differential alone"
        ),
        "AI Rule Candidate": (
            "IF insurer_relies_on_lower_utility_usage_compared_to_a_prior_period_to_infer_breach_of_a_minimum_continuous_heating_condition "
            "AND the_actual_compliance_window_before_the_loss_was_short_and_a_thermostatic_system_would_not_show_continuous_draw "
            "AND there_are_credible_non_fraudulent_explanations_for_any_inspection_or_reconnection_delay "
            "THEN the_bill_comparison_alone_does_not_establish_non_compliance_and_the_claim_should_be_settled"
        ),
        "Source PDF": "DRN7122779.pdf",
    },
    {
        "Case ID": "UNOC-052",
        "FOS Decision ID": "DRN7223290",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "5 Jun 2020",
        "Claim Type": (
            "Joint home insurance — escape of water discovered February 2018; "
            "insurer declined citing the unoccupied-property exclusion (not "
            "permanently lived in for more than 60 consecutive days), after an initial "
            "fraud allegation (later retracted) that the policyholder had knowingly "
            "misrepresented the property as occupied"
        ),
        "Unoccupied Period / Circumstance": (
            "Policyholder said he'd been the sole occupant since the property was "
            "refurbished in August 2017, working shifts and having caring "
            "responsibilities elsewhere; insurer's loss adjuster/investigator found a "
            "lack of furniture and food, a council tax record showing the property as "
            "unfurnished/unoccupied since purchase, extremely low water usage "
            "(£3.27 over six months), the policyholder still listed at a different "
            "address on his driving licence, car insurance and electoral roll, and no "
            "evidence he was registered with a doctor or had utilities/broadband near "
            "the insured address"
        ),
        "Property Type": "Residential property (occupancy status disputed; not the policyholder's confirmed sole residence)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Multiple evidentiary indicators (utility usage, council tax status, "
            "driving licence/insurance address, electoral roll, absence of local "
            "medical registration or utility contracts) showed the property wasn't "
            "permanently lived in by the policyholder for more than 60 consecutive "
            "days"
        ),
        "Evidence Dispute": (
            "Mr J: said he lived there normally and had furnished the kitchen, "
            "bedroom and other rooms, was registered to vote there, and used shower "
            "facilities at work explaining low water usage; he wasn't present at the "
            "time of the incident due to bad weather. AXA: retracted its fraud "
            "allegation but maintained the occupancy exclusion applied, citing the "
            "cumulative weight of independent evidence (utility usage, council tax, "
            "driving licence, electoral/GP registration, absence of local utility "
            "contracts) rather than any single factor. FOS: while any one indicator "
            "might be explained individually (e.g. showering at work), the cumulative "
            "pattern — very low water usage, driving licence and car insurance only "
            "updated to the insured address after the incident, no council tax "
            "liability being collected since the purchase date, no local GP "
            "registration, no TV or broadband contracts at the address, and legal "
            "advice sought near his other address rather than the insured one — was "
            "more consistent with him not permanently living there"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — AXA's reliance on the unoccupied-property "
            "exclusion was reasonable given the weight of independent evidence, though "
            "any reference to the retracted fraud allegation should be removed from "
            "the policyholder's insurance records"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "When assessing whether a policyholder was 'permanently living' at a "
            "property for occupancy-exclusion purposes, no single piece of "
            "circumstantial evidence (e.g. low water usage alone, or an "
            "out-of-date driving licence alone) needs to be individually conclusive — "
            "an insurer can rely on the cumulative, mutually-reinforcing weight of "
            "several independent indicators (utility usage, council tax liability "
            "status, official address records across multiple institutions, absence "
            "of local medical/utility registrations, and even where the policyholder "
            "sought legal advice) even where the policyholder offers plausible "
            "individual explanations for some of them"
        ),
        "Missing Evidence": (
            "Contemporaneous evidence corroborating the policyholder's claimed sole "
            "occupancy (e.g. updated official address records, GP registration, local "
            "utility contracts) predating the escape of water — the documents he did "
            "provide (including his driving licence) were dated after the incident"
        ),
        "Ombudsman Reasoning": (
            "Individually, low water usage might be explained by showering at work, "
            "and forgetting to update a driving licence is common — but taken "
            "together with the council tax record (unoccupied since purchase, no "
            "liability collected), the absence of any local GP registration or "
            "utility/broadband contracts, and the fact he sought legal advice and had "
            "his car insurance registered near his other address rather than the "
            "insured one, the cumulative picture was more consistent with him not "
            "permanently living at the insured property; AXA's retraction of the fraud "
            "allegation was appropriate but didn't change the underlying occupancy "
            "conclusion"
        ),
        "Workflow Insight": (
            "When assessing a disputed occupancy claim, don't evaluate each "
            "circumstantial indicator (utility usage, address records, registrations) "
            "in isolation against an innocent explanation — assess the cumulative "
            "pattern across multiple independent sources; a policyholder providing "
            "innocent explanations for individual data points doesn't defeat an "
            "occupancy exclusion if the totality of independent evidence still points "
            "the other way"
        ),
        "AI Rule Candidate": (
            "IF multiple_independent_occupancy_indicators_eg_utility_usage_council_tax_status_official_address_records_local_registrations "
            "ALL_POINT_AWAY_FROM_the_policyholders_claimed_occupancy "
            "THEN the_cumulative_weight_of_evidence_can_support_the_unoccupancy_exclusion_even_where_each_individual_indicator_has_a_plausible_innocent_explanation"
        ),
        "Source PDF": "DRN7223290.pdf",
    },
    {
        "Case ID": "UNOC-053",
        "FOS Decision ID": "DRN7626527",
        "Insurer Name": "Ageas 50 Limited (trading as RIAS plc)",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": (
            "Home insurance (broker mis-sale complaint) — theft claim declined by the "
            "underlying insurer citing the unoccupied-property exclusion (not lived in "
            "for more than 60 consecutive days) after the policyholder returned from "
            "time outside the UK to find contents missing; complaint concerned whether "
            "the broker had adequately brought the exclusion to his attention at sale"
        ),
        "Unoccupied Period / Circumstance": (
            "Policyholder was outside the UK for an extended period; on application "
            "and at each renewal he was asked whether the home would be left "
            "unoccupied for more than 60 consecutive days and answered 'No'; he was "
            "sent a 'Policy Summary' booklet (which he says was the only documentation "
            "he received) confirming exclusions applied for unoccupied properties, and "
            "the broker said a full policy booklet was also sent though the "
            "policyholder said he didn't receive it"
        ),
        "Property Type": "Residential property (occupancy status disputed while policyholder travelled)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a decision by the broker directly — the underlying "
            "insurer declined the theft claim under the unoccupied-property exclusion; "
            "this complaint concerned only whether the broker adequately brought that "
            "exclusion to the policyholder's attention at the point of sale and "
            "renewal"
        ),
        "Evidence Dispute": (
            "Mr G: the 'Policy Summary' booklet he received didn't adequately bring "
            "the unoccupied-property exclusions to his attention, and even though he'd "
            "confirmed at application that the home wouldn't be left empty for more "
            "than 60 days, the significance of that answer hadn't been explained to "
            "him. RIAS: it had asked him directly whether the property would be "
            "unoccupied for more than 60 days (he said no) and sent documentation at "
            "inception and each renewal referencing the unoccupied-property "
            "exclusions. FOS: the 'Policy Summary' itself stated it didn't contain the "
            "full policy conditions and that these were in the separate policy "
            "wording; if Mr G hadn't received that fuller wording, he could reasonably "
            "have called RIAS to request it; the Policy Summary he did receive still "
            "confirmed exclusions applied to unoccupied properties, and the direct "
            "question-and-answer at application/renewal indicated to RIAS that the "
            "unoccupied exclusion wouldn't be a significant limitation for him, given "
            "his own repeated 'No' answers"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — RIAS sufficiently brought the unoccupied-property "
            "exclusion to Mr G's attention through the direct application/renewal "
            "questions and the Policy Summary documentation; no award made"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "Where a broker directly asks a policyholder at both application and each "
            "renewal whether the property will be left unoccupied for more than a "
            "specified number of days, and the policyholder repeatedly confirms it "
            "won't be, this exchange itself is evidence the broker sufficiently "
            "brought the significance of the unoccupied-property exclusion to the "
            "policyholder's attention — even where a summary document (rather than "
            "the full policy wording) is the only documentation the policyholder says "
            "they actually received, provided that summary itself flags that "
            "exclusions apply to unoccupied properties and that fuller wording is "
            "available on request"
        ),
        "Missing Evidence": (
            "Not applicable — whether the full policy booklet was actually received "
            "by Mr G was disputed but not determinative, since the Policy Summary and "
            "the direct application questions were sufficient on their own"
        ),
        "Ombudsman Reasoning": (
            "The direct question asked at application and each subsequent renewal — "
            "whether the property would be left unoccupied for more than 60 "
            "consecutive days — put Mr G on notice that this was a relevant factor for "
            "his cover, and his repeated 'No' answers meant RIAS reasonably understood "
            "the exclusion wasn't a live concern for him; even accepting he only "
            "received the Policy Summary (not the full wording), that document itself "
            "flagged unoccupied-property exclusions existed and directed him to the "
            "full wording if he wanted more detail, which he could have requested but "
            "didn't"
        ),
        "Workflow Insight": (
            "When a broker asks a direct, unambiguous question about a policy-relevant "
            "circumstance (e.g. extended unoccupancy) at both sale and every renewal, "
            "and the policyholder consistently answers in a way that would rule out "
            "the exclusion applying, this exchange itself can satisfy the broker's "
            "disclosure duty — a summary document that flags an exclusion exists and "
            "signposts fuller wording on request need not itself contain every detail, "
            "provided the policyholder wasn't prevented from requesting more "
            "information"
        ),
        "AI Rule Candidate": (
            "IF broker_asks_a_direct_unambiguous_question_about_a_time_limited_unoccupancy_threshold_at_both_sale_and_every_renewal "
            "AND policyholder_consistently_answers_in_a_way_ruling_out_the_exclusion_applying "
            "THEN this_exchange_together_with_a_summary_document_flagging_the_exclusions_exist_satisfies_the_brokers_disclosure_duty_even_if_the_full_policy_wording_was_not_confirmed_received"
        ),
        "Source PDF": "DRN7626527.pdf",
    },
    {
        "Case ID": "UNOC-054",
        "FOS Decision ID": "DRN7714548",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": (
            "Landlords household / property insurance — escape of water from a burst "
            "loft pipe during refurbishment between tenancies; insurer declined citing "
            "the property had been unoccupied for more than 30 days (and separately, "
            "breach of a 14-day unoccupancy endorsement requiring services turned off "
            "and drained), amid significant confusion over which of several "
            "conflicting policy documents actually applied"
        ),
        "Unoccupied Period / Circumstance": (
            "Following the end of a tenancy, the policyholder arranged extensive "
            "refurbishment/maintenance to re-let the property and kept the heating on "
            "throughout (both to protect the property from cold and because "
            "contractors were working in it); more than two months after the tenant "
            "left, during exceptionally cold weather, the boiler failed and a loft "
            "pipe burst; the policyholder and letting agent visited regularly and "
            "contractors were present on a scheduled basis throughout, though RSA said "
            "the property had nonetheless been 'unoccupied' since the tenant moved out"
        ),
        "Property Type": "Rental property between tenancies, undergoing refurbishment",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "RSA said the property had been unoccupied (defined in its policy "
            "documentation as 'the part or whole of the property not lived in by a "
            "person authorised by You') for more than 30/60 days, and separately that "
            "the policyholder hadn't complied with a 14-day unoccupancy endorsement "
            "requiring gas and water supplies to be turned off and the system drained"
        ),
        "Evidence Dispute": (
            "Policyholder: the property was effectively never left unoccupied for "
            "more than a few days at a time given the frequency of contractor visits "
            "and her and her partner's own visits to check progress and collect mail; "
            "keeping the heating on for the workmen and to protect the property "
            "against cold weather was reasonable in the circumstances. RSA: supplied "
            "conflicting, inconsistent policy documentation across several requests "
            "(a 'Property Insurance Policy' booklet, a 'Summary of Cover' referencing "
            "a combined 'Property Policy'/'Unoccupied Property Policy', and at one "
            "point a different 'Let Property Flexilet' booklet entirely), none of "
            "which matched the endorsement referenced on the policyholder's own "
            "certificate. FOS: given the unresolved documentation confusion, the "
            "policyholder's own certificate (specific to her policy, correct property "
            "and cover period) was the most reliable indicator of the actual "
            "endorsements in place, and RSA's supplied booklets likely weren't the "
            "correct documents; but even assuming they were correct, the policy's "
            "'unoccupied' definition ('not lived in by a person authorised by You') "
            "was a significantly unusual definition that wasn't adequately drawn to "
            "the policyholder's attention despite being prominently set out in the "
            "Summary of Cover — without knowing that specific, unusual definition, she "
            "couldn't appreciate the significance of the exclusions and limitations "
            "referencing it, so the ordinary meaning of 'unoccupied' should apply "
            "instead, under which the property (given the frequency of contractor and "
            "personal visits) shouldn't be treated as unoccupied at all"
        ),
        "Outcome Category": "Upheld",
        "Outcome": (
            "RSA required to settle the claim in accordance with the remaining policy "
            "terms and conditions, and pay 8% simple interest per year on any cash "
            "payment from the date of loss to the date of payment"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "Where an insurer cannot reliably establish which of several conflicting "
            "policy documents it has supplied actually governs the claim, the "
            "policyholder's own certificate (specific to their policy, property and "
            "cover period) is the more reliable indicator of the endorsements "
            "actually in place; separately, where a policy defines 'unoccupied' in a "
            "way significantly different from its ordinary meaning (e.g. 'not lived "
            "in by a person authorised by You', rather than the Service's usual "
            "approach recognising regular visits/contractor presence as occupation), "
            "the insurer must adequately draw that unusual definition to the "
            "policyholder's attention — merely including it in a summary of cover "
            "alongside numerous other unoccupancy exclusions and limitations is not "
            "necessarily adequate; without such notice, the ordinary meaning of "
            "'unoccupied' should be applied instead"
        ),
        "Missing Evidence": (
            "A single, consistent, agreed-correct policy document — RSA identified "
            "three different, mutually-inconsistent booklets across its various "
            "responses as 'the' correct policy document"
        ),
        "Ombudsman Reasoning": (
            "The significant confusion in RSA's own document production (three "
            "different, non-matching policy booklets identified as correct at "
            "different points) meant the policyholder's own certificate — which "
            "matched the property, cover period and premium — was the more reliable "
            "record of her actual policy terms; even if RSA's supplied booklets were "
            "assumed correct, the unusual 'unoccupied' definition they contained was "
            "significantly different from the ordinary meaning and wasn't adequately "
            "highlighted, so it would not be reasonable to allow RSA to rely on it; "
            "applying the ordinary meaning instead, the frequency of contractor and "
            "personal visits meant the property shouldn't be treated as unoccupied"
        ),
        "Workflow Insight": (
            "When an insurer supplies inconsistent or conflicting policy documentation "
            "across a complaint investigation, treat the policyholder's own "
            "certificate (matching the specific property, cover period, sums insured "
            "and premium) as the more reliable record of actual policy terms; "
            "separately, always check whether a policy's definition of 'unoccupied' "
            "departs significantly from the ordinary meaning of the word — if so, it "
            "must be clearly and specifically highlighted (not merely present "
            "somewhere in a summary of cover) or the ordinary meaning should govern "
            "instead"
        ),
        "AI Rule Candidate": (
            "IF insurer_supplies_multiple_inconsistent_policy_documents_during_a_claim_investigation "
            "THEN the_policyholders_own_certificate_matching_the_specific_property_cover_period_and_premium_is_the_more_reliable_record_of_actual_terms; "
            "IF policy_defines_unoccupied_significantly_differently_from_its_ordinary_meaning "
            "AND this_unusual_definition_was_not_specifically_and_clearly_highlighted_to_the_policyholder "
            "THEN the_ordinary_meaning_of_unoccupied_should_govern_instead_and_regular_visits_or_contractor_presence_may_mean_the_property_is_not_unoccupied"
        ),
        "Source PDF": "DRN7714548.pdf",
    },
    {
        "Case ID": "UNOC-055",
        "FOS Decision ID": "DRN7954504",
        "Insurer Name": "International Insurance Company of Hannover SE",
        "FOS Decision Date": "16 Feb 2018",
        "Claim Type": (
            "Home insurance — burglary claim; insurer voided the policy for "
            "misrepresentation about main residence status, after evidence showed the "
            "policyholder didn't actually live at the insured address as declared at "
            "inception and each renewal"
        ),
        "Unoccupied Period / Circumstance": (
            "Insurer's loss adjuster reported the policyholder told them the property "
            "had been unoccupied since 2013 (the policyholder later called this a "
            "'clerical error'); electoral roll entries for the policyholder and his "
            "wife showed a different address since 2013; a signed statement to the "
            "loss adjuster said the family lived at the insured address until 2012 "
            "and mainly at another address since 2013, visiting the insured address "
            "roughly every six weeks"
        ),
        "Property Type": "Residential property (main residence status disputed)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Policyholder was asked clear questions at inception (confirming the "
            "property was his main residence, unoccupied no more than 30 days a year) "
            "and this was untrue based on his own signed statement, electoral roll "
            "registration elsewhere, and the loss adjuster's account — a qualifying "
            "misrepresentation entitling the insurer to void the policy, since it "
            "wouldn't offer cover for a property that wasn't the policyholder's main "
            "residence"
        ),
        "Evidence Dispute": (
            "Mr E: said whether the property was his 'main residence' was irrelevant "
            "since he'd complied with the 45-day unoccupancy requirement he "
            "understood applied, visited weekly, and took safety precautions; said it "
            "was IICH's 'clerical error' that it recorded him telling the loss "
            "adjuster the property had been unoccupied since 2013. IICH: clear, "
            "specific questions were asked at the outset about main residence status "
            "and expected unoccupancy, confirmed again via a Statement of Insurance "
            "sent for him to check and at each renewal without amendment; his own "
            "signed statement to the loss adjuster (that he and his family lived "
            "mainly elsewhere since 2013, visiting the insured address roughly every "
            "six weeks, with his wife not visiting for over six months) contradicted "
            "his 'main residence' declaration; electoral roll records supported this. "
            "FOS: Mr E's own account across different communications was "
            "contradictory (variously saying the property was never unoccupied for "
            "more than 45 days, that main residence status was irrelevant since he "
            "visited regularly, and later that it was his main residence at the time "
            "of the burglary), but his signed statement to the loss adjuster — the "
            "most contemporaneous and detailed account — was the more reliable "
            "evidence, and it clearly showed the insured address wasn't his main "
            "residence"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — IICH acted fairly and in accordance with the "
            "policy terms in voiding the policy for misrepresentation about main "
            "residence status; no further action required"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "Where a policy is specifically conditioned on the insured property being "
            "the policyholder's main residence (a materially different, and often "
            "more significant, question than simply whether unoccupancy exceeded a "
            "day threshold), and the policyholder's own signed, contemporaneous "
            "statement to a loss adjuster describes an actual living pattern "
            "inconsistent with that declaration (e.g. living mainly elsewhere, "
            "visiting only periodically), that statement is generally more reliable "
            "than the policyholder's later, shifting explanations to the insurer or "
            "this Service; a general unoccupancy day-count compliance argument (e.g. "
            "'never unoccupied for more than 45 days') doesn't answer a separate "
            "'main residence' misrepresentation question"
        ),
        "Missing Evidence": (
            "Not applicable — the signed statement to the loss adjuster and the "
            "electoral roll records were treated as sufficiently reliable, "
            "contemporaneous evidence despite Mr E's later attempts to characterise "
            "parts of it as a 'clerical error'"
        ),
        "Ombudsman Reasoning": (
            "Mr E's account across different points in the process was contradictory "
            "and shifted depending on what argument he was making, whereas his signed "
            "statement to the loss adjuster gave a detailed, specific account "
            "(mainly living elsewhere since 2013, visiting the insured address every "
            "six weeks, wife not visiting for over six months) that was corroborated "
            "by independent electoral roll records; this showed the insured address "
            "wasn't genuinely his main residence, contrary to what was declared at "
            "inception and confirmed unamended at each renewal; IICH's evidence that "
            "it wouldn't have offered cover on the correct facts (since it doesn't "
            "insure unoccupied properties) supported voiding being fair and "
            "reasonable"
        ),
        "Workflow Insight": (
            "When a policyholder's account of their living arrangements shifts across "
            "different stages of a complaint (to the loss adjuster, then in "
            "correspondence, then to this Service), give the most weight to the "
            "earliest, most detailed, contemporaneous and specific account (e.g. a "
            "signed statement to a loss adjuster) rather than later, more generalised "
            "or self-serving reformulations; distinguish a 'main residence' "
            "declaration question from a simple unoccupancy day-count question, since "
            "compliance with one doesn't answer the other"
        ),
        "AI Rule Candidate": (
            "IF policyholders_account_of_living_arrangements_shifts_across_different_stages_of_a_complaint "
            "THEN give_most_weight_to_the_earliest_most_detailed_contemporaneous_signed_statement_over_later_reformulations; "
            "policy_conditioned_on_main_residence_status_is_a_distinct_question_from_a_general_unoccupancy_day_count_requirement_and_compliance_with_one_does_not_establish_compliance_with_the_other"
        ),
        "Source PDF": "DRN7954504.pdf",
    },
    {
        "Case ID": "UNOC-056",
        "FOS Decision ID": "DRN8194702",
        "Insurer Name": "UK Insurance Limited",
        "FOS Decision Date": "14 Jan 2018",
        "Claim Type": (
            "Landlord insurance (two properties on one policy) — escape of water "
            "claim on the unoccupied property was accepted, but the policyholder "
            "disputed the quantum of the loss-of-rent payment, non-renewal of the "
            "whole policy because one property was unoccupied, and a service error "
            "(excess withheld twice)"
        ),
        "Unoccupied Period / Circumstance": (
            "One of two properties on a joint landlord policy became unoccupied from "
            "February 2016; leak damage reported September 2016 was initially "
            "repudiated for unoccupancy but reconsidered once it was shown the leak "
            "occurred during an earlier claim within 30 days of the property becoming "
            "unoccupied; UKI declined to renew the whole two-property policy at its "
            "next renewal because of the unoccupied property, rather than removing "
            "only that property from cover"
        ),
        "Property Type": "Rented residential properties (two properties on one landlord policy; one became unoccupied)",
        "Dispute Type": "Claim Quantum Dispute",
        "Coverage Decision": "Accepted — With Deductions",
        "Rejection Reason": (
            "Not applicable to a full decline — the escape of water claim was "
            "accepted, with some damage attributed to uncovered gradual-deterioration "
            "maintenance issues rather than the leak itself; the loss-of-rent element "
            "was paid at the policy's maximum sum insured (£14,400) rather than the "
            "higher monthly rate the policyholder said she could have achieved on the "
            "open market"
        ),
        "Evidence Dispute": (
            "Miss A: disputed the loss-of-rent rate (£1,200pcm applied versus "
            "£1,300-1,400pcm she believed achievable, supported by a letting agent's "
            "estimate), wanted rent paid until repairs were actually completed (not "
            "capped at the 12-month indemnity period), was unhappy the whole "
            "two-property policy wasn't renewed rather than just the unoccupied "
            "property, and was unhappy a £200 excess was mistakenly withheld twice. "
            "UKI: the sum insured (£14,400, based on £1,200pcm) was the contractual "
            "maximum for the 12-month indemnity period regardless of a higher "
            "achievable rate; it wasn't obliged to renew cover for an unoccupied "
            "property, and since both properties were on one combined contract, "
            "ending cover for one meant it wasn't unreasonable to end the whole "
            "contract rather than assume she wanted to renew only the unaffected "
            "property; the double excess withholding was a system error, corrected "
            "and refunded promptly. FOS: the loss-of-rent sum insured was a "
            "contractual cap properly applied; investigating a complex claim "
            "(including potential unoccupancy grounds to decline it) before accepting "
            "liability wasn't inherently unreasonable delay; declining to renew "
            "unoccupied-property cover is a legitimate commercial underwriting "
            "decision this Service won't interfere with, and ending the whole "
            "combined-property contract (rather than assuming a split was wanted) "
            "wasn't unreasonable; the excess error was promptly corrected"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — the loss-of-rent payment was correctly capped at "
            "the policy limit, non-renewal of the combined two-property policy "
            "because of the unoccupied property was a reasonable commercial decision, "
            "and the excess/payment-timing errors were promptly and adequately "
            "corrected"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Quantum Only",
        "Key Policy Clause": (
            "A 'loss of rent receivable' indemnity is capped at the policy's stated "
            "sum insured for the maximum indemnity period, regardless of a higher "
            "market rent the policyholder believes could have been achieved, unless "
            "the insurer's own breach or negligence caused a reasonably foreseeable "
            "uninsured loss beyond that; where two properties are combined under one "
            "landlord insurance contract and one becomes unoccupied, an insurer "
            "declining to renew the whole combined contract (rather than assuming the "
            "policyholder wants to split cover and retain only the unaffected "
            "property) is a legitimate commercial underwriting decision"
        ),
        "Missing Evidence": (
            "Not applicable — the facts (rent achieved historically, letting agent's "
            "estimate, the excess double-withholding) were undisputed; the dispute was "
            "over the correct contractual quantum and remedy"
        ),
        "Ombudsman Reasoning": (
            "The policy's loss-of-rent section indemnifies up to the stated sum "
            "insured for the indemnity period (12 months), which UKI paid in full "
            "(£14,400, at £1,200pcm); a letting agent's estimate of a higher "
            "achievable rent didn't change the contractual cap, and rent couldn't be "
            "paid beyond the indemnity period simply because repairs were delayed "
            "while UKI investigated a complex, potentially unoccupancy-affected claim; "
            "investigating before accepting liability on a claim with potential "
            "grounds to decline wasn't unreasonable delay; ending the whole combined "
            "policy (rather than only the unoccupied property) was UKI's legitimate "
            "commercial decision, since it covered two properties under one contract "
            "and total premium; the double-excess error was a system glitch, corrected "
            "promptly upon discovery"
        ),
        "Workflow Insight": (
            "For loss-of-rent disputes, check whether the payment matches the "
            "policy's stated sum insured and indemnity period cap — a higher "
            "achievable market rent (even if independently estimated by a letting "
            "agent) doesn't override a contractual limit; when an insurer declines to "
            "renew a multi-property policy because one property became unoccupied, "
            "treat this as a legitimate single commercial decision rather than "
            "requiring the insurer to have offered a split renewal for the unaffected "
            "property"
        ),
        "AI Rule Candidate": (
            "IF loss_of_rent_claim_payment_matches_the_policys_stated_sum_insured_and_indemnity_period "
            "THEN a_higher_market_estimate_from_a_letting_agent_does_not_entitle_the_policyholder_to_more_than_the_contractual_cap; "
            "IF multiple_properties_are_combined_under_one_landlord_policy_and_one_becomes_unoccupied "
            "THEN insurer_declining_to_renew_the_whole_combined_contract_rather_than_offering_a_split_renewal_is_a_legitimate_commercial_decision"
        ),
        "Source PDF": "DRN8194702.pdf",
    },
    {
        "Case ID": "UNOC-057",
        "FOS Decision ID": "DRN8309060",
        "Insurer Name": "Lloyds Bank General Insurance Limited",
        "FOS Decision Date": "11 Apr 2016",
        "Claim Type": (
            "Home insurance — malicious damage and arson claims; insurer found "
            "evidence the policyholder hadn't lived at the property for a long time "
            "and backdated cancellation of cover to a year after it believed he "
            "stopped living there, declining the two claims that arose after that "
            "backdated cancellation date"
        ),
        "Unoccupied Period / Circumstance": (
            "Insurer believed the policyholder hadn't lived at the property since "
            "late 2012 or earlier, based on extremely low electricity usage (2-7% of "
            "an average household from 2011-2013) and a post-fire assessor's report "
            "indicating the property had been unoccupied and used as a meeting place "
            "by local youths for some time; insurer's stated approach was to provide a "
            "lower level of cover for unoccupied properties for one year, then no "
            "cover after that, and it backdated cancellation to 1 December 2013 "
            "(a year after the date it believed occupancy ceased)"
        ),
        "Property Type": "Residential property (occupancy status disputed over an extended period)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Policyholder hadn't told the insurer he'd stopped living at the property "
            "as required by the policy, and the insurer's evidence indicated he'd "
            "stopped living there more than a year before the two declined claims — "
            "beyond the one year of reduced unoccupied-property cover it said it would "
            "have provided"
        ),
        "Evidence Dispute": (
            "Mr C: said he was living at the property until a burglary in late "
            "November 2013 (paid by Lloyds, as it preceded the cancellation date) made "
            "it uninhabitable, and offered to provide witness statements supporting "
            "this. Lloyds: electricity usage during 2012 was only 2-7% of an average "
            "household, a post-fire assessor's report described the property as "
            "unoccupied and used by local youths as a meeting place, and earlier 2013 "
            "theft claims hadn't mentioned any change in occupancy despite Lloyds "
            "believing he'd already moved out by then. FOS: the electricity usage "
            "evidence, considered against Mr C's explanation (long hours as a taxi "
            "driver), still indicated no more than occasional visits rather than "
            "genuine residence; the assessor's report of youths using the property as "
            "a meeting place independently corroborated non-occupancy; witness "
            "statements weren't accepted as equivalent to independent evidence like "
            "utility bills, and Lloyds had already given Mr C a list of acceptable "
            "evidence types which he hadn't otherwise provided"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — Lloyds acted reasonably in backdating "
            "cancellation to a year after it believed occupancy ceased, based on "
            "persuasive independent evidence (utility usage, assessor's report), and "
            "in declining the two claims that fell after that date"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "An insurer's stated commercial approach of providing a reduced level of "
            "cover for an unoccupied property for a fixed period (e.g. one year) and "
            "no cover thereafter is a legitimate underwriting decision; where "
            "independent evidence (extremely low utility usage compared to an average "
            "household, and an independent assessor's observations, e.g. of "
            "unauthorised use by local youths) points to a specific date occupancy "
            "ceased, an insurer can fairly backdate a policy cancellation to that "
            "date plus its stated grace period, and decline claims falling after it; "
            "witness statements from the policyholder's own contacts are not treated "
            "as equivalent to independent documentary evidence (e.g. utility bills) "
            "when assessing disputed occupancy"
        ),
        "Missing Evidence": (
            "Independent documentary evidence (of the type Lloyds' agent had "
            "previously told Mr C would be acceptable) corroborating his claimed "
            "occupancy through to late November 2013 — he offered only witness "
            "statements, which weren't treated as sufficient"
        ),
        "Ombudsman Reasoning": (
            "Electricity usage in 2012 at only 2-7% of an average household "
            "consumption was not adequately explained by Mr C's long working hours as "
            "a taxi driver, and indicated no more than occasional visits; an "
            "independent post-fire assessor separately reported signs the property had "
            "been unoccupied and used as a meeting place by local youths, "
            "corroborating the low-usage finding; witness statements from Mr C's own "
            "contacts weren't independent in the way utility bills are, and he hadn't "
            "provided any of the other acceptable evidence types Lloyds' agent had "
            "already identified for him; on balance, it was reasonable for Lloyds to "
            "conclude occupancy had ceased around November 2012 and to backdate "
            "cancellation a year from that date"
        ),
        "Workflow Insight": (
            "When assessing disputed long-term occupancy, weigh independent "
            "documentary evidence (utility usage compared to household averages, "
            "independent assessor observations) more heavily than witness statements "
            "from the policyholder's own contacts, which lack independence; where an "
            "insurer's stated policy is to provide reduced unoccupied-property cover "
            "for a fixed grace period before cover lapses entirely, backdating "
            "cancellation to that period's end from the evidenced date occupancy "
            "ceased is a reasonable approach to claims arising after that point"
        ),
        "AI Rule Candidate": (
            "IF independent_evidence_eg_utility_usage_far_below_household_average_and_independent_assessor_observations_indicate_a_specific_date_occupancy_ceased "
            "THEN insurer_may_backdate_cancellation_to_that_date_plus_its_stated_unoccupied_property_grace_period_and_decline_claims_arising_after_it; "
            "witness_statements_from_the_policyholders_own_contacts_are_not_equivalent_to_independent_documentary_evidence_when_assessing_disputed_long_term_occupancy"
        ),
        "Source PDF": "DRN8309060.pdf",
    },
    {
        "Case ID": "UNOC-058",
        "FOS Decision ID": "DRN8328581",
        "Insurer Name": "Ageas 50 Limited",
        "FOS Decision Date": "Not stated in document",
        "Claim Type": (
            "Home insurance (broker administration complaint) — escape of water "
            "claim from a burst loft pipe declined by the underlying insurer citing "
            "the unoccupied-property exclusion, after the broker added a specific "
            "term excluding escape-of-water cover for the entire duration of the "
            "policyholder's 90-day holiday but never called back to explain this as "
            "promised"
        ),
        "Unoccupied Period / Circumstance": (
            "Policyholder proactively called her broker to disclose a 90-day holiday "
            "during which her home would be unoccupied; broker's in-house "
            "underwriters added a term excluding escape-of-water damage for the "
            "entire holiday (not just the usual 60-day threshold), and sent an "
            "amended schedule, but never called back as twice promised to explain the "
            "new term; policyholder's daughter discovered a burst loft pipe during "
            "the holiday"
        ),
        "Property Type": "Residential property (long holiday absence, proactively disclosed)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a decision by the broker directly — the underlying "
            "insurer declined the escape of water claim under the newly-added "
            "unoccupancy term; this complaint concerned whether the broker adequately "
            "explained that new term to the policyholder after she proactively "
            "disclosed her holiday"
        ),
        "Evidence Dispute": (
            "Ms F: she specifically rang her broker to disclose the extended holiday "
            "and was told twice she'd be called back to confirm what terms, if any, "
            "would apply — that call never came, and she was left with only an "
            "amended schedule she didn't realise contained a new, more restrictive "
            "term; had she known, she'd have paid extra for full cover. Ageas: there "
            "was no change to the general unoccupancy position (she'd never been "
            "covered for escape of water beyond 60 days), she'd taken an extended "
            "holiday before, and it always communicated in writing without any "
            "indication this didn't suit her. FOS: Ageas's response missed the point "
            "— the actual term added was materially different from the usual 60-day "
            "threshold, excluding escape-of-water cover for the entire 90-day holiday "
            "from day one, and this specific alteration wasn't adequately drawn to "
            "Ms F's attention; having proactively called to disclose her holiday and "
            "been promised a follow-up call that never came, she was entitled to "
            "assume the information had been noted and cover would continue as "
            "before"
        ),
        "Outcome Category": "Upheld",
        "Outcome": (
            "Ageas required to deal with the escape of water claim as if it were the "
            "insurer (instructing an independent loss adjuster, cash-settling at open "
            "market reinstatement rates if applicable, with 8% simple interest from "
            "the date of claim), and pay Ms F £400 compensation for distress and "
            "inconvenience"
        ),
        "Compensation Awarded (£)": 400,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "Where a policyholder proactively discloses an extended absence to their "
            "broker and is explicitly told they'll receive a follow-up call to "
            "confirm what terms will apply, a broker's failure to make that promised "
            "call — instead relying on an amended schedule alone to convey a newly "
            "added, materially more restrictive exclusion (here, escape-of-water "
            "cover excluded for the entire holiday rather than only after the usual "
            "60-day threshold) — does not adequately draw the new term to the "
            "policyholder's attention; a general awareness of a standard unoccupancy "
            "exclusion doesn't put a policyholder on notice of a specific, "
            "non-standard alteration made in response to their particular disclosure"
        ),
        "Missing Evidence": (
            "Not applicable — that no follow-up call was made, despite being promised "
            "twice, was not disputed"
        ),
        "Ombudsman Reasoning": (
            "The broker's argument that Ms F was already generally aware of the "
            "standard 60-day unoccupancy exclusion missed the point, since the actual "
            "term added for this holiday was materially different (excluding "
            "escape-of-water for the whole 90 days, not just beyond 60); having "
            "proactively disclosed her holiday and been told twice she'd receive a "
            "call to confirm applicable terms, Ms F was entitled to assume, in the "
            "absence of that call, that her disclosure had simply been noted and "
            "standard cover would continue; had she known the true position she would "
            "likely have sought alternative unoccupied-property cover or taken "
            "precautions like draining the water system, so the broker's failure "
            "caused real detriment"
        ),
        "Workflow Insight": (
            "When a policyholder proactively discloses a change in circumstances "
            "(e.g. an extended absence) and is promised a follow-up call to confirm "
            "resulting policy changes, a broker cannot rely on a passively-sent "
            "amended schedule alone to satisfy that promise — especially where the "
            "resulting term is materially more restrictive than the policyholder's "
            "prior, general understanding of the standard exclusion; assess what "
            "specifically changed against what the policyholder could reasonably have "
            "already known, not just whether an unoccupancy exclusion of some kind "
            "existed generally"
        ),
        "AI Rule Candidate": (
            "IF policyholder_proactively_discloses_an_extended_absence_to_a_broker_and_is_promised_a_follow_up_call_to_confirm_applicable_terms "
            "AND broker_fails_to_make_that_call_and_instead_relies_only_on_a_passively_sent_amended_schedule_containing_a_materially_more_restrictive_new_term "
            "THEN the_new_term_was_not_adequately_disclosed_regardless_of_the_policyholders_general_prior_awareness_of_a_standard_unoccupancy_exclusion"
        ),
        "Source PDF": "DRN8328581.pdf",
    },
    {
        "Case ID": "UNOC-059",
        "FOS Decision ID": "DRN8482133",
        "Insurer Name": "Society of Lloyd's",
        "FOS Decision Date": "14 Dec 2015",
        "Claim Type": (
            "Unoccupied property insurance (two rental flats) — theft and malicious "
            "damage claim declined citing a restricted 'FLEE' (fire, lightning, "
            "earthquake, explosion) cover term issued when the flats were unoccupied "
            "for renovation, even though the insurer's own published policy change "
            "meant full cover should have applied by the time of the loss"
        ),
        "Unoccupied Period / Circumstance": (
            "Flats insured in 2013 while undergoing internal (non-structural) "
            "renovation for a planned 12-week period, with a restricted FLEE-only "
            "policy issued on that basis; by 2014 (when the theft/malicious damage "
            "claim arose) the properties had become tenanted as planned; insurer "
            "declined citing the restricted cover, saying it hadn't been told the "
            "flats had become tenanted"
        ),
        "Property Type": "Rental flats (unoccupied for renovation, later re-let as planned)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Only restricted 'FLEE' cover applied to the policy (issued while the "
            "flats were unoccupied for renovation), and the insurer hadn't been "
            "informed the properties had since become tenanted"
        ),
        "Evidence Dispute": (
            "Mr R: pointed to the insurer's own 2012 customer newsletter, which said "
            "that from 1 April 2012 its Unoccupied Property cover would provide full "
            "cover at all times unless restricted cover was specifically requested or "
            "structural work was being undertaken — neither of which applied to him "
            "(he'd only requested cover for minor, non-structural renovation and "
            "hadn't asked for restricted cover); there was also no policy requirement "
            "to notify the insurer once the properties became tenanted. Lloyd's: said "
            "its underwriter wouldn't have offered full cover in Mr R's "
            "circumstances, and that the newsletter's increased cover only applied to "
            "subsidence or accidental damage. FOS: Lloyd's hadn't explained how its "
            "position accorded with the newsletter's clear wording, which imposed no "
            "such peril-specific limitation and applied to Mr R's exact situation "
            "(non-structural renovation, no restricted cover requested); since the "
            "flats becoming tenanted didn't need to be reported and there was no "
            "evidence full cover would have been refused once occupied, the claim "
            "should be dealt with as if full (unoccupied property) cover applied "
            "throughout"
        ),
        "Outcome Category": "Upheld",
        "Outcome": (
            "Society of Lloyd's required to deal with Mr R's claim as if full cover "
            "applied, subject to the remaining policy terms and conditions, and pay "
            "£200 compensation for the inconvenience caused by its handling of the "
            "claim"
        ),
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "Where an insurer has published a general customer communication (e.g. a "
            "newsletter) stating that its unoccupied-property cover will provide full "
            "cover at all times unless restricted cover is specifically requested or "
            "structural work is being undertaken, and a policyholder's actual "
            "circumstances (non-structural renovation, no restricted cover requested) "
            "fall squarely within that stated full-cover scenario, the insurer cannot "
            "later assert an unexplained, undocumented limitation (e.g. that the "
            "increased cover applied only to certain perils) to justify declining a "
            "claim under a narrower restricted-cover term — it must show how its "
            "position accords with what it told policyholders generally it would do"
        ),
        "Missing Evidence": (
            "Any explanation from Lloyd's of how its asserted peril-specific "
            "limitation (that increased unoccupied-property cover applied only to "
            "subsidence or accidental damage) was consistent with, or derived from, "
            "the newsletter's unqualified statement of full cover — none was provided "
            "despite being asked"
        ),
        "Ombudsman Reasoning": (
            "The 2012 newsletter's language was unqualified: full cover applies to "
            "unoccupied properties at all times unless restricted cover is requested "
            "or structural work is underway; Mr R's disclosed circumstances (minor, "
            "non-structural renovation, no restricted cover requested) matched this "
            "exactly, so he was entitled to expect full cover; Lloyd's assertion that "
            "it wouldn't have offered full cover, and that the newsletter's increase "
            "was peril-limited, was unsupported and inconsistent with the newsletter's "
            "actual wording; there being no requirement to report the flats becoming "
            "tenanted, and no evidence full cover would have been withdrawn once they "
            "were, the claim should be dealt with as though full cover had applied "
            "throughout"
        ),
        "Workflow Insight": (
            "When an insurer relies on a restricted-cover term to decline a claim, "
            "check whether the insurer has separately published general customer "
            "communications (newsletters, cover-change notices) describing a more "
            "generous default position that would apply to the policyholder's actual "
            "disclosed circumstances — an insurer cannot decline a claim under a "
            "narrower term without explaining how that's consistent with what it told "
            "policyholders generally it would do in that scenario"
        ),
        "AI Rule Candidate": (
            "IF insurer_has_published_a_general_customer_communication_stating_a_more_generous_default_unoccupied_property_cover_position "
            "AND policyholders_actual_disclosed_circumstances_fall_within_the_scenario_described_in_that_communication "
            "THEN insurer_cannot_rely_on_a_narrower_restricted_cover_term_to_decline_the_claim_without_explaining_how_this_is_consistent_with_the_published_communication"
        ),
        "Source PDF": "DRN8482133.pdf",
    },
    {
        "Case ID": "UNOC-060",
        "FOS Decision ID": "DRN9942137",
        "Insurer Name": "L. P. Dawe",
        "FOS Decision Date": "10 Jul 2015",
        "Claim Type": (
            "Home insurance (broker mis-sale complaint) — inherited property "
            "insured via a broker; escape of water claim declined and the policy "
            "cancelled by the insurer, after the broker recommended a renewal policy "
            "without updating stale information about the policyholder's living "
            "circumstances, resulting in a policy unsuited to the property being a "
            "second (frequently empty) home rather than her permanent residence"
        ),
        "Unoccupied Period / Circumstance": (
            "Policyholder inherited the property from her late mother in 2010; "
            "broker's original 2010 file note recorded she would 'eventually be "
            "living in' it once she sold her own home, with the property unoccupied "
            "at times in the meantime; when her original insurer stopped offering "
            "this type of policy, the broker recommended a replacement with a second "
            "insurer in 2011/2012 using the same year-old assumptions, without "
            "checking whether she'd actually moved in; by the July 2013 escape of "
            "water claim, she was still living at her own home with her mother's "
            "house standing empty for extended periods"
        ),
        "Property Type": "Inherited residential property (never became the policyholder's permanent residence as originally assumed)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a decision by the broker directly — the second "
            "insurer refused to consider the escape of water claim and cancelled the "
            "policy; this complaint concerned only whether the broker had sold/"
            "renewed her into an unsuitable policy"
        ),
        "Evidence Dispute": (
            "Miss B: the broker sold her an inappropriate policy that didn't cover "
            "her actual circumstances (the property being a second home, empty for "
            "significant periods) rather than her main residence. L P Dawe: it had "
            "originally been given information (via the solicitors administering the "
            "estate) that she intended to move into the property once she sold her "
            "own home, justifying a policy suited to that plan; the policy's "
            "statement of facts (listing the property as her 'permanent residence') "
            "was sent to her and she never queried or corrected it. FOS: the 2010 "
            "recommendation was suitable based on the information available at the "
            "time (her stated intention to move in once her own home sold); however, "
            "when the original insurer stopped offering the product and L P Dawe had "
            "to recommend a replacement policy for the 2011/2012 renewal, it simply "
            "reused the year-old assumption that she'd moved in as planned, without "
            "checking her actual current circumstances — this fell short of what an "
            "advising broker should do at each renewal, especially given a change of "
            "insurer/product was involved, not just a routine like-for-like renewal"
        ),
        "Outcome Category": "Upheld",
        "Outcome": (
            "L P Dawe required to pay Miss B the amount an insurer would probably "
            "have paid had she been sold a policy suited to the property being a "
            "second home left empty for significant periods (via agreement or a "
            "jointly-instructed loss adjuster's assessment, at L P Dawe's expense), "
            "arrange removal of the cancellation record from the insurer's internal "
            "and external databases (or provide a letter confirming the cancellation "
            "was its error), and pay Miss B £200 for distress and inconvenience"
        ),
        "Compensation Awarded (£)": 200,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "A broker's recommendation for an initial policy can be suitable when "
            "based on the customer's stated intentions at the time (e.g. an intention "
            "to move into an inherited property once a current home is sold) — but at "
            "each subsequent renewal, particularly where a change of insurer or "
            "product is required (not simply a routine like-for-like renewal), the "
            "broker must update its understanding of the customer's actual current "
            "circumstances rather than reusing stale assumptions; a customer's "
            "failure to proactively correct an inaccurate 'permanent residence' "
            "assumption printed in a statement of facts does not excuse the broker "
            "from its own duty to actively re-verify suitability when arranging a new "
            "policy at renewal"
        ),
        "Missing Evidence": (
            "Any updated information-gathering by L P Dawe at the 2011/2012 renewal "
            "about whether Miss B had actually moved into the inherited property as "
            "originally planned — none was obtained; it simply carried forward the "
            "2010 assumption"
        ),
        "Ombudsman Reasoning": (
            "L P Dawe's 2010 file note showed a reasonable, suitable recommendation "
            "based on Miss B's stated plan to eventually live in the inherited "
            "property; but when the original insurer withdrew from this product and a "
            "new policy with a different insurer had to be arranged for the next "
            "renewal, L P Dawe simply carried forward the same year-old assumption "
            "without re-checking whether she'd actually moved in — this was not good "
            "enough advising practice, particularly given a change of insurer/product "
            "(as opposed to a routine renewal) specifically calls for re-verification; "
            "as a result she was sold a policy that didn't suit her real "
            "circumstances (property remaining a second, often-empty home), and when "
            "she came to claim, this mismatch caused her claim to be refused and her "
            "policy cancelled"
        ),
        "Workflow Insight": (
            "When a broker has to move a customer to a new insurer or product at "
            "renewal (rather than a routine like-for-like renewal with the same "
            "insurer), treat this as a trigger to actively re-verify the customer's "
            "current circumstances rather than carrying forward assumptions from the "
            "original sale — this is especially important for unoccupied/second-home "
            "situations where an original stated intention (e.g. to move in) may not "
            "have materialised by the time of a later renewal"
        ),
        "AI Rule Candidate": (
            "IF broker_must_arrange_a_new_policy_with_a_different_insurer_or_product_at_renewal_rather_than_a_routine_like_for_like_renewal "
            "AND broker_carries_forward_a_stale_assumption_about_the_customers_occupancy_circumstances_from_the_original_sale_without_re_verifying_it "
            "THEN broker_is_responsible_for_any_resulting_unsuitable_policy_and_the_consequences_of_a_claim_being_refused_on_that_basis"
        ),
        "Source PDF": "DRN9942137.pdf",
    },
]


def validate_case(case: dict) -> None:
    for field, allowed in CONTROLLED_VOCAB.items():
        value = case.get(field, "")
        if value not in allowed:
            raise ValueError(
                f"Case {case['Case ID']}: '{field}' = '{value}' not in controlled vocab {allowed}"
            )


def append_cases(ws, cases: list) -> None:
    next_row = ws.max_row + 1
    for case_idx, case in enumerate(cases):
        row_fill = ROW_FILL_ODD if (next_row % 2 == 1) else ROW_FILL_EVEN
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=case.get(col_name, ""))
            cell.font      = ROW_FONT
            cell.fill      = row_fill
            cell.border    = ROW_BORDER
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        ws.row_dimensions[next_row].height = 80
        next_row += 1


def main() -> None:
    for case in NEW_CASES:
        validate_case(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases",
        "Unoccupied_Property_Case_Database.xlsx"
    )

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(
            f"Database not found: {xlsx_path}\n"
            "Run create_unoccupied_property_case_db.py first to create it."
        )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Cases"]

    rows_before = ws.max_row - 1  # exclude header
    append_cases(ws, NEW_CASES)
    rows_after = ws.max_row - 1

    wb.save(xlsx_path)

    print(f"Appended : {rows_after - rows_before} cases")
    print(f"Total    : {rows_after} data rows")
    print(f"Last ID  : {ws.cell(row=ws.max_row, column=1).value}")
    print(f"Saved    : {xlsx_path}")


if __name__ == "__main__":
    main()
