"""
Standard append script for Flood Case Database — Schema v1 (21 columns).

Usage
-----
1. Read the source PDF(s) and extract the fields listed in NEW_CASES below.
2. Add one dict per case to NEW_CASES following the extraction rules.
3. Run from the repo root:
       py scripts/append_flood_v1.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Flood_Case_Database.xlsx

===========================================================================
FIELD EXTRACTION RULES
===========================================================================

Case ID         : Format FLOOD-NNN (zero-padded to 3 digits)
FOS Decision ID : DRN-XXXXXXX or DRNXXXXXXX as printed in the PDF
Insurer Name    : Formal registered name from the FOS decision
FOS Decision Date : DD Mon YYYY — accept-or-reject deadline in final paragraph
Claim Type      : Physical incident and nature of dispute in one sentence
Leak Source     : Flood mechanism — physical source and route of water ingress
                  e.g. "Heavy rainfall overwhelming drainage — water entered at ground level"
                       "Watercourse overflow — river/beck burst banks"
                       "Backed-up drain — water accumulated in basement"
                       "Raised water table post-flood penetrating sub-floor"
Property Type   : "Residential home" / "Unoccupied residential property" /
                  "Leasehold flat" / "Commercial" / etc.
Dispute Type    : Controlled vocab (7 values)
Coverage Decision : Controlled vocab (5 values)
Rejection Reason  : Insurer's stated reason for declining
Evidence Dispute  : What evidence each party relied on
Outcome Category  : Controlled vocab (4 values)
Outcome           : Full FOS remedy instructions
Compensation Awarded (£) : Integer — D&I only; 0 if none
Is Core Case      : Controlled vocab (5 values)
Key Policy Clause : Policy wording or FOS/FCA principle applied
Missing Evidence  : Evidence that was absent and affected the outcome
Ombudsman Reasoning : How the ombudsman weighed the evidence
Workflow Insight  : Operational rule for the claims workflow
AI Rule Candidate : Machine-evaluable rule for the rules engine
Source PDF        : Filename only (e.g. DRN0070249.pdf)
===========================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 1: FLOOD-001 to FLOOD-010
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "FLOOD-001",
        "FOS Decision ID": "DRN0070249",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "Not stated",
        "Claim Type": "Flood claim accepted; insurer refused to treat post-flood damp as part of the flood claim, attributing it to a defective floor rather than flood water re-entering via raised water table",
        "Leak Source": "Property flooded twice; post-drying damp attributed by FOS to flood-raised water table penetrating sub-floor through defective damp-proof membrane (DPM) junction at extension",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "RSA refused to treat damp as flood consequence — cited drying certificate showing property certified dry, and argued defective floor (not flood) was the cause of the damp",
        "Evidence Dispute": "RSA's independent report noted defective DPM but also said flood water 'could have' penetrated the sub-floor; RSA relied on drying certificate; policyholder relied on same independent report and absence of any pre-existing damp history",
        "Outcome Category": "Upheld",
        "Outcome": "RSA required to deal with damp as part of flood claim; RSA not obliged to repair the structural floor defect if it can carry out a lasting and effective repair to the damp damage; RSA to pay £200 compensation for distress and inconvenience",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Where no alternative moisture source is identified and no pre-existing damp history exists, a flood-raised water table re-entering a property after drying is accepted as the cause; insurer must produce evidence to disprove flood causation — a floor defect does not automatically exclude flood as the moisture source",
        "Missing Evidence": "No evidence from RSA of an alternative moisture source; no evidence of pre-existing damp before either flood event",
        "Ombudsman Reasoning": "Property was certified dry but damp reappeared shortly after; independent report stated flood water could have penetrated sub-floor via defective DPM; in absence of any alternative explanation for moisture source, most likely cause is flood raising the water table; RSA failed to produce evidence showing moisture did not come from the flood",
        "Workflow Insight": "When an insurer disputes secondary flood damage (post-drying damp) as unrelated to the flood, the insurer must demonstrate an alternative moisture source; absence of such evidence tips the balance in the policyholder's favour; a drying certificate does not conclusively rule out subsequent water table ingress",
        "AI Rule Candidate": "IF post_drying_damp_present AND no_pre_existing_damp_history AND insurer_cannot_identify_alternative_moisture_source THEN attribute_secondary_damp_to_flood AND require_insurer_to_include_in_claim",
        "Source PDF": "DRN0070249.pdf",
    },
    {
        "Case ID": "FLOOD-002",
        "FOS Decision ID": "DRN0420936",
        "Insurer Name": "U K Insurance Limited",
        "FOS Decision Date": "5 Oct 2015",
        "Claim Type": "Flood claim declined — backed-up drain caused approximately 1.5 inches of water in basement; insurer attributed damage to pre-existing rising damp rather than flood",
        "Leak Source": "Backed-up drain — water accumulated in basement (approximately 1.5 inches) following drain backup; loss adjuster attended 4 weeks after flood when water had gone",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "UKI said there had not been a flood and damage was consistent with pre-existing rising damp throughout walls; loss adjuster found no evidence consistent with flood damage",
        "Evidence Dispute": "Loss adjuster observed rising damp throughout walls 4 weeks post-flood but conducted no moisture testing; policyholder produced a report from his employer who attended during the flood, and a letter from a kitchen fitter confirming water damage was visible 2 weeks after the incident",
        "Outcome Category": "Upheld",
        "Outcome": "UKI directed to reconsider claim after properly investigating whether property had pre-existing rising damp — specifically to conduct rising damp testing; UKI to pay £100 compensation for delay and handling failures; FOS declined to order immediate payment pending proper investigation",
        "Compensation Awarded (£)": 100,
        "Is Core Case": "Yes",
        "Key Policy Clause": "A flood does not have to be a sudden and violent event — water building up from a backed drain qualifies; where insurer declines on pre-existing damp grounds, it must demonstrate proper moisture testing, not merely visual observation weeks after the event; onus on insurer to prove pre-existing condition before declining",
        "Missing Evidence": "Formal rising damp test results from before the flood or shortly after; loss adjuster conducted no moisture testing — visual assessment 4 weeks post-flood insufficient to establish pre-existing rising damp",
        "Ombudsman Reasoning": "Loss adjuster attended when water had gone and conducted no testing; visual observation of damp weeks after a flood is insufficient to conclude it was pre-existing; insurer not satisfied that pre-existing condition proven on balance of probabilities; kitchen fitter letter, though late, cast sufficient doubt on loss adjuster's conclusion; proper testing required before claim can be validly declined on pre-existing damp grounds",
        "Workflow Insight": "A loss adjuster attending property weeks after a flood cannot rely on visual damp observation to decline; formal moisture testing is required to distinguish pre-existing from flood-caused damp; delay in attending the property is a factor weighing against the insurer when causation is disputed",
        "AI Rule Candidate": "IF insurer_declines_on_pre_existing_damp AND no_moisture_testing_conducted AND loss_adjuster_attended_weeks_post_flood THEN decline_insufficient_evidence AND direct_insurer_to_conduct_proper_damp_testing_before_redecision",
        "Source PDF": "DRN0420936.pdf",
    },
    {
        "Case ID": "FLOOD-003",
        "FOS Decision ID": "DRN1043965",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "11 Jan 2016",
        "Claim Type": "Peril classification dispute — AXA recorded claim as flood after garden drainage pump failure; policyholders argued pump failure constituted escape of water from apparatus. NOTE: FOS reclassified this claim as Escape of Water — retained in Flood database as a flood/EOW boundary case.",
        "Leak Source": "Garden drainage pump failure — pump failed to remove groundwater causing rainwater to back up through the drainage system and enter the property; policyholder confirmed the property had survived heavier rainfall without incident when pump was working",
        "Property Type": "Residential home",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "Not applicable — AXA accepted and settled the claim under the flood peril; policyholders dispute the flood classification because it increased their flood excess to £10,000 at renewal",
        "Evidence Dispute": "AXA file notes confirmed drainage system installed specifically for groundwater; pump failure caused rainwater to flow back up system and enter the house; policyholders said they had experienced heavier and prolonged rainfall before without water entry — pump failure was the differentiating factor; AXA argued drain is not a 'water apparatus' and water arrived from external source",
        "Outcome Category": "Upheld",
        "Outcome": "AXA required to reclassify October 2013 claim from flood to escape of water; FOS held that a garden drainage pump qualifies as 'apparatus' under the escape of water peril; no new monetary award (claim already settled)",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Escape of water from 'fixed water tanks, apparatus and pipes' — a garden drainage pump constitutes 'apparatus'; policy does not restrict escape of water to the domestic water system; pump failure is the proximate cause of water entry regardless of rainfall volume; no policy exclusion for water arriving from an external source",
        "Missing Evidence": "Not applicable — case decided on policy wording interpretation",
        "Ombudsman Reasoning": "Policy defines escape of water as from fixed water tanks, apparatus and pipes; no policy definition of apparatus; no restriction to domestic water system; garden drainage pump's sole purpose is to remove water from the home — it is reasonably classified as apparatus; pump failure caused water to escape the drain and enter the property; amount of rainfall was irrelevant — pump failure was the proximate cause regardless of how much rain fell",
        "Workflow Insight": "When water enters via a mechanical failure (pump, valve, apparatus), classify as escape of water not flood even if rainfall was a contributing factor; the proximate cause is the mechanical failure, not the external water volume; absence of a policy definition of apparatus should be interpreted in the customer's favour",
        "AI Rule Candidate": "IF water_entry_caused_by_apparatus_failure (pump, valve, mechanical device) AND pump_failure_is_proximate_cause THEN classify_as_escape_of_water NOT flood REGARDLESS_OF_concurrent_rainfall",
        "Source PDF": "DRN1043965.pdf",
    },
    {
        "Case ID": "FLOOD-004",
        "FOS Decision ID": "DRN-1611818",
        "Insurer Name": "Fairmead Insurance Limited",
        "FOS Decision Date": "8 Apr 2021",
        "Claim Type": "Peril classification dispute — L&G classified claim as flood after heavy rain caused water build-up entering through patio doors; policyholders argued storm was proximate cause and sought storm classification to avoid flood risk register and premium increase",
        "Leak Source": "Storm-induced surface flooding — heavy rainfall (confirmed storm conditions meeting L&G's 25mm/hr threshold) overwhelmed drainage; water built up and entered property at ground level through patio doors damaging flooring",
        "Property Type": "Residential home",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "Not applicable — L&G accepted and settled the claim under the flood peril; policyholders dispute the flood classification due to premium increase and placement on flood risk register",
        "Evidence Dispute": "Rain radar reports confirmed localised torrential downpours likely meeting L&G's 25mm/hr storm definition; policyholders argued storm volume of water overwhelmed drains causing damage — storm was the cause; L&G argued water entered from ground level indicating flood as proximate cause",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — L&G correct to classify claim under flood peril; storm conditions confirmed but storm was background occasion not dominant cause; flood (build-up and ground-level ingress of water) was the proximate cause of the flooring damage",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Proximate cause test — storm may generate the rainfall but if flood (water build-up and ingress at ground level) is the direct mechanism of damage, the claim correctly falls under the flood peril; storm is the background occasion not the dominant and effective cause; it would be unusual for rainfall alone to cause damage typical of a storm — heavy rainfall more commonly leads to flooding",
        "Missing Evidence": "Not applicable — case decided on proximate cause analysis",
        "Ombudsman Reasoning": "Storm conditions confirmed (likely 25mm/hr); storm caused water to build up; it was the accumulated water entering at ground level (flood) that caused damage to flooring, not direct storm action such as wind pressure or wind-driven rain; storm was the background occasion; flood was the proximate cause; classification as flood is accurate regardless of how the flood was triggered",
        "Workflow Insight": "Even where storm conditions are confirmed, if water enters via ground-level build-up and accumulation rather than direct storm action, classify as flood not storm; the question is not what caused the flood but what was the proximate cause of the damage — accumulation and ingress at ground level = flood",
        "AI Rule Candidate": "IF storm_conditions_confirmed AND water_enters_at_ground_level_via_accumulation AND no_direct_storm_action_damage (wind_pressure, wind_driven_rain) THEN proximate_cause = flood NOT storm",
        "Source PDF": "DRN-1611818.pdf",
    },
    {
        "Case ID": "FLOOD-005",
        "FOS Decision ID": "DRN-1846883",
        "Insurer Name": "U K Insurance Limited",
        "FOS Decision Date": "12 Jan 2021",
        "Claim Type": "Flood claim accepted; dispute over repair scope and cash settlement amount; policyholder sought improvement works (tanking, external drainage) and disputed that settlement was limited to insurer's own supplier costs",
        "Leak Source": "Heavy rainfall — rainwater penetrated through brickwork and rose through concrete basement floors due to raised water table; Section 19 Flood Investigation invoked by local Flood Risk Manager; property in flood zone with internal floor levels lower than external ground levels",
        "Property Type": "Residential home (unusual design — habitable space on first floor; ground floor set to storage and casual use)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "Not applicable — claim accepted; UKI limited cash settlement to own contractor cost (£3,970.83) and declined to fund improvement works (tanking, external drainage sumps) on the basis that these were betterment not covered under the policy",
        "Evidence Dispute": "UKI building consultant said property design was built to cope with localised flooding (internal floor levels below external); structural engineer (Mr R) said cosmetic repairs proposed by UKI were inadequate and flooding was inconceivable as a one-off; contractor quote included £61,000 for external prevention and tanking works; contractor confirmed improvements wouldn't stop flooding due to floor level differential",
        "Outcome Category": "Not Upheld",
        "Outcome": "UKI cash settlement of £3,970.83 upheld as fair — limited to own contractor cost in line with policy terms; improvement works (tanking, external drainage) not required as property was not demonstrably watertight before the flood given its design and location; £390 compensation already paid by UKI for drying company failures and unannounced visits endorsed as adequate; no additional award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Cash settlement may be capped at insurer's own supplier cost under policy terms; insurer is not required to fund betterment (tanking, flood prevention) unless: (1) the property was demonstrably watertight before the flood, and (2) repairs would be ineffective without the improvement; where structural evidence shows property was not watertight before the flood and flooding is likely recurring, betterment principle applies and insurer need not fund prevention works",
        "Missing Evidence": "Evidence that property was watertight before the initial flooding — structural engineer concluded flooding was inconceivable as a one-off given floor levels, flood zone location, and drainage constraints; this precluded the tanking requirement",
        "Ombudsman Reasoning": "Structural engineer confirmed flooding was inconceivable as a one-off; UKI building consultant noted property design accommodated local flooding; contractor confirmed improvements would not stop flooding due to floor level differential; cannot conclude property was watertight before flood; betterment principle applies — UKI not required to fund prevention; cash settlement correctly limited to own supplier cost under policy terms",
        "Workflow Insight": "Insurer may limit cash settlement to own-supplier cost; tanking and flood prevention works are only required where property was demonstrably watertight before the flood and repairs would be ineffective without the improvement; where structural evidence shows inherent vulnerability (floor levels, flood zone), betterment principle applies",
        "AI Rule Candidate": "IF property_not_demonstrably_watertight_before_flood (structural_evidence, design_factors, flood_zone) AND improvement_works_would_not_prevent_future_flooding THEN insurer_not_required_to_fund_betterment AND cash_settlement_capped_at_own_supplier_cost",
        "Source PDF": "DRN-1846883.pdf",
    },
    {
        "Case ID": "FLOOD-006",
        "FOS Decision ID": "DRN-2024904",
        "Insurer Name": "Ageas Insurance Limited",
        "FOS Decision Date": "31 Dec 2020",
        "Claim Type": "Dispute over which policy excess applies — policyholder argued storm excess (£350) should apply rather than flood excess (£1,000) after ground-level water ingress during a storm",
        "Leak Source": "Storm-induced surface flooding — heavy storm rainfall overwhelmed drainage and possibly contributed to a local river overflowing; water entered property at ground floor level from external source in substantial volume",
        "Property Type": "Residential home",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "Not applicable — claim accepted under flood peril; Ageas applied £1,000 flood excess rather than £350 buildings excess covering storm",
        "Evidence Dispute": "Ageas argued water entered at ground level from external source in substantial volume meeting policy flood definition; policyholder argued storm was the cause of the water and no river burst its banks near his home — only rainfall from the storm caused the water; policy flood definition cited by Ageas",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Ageas correct to apply flood excess; water entering at ground level from external source in substantial volume meets policy flood definition; even if both storm and flood sections applied, policy clause requires the higher excess to be charged; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood defined as 'overflowing or movement of a body of water (volumes, weight or force substantial and exceptional beyond normal limits) which enters a property rapidly from an external source from outside the Buildings and which enters at the ground floor or below'; where both storm and flood sections are engaged by the same claim, only one excess is deducted — if excesses differ, the higher excess applies",
        "Missing Evidence": "Not applicable — decided on policy wording and excess clause interpretation",
        "Ombudsman Reasoning": "Water entered property at ground level from external source in substantial volume — meets policy definition of flood; storm was the cause of the heavy rainfall but the rainfall itself did not damage the property — it was the overflowing water entering at ground level that caused damage; flood was the proximate cause; even if both storm and flood sections applied, policy requires higher (flood) excess; outcome unchanged whether or not river contribution is confirmed",
        "Workflow Insight": "When determining which excess to apply, assess the mechanism of water entry not its ultimate trigger; ground-level entry from an external source in substantial volume meets the flood excess threshold regardless of whether a storm caused the water; where both storm and flood sections are engaged, always apply the higher excess under the standard multi-section excess clause",
        "AI Rule Candidate": "IF water_enters_at_ground_floor_or_below AND source_is_external AND volume_substantial THEN apply_flood_excess; IF both_storm_and_flood_sections_engaged THEN apply_higher_of_the_two_excesses",
        "Source PDF": "DRN-2024904.pdf",
    },
    {
        "Case ID": "FLOOD-007",
        "FOS Decision ID": "DRN-2075105",
        "Insurer Name": "The Salvation Army General Insurance Corporation Ltd",
        "FOS Decision Date": "29 Oct 2020",
        "Claim Type": "No flood event — dispute over flood exclusion endorsement applied at policy inception; policyholder argued his postcode carries low flood risk per government agency data and endorsement was unfair",
        "Leak Source": "Not applicable — no flood event occurred; dispute is about underwriting endorsement at inception",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "SAGIC's postcode flood mapping tool flagged the policyholder's postcode as 'Flooding: decline flood cover'; SAGIC said it cannot change its underwriting criteria; endorsement communicated and agreed before purchase",
        "Evidence Dispute": "Policyholder held government agency confirmation that his area is at low risk of flooding; SAGIC's internal postcode checker showed 'Flooding: decline flood cover' for that postcode — applied consistently to all policyholders at that postcode for this product",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — SAGIC entitled to apply its underwriting criteria; postcode checker applied consistently; flood endorsement clearly disclosed and explained before purchase; policyholder could have declined or shopped elsewhere; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Administrative",
        "Key Policy Clause": "An insurer may decide what risks it wishes to take on — this is a commercial decision FOS cannot override; FOS can only intervene if underwriting criteria have been applied inconsistently (treating the policyholder differently from others in the same position) or the endorsement was not disclosed before purchase; government flood risk data does not override an insurer's internal postcode mapping tool if that tool is applied consistently",
        "Missing Evidence": "Not applicable — no flood claim; underwriting decision challenge only",
        "Ombudsman Reasoning": "Any policyholder at the same postcode applying for this product would have received the same endorsement — consistent application confirmed; endorsement wording was clear and communicated before purchase; policyholder had the option to cancel or shop around; FOS cannot interfere with a commercial underwriting decision applied fairly and consistently",
        "Workflow Insight": "Flood exclusion endorsements applied at inception by postcode mapping are valid provided: (1) the criteria are applied consistently to all policyholders at that postcode, and (2) the endorsement is clearly disclosed before purchase; government agency flood risk data does not supersede an insurer's internal underwriting tool",
        "AI Rule Candidate": "IF flood_endorsement_applied_at_inception AND postcode_checker_applied_consistently AND endorsement_disclosed_before_purchase THEN endorsement_valid AND fos_will_not_override; government_flood_risk_data DOES_NOT override insurer_postcode_underwriting_tool",
        "Source PDF": "DRN-2075105.pdf",
    },
    {
        "Case ID": "FLOOD-008",
        "FOS Decision ID": "DRN-2101511",
        "Insurer Name": "Accredited Insurance (Europe) Ltd",
        "FOS Decision Date": "16 Sep 2020",
        "Claim Type": "Flood claim declined (flood cover excluded from policy schedule); policyholder argued Storm Ciara was the proximate cause so damage should be covered under the storm section of the policy",
        "Leak Source": "Watercourse overflow — Storm Ciara caused a local beck to accumulate excessive water and burst its banks; floodwater from the beck entered ground floor of property damaging floor coverings, kitchen units and soft furnishings",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Policy schedule excluded flood cover for both buildings and contents; surveyor and local press reports confirmed damage was caused by floodwater from the beck overflowing — this constitutes flood under the policy definition; storm was background cause not proximate cause",
        "Evidence Dispute": "Weather and news reports confirming Storm Ciara caused significant damage in the area; surveyor confirmed beck overflow as cause of ground floor flooding; local press confirmed beck flooded gardens; policyholder argued storm was proximate cause as drains could not cope with Ciara's rainfall; policy storm section did not explicitly exclude floods caused by storms",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — flood cover was validly excluded from the policy; floodwater from the beck overflowing was the proximate cause of damage, not direct storm action; where flood exclusion applies and flood is the proximate cause, storm cover does not substitute even if storm caused the flood; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Where flood cover is excluded from the policy and damage is caused by floodwater from an external watercourse entering at ground level, the flood exclusion applies regardless of what caused the watercourse to overflow; storm cover does not extend to flood damage even where storm caused the flood — when one cause is covered (storm) and the other is the direct primary cause and is excluded (flood), the insurer may decline; proximate cause determines which peril applies",
        "Missing Evidence": "Not applicable — both parties accepted the facts; case decided on proximate cause and policy interpretation",
        "Ombudsman Reasoning": "Flood cover excluded from policy schedule — confirmed; Storm Ciara caused excessive rainfall filling the beck; beck burst its banks and floodwater entered the property at ground level — this is flood as defined in the policy; storm was background cause; beck flooding (not direct storm rainfall) was the direct and primary cause of damage; where one cause is excluded and is the dominant proximate cause, insurer entitled to decline; storm cover cannot be used as a substitute for excluded flood cover when flood was the mechanism of damage",
        "Workflow Insight": "If flood is excluded from the policy, conduct proximate cause analysis to determine whether storm cover could apply; if an external watercourse overflow (not direct storm action) caused the damage, flood exclusion takes precedence over storm cover; storm cover does not substitute for excluded flood cover when flood was the proximate cause",
        "AI Rule Candidate": "IF flood_cover_excluded AND external_watercourse_overflow_caused_damage AND storm_was_background_not_proximate_cause THEN decline_under_flood_exclusion AND storm_cover_does_not_apply",
        "Source PDF": "DRN-2101511.pdf",
    },
    {
        "Case ID": "FLOOD-009",
        "FOS Decision ID": "DRN2337599",
        "Insurer Name": "UK Insurance Limited",
        "FOS Decision Date": "14 Mar 2016",
        "Claim Type": "Flood claim accepted; multiple handling failures including disputed repair scope (quotes ranging from £25,937 to £161,617), alternative accommodation costs withheld, fraud suspicions raised without proper investigation, first-floor items disposed of without insurer inspection",
        "Leak Source": "Drainage system failure — torrential rainfall caused drains to back up and forced off manhole cover; sewage flooded garden and entered rear of property through air bricks",
        "Property Type": "Residential home",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "Not applicable — claim accepted; disputes about repair scope, ALE payment, and alleged fraud",
        "Evidence Dispute": "Repair quotes wildly divergent: £25,937 initial, revised to approximately £260,000 via structural engineer, then formal tender produced £161,617.90 and £59,272 — UKI preferred lowest quote from contractor who had not inspected interior; UKI fraud suspicions based on prior planning permission and unusual ALE features; policyholders disputed fraud allegations and claimed ALE costs were legitimate; first-floor items disposed of before insurer inspection",
        "Outcome Category": "Upheld in Part",
        "Outcome": "UKI to appoint independent RICS-accredited surveyor (selected by policyholders from UKI list of three) to report on flood repair scope and outstanding work; UKI to settle claim in line with surveyor's report; UKI to pay all outstanding insurance-related costs (ALE, travel, postal redirection) without relying on time limits; if cash settlement, interest at 8% simple per annum from date of loss; £2,250 compensation already paid by UKI endorsed as adequate — no additional award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Where repair quotes diverge significantly and a contractor has not inspected the property interior, insurer cannot insist on the lower quote; ALE time limits cannot be invoked where claim delay was substantially caused by the insurer's own handling failures; fraud suspicions must be actively investigated (site visits) not based on background checks or speculation; 'legal fees' clause restricted to fees arising directly from damage — does not cover third-party litigation costs; items disposed of before insurer inspection are excluded from claim",
        "Missing Evidence": "Interior inspection by contractor who submitted the lower repair quote; UKI should have visited the alternative accommodation to verify occupancy rather than conducting background checks",
        "Ombudsman Reasoning": "Repair quote process unreliable — 10x variance and lower quote from contractor who had not inspected interior; UKI wrong to insist on lower quote in those circumstances; planning permission history irrelevant to fraud — insurer should have visited ALE property instead of speculating; ALE time limits cannot be applied where delay was partly UKI's fault; first-floor items disposed of without inspection — correctly excluded; legal fees for suing water company not covered; UKI's own loss adjuster described cause as flood (not escape of water) — UKI wrong to deny grant eligibility on that basis",
        "Workflow Insight": "Where repair quotes differ by more than 100%, appoint an independent RICS surveyor to establish scope rather than insisting on the lower quote; fraud suspicions require active investigation (visit the property, inspect the ALE) not background checks; ALE time limits are suspended where claim delay is substantially attributable to insurer conduct; do not allow items to be disposed of before insurer inspection",
        "AI Rule Candidate": "IF repair_quotes_differ_by_more_than_100_percent AND lower_quote_contractor_has_not_inspected_property THEN direct_independent_RICS_surveyor_appointment; IF claim_delay_substantially_caused_by_insurer THEN ALE_time_limits_suspended; IF fraud_suspected THEN active_site_investigation_required NOT background_checks_only",
        "Source PDF": "DRN2337599.pdf",
    },
    {
        "Case ID": "FLOOD-010",
        "FOS Decision ID": "DRN-2339494",
        "Insurer Name": "Aviva Insurance Limited",
        "FOS Decision Date": "14 Dec 2020",
        "Claim Type": "Flood claim accepted; dispute over whether installation of a tanking (waterproofing) system must be included in basement reinstatement; insurer argued waterproofing was not damaged by flood and upgrade was betterment",
        "Leak Source": "Heavy rainfall — drainage unable to cope with volume; water gathered beneath property and rose round perimeter of concrete basement slabs; basement bedrooms flooded; surveyor noted no previous records of flooding beneath the property",
        "Property Type": "Residential home (converted basement used as children's bedrooms)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "Not applicable — claim accepted; Aviva refused to include tanking repair/installation, arguing the existing waterproofing system was not damaged by the flood — merely overwhelmed — and that installing tanking would constitute betterment not covered under the policy",
        "Evidence Dispute": "Surveyor's report made no reference to inspecting the waterproofing system; Aviva relied on post-inspection email exchange with policyholder to conclude no tanking damage; policyholder received advice from a basement waterproofing company that once water finds its way through a waterproofing system it compromises that system's future adequacy; basement flooded again in August 2020; policyholder provided 1999 builder invoice confirming tanking had been installed when basement was converted",
        "Outcome Category": "Upheld",
        "Outcome": "Aviva required to include installation of tanking (waterproofing system) in the repairs under the flood claim; surveyor failed to inspect or report on waterproofing system — reliance on post-inspection email exchange with policyholder insufficient; basement was watertight for 20 years before flood; tanking is necessary for effective and lasting repair, not betterment; no compensation award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Insurer must make effective and long-lasting repairs to flood damage; where a basement was demonstrably watertight before the flood (no prior ingress history) and a repair without tanking would not be effective or lasting, tanking is a necessary element of reinstatement not betterment; a surveyor must explicitly inspect and report on the waterproofing system before opining that it was undamaged; reliance on policyholder email discussion post-inspection is insufficient to establish no damage to tanking",
        "Missing Evidence": "Surveyor's report did not address the waterproofing system inspection at all; Aviva provided no expert evidence post-provisional-decision establishing that the tanking had not been compromised by the flood",
        "Ombudsman Reasoning": "Surveyor made no reference to inspecting the waterproofing system in his report; cannot conclude tanking undamaged based on email discussions with an unqualified policyholder; basement waterproofing had been in place for approximately 20 years; surveyor noted no previous records of flooding — basement was watertight before the flood; repair without tanking would not be effective or lasting; subsequent flooding in August 2020 confirmed this; tanking is not betterment where it is necessary to make the repair lasting",
        "Workflow Insight": "If insurer declines to include tanking in basement flood repairs, verify whether surveyor explicitly inspected and reported on the waterproofing system; if not, the surveyor's opinion that no damage occurred to waterproofing is unsupported; if basement was demonstrably watertight before the flood (documented by no prior ingress and/or installation records), require tanking as part of lasting reinstatement",
        "AI Rule Candidate": "IF basement_flooded AND basement_was_watertight_before_flood (no_prior_ingress_history) AND surveyor_did_not_inspect_waterproofing_system THEN surveyor_opinion_on_tanking_unsupported AND tanking_required_as_part_of_lasting_repair AND not_betterment",
        "Source PDF": "DRN-2339494.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Flood_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
