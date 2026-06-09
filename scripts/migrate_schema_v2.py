"""
Schema migration v2 — adds 7 new columns to Escape_of_Water_Case_Database.xlsx
and backfills them for EOW-001 through EOW-015.

New columns inserted:
  Insurer Name          (after FOS Decision ID)
  FOS Decision Date     (after Insurer Name)
  Dispute Type          (after Property Type)
  Coverage Decision     (after Dispute Type)
  Outcome Category      (after Evidence Dispute)
  Compensation Awarded  (after Outcome)
  Is Core Case          (after Compensation Awarded)

Run from repo root:  py scripts/migrate_schema_v2.py
"""

import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Target column definition  (order matters)
# ---------------------------------------------------------------------------
COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",           # NEW
    "FOS Decision Date",      # NEW
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",           # NEW
    "Coverage Decision",      # NEW
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",       # NEW
    "Outcome",
    "Compensation Awarded (£)",  # NEW
    "Is Core Case",           # NEW
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

COL_WIDTHS = {
    "Case ID": 12,
    "FOS Decision ID": 16,
    "Insurer Name": 30,
    "FOS Decision Date": 18,
    "Claim Type": 40,
    "Leak Source": 40,
    "Property Type": 22,
    "Dispute Type": 30,
    "Coverage Decision": 25,
    "Rejection Reason": 50,
    "Evidence Dispute": 55,
    "Outcome Category": 20,
    "Outcome": 55,
    "Compensation Awarded (£)": 22,
    "Is Core Case": 14,
    "Key Policy Clause": 55,
    "Missing Evidence": 45,
    "Ombudsman Reasoning": 70,
    "Workflow Insight": 70,
    "AI Rule Candidate": 70,
    "Source PDF": 22,
}

# Columns that get centred (not wrapped)
CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Outcome Category", "Coverage Decision",
    "Compensation Awarded (£)", "Is Core Case", "Source PDF",
}

# ---------------------------------------------------------------------------
# Backfill data — new fields only, keyed by Case ID
# ---------------------------------------------------------------------------
BACKFILL = {
    "EOW-001": {
        "Insurer Name": "Fairmead Insurance Limited",
        "FOS Decision Date": "11 Sep 2020",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Upheld in Part",
        "Compensation Awarded (£)": 150,
        "Is Core Case": "Yes",
    },
    "EOW-002": {
        "Insurer Name": "U K Insurance Limited (UKI)",
        "FOS Decision Date": "30 Mar 2022",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Outcome Category": "Upheld",
        "Compensation Awarded (£)": 400,
        "Is Core Case": "Yes",
    },
    "EOW-003": {
        "Insurer Name": "QIC Europe Ltd",
        "FOS Decision Date": "08 Jan 2024",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Upheld",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
    },
    "EOW-004": {
        "Insurer Name": "UK Insurance Limited (UKI)",
        "FOS Decision Date": "27 Feb 2015",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Not Applicable",
        "Outcome Category": "Upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Administrative",
    },
    "EOW-005": {
        "Insurer Name": "Royal & Sun Alliance Insurance Ltd (RSA)",
        "FOS Decision Date": "07 Dec 2023",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Upheld",
        "Compensation Awarded (£)": 150,
        "Is Core Case": "Yes",
    },
    "EOW-006": {
        "Insurer Name": "Covea Insurance plc",
        "FOS Decision Date": "08 Apr 2020",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Not Upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
    },
    "EOW-007": {
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "29 Jul 2020",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Not Upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
    },
    "EOW-008": {
        "Insurer Name": "Insurers at Lloyd's (Society of Lloyd's)",
        "FOS Decision Date": "06 May 2020",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Not Upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
    },
    "EOW-009": {
        "Insurer Name": "U K Insurance Limited (UKI)",
        "FOS Decision Date": "10 Mar 2021",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Outcome Category": "Not Upheld",
        "Compensation Awarded (£)": 750,
        "Is Core Case": "No — Handling Dispute",
    },
    "EOW-010": {
        "Insurer Name": "Zurich Insurance PLC",
        "FOS Decision Date": "28 Feb 2019",
        "Dispute Type": "Pre-Inception Damage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Not Upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
    },
    "EOW-011": {
        "Insurer Name": "Fairmead Insurance Limited",
        "FOS Decision Date": "01 Mar 2021",
        "Dispute Type": "Pre-Inception Damage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Upheld",
        "Compensation Awarded (£)": 400,
        "Is Core Case": "Yes",
    },
    "EOW-012": {
        "Insurer Name": "Admiral Insurance (Gibraltar) Limited",
        "FOS Decision Date": "20 Apr 2021",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Upheld",
        "Compensation Awarded (£)": 400,
        "Is Core Case": "Yes",
    },
    "EOW-013": {
        "Insurer Name": "Covea Insurance Plc",
        "FOS Decision Date": "29 Jun 2021",
        "Dispute Type": "Pre-Inception Damage Dispute",
        "Coverage Decision": "Declined — Full",
        "Outcome Category": "Compensation Only",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
    },
    "EOW-014": {
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "20 May 2021",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Outcome Category": "Upheld",
        "Compensation Awarded (£)": 150,
        "Is Core Case": "No — Commercial",
    },
    "EOW-015": {
        "Insurer Name": "Endsleigh Insurance Services Ltd",
        "FOS Decision Date": "01 Jul 2021",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Accepted",
        "Outcome Category": "Not Upheld",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
    },
}

NEW_COLS = {
    "Insurer Name", "FOS Decision Date", "Dispute Type",
    "Coverage Decision", "Outcome Category",
    "Compensation Awarded (£)", "Is Core Case",
}


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def header_fill():
    return PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

def row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases",
        "Escape_of_Water_Case_Database.xlsx"
    )

    # --- read existing data -------------------------------------------------
    wb_old = openpyxl.load_workbook(xlsx_path)
    ws_old = wb_old.active

    old_headers = [c.value for c in next(ws_old.iter_rows(min_row=1, max_row=1))]
    existing_rows = []
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        existing_rows.append(dict(zip(old_headers, row)))
    wb_old.close()

    # --- rebuild workbook ---------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Escape of Water Cases"

    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Calibri", size=10)
    border    = thin_border()
    hdr_fill  = header_fill()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # write header row
    for ci, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.border = border
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 20)
    ws.row_dimensions[1].height = 30

    # write data rows
    for ri, row_data in enumerate(existing_rows, start=2):
        case_id = row_data.get("Case ID", "")
        backfill = BACKFILL.get(case_id, {})
        fill = row_fill(ri % 2 == 0)

        for ci, col in enumerate(COLUMNS, start=1):
            if col in NEW_COLS:
                value = backfill.get(col, "")
            else:
                value = row_data.get(col, "")

            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font   = cell_font
            cell.fill   = fill
            cell.border = border
            cell.alignment = centre if col in CENTRED_COLS else wrap

        ws.row_dimensions[ri].height = 120

    ws.freeze_panes = "A2"

    wb.save(xlsx_path)

    # --- report -------------------------------------------------------------
    total_data_rows = len(existing_rows)
    print(f"Saved: {xlsx_path}")
    print(f"Total columns : {len(COLUMNS)}")
    print(f"Total data rows: {total_data_rows}")
    print()
    print("Column list:")
    for i, c in enumerate(COLUMNS, 1):
        tag = " [NEW]" if c in NEW_COLS else ""
        print(f"  {i:>2}. {c}{tag}")
    print()
    print("New field values (first 5 rows):")
    wb2 = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws2 = wb2.active
    new_col_indices = [i for i, c in enumerate(COLUMNS, 1) if c in NEW_COLS]
    for row in ws2.iter_rows(min_row=2, max_row=6, values_only=True):
        case_id = row[0]
        new_vals = {COLUMNS[i-1]: row[i-1] for i in new_col_indices}
        print(f"  {case_id}: {new_vals}")
    wb2.close()


if __name__ == "__main__":
    main()
