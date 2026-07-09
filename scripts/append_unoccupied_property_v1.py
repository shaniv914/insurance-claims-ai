"""
Standard append script for all Unoccupied Property batches (schema v1, 21
columns — same schema as EOW v2 / Storm v1 / Flood v1 / Subsidence v1 /
Theft v1; column 6 = "Unoccupied Period / Circumstance").

Active — reuse this script for every future batch: replace NEW_CASES below
with the next batch's cases and run again. Appends only — never modifies
existing rows in knowledge/case-databases/Unoccupied_Property_Case_Database.xlsx.

Current contents: Batch 3 (UNOC-021 to UNOC-030).
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Unoccupied Period / Circumstance",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": [
        "Coverage Dispute",
        "Endorsement / Exclusion Challenge",
        "Broker Conduct Dispute",
        "Claim Recording / Administrative Dispute",
        "Claim Quantum Dispute",
        "Third-Party / Liability Dispute",
        "Causation Dispute",
    ],
    "Coverage Decision": [
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — With Deductions",
        "Not Applicable",
    ],
    "Outcome Category": [
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Not Applicable",
    ],
    "Is Core Case": [
        "Yes",
        "No — Administrative",
        "No — Broker Dispute",
        "No — Quantum Only",
        "No — Time-Barred",
    ],
}

ROW_FILL_ODD  = PatternFill(start_color="F9EDF2", end_color="F9EDF2", fill_type="solid")
ROW_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_FONT      = Font(name="Calibri", size=10)
ROW_BORDER    = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

NEW_CASES = [
    {
        "Case ID": "UNOC-021",
        "FOS Decision ID": "DRN3308022",
        "Insurer Name": "Ageas Insurance Limited",
        "FOS Decision Date": "4 Nov 2018",
        "Claim Type": (
            "Home insurance (rental property let to family) — escape of water discovered "
            "~1 month after policy inception; insurer declined citing numerous "
            "inconsistencies in the policyholders' account of occupancy and compliance "
            "with policy conditions, without conclusively establishing the property had "
            "actually been unoccupied or unfurnished for the qualifying period"
        ),
        "Unoccupied Period / Circumstance": (
            "Property let to family; policy incepted only about a month before the escape "
            "of water was discovered; insurer suspected the property was not genuinely "
            "occupied as claimed (citing a friend's inconsistent account of how often he "
            "stayed, and the property being under renovation) but could not show "
            "unoccupancy or lack of furnishing had lasted more than 90 consecutive days, "
            "since the policy itself had only been in force for about a month"
        ),
        "Property Type": "Residential rental property (let to family)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Insurer said there were too many inconsistencies in the claim and it hadn't "
            "been shown the property was occupied as described or that policy conditions "
            "had been complied with — including doubts over the intended letting to "
            "family, a friend's account of overnight stays, renovation works in progress, "
            "and a three-month historical gap in insurance cover"
        ),
        "Evidence Dispute": (
            "Mr and Mrs H: property was going to be let to family, doubts raised by the "
            "insurer were misunderstandings or explicable (e.g. quoting the plumber a "
            "repair estimate as an experienced builder wanting an accurate figure); gap in "
            "insurance history was inadvertent given Mr H's serious health issues. Ageas: "
            "multiple inconsistencies (letting intentions, friend's stay frequency, "
            "ongoing renovation, plumber quote conversation, insurance history) meant it "
            "hadn't been shown the property was occupied as claimed or that conditions "
            "were met. FOS: the 90-day unoccupied/unfurnished exclusion operates as an "
            "exclusion, so the burden is on the insurer to show, on the balance of "
            "probabilities, that the property was unoccupied or unfurnished for more than "
            "90 days — since the policy had only started about a month before the loss, "
            "that could not have been the case regardless of the other inconsistencies; "
            "inconsistencies alone are not usually sufficient grounds to refuse a claim "
            "absent strong grounds that doing so is fair and reasonable"
        ),
        "Outcome Category": "Upheld",
        "Outcome": (
            "Ageas required to settle Mr and Mrs H's claim in accordance with the "
            "remaining policy terms and conditions"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "Where an unoccupied/unfurnished exclusion requires the property to have been "
            "unoccupied or unfurnished for more than a specified number of consecutive "
            "days (here 90), that condition operates as an exclusion rather than a "
            "coverage precondition, meaning the burden of proof lies with the insurer to "
            "show, on the balance of probabilities, that the qualifying period was met — "
            "where the policy itself has only been in force for a shorter period than the "
            "qualifying threshold, the insurer cannot discharge that burden regardless of "
            "how many other inconsistencies exist in the policyholder's account; general "
            "inconsistencies in a policyholder's evidence, without more, do not meet this "
            "service's threshold of 'strong grounds' required to justify refusing an "
            "otherwise valid claim"
        ),
        "Missing Evidence": (
            "Evidence that the property had actually been unoccupied or unfurnished for "
            "more than 90 consecutive days — could not exist given the policy had only "
            "been in force for about a month before the loss was discovered"
        ),
        "Ombudsman Reasoning": (
            "Policy covered EOW unless the property had been left unfurnished or "
            "unoccupied for more than 90 consecutive days; this operates as an exclusion, "
            "so it fell to Ageas to prove the qualifying period on the balance of "
            "probabilities; since the policy had only incepted about a month earlier, this "
            "was structurally impossible regardless of the unresolved inconsistencies "
            "about furnishing, letting intentions, or the plumber conversation; the "
            "insurance-history gap was more likely an oversight given Mr H's health issues "
            "than a deliberate misrepresentation; overall, inconsistencies are not "
            "themselves strong grounds to decline a claim"
        ),
        "Workflow Insight": (
            "Where an unoccupied/unfurnished exclusion is time-qualified (e.g. 90 "
            "consecutive days), claims handlers must first check whether the policy has "
            "even been in force long enough for the qualifying period to have elapsed "
            "before investigating disputed factual inconsistencies about occupancy — if "
            "the policy is younger than the qualifying period, the exclusion cannot apply "
            "regardless of what is eventually established about the property's actual "
            "use; unresolved inconsistencies in a policyholder's account should be weighed "
            "against this service's 'strong grounds' threshold rather than treated as "
            "automatically fatal to a claim"
        ),
        "AI Rule Candidate": (
            "IF unoccupied_or_unfurnished_exclusion_requires_qualifying_period_of_N_consecutive_days "
            "AND policy_has_been_in_force_for_fewer_than_N_days_since_inception "
            "THEN exclusion_cannot_apply_regardless_of_disputed_occupancy_evidence; "
            "unoccupied_unfurnished_exclusion_operates_as_an_exclusion_so_burden_of_proof_is_on_insurer_not_policyholder; "
            "general_unresolved_inconsistencies_in_policyholder_account_do_not_alone_meet_strong_grounds_threshold_to_decline_a_claim"
        ),
        "Source PDF": "DRN3308022.pdf",
    },
    {
        "Case ID": "UNOC-022",
        "FOS Decision ID": "DRN3390381",
        "Insurer Name": "Swinton Group Ltd",
        "FOS Decision Date": "7 Sep 2019",
        "Claim Type": (
            "Broker complaint (not a declined claim) — policyholder alleged Swinton "
            "misled her into believing unoccupied buildings cover was only available "
            "through it and overcharged her for short-term (3/6-month) unoccupied "
            "property policies compared to a cheaper annual policy she later obtained "
            "from another provider"
        ),
        "Unoccupied Period / Circumstance": (
            "Property insured on a rolling 6-month then 3-month unoccupied buildings "
            "basis from April 2015 until Swinton would no longer renew short-term cover "
            "from January 2017 (requiring a 12-month term instead), at which point she "
            "obtained a cheaper annual policy elsewhere"
        ),
        "Property Type": "Unoccupied property (type not further specified in decision)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a coverage decline — dispute concerned whether Swinton "
            "misrepresented itself as the sole provider of unoccupied buildings cover and "
            "whether its short-term premiums were unfairly high compared to annual cover "
            "available elsewhere"
        ),
        "Evidence Dispute": (
            "Miss G: was told she could only get unoccupied buildings cover through "
            "Swinton, and it had overcharged her (over £1,200/year for short-term cover "
            "vs £170/year annual cover elsewhere). Swinton: found no evidence it told her "
            "it was the sole provider; it could only advise on its own panel prices, not "
            "the wider market; premium was accurately calculated to reflect risk plus a "
            "reasonable profit margin, and short-term policies generate more "
            "administrative work reflected in price; she had a 14-day cooling-off period "
            "each renewal to shop around. FOS: no recording existed of the original call, "
            "but it was more likely Swinton said only one insurer on its panel offered the "
            "product, not that it was the only provider in the market; premium "
            "calculation checked and found accurate and fair; even had Swinton flagged "
            "that an annual policy would be cheaper, Miss G was hoping to sell the "
            "property before each policy expired and had chosen short-term cover "
            "deliberately, so she would likely not have acted differently — this was "
            "confirmed when, even after Swinton stopped offering short-term cover, she "
            "sought another short-term provider rather than take a 12-month policy "
            "through Swinton"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — no evidence Swinton misrepresented itself as the sole "
            "provider, its premiums were fairly calculated, and Miss G would not have "
            "acted differently even with fuller information about annual pricing"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "A broker is entitled to price short-term (e.g. 3 or 6-month) unoccupied "
            "property cover higher on a pro-rata basis than an equivalent annual policy, "
            "reflecting the additional administrative burden of more frequent renewals, "
            "provided the pricing methodology is shown to be applied consistently and "
            "fairly; a broker is not obliged to advise on whether a competitor or an "
            "alternative policy term would offer better value, only to price its own "
            "panel accurately — where a consumer's own stated circumstances (e.g. an "
            "intention to sell the property imminently) make it unlikely she would have "
            "chosen a different term even with fuller comparative information, a failure "
            "to proactively highlight cheaper alternatives does not cause a compensable "
            "loss"
        ),
        "Missing Evidence": (
            "A recording of the original 2015 sales call — absent, requiring the "
            "ombudsman to decide on the balance of probabilities what was most likely "
            "said"
        ),
        "Ombudsman Reasoning": (
            "No call recording existed for the initial 2015 contact; more likely Swinton "
            "said only one panel insurer offered the product rather than claiming market "
            "exclusivity, since a simple check would have disproved a false exclusivity "
            "claim; premium calculation reviewed confidentially and found accurate, "
            "reflecting risk plus profit and short-term administrative costs; Miss G "
            "renewed on a short-term basis each time because she hoped to sell soon, and "
            "even after Swinton withdrew short-term cover she sought another short-term "
            "provider rather than accept a cheaper annual Swinton policy — demonstrating "
            "she valued the short-term flexibility over price and would not have acted "
            "differently with better information"
        ),
        "Workflow Insight": (
            "Where a broker complaint alleges a misleading claim of market exclusivity "
            "but no call recording exists, assess plausibility against what a reasonable "
            "broker would say and what a quick market check would reveal — an "
            "exclusivity claim that could be trivially disproved is inherently less "
            "likely to have been made; when assessing detriment from an alleged pricing "
            "non-disclosure, look at what the consumer's subsequent actual choices reveal "
            "about what she would have done with fuller information, not just her stated "
            "hindsight preference"
        ),
        "AI Rule Candidate": (
            "IF broker_prices_short_term_unoccupied_property_cover_higher_per_month_than_equivalent_annual_cover "
            "AND pricing_methodology_is_shown_to_be_accurate_and_consistently_applied "
            "THEN higher_short_term_pricing_is_not_unfair_even_without_proactive_comparison_to_annual_cover; "
            "IF consumers_subsequent_choices_show_she_continued_to_prefer_short_term_cover_even_after_being_offered_a_cheaper_annual_alternative "
            "THEN broker_failure_to_earlier_highlight_annual_cover_savings_did_not_cause_a_compensable_loss"
        ),
        "Source PDF": "DRN3390381.pdf",
    },
    {
        "Case ID": "UNOC-023",
        "FOS Decision ID": "DRN-3516774",
        "Insurer Name": "Society of Lloyd's",
        "FOS Decision Date": "4 Jul 2022",
        "Claim Type": (
            "Unoccupied Property insurance — escape of water from a loft leak discovered "
            "by policyholder's daughter during a routine check; insurer avoided the "
            "policy from inception (and separately from renewal) for misrepresentation "
            "about intended occupancy, and identified alternative grounds (Water Turn "
            "Off Clause breach, 30-day inspection condition breach) that would likely "
            "also have defeated the claim"
        ),
        "Unoccupied Period / Circumstance": (
            "Policyholder took out Unoccupied Property insurance in December 2019 while "
            "working abroad for about a year; Statement of Fact recorded the property "
            "would become occupied in 9-12 months and was not occupied from inception; "
            "at the December 2020 renewal, the Statement of Fact recorded the property "
            "had been unoccupied for 'more than 1 year', but the policyholder was in fact "
            "staying at the property around that renewal date (returned to the UK "
            "November 2020, remained until 5 February 2021) — contrary to the "
            "Unoccupied Property policy's definition of 'occupied' as lived in for more "
            "than 7 consecutive days"
        ),
        "Property Type": "Residential property intended to be unoccupied while owner worked abroad",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Lloyd's avoided the policy from inception, concluding Mr M hadn't disclosed "
            "at the 2019 inception that he intended to occupy the property every 3-4 "
            "months for more than 7 days at a time, and separately that the 2020 renewal "
            "Statement of Fact incorrectly recorded the property as unoccupied for 'more "
            "than 1 year' when he was in fact living there around the renewal date; "
            "Lloyd's said it would only have offered a Holiday Home Policy (which "
            "excludes escape of water) had accurate information been given"
        ),
        "Evidence Dispute": (
            "Mr M: told his broker at inception he'd be working abroad for a year; told "
            "the broker at the 2020 renewal he was temporarily occupying the house but "
            "was renewing on the basis he was about to leave the UK again; renewed in "
            "good faith and the leak occurred while the property was genuinely "
            "unoccupied. Lloyd's: no recording exists of either the 2019 inception or "
            "2020 renewal calls; the written Statement of Fact for each year is the best "
            "contemporaneous record and recorded (2019) occupation expected in 9-12 "
            "months and (2020) unoccupied for 'more than 1 year' — both materially "
            "inaccurate given the policy's 7-consecutive-day 'occupied' definition and "
            "Mr M's own account of returning every 3-4 months and being present around "
            "the 2020 renewal date; an internal business-file note appearing to support "
            "Mr M was subsequently clarified by Lloyd's as a typographical error. FOS "
            "(provisional and final): without a recording, the Statement of Fact — which "
            "Mr M confirmed reading, understanding and agreeing to a disclosure "
            "declaration on — is the more reliable record; on that record Mr M did not "
            "disclose an intention to occupy for more than 7 days at a time at either "
            "inception or renewal, and the 2020 Statement was objectively wrong since his "
            "own account showed he was living in the property around the renewal date; "
            "this was a careless misrepresentation entitling Lloyd's to avoid the policy "
            "under CIDRA; even had the policy been valid, a screen report found the stop "
            "cock may not have been fully closed (Water Turn Off Clause breach) and "
            "there was no record of the required 30-day inspections, either of which "
            "would likely also have defeated the claim"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — Lloyd's fairly avoided the policy from inception for "
            "careless misrepresentation about intended and actual occupancy, and "
            "correctly declined the claim; premiums for both policy years were refunded, "
            "which was fair"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "An Unoccupied Property policy defining 'occupied' as lived in for more than "
            "7 consecutive days is designed specifically to exclude anyone who intends, "
            "or turns out, to stay at the property for extended periods — a "
            "policyholder's own evidence that they intended to return every 3-4 months "
            "for stays exceeding 7 days, or that they were actually staying at the "
            "property around a renewal date, is itself sufficient to show the Unoccupied "
            "Property policy was unsuitable and that a misrepresentation as to occupancy "
            "occurred; where no recording exists of a sales or renewal call, a written "
            "Statement of Fact that the policyholder has confirmed reading, understanding "
            "and agreeing to (as part of a CIDRA-compliant disclosure declaration) is the "
            "more reliable contemporaneous record of what was disclosed, and will "
            "generally be preferred over a policyholder's later recollection where the "
            "two conflict; for a misrepresentation to justify avoidance under CIDRA, the "
            "insurer must show it is a qualifying misrepresentation — i.e., that accurate "
            "disclosure would have changed whether or on what terms cover was offered — "
            "and an insurer's evidence that it would only have offered a different "
            "product (here, a Holiday Home Policy excluding escape of water) satisfies "
            "this"
        ),
        "Missing Evidence": (
            "Recordings of the 2019 inception call and the 2020 renewal call — neither "
            "was available, requiring reliance on the Statements of Fact as the best "
            "contemporaneous written record"
        ),
        "Ombudsman Reasoning": (
            "2019 Statement of Fact recorded the property not occupied from inception and "
            "expected to become occupied in 9-12 months; this is inconsistent with Mr M's "
            "claimed intention (later stated) to return every 3-4 months for over 7 days "
            "at a time; 2020 renewal Statement of Fact recorded unoccupied for 'more than "
            "1 year', but Mr M's own account confirmed he was staying at the property "
            "from November 2020 until 5 February 2021, spanning the December 2020 "
            "renewal date — objectively incorrect; Lloyd's underwriters confirmed they "
            "would not have offered the Unoccupied Property product (or would have "
            "offered a Holiday Home Policy instead) had accurate information been given, "
            "satisfying the qualifying misrepresentation test under CIDRA; fair for "
            "Lloyd's to avoid the policy at both inception and renewal and decline the "
            "claim, while returning all premiums paid; additionally, the stop cock may "
            "not have been fully closed and 30-day inspection records were unavailable, "
            "either of which would likely have independently defeated the claim even "
            "absent misrepresentation"
        ),
        "Workflow Insight": (
            "When assessing a suspected misrepresentation about intended occupancy for an "
            "Unoccupied Property policy, compare the policyholder's own account of their "
            "actual movements (dates present at the property) against the policy's "
            "numeric 'occupied' definition (e.g. more than 7 consecutive days) — a "
            "policyholder's own narrative can itself supply the evidence of "
            "misrepresentation without need for a call recording; renewal Statements of "
            "Fact should be checked against events the policyholder has independently "
            "confirmed (e.g. return travel dates) since a renewal declaration can become "
            "inaccurate due to intervening circumstances (such as pandemic travel "
            "restrictions) even if accurate at a prior renewal; where alternative "
            "independent grounds for declining a claim exist (e.g. a specific safety "
            "condition breach), documenting these strengthens the fairness of an "
            "avoidance decision even where the primary ground is misrepresentation"
        ),
        "AI Rule Candidate": (
            "IF unoccupied_property_policy_defines_occupied_as_lived_in_for_more_than_N_consecutive_days "
            "AND policyholders_own_account_shows_they_were_present_at_the_property_for_periods_exceeding_N_days_around_inception_or_renewal "
            "THEN this_is_evidence_of_a_qualifying_misrepresentation_about_intended_or_actual_occupancy_justifying_avoidance; "
            "IF no_call_recording_exists_for_inception_or_renewal "
            "THEN a_statement_of_fact_the_policyholder_confirmed_reading_and_agreeing_to_is_the_preferred_contemporaneous_record_over_later_recollection; "
            "insurer_avoiding_policy_for_misrepresentation_must_show_it_would_have_offered_different_terms_or_a_different_product_if_accurately_informed_to_satisfy_the_qualifying_misrepresentation_test"
        ),
        "Source PDF": "DRN-3516774.pdf",
    },
    {
        "Case ID": "UNOC-024",
        "FOS Decision ID": "DRN3646077",
        "Insurer Name": "Adrian Flux Insurance Services Group",
        "FOS Decision Date": "3 Nov 2019",
        "Claim Type": (
            "Broker mis-sale complaint — buildings insurance for an inherited unoccupied "
            "property; escape of water from a burst loft pipe in March 2018 declined "
            "because the property wasn't kept heated above 15°C (or mains drained) "
            "during the winter exclusion period; executors/solicitor-representative said "
            "they were unaware of this requirement"
        ),
        "Unoccupied Period / Circumstance": (
            "Property inherited January 2017, stood empty while the estate was "
            "administered; unoccupied buildings cover arranged through the broker, "
            "renewed January 2018; water supply remained connected but heating was "
            "turned off during the relevant winter period"
        ),
        "Property Type": "Residential property, unoccupied pending estate administration",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a coverage decline — dispute concerned whether the broker "
            "adequately notified the policyholders' representative (AL) of the winter "
            "heating/mains-drain condition, not whether the insurer correctly applied it"
        ),
        "Evidence Dispute": (
            "AL (executors' solicitor): unaware of the 15°C heating/mains-drain "
            "requirement at inception or renewal; would have kept the heating on had "
            "they known. Flux: the requirement was clearly set out as an 'Additional "
            "Notes' item in the January 2017 quotation ('Important Information' section) "
            "sent the day before AL phoned to accept the policy, and repeated in the "
            "same location in the January 2018 renewal documentation; AL confirmed on "
            "the renewal call that he was happy with the terms and conditions. FOS: the "
            "clause itself is not unusual for an unoccupied property policy (it "
            "mitigates a well-known frost-damage risk) but the requirement to maintain "
            "constant heating for five winter months is onerous, so it needed to be "
            "clearly notified — it was clearly highlighted, in the same place, at both "
            "the original quotation stage and the renewal stage, and AL confirmed "
            "acceptance of terms and conditions on the renewal call; Flux was entitled to "
            "assume the policy was purchased after AL read and agreed to that "
            "information"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — Flux did enough to clearly highlight the onerous "
            "winter heating/mains-drain condition at both inception and renewal; it was "
            "not responsible for the reduced property sale price"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "A winter escape-of-water exclusion for unoccupied properties, requiring "
            "either mains water to be turned off and drained or the property to be kept "
            "heated at a minimum temperature (e.g. 15°C) throughout a defined winter "
            "period, is not itself an unusual clause for this type of policy since it "
            "mitigates a well-recognised frost-damage risk — but the heating-maintenance "
            "option is an onerous condition (given the potential cost of five months' "
            "continuous heating) and so must be clearly notified to the policyholder or "
            "their representative; specifically highlighting the requirement in an "
            "'Important Information' or equivalent section of both the original "
            "quotation and each renewal document, in the same consistent location, is "
            "sufficient to discharge the broker's duty to notify — even where the same "
            "wording is not repeated in every subsequent communication (e.g. the policy "
            "schedule itself)"
        ),
        "Missing Evidence": (
            "Evidence that AL specifically read the 'Additional Notes' section of the "
            "quotation and renewal documents — not directly available, but AL's "
            "acceptance of the policy the day after receiving the quotation, and his "
            "confirmation of being happy with 'terms and conditions' on the renewal "
            "call, supported an inference that the notification had been received and "
            "accepted"
        ),
        "Ombudsman Reasoning": (
            "Clause required mains off/drained or heating maintained at 15°C from 1 "
            "November to 31 March; turning mains off is not onerous but constant heating "
            "for five months is; clause was set out clearly under 'Additional "
            "Notes'/'Important Information' at both the January 2017 quotation (a day "
            "before AL accepted the policy by phone) and the January 2018 renewal "
            "documentation (in the 'Excesses, endorsements and conditions' section); AL "
            "confirmed on the renewal call he was happy with the terms and conditions; "
            "taken together this was sufficient to discharge Flux's duty to notify AL of "
            "the condition at both inception and renewal; Flux therefore not responsible "
            "for the declined claim or the reduced property sale price"
        ),
        "Workflow Insight": (
            "For onerous winter heating/mains-drain conditions on unoccupied property "
            "policies, brokers should highlight the condition in a consistently-located "
            "'Important Information' or 'Additional Notes' section at both initial "
            "quotation and every renewal — doing so in the same place each time, "
            "combined with a policyholder's acceptance of 'terms and conditions' on a "
            "renewal call, is generally sufficient evidence of adequate notification "
            "even without the condition being separately repeated in every other "
            "communication (e.g. the policy schedule or a covering email)"
        ),
        "AI Rule Candidate": (
            "IF onerous_winter_heating_or_mains_drain_condition_is_clearly_set_out_in_a_consistently_located_important_information_section_at_both_quotation_and_each_renewal "
            "AND policyholder_or_representative_confirms_acceptance_of_terms_and_conditions_on_the_renewal_call "
            "THEN broker_has_discharged_its_duty_to_notify_the_condition_even_if_not_repeated_in_every_other_document; "
            "unoccupied_property_winter_escape_of_water_conditions_requiring_mains_off_or_maintained_heating_are_not_unusual_but_the_heating_maintenance_option_is_onerous_and_must_be_clearly_highlighted"
        ),
        "Source PDF": "DRN3646077.pdf",
    },
    {
        "Case ID": "UNOC-025",
        "FOS Decision ID": "DRN-3848349",
        "Insurer Name": "T&R Direct Ltd",
        "FOS Decision Date": "10 Mar 2023",
        "Claim Type": (
            "Broker mis-sale complaint — unoccupied property owner's policy arranged "
            "(non-advised, per the broker) for a property undergoing renovation; theft "
            "of new kitchen appliances (still boxed, delivered but not yet installed) "
            "declined because the policyholder had selected 'standard' cover "
            "(fire/lightning/explosion/aircraft only) on the proposal form and had not "
            "opted for the available theft extension"
        ),
        "Unoccupied Period / Circumstance": (
            "Property unoccupied and undergoing renovation from around June 2021; theft "
            "of appliances occurred December 2021 while the property remained empty "
            "pending completion of works"
        ),
        "Property Type": "Residential property under renovation (unoccupied)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a coverage decline by the insurer — the underlying theft "
            "coverage decision was the subject of a separate complaint; this complaint "
            "concerned only whether the broker's conduct in arranging the policy "
            "amounted to a mis-sale"
        ),
        "Evidence Dispute": (
            "Mr F: expected the new policy to match his previous (occupied) contents "
            "cover; felt T&R had sent him a 'misleading and poorly structured' proposal "
            "form and effectively directed him to take the policy without adequately "
            "explaining that 'standard' cover excluded theft, a major exclusion; his own "
            "insurance adviser considered a policy clause showed theft was included and "
            "that T&R should have highlighted theft as a major exclusion. T&R: the "
            "proposal form clearly defined 'standard' cover (fire/lightning/explosion/"
            "aircraft only) and offered a separate theft extension Mr F did not select; "
            "the policy was arranged on a non-advised basis — Mr F completed and signed "
            "the form himself, and a recorded call showed he had actively turned his mind "
            "to security, stating there was 'nothing of substantial value' left at the "
            "property besides office equipment, furniture and books, locked in a "
            "third-floor room. FOS: the policy itself was not inherently unsuitable — it "
            "did provide unoccupied-property cover and a theft extension was available; "
            "Mr F made clear, hand-written elections on the form for a new and different "
            "type of policy; the word 'only' in the standard-cover definition "
            "emphasised its limits; Mr F was capable and engaged (adding handwritten "
            "comments, asking specific questions about excess and financial cover) but "
            "did not ask about theft; the sale was non-advised since no cover-level "
            "advice was given and the declaration/documentation did not indicate an "
            "advised sale; it would have been better practice for T&R to flag at the "
            "point of claim that theft cover hadn't been selected, but this did not "
            "amount to mis-sale or attract compensation"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — T&R didn't mis-sell the policy; it was arranged on a "
            "non-advised basis, the proposal form clearly defined the standard cover "
            "limits and offered a theft extension Mr F chose not to select, and Mr F's "
            "own engagement with the form and call showed an informed choice"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "Where an unoccupied property owner's policy proposal form clearly defines "
            "'standard' cover as limited to specific perils (e.g. 'fire, lightning, "
            "explosion, aircraft only') and separately offers an optional theft "
            "extension that the policyholder does not select, a broker arranging the "
            "policy on a non-advised basis is not required to specifically flag that "
            "theft is excluded from standard cover — the onus falls on the consumer to "
            "review the cover options and select accordingly, particularly where the "
            "policyholder's own recorded comments show they turned their mind to the "
            "security of contents at the property; a sale is non-advised where the "
            "broker does not recommend specific cover-level choices and the policyholder "
            "independently completes and signs the proposal form; it would be good "
            "practice, though not a mis-sale, for a broker to flag at the point of a "
            "claim that a relevant extension wasn't selected, to manage the "
            "policyholder's expectations before the insurer's claims decision"
        ),
        "Missing Evidence": (
            "Clear evidence that Mr F specifically requested theft cover matching his "
            "previous occupied-property policy — absent; his own recorded comments about "
            "there being 'nothing of substantial value' at the property were, if "
            "anything, evidence he had actively (if perhaps mistakenly) turned his mind "
            "away from needing theft cover"
        ),
        "Ombudsman Reasoning": (
            "Proposal form clearly defined standard cover (fire/lightning/explosion/"
            "aircraft 'only') separately from an extended-cover option that included "
            "theft, which Mr F did not tick; recorded call showed Mr F actively "
            "considered and downplayed the value of contents left at the property; Mr F "
            "engaged meaningfully with the form (handwritten comments, specific "
            "questions on excess and financial cover level) but not on theft; sale was "
            "non-advised — no cover-level recommendation from T&R, form completed and "
            "signed by Mr F, declaration onus on him to check completeness/accuracy or "
            "consult an adviser if in doubt; T&R could have done more at the claims "
            "stage to flag the missing theft cover but this is a best-practice "
            "observation, not a basis for compensation or a mis-sale finding"
        ),
        "Workflow Insight": (
            "For non-advised sales of unoccupied property owner's policies, brokers "
            "should still consider flagging, at the point a claim is first reported, "
            "whether the relevant peril extension was selected on the original "
            "proposal — this manages the policyholder's expectations and avoids "
            "unnecessary distress even where it is not required to avoid a mis-sale "
            "finding; a policyholder's own recorded statements about the value or "
            "nature of contents at a property can be persuasive evidence that they made "
            "an informed (if potentially mistaken) choice not to select an available "
            "cover extension"
        ),
        "AI Rule Candidate": (
            "IF proposal_form_clearly_defines_standard_cover_perils_and_separately_offers_an_optional_extension_covering_the_peril_later_claimed_for "
            "AND policyholder_completed_and_signed_the_form_without_broker_recommendation_on_cover_level "
            "THEN sale_is_non_advised_and_broker_is_not_liable_for_policyholder_not_selecting_the_extension; "
            "policyholders_own_recorded_statements_about_the_value_of_contents_at_the_property_can_evidence_an_informed_choice_not_to_select_an_available_extension; "
            "broker_failing_to_flag_at_claim_stage_that_a_relevant_extension_was_not_selected_is_a_best_practice_gap_not_a_mis_sale"
        ),
        "Source PDF": "DRN-3848349.pdf",
    },
    {
        "Case ID": "UNOC-026",
        "FOS Decision ID": "DRN-3999409",
        "Insurer Name": "Affinity Insurance Solutions Limited",
        "FOS Decision Date": "1 May 2023",
        "Claim Type": (
            "Broker mis-sale complaint — home insurance arranged by phone for an "
            "inherited unoccupied property; escape of water declined by the underwriter "
            "because the property had been unoccupied for more than 60 consecutive "
            "days, which the policyholder said she wasn't warned about at the point of "
            "sale"
        ),
        "Unoccupied Period / Circumstance": (
            "Property inherited following a bereavement in April 2021; insured as "
            "unoccupied from inception (June 2021), renewed July 2022; escape of water "
            "discovered mid-December 2022 while the property remained unoccupied, "
            "though visited regularly by the family for upkeep"
        ),
        "Property Type": "Residential property (inherited, unoccupied)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a coverage decline by the underwriter directly — this "
            "complaint concerned whether the broker adequately explained the 60-day "
            "unoccupied-property escape-of-water exclusion at the point of sale, not the "
            "underwriter's application of it"
        ),
        "Evidence Dispute": (
            "Mrs K (via her husband, who arranged the policy on her behalf): wasn't made "
            "aware an escape of water wouldn't be covered if the property was unoccupied "
            "for more than 60 consecutive days; both she and her husband were vulnerable "
            "at the time, dealing with bereavement, grief and ill health. Affinity: its "
            "call handler told Mr K on at least two occasions during the sales call that "
            "a restriction applied to cover where a property was unoccupied for more "
            "than 60 consecutive days, and Mr K confirmed 'that's fine' and proceeded; "
            "the subsequent 46-page policy booklet, sent by email, also clearly set out "
            "the exclusion and advised the policyholder to check the cover was right for "
            "her, consistent with a non-advised sale. FOS (having listened to the "
            "recorded call): the restriction was clearly stated at least three times "
            "during the call and Mr K confirmed acceptance each time; the policy "
            "documentation was significantly more detailed than what was said on the "
            "call (a 46-page booklet, not unusual) and clearly excluded EOW after 60 "
            "days' unoccupancy — a common feature of unoccupied-property policies "
            "generally; Mrs K had a cooling-off period to review and cancel if the cover "
            "wasn't suitable; while sympathetic to the bereavement and ill health, there "
            "wasn't enough evidence that the couple's circumstances precluded "
            "understanding the clearly-stated restriction"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — Affinity clearly explained the 60-day "
            "unoccupied-property restriction on the recorded sales call (at least three "
            "times) and in the subsequent policy documentation; the policy was not "
            "mis-sold"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "Where a call recording shows an unoccupied-property time-limited exclusion "
            "(e.g. 60 consecutive days for escape of water) was clearly stated multiple "
            "times during a sales call and acknowledged by the person arranging the "
            "policy, and the subsequent policy documentation also clearly sets out the "
            "same restriction, a broker has discharged its duty to provide clear, fair "
            "and not misleading information — this remains the case even where the "
            "policy was arranged by a family member on behalf of a vulnerable consumer "
            "affected by bereavement and ill health, provided there's no evidence the "
            "vulnerability precluded understanding what was communicated at the time; a "
            "cooling-off period during which the policyholder could review documentation "
            "and cancel without penalty is a further opportunity for the consumer to "
            "identify unsuitable terms, reducing the broker's residual responsibility"
        ),
        "Missing Evidence": (
            "Evidence that Mr and Mrs K's vulnerability (grief, ill health) at the time "
            "of the June 2021 sales call was so severe as to preclude understanding the "
            "restriction that was repeated multiple times and expressly acknowledged — "
            "absent"
        ),
        "Ombudsman Reasoning": (
            "Recorded call (good quality, provided to Mrs K in the interests of candour) "
            "showed the 60-day unoccupied restriction stated at least three times, with "
            "Mr K responding 'that's fine' and confirming he was happy to proceed; "
            "46-page policy booklet (not unusually long) and covering letter advising "
            "the policyholder to check cover was 'right for you' were sent after "
            "inception; this pattern is consistent with a non-advised sale where the "
            "onus fell on Mrs K to review the documentation; while sympathetic to the "
            "family's grief and Mr K's evident concern for his wife's wellbeing, there "
            "was no evidence their circumstances prevented them from understanding a "
            "restriction repeated multiple times on the call; not mis-sold"
        ),
        "Workflow Insight": (
            "When a family member arranges an unoccupied-property policy on behalf of a "
            "bereaved or vulnerable consumer, a broker's obligations are assessed "
            "against what was communicated to the person actually on the call and in "
            "subsequent documentation — repeating a material time-limited exclusion "
            "multiple times during the call and echoing it in the policy booklet, "
            "combined with a cooling-off period, is strong evidence of a fair sale even "
            "in circumstances of bereavement, absent specific evidence that the "
            "vulnerability itself impaired understanding of what was said at the time"
        ),
        "AI Rule Candidate": (
            "IF recorded_sales_call_shows_time_limited_unoccupied_property_exclusion_stated_multiple_times_and_acknowledged_by_the_person_arranging_the_policy "
            "AND subsequent_policy_documentation_also_clearly_sets_out_the_same_exclusion "
            "THEN broker_has_provided_clear_fair_and_not_misleading_information_regardless_of_the_policyholders_bereavement_or_vulnerability_absent_specific_evidence_the_vulnerability_impaired_understanding_at_the_time; "
            "cooling_off_period_availability_further_reduces_broker_residual_responsibility_for_a_clearly_disclosed_unoccupancy_restriction"
        ),
        "Source PDF": "DRN-3999409.pdf",
    },
    {
        "Case ID": "UNOC-027",
        "FOS Decision ID": "DRN4133004",
        "Insurer Name": "St Andrew's Insurance Plc",
        "FOS Decision Date": "16 Oct 2019",
        "Claim Type": (
            "Buildings insurance — escape of water discovered by policyholder's son "
            "(holding power of attorney) during a maintenance visit; property had been "
            "empty for around 18 months (owner in a care home) while being maintained/"
            "decorated by family with a view to renting or selling; insurer declined "
            "citing the 30-day unoccupied exclusion and separately cancelled/backdated "
            "the policy to its one-year maximum unoccupied-cover period"
        ),
        "Unoccupied Period / Circumstance": (
            "Owner entered a care home in August 2016; property unoccupied thereafter "
            "for around 18 months by the time of the February 2018 loss, though visited "
            "every other day by family for maintenance and decoration; property retained "
            "essential living items (cooking appliances, sanitary fittings, beds)"
        ),
        "Property Type": "Residential property (owner in care home; family maintaining for future rental or sale)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Escape of water excluded as the property had been unoccupied (not lived in) "
            "for more than 30 days; separately, St Andrew's said it would only have "
            "insured an unoccupied property for a maximum of one year and so backdated "
            "cancellation of the policy to August 2017, refunding premiums paid since "
            "then"
        ),
        "Evidence Dispute": (
            "Mr G (via his son/attorney): family visited every other day to maintain and "
            "decorate the property with a view to letting or selling it, so it shouldn't "
            "be classed as unoccupied; the broker (not St Andrew's) had been told about "
            "the change in 2016 and hadn't passed this on or warned of the unoccupancy "
            "consequences — a separate complaint. St Andrew's: to meet the policy's "
            "'occupied' definition the property must be lived in by the owner or a "
            "family member and contain essential living items; regular maintenance "
            "visits, however frequent, are not the same as living there; had it known "
            "the property was unoccupied, it would only have covered EOW-type damage if "
            "additional risk-reduction steps had been taken (water turned off, regular "
            "inspections recorded), and would in any event have limited cover to one "
            "year from the start of unoccupancy, consistent with its underwriting "
            "notes. FOS: property had the essential items for normal living (satisfying "
            "that limb of the definition) but visiting to carry out maintenance, however "
            "frequently, does not amount to 'living' there; the water hadn't been turned "
            "off, so the breach of the unoccupied-property risk-reduction condition was "
            "directly connected to the type of loss claimed (an EOW-prevention measure "
            "not taken); backdating cancellation to August 2017 (one year from actual "
            "unoccupancy) and refunding premiums paid thereafter was a fair reflection "
            "of what St Andrew's would have done had it known the true position at the "
            "time"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — St Andrew's fairly declined the EOW claim under the "
            "30-day unoccupied exclusion and fairly cancelled/backdated the policy to "
            "its one-year unoccupied-cover maximum, refunding premiums paid after that "
            "date"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "An 'occupied'/'lived in' definition requiring both (a) actual habitation by "
            "the owner or an authorised family member and (b) the presence of essential "
            "items for normal living (cooking appliances, sanitary fittings, beds) is a "
            "conjunctive test — satisfying the furnishing limb alone does not meet the "
            "definition if no one is actually living there; family members visiting "
            "every other day to carry out maintenance or decoration, however frequently, "
            "does not constitute 'living' in a property for this purpose; where an "
            "insurer's maximum unoccupied-property cover period (e.g. one year) is a "
            "documented underwriting practice, it is fair for the insurer to backdate a "
            "cancellation to the actual date unoccupancy began (as established by the "
            "evidence) once it discovers the true position, and to refund premiums paid "
            "for the period it would not otherwise have covered; a breach of a "
            "risk-reduction condition specific to unoccupied properties (e.g. failing to "
            "turn off the water supply) is materially connected to, and can bar, a claim "
            "for the very type of damage (escape of water) that the condition was "
            "designed to prevent"
        ),
        "Missing Evidence": (
            "Evidence that the water supply had been turned off, or an alternative "
            "continuous-heating arrangement maintained, during the unoccupied period — "
            "absent; this breach was directly connected to the escape of water loss"
        ),
        "Ombudsman Reasoning": (
            "Property contained essential living items but was only visited, not lived "
            "in, by family members carrying out maintenance and decoration over roughly "
            "18 months; the 'occupied' definition requires both limbs to be met and "
            "'living in' was not established by visits alone; had St Andrew's known of "
            "the unoccupancy it would only have covered EOW-type losses subject to "
            "additional conditions (water off, regular recorded inspections) which were "
            "not met, and the failure to turn off the water was directly connected to "
            "the loss claimed; St Andrew's underwriting notes confirmed a one-year "
            "maximum unoccupied-cover policy, so backdating cancellation to August 2017 "
            "(the true start of unoccupancy) and refunding premiums paid since was the "
            "fair outcome reflecting what would have happened had it been told at the "
            "time; broker's alleged failure to pass on notification was a separate "
            "matter not attributable to St Andrew's"
        ),
        "Workflow Insight": (
            "When assessing an 'occupied'/'lived in' definition requiring both "
            "habitation and essential furnishing, confirm both limbs separately — a "
            "well-furnished, well-maintained property is not thereby 'occupied' if no "
            "one is actually residing there; regular family visits for maintenance or "
            "decoration purposes should be distinguished from residence when assessing "
            "unoccupancy thresholds; where an insurer discovers post-loss that a "
            "property was unoccupied for longer than its maximum unoccupied-cover "
            "period, backdating cancellation to the actual (not notified) start date of "
            "unoccupancy and refunding the resulting unearned premium is a fair remedy, "
            "provided supported by documented underwriting practice"
        ),
        "AI Rule Candidate": (
            "IF policy_occupied_definition_requires_both_actual_habitation_AND_presence_of_essential_living_items "
            "AND only_the_furnishing_limb_is_satisfied "
            "THEN property_is_not_occupied_for_policy_purposes_regardless_of_frequency_of_maintenance_visits; "
            "IF insurer_has_a_documented_maximum_unoccupied_cover_period "
            "AND discovers_post_loss_that_property_exceeded_that_period "
            "THEN backdating_cancellation_to_the_actual_start_of_unoccupancy_and_refunding_premiums_paid_thereafter_is_a_fair_remedy; "
            "breach_of_an_unoccupied_property_risk_reduction_condition_directly_connected_to_the_type_of_loss_claimed_bars_the_claim"
        ),
        "Source PDF": "DRN4133004.pdf",
    },
    {
        "Case ID": "UNOC-028",
        "FOS Decision ID": "DRN-4195324",
        "Insurer Name": "Accredited Insurance (Europe) Ltd",
        "FOS Decision Date": "22 Aug 2023",
        "Claim Type": (
            "Home insurance — water leak claim; insurer investigated and concluded the "
            "policyholder didn't actually live at the insured property (contrary to "
            "what was declared as her main residence), avoided the policy for "
            "deliberate/reckless misrepresentation, and separately suggested the claim "
            "would also be excluded under an unoccupancy condition"
        ),
        "Unoccupied Period / Circumstance": (
            "Insurer's investigator concluded the property wasn't the policyholder's "
            "main residence and may have been used for business purposes or otherwise "
            "unoccupied for a period before the water damage was discovered (evidenced, "
            "it said, by mould growth and discrepancies over when the loss occurred) — "
            "but the ombudsman found the insurer's occupancy investigation itself was "
            "flawed and unproven"
        ),
        "Property Type": "Residential property (occupancy status disputed)",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": (
            "Accredited avoided the policy (treated it as if it never existed) on the "
            "basis that Miss V had misrepresented that the property was her main "
            "residence, which it said was a deliberate/reckless qualifying "
            "misrepresentation entitling it to refuse all claims; it separately argued "
            "that even if the policy were valid, the claim would be excluded under the "
            "unoccupancy condition given delay in the damage being discovered"
        ),
        "Evidence Dispute": (
            "Miss V: disputed the investigator's conclusions about where she lived and "
            "the fairness/conduct of the investigation (including an allegation of "
            "bullying during an interview). Accredited: its investigator compared her "
            "circumstances to another property it believed she lived at, and relied on "
            "a neighbour's account and inconsistencies it identified; said it wouldn't "
            "have offered a household policy had it known the property wasn't her main "
            "residence, and that even if cover existed, unoccupancy (evidenced by mould "
            "growth and timeline discrepancies) would exclude the claim. FOS "
            "(provisional and final): repeatedly asked Accredited to justify its "
            "investigation methodology and to produce its underwriting criteria to show "
            "the alleged misrepresentation was a 'qualifying' one under CIDRA (i.e., "
            "that it would have changed whether or on what terms cover was offered); "
            "Accredited initially said it couldn't access its underwriting information, "
            "then supplied a document that turned out not to be the underwriting "
            "criteria, and only after repeated requests produced its actual "
            "underwriting guide and endorsements; that document's list of decline "
            "reasons did not include unoccupancy, its general guidelines said it would "
            "cover most non-standard risks including unoccupied properties, and none of "
            "the unoccupied-property endorsements referenced said cover would be "
            "declined outright (some referred to reduced cover and protective measures "
            "only); Accredited therefore failed to show the alleged misrepresentation "
            "was a qualifying one, and also failed to justify the reasonableness of its "
            "occupancy investigation itself"
        ),
        "Outcome Category": "Upheld",
        "Outcome": (
            "Accredited required to remove references to the policy avoidance from "
            "internal and external databases, reinstate the policy, consider the claim, "
            "and pay Miss V £200 compensation"
        ),
        "Compensation Awarded (£)": 200,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "To avoid a policy for misrepresentation under CIDRA, an insurer must show "
            "not only that the policyholder failed to take reasonable care in answering "
            "a question (e.g. about main residence status) but also that the "
            "misrepresentation was a 'qualifying' one — meaning the insurer would have "
            "offered the policy on different terms, or not at all, had accurate "
            "information been given; an insurer's unoccupancy-related underwriting "
            "guidance that identifies reduced cover and protective measures (rather "
            "than outright decline) for unoccupied or non-standard-occupancy properties "
            "undermines a claim that a misrepresentation about occupancy/residence "
            "status was 'qualifying' in the CIDRA sense; an insurer relying on avoidance "
            "for misrepresentation bears the burden of promptly producing its actual "
            "underwriting criteria to substantiate that claim — repeated failure or "
            "delay in producing this evidence, or producing the wrong document, "
            "undermines the insurer's position and supports a finding against it"
        ),
        "Missing Evidence": (
            "Accredited's genuine underwriting criteria — initially claimed "
            "unavailable, then a non-underwriting document was supplied, and only after "
            "repeated requests was the correct document produced; even then, it did not "
            "support the claim that the alleged misrepresentation was a qualifying one"
        ),
        "Ombudsman Reasoning": (
            "Multiple issues identified with how Accredited investigated occupancy "
            "(including the investigator personally following up a bullying allegation "
            "raised against himself), and Accredited did not adequately justify the "
            "conclusions reached; even setting aside the reasonable-care question, "
            "Accredited failed — despite repeated requests and after supplying an "
            "incorrect document — to show the misrepresentation was a qualifying one; "
            "the underwriting guide's list of decline reasons did not include "
            "unoccupancy, its general position was to cover most non-standard risks "
            "including unoccupied properties, and none of the referenced "
            "unoccupied-property endorsements provided for outright decline; therefore "
            "avoidance and claim decline were not reasonable; a related, "
            "not-yet-actioned suggestion that endorsements might have been applied for "
            "unoccupancy was irrelevant since that action was never actually taken"
        ),
        "Workflow Insight": (
            "When an insurer relies on misrepresentation to avoid a policy connected to "
            "disputed occupancy or residence status, it must be prepared to promptly "
            "produce its actual underwriting criteria to demonstrate the "
            "misrepresentation was 'qualifying' — repeated failure to do so, or "
            "supplying incorrect documents, will be treated as undermining the "
            "insurer's position rather than neutral; underwriting guidance that "
            "provides for reduced cover or protective conditions (rather than outright "
            "decline) for unoccupied or non-standard-occupancy properties is strong "
            "evidence against a claim that misrepresenting occupancy status was a "
            "qualifying misrepresentation; claims investigations into disputed "
            "occupancy should not be conducted or self-validated by the same "
            "individual against whom a conduct complaint (e.g. bullying) has been "
            "raised"
        ),
        "AI Rule Candidate": (
            "IF insurer_avoids_policy_for_misrepresentation_about_occupancy_or_residence_status "
            "AND insurer_cannot_promptly_produce_underwriting_criteria_showing_the_misrepresentation_would_have_changed_terms_or_offer "
            "THEN avoidance_is_not_a_qualifying_misrepresentation_under_cidra_and_should_not_be_upheld; "
            "IF underwriting_guidance_for_unoccupied_or_non_standard_occupancy_properties_provides_for_reduced_cover_or_protective_conditions_rather_than_outright_decline "
            "THEN this_undermines_a_claim_that_misrepresenting_occupancy_status_was_a_qualifying_misrepresentation; "
            "insurer_repeatedly_failing_to_produce_or_misidentifying_its_own_underwriting_document_when_requested_weighs_against_the_insurers_position"
        ),
        "Source PDF": "DRN-4195324.pdf",
    },
    {
        "Case ID": "UNOC-029",
        "FOS Decision ID": "DRN4242877",
        "Insurer Name": "Stride Limited",
        "FOS Decision Date": "25 Oct 2018",
        "Claim Type": (
            "Broker mis-sale complaint — home insurance for an empty property; "
            "vandalism damage claim (paint splashed from an incident targeting a "
            "neighbour's car) attracted a £1,000 increased excess for unoccupied-"
            "property risks (theft, vandalism, malicious damage, accidental damage) "
            "under the policy; policyholders felt the policy had been mis-sold as "
            "unsuitable for an unoccupied property, though the insurer reduced the "
            "excess to £100 given the collateral nature of the damage and the small "
            "claim (£300) was not pursued"
        ),
        "Unoccupied Period / Circumstance": (
            "Property already unoccupied for some time when the policy was taken out in "
            "November 2017, as recorded on the statement of fact"
        ),
        "Property Type": "Residential property, unoccupied",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a coverage decline — the underlying vandalism claim was "
            "not declined (a reduced £100 excess was offered and the claim, worth £300, "
            "was simply not pursued by the policyholders); the complaint concerned "
            "whether the policy itself was mis-sold as unsuitable for an unoccupied "
            "property"
        ),
        "Evidence Dispute": (
            "Mr and Mrs H: believed the policy became entirely invalid after 45 days' "
            "unoccupancy, and hadn't received the policy wording until after the claim, "
            "making it an unsuitable/mis-sold policy. Stride: statement of facts and "
            "policy schedule (with a link to the full policy wording) were emailed at "
            "inception and referenced by Mr and Mrs H in their own correspondence, so "
            "were likely received; the schedule itself stated cover for certain losses "
            "(theft, vandalism, malicious damage, accidental damage) while unoccupied "
            "for more than 45 days continued but with an increased £1,000 excess, not "
            "that cover was withdrawn entirely; this was the most competitive policy "
            "Stride could source for an unoccupied property. FOS: the policy schedule "
            "made clear the property was insured as unoccupied and that certain risks "
            "carried a higher excess (not exclusion) after 45 days — so the policy was "
            "not invalid or unsuitable for an unoccupied property, contrary to Mr and "
            "Mrs H's belief; even accepting they weren't specifically told about the "
            "increased excess in advance, the increased excess and inspection/"
            "drain-down conditions are common unoccupied-property policy terms, and it "
            "was unlikely they'd have obtained materially better terms elsewhere at a "
            "comparable price"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — the policy was a suitable, valid unoccupied-"
            "property policy (not invalidated after 45 days, merely subject to an "
            "increased excess for certain perils); no mis-sale established and no "
            "refund or claim-record amendment required"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": (
            "An unoccupied-property home insurance policy that imposes an increased "
            "excess (rather than a blanket exclusion) for certain perils such as theft, "
            "vandalism, malicious damage and accidental damage once the property has "
            "been unoccupied for more than a stated period (e.g. 45 days) remains a "
            "valid and suitable policy for an unoccupied property — a policyholder's "
            "mistaken belief that the policy became wholly invalid after that period "
            "does not, without more, establish mis-sale; an increased excess and "
            "standard unoccupied-property conditions (regular inspection, mains "
            "drain-down over winter) are common industry terms, and a broker is not "
            "required to show it could have sourced materially better terms at a "
            "comparable price to defeat a suitability complaint"
        ),
        "Missing Evidence": (
            "Clear evidence Mr and Mrs H did not in fact receive the statement of facts "
            "and policy schedule (with link to policy wording) at inception — "
            "undermined by their own correspondence referencing that document, and by "
            "the absence of any record of them chasing Stride for a missing link to the "
            "full wording"
        ),
        "Ombudsman Reasoning": (
            "Statement of facts and schedule sent by email at inception and referenced "
            "by the policyholders themselves; schedule clearly stated the property was "
            "unoccupied and that certain perils carried an increased £1,000 excess (not "
            "exclusion) once unoccupied more than 45 days — this contradicted the "
            "policyholders' belief that the policy became entirely invalid; even "
            "assuming they weren't specifically warned of the increased excess in "
            "advance, the terms were common for this type of policy and it was unlikely "
            "a materially better-value alternative existed; the insurer's willingness to "
            "reduce the excess to £100 given the collateral nature of the damage, and "
            "the policyholders' own decision not to pursue the £300 claim, meant there "
            "was no declined claim to remedy; no basis to require a premium refund or "
            "claim-record amendment since the policy remained valid throughout"
        ),
        "Workflow Insight": (
            "When a policyholder alleges an unoccupied-property policy is invalid or "
            "unsuitable because of a time-triggered condition, check whether that "
            "condition is actually an exclusion (no cover at all) or merely an "
            "increased excess/reduced-scope condition (cover continues, on different "
            "terms) — many unoccupancy complaints stem from this misunderstanding "
            "rather than an actual coverage gap; where a small claim is voluntarily not "
            "pursued because a reduced excess still exceeds the claim value, this "
            "should not be treated as an unresolved declined claim when assessing a "
            "related mis-sale complaint"
        ),
        "AI Rule Candidate": (
            "IF unoccupied_property_policy_schedule_states_certain_perils_carry_an_increased_excess_rather_than_exclusion_after_a_stated_unoccupancy_period "
            "THEN policy_remains_valid_and_suitable_for_an_unoccupied_property_and_is_not_thereby_mis_sold; "
            "policyholder_mistaken_belief_that_a_time_triggered_excess_increase_equals_total_policy_invalidity_does_not_by_itself_establish_broker_mis_sale; "
            "IF policyholder_voluntarily_declines_to_pursue_a_small_claim_because_the_reduced_excess_exceeds_the_claim_value "
            "THEN this_is_not_a_declined_claim_requiring_remedy_in_a_related_mis_sale_complaint"
        ),
        "Source PDF": "DRN4242877.pdf",
    },
    {
        "Case ID": "UNOC-030",
        "FOS Decision ID": "DRN-4660265",
        "Insurer Name": "Amtrust Europe Limited",
        "FOS Decision Date": "3 May 2024",
        "Claim Type": (
            "Holiday homes property owners insurance — escape of water claim; insurer "
            "asked the policyholder (a property business) to provide evidence of "
            "compliance with the policy's unoccupied-property endorsement (including "
            "that the water supply was turned off at the mains, or heating maintained) "
            "before considering the claim further, since it believed the property may "
            "not have been lived in throughout the run-up to the loss, spanning a "
            "switch from an unoccupied property policy to a holiday home owners policy"
        ),
        "Unoccupied Period / Circumstance": (
            "Property let to a company that appeared to intend renting it out via "
            "Airbnb, under a lease/arrangement commencing before the November 2022 "
            "policy switch; insurer's loss adjusters found evidence the property may "
            "not have been lived in at any point between that arrangement starting and "
            "the mid-December 2022 escape of water; policyholder argued the "
            "30-consecutive-day unoccupancy clock should only start running from the "
            "November 2022 policy inception date, not before"
        ),
        "Property Type": "Holiday home / rental property (occupancy status disputed around a policy switch)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": (
            "Not applicable to a final coverage decline — Amtrust had not declined the "
            "claim but was requesting evidence of compliance with the unoccupied-"
            "property endorsement (or evidence the property had been lived in within "
            "the 30 days before the loss) before resuming consideration of the claim"
        ),
        "Evidence Dispute": (
            "G (property business): it was unreasonable to require proof of compliance "
            "with the unoccupied endorsement because the 30 consecutive unoccupied days "
            "could only reasonably be counted from the date the holiday home owners "
            "policy was taken out (30 November 2022), not from any earlier point under "
            "the prior unoccupied property policy. Amtrust: the property may not have "
            "been lived in at any point since the letting/Airbnb arrangement began, well "
            "before the policy switch; the policy wording placed no restriction on when "
            "the 30-day unoccupancy count could start, and requiring renewed occupation "
            "evidence made sense given the risk the endorsement was designed to guard "
            "against continued regardless of which policy was technically in force on a "
            "given date. FOS: nothing in the policy wording supported G's argument that "
            "the inception date reset the 30-day clock — the endorsement simply "
            "requires compliance if the property 'isn't lived in for 30 days or more', "
            "with no stated limitation on when that period can start; it would make no "
            "sense from a risk-appetite perspective for the 30 days to only start at "
            "policy inception, since the underlying risk (an unoccupied property) is "
            "unaffected by a mid-run change of insurer or product; it was therefore "
            "reasonable for Amtrust to ask G for evidence of occupation in the 30 days "
            "before the loss, or compliance with the endorsement (e.g. mains water off, "
            "or heating maintained), before resuming its consideration of the claim"
        ),
        "Outcome Category": "Not Upheld",
        "Outcome": (
            "Complaint not upheld — Amtrust acted reasonably in requiring evidence of "
            "occupation or endorsement compliance before resuming consideration of the "
            "claim; the policy inception date does not reset the 30-day unoccupancy "
            "clock"
        ),
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": (
            "Where a policy's unoccupied-property endorsement is triggered by the "
            "property 'not being lived in for 30 days or more' with no wording limiting "
            "when that period can start, the count is not reset by the inception date "
            "of a new policy or a mid-term switch between policy types (e.g. from an "
            "unoccupied property policy to a holiday home owners policy) — the risk the "
            "endorsement addresses (an unoccupied property being more vulnerable to "
            "loss) exists independently of which specific policy or insurer happens to "
            "be on risk at a given moment, so an insurer is entitled to look at the full "
            "run of consecutive unoccupied days, including periods before the current "
            "policy's inception, when assessing endorsement compliance"
        ),
        "Missing Evidence": (
            "Evidence, at the time of the ombudsman's decision, that the property had "
            "been lived in at some point in the 30 days before the 16 December 2022 "
            "loss, or that G had complied with the endorsement's requirements (e.g. "
            "mains water turned off, or heating maintained at a steady level) — not yet "
            "provided by G; the ombudsman noted Amtrust would be expected to resume "
            "considering the claim if such evidence were subsequently produced"
        ),
        "Ombudsman Reasoning": (
            "Property let to a company that appeared to intend Airbnb-style rental from "
            "before the November 2022 policy switch; Amtrust's loss adjusters found "
            "evidence the property may not have been lived in at any point during that "
            "arrangement; G's argument that the policy inception date restarts the "
            "30-day unoccupancy clock is not supported by anything in the policy "
            "wording, and makes no sense in terms of the insurer's risk appetite, since "
            "the underlying unoccupancy risk is unaffected by which policy is "
            "technically on risk on a given date; therefore reasonable for Amtrust to "
            "ask for evidence of occupation or endorsement compliance (mains water off/"
            "heating maintained) before resuming consideration of the claim; complaint "
            "not upheld, but the claim itself remains open pending that evidence"
        ),
        "Workflow Insight": (
            "When a property changes insurer or policy type (e.g. switching from an "
            "unoccupied property policy to a holiday home owners policy) without a "
            "change in actual physical occupancy, claims handlers assessing a "
            "subsequent unoccupancy-endorsement question should look at the full "
            "continuous run of unoccupied days across the switch, not just days since "
            "the current policy's inception — a policyholder cannot rely on a policy "
            "switch to reset an unoccupancy clock where there was no actual change in "
            "whether the property was being lived in; where a claim is not yet formally "
            "declined but is paused pending evidence of a condition (such as "
            "endorsement compliance), this should be recorded as an open evidentiary "
            "request rather than a coverage decision"
        ),
        "AI Rule Candidate": (
            "IF unoccupied_property_endorsement_is_triggered_by_property_not_being_lived_in_for_N_or_more_consecutive_days "
            "AND policy_wording_places_no_limitation_on_when_that_period_can_start "
            "THEN the_count_includes_days_before_the_current_policys_inception_and_is_not_reset_by_a_mid_run_policy_or_insurer_switch; "
            "IF property_changes_insurer_or_product_type_without_any_actual_change_in_physical_occupancy "
            "THEN insurer_may_require_evidence_of_occupation_or_endorsement_compliance_covering_the_full_continuous_unoccupied_period_not_just_the_period_since_the_current_policys_inception"
        ),
        "Source PDF": "DRN-4660265.pdf",
    },
]


def validate_case(case: dict) -> None:
    for field, allowed in CONTROLLED_VOCAB.items():
        value = case.get(field, "")
        if value not in allowed:
            raise ValueError(
                f"Case {case['Case ID']}: '{field}' = '{value}' not in controlled vocab {allowed}"
            )


def append_cases(ws, cases: list) -> None:
    next_row = ws.max_row + 1
    for case_idx, case in enumerate(cases):
        row_fill = ROW_FILL_ODD if (next_row % 2 == 1) else ROW_FILL_EVEN
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=case.get(col_name, ""))
            cell.font      = ROW_FONT
            cell.fill      = row_fill
            cell.border    = ROW_BORDER
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        ws.row_dimensions[next_row].height = 80
        next_row += 1


def main() -> None:
    for case in NEW_CASES:
        validate_case(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases",
        "Unoccupied_Property_Case_Database.xlsx"
    )

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(
            f"Database not found: {xlsx_path}\n"
            "Run create_unoccupied_property_case_db.py first to create it."
        )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Cases"]

    rows_before = ws.max_row - 1  # exclude header
    append_cases(ws, NEW_CASES)
    rows_after = ws.max_row - 1

    wb.save(xlsx_path)

    print(f"Appended : {rows_after - rows_before} cases")
    print(f"Total    : {rows_after} data rows")
    print(f"Last ID  : {ws.cell(row=ws.max_row, column=1).value}")
    print(f"Saved    : {xlsx_path}")


if __name__ == "__main__":
    main()
