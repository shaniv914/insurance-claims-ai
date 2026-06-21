"""
Standard append script for Flood Case Database — Schema v1 (21 columns).
Batch 6: FLOOD-051 to FLOOD-056 (Final batch — completes 56/56 cases)

Usage
-----
Run from the repo root:
    py scripts/append_flood_v6.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Flood_Case_Database.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 6: FLOOD-051 to FLOOD-056 (Final)
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "FLOOD-051",
        "FOS Decision ID": "DRN7363961",
        "Insurer Name": "Lloyds Bank General Insurance Limited",
        "FOS Decision Date": "2014 (est.)",
        "Claim Type": "Home insurance — 2012 flood damage claim; insurer avoided policy from commencement (2010) for alleged non-disclosure at inception of previous flood claims (2004) and prior insurance refusals; consumer (Ms B) states she disclosed both facts to a branch representative during a 2010 in-branch meeting",
        "Leak Source": "External flood water — 2012 flood event at residential home; nature of flood source not specified; flood claim triggered discovery of alleged non-disclosure at 2010 policy inception",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Lloyds Bank: Ms B did not disclose previous flood claims or prior insurance refusals during the 2010 branch meeting; branch representative (a branch manager) would have been unable to provide a quotation if such facts had been disclosed; application records show no such disclosure; policy documentation sent post-inception recorded no prior flood or refusal; therefore policy validly avoided from commencement",
        "Evidence Dispute": "Lloyds Bank: application records show no prior flood or refusal; branch manager had no reason to falsify records; NCD documentation reflects 6-year claims-free period as a maximal NCD band, not as evidence of 2004 flood disclosure; 2011 flood incident not disclosed at 2012 renewal. Ms B: disclosed prior 2004 flood and prior refusals to branch representative; visited branch for unrelated matter and was approached about home insurance; provides detailed account of 2010 meeting. FOS: Ms B's consistent prior practice of disclosing flood claims (confirmed by refusals from multiple insurers including Lloyds Bank itself); account of meeting credible; 6-year NCD documentation consistent with 2004 flood disclosure (disclosing a 2004 claim in 2010 = 6 years); 2011 unclaimed flood need not be disclosed at renewal in response to 'circumstances changed' question.",
        "Outcome Category": "Upheld",
        "Outcome": "Policy must be reinstated (subject to premiums being paid) and 2012 flood claim reconsidered on its merits; records of policy voidance removed from internal and external databases; £200 D&I compensation; if Mr A and Ms B paid increased premiums due to the voidance, Lloyds Bank must reimburse the amount attributable solely to the voidance with 8% simple interest from date paid to settlement",
        "Compensation Awarded (£)": 200,
        "Is Core Case": "No — Administrative",
        "Key Policy Clause": "Balance of probabilities standard applies to determine what was disclosed at policy inception where no recording exists; consumer's consistent prior disclosure practice across multiple insurers is strong circumstantial evidence of disclosure to this insurer; consumer reasonably entitled to believe disclosed information was accurately recorded without verifying post-sale documentation; renewal question asking if 'circumstances had changed' does not necessarily require disclosure of a prior flood event that was not claimed for (especially where third party accepted liability)",
        "Missing Evidence": "Recording or contemporaneous written record of the 2010 branch meeting discussion",
        "Ombudsman Reasoning": "Pattern of disclosure established: Ms B was previously refused cover due to disclosing flood claims to multiple insurers including Lloyds Bank, proving consistent disclosure practice. Her account of the 2010 meeting credible: she visited for an unrelated reason with no incentive to withhold material facts when approached about insurance. NCD documentation recording 6 years claims-free is consistent with a disclosed 2004 flood claim (exactly 6 years prior). On 2011 non-claimed flood: 'circumstances changed' renewal question does not necessarily lead a consumer to disclose an event not claimed for, particularly where the council accepted liability. Balance of probabilities favours Ms B having disclosed both material facts.",
        "Workflow Insight": "Policy voidance for non-disclosure must be proven by insurer on balance of probabilities. Circumstantial evidence of consistent prior disclosure practice can outweigh application records where no recording of the inception conversation exists. Consumer's consistent pattern of disclosing material facts to prior insurers (evidenced by refusals arising from those disclosures) is strong evidence of continued disclosure. Post-sale documentation errors do not necessarily override what was verbally disclosed at inception. Renewal 'circumstances changed' question has limited scope — it does not necessarily capture every event that occurred since inception, particularly unclaimed events where a third party accepted responsibility.",
        "AI Rule Candidate": "IF insurer_avoids_policy_for_non_disclosure AND consumer_has_consistent_prior_disclosure_pattern_evidenced_by_prior_refusals AND no_recording_of_inception_meeting THEN FOS_applies_balance_of_probabilities_favouring_consumer_disclosure; IF renewal_documentation_asks_if_circumstances_changed AND consumer_experienced_flood_but_did_not_claim_AND_third_party_accepted_liability THEN non_disclosure_of_that_flood_at_renewal_is_reasonable; policy_voidance_requires_insurer_to_prove_non_disclosure_on_balance_of_probabilities",
        "Source PDF": "DRN7363961.pdf",
    },
    {
        "Case ID": "FLOOD-052",
        "FOS Decision ID": "DRN7937294",
        "Insurer Name": "UK General Insurance (Ireland) Limited",
        "FOS Decision Date": "Sep 2015",
        "Claim Type": "Home insurance — cellar flooded with a few inches of water; insurer's surveyor confirmed flood occurred and attributed cause to temporary rise in water table from prolonged heavy rainfall; policy contained specific exclusion for 'loss or damage caused by rising water table levels'; claim declined under exclusion; consumer also complained about poor service",
        "Leak Source": "Groundwater — water table rose following weeks of heavy rainfall, causing water to rise through cellar floor; water receded within days after rainfall ceased; surveyor ruled out burst pipes or above-ground water ingress; Victorian cellar without adequate damp-proof membrane made property susceptible to water table ingress",
        "Property Type": "Residential home (with cellar)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "UKII: surveyor confirmed flooding caused by rising water table following heavy rainfall; cellar lacks adequate damp-proof membrane; policy specifically excluded 'loss or damage caused by rising water table levels'; exclusion clearly set out in policy documentation; exclusion fair and reasonable to apply",
        "Evidence Dispute": "UKII: surveyor inspected cellar; ruled out burst pipes and above-ground ingress; water table rise confirmed as sole cause; property not flooded in 8+ years of heavy rain suggesting general water table rise. Consumer: property built into hillside (only two of four cellar walls fully underground) making water table rise unlikely; UKII used incorrect rainfall data (Met Office SE regional = 64.84mm, not UKII's 118.8mm); if water table caused it, should have stayed flooded during subsequent 95mm of rain. FOS: checked UKII rainfall data against postcode-specific records — found UKII's figures correct; water table rise can occur on hillsides and quick drainage is consistent with hillside; exclusion clear and not unusual; 8+ years without flooding during other heavy rain periods suggests water table may have generally risen.",
        "Outcome Category": "Compensation Only",
        "Outcome": "Flood claim decline upheld as reasonable; UKII must pay £100 compensation to Mr and Mrs S for distress and inconvenience caused by UKII's poor service (agent's slow and unclear responses; surveyor misunderstanding about who diagnosed the cause)",
        "Compensation Awarded (£)": 100,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Explicit 'rising water table levels' exclusion is valid, clear and not unusual in home insurance policies; insurer who agrees flooding occurred can still decline under rising water table exclusion; exclusion applies to temporary water table rises from heavy rain, not only permanent rises; water table rises can occur on hillsides, not only in low-lying areas; postcode-specific Met Office rainfall data is more reliable than regional averages; exclusion may apply even where flooding has not occurred previously",
        "Missing Evidence": "N/A — exclusion was clear in policy; rainfall data verified by FOS using postcode-specific records; surveyor confirmed sole cause",
        "Ombudsman Reasoning": "UKII's surveyor confirmed water table rise as sole cause; UKII agreed it was a flood. FOS initially (first provisional) thought exclusion unreasonable for a temporary rainfall-driven water table rise. After UKII challenged, FOS reconsidered: property hadn't flooded in 8+ years despite other heavy rain, suggesting water table may have generally risen — making the exclusion logically applicable. Consumer's postcode-specific rainfall data argument rejected after FOS independently verified UKII's figures were correct. Hillside objection rejected: water table rises can occur on hillsides and quick drainage is explained by topography. Exclusion is clearly drafted, not unusual, and fair to apply. Service failures by UKII's agent and surveyor misunderstanding warranted modest £100 D&I.",
        "Workflow Insight": "A clearly drafted 'rising water table levels' exclusion is distinct from having no flood definition. Where such exclusion exists: (a) insurer can concede flood occurred and still decline under the exclusion; (b) exclusion applies to temporary weather-driven rises, not just permanent table rises; (c) consumer cannot defeat the exclusion by pointing to hillside topography or challenging rainfall data without postcode-specific evidence. Contrast with Churchill-type cases where no policy flood definition existed — here the exclusion is express. Service failures (slow responses, surveyor attribution errors) warrant modest D&I even where underlying claim decline is upheld.",
        "AI Rule Candidate": "IF policy_has_explicit_rising_water_table_exclusion AND surveyor_confirms_rising_water_table_as_sole_cause THEN decline_justified_regardless_of_flood_definition; rising_water_table_exclusion_applies_to_temporary_rainfall_driven_rises_not_just_permanent; water_table_rise_can_occur_on_hillsides; postcode_specific_rainfall_data_required_to_challenge_insurer_rainfall_figures; distinguish_express_water_table_exclusion_from_policies_with_no_flood_definition",
        "Source PDF": "DRN7937294.pdf",
    },
    {
        "Case ID": "FLOOD-053",
        "FOS Decision ID": "DRN7939869",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "Jan 2017",
        "Claim Type": "Home insurance — basement flat flood from rising water table following heavy rain; RSA accepted flood cover but declined payment on basis contractor's invoice showed only preventative work (drain laying); consumer disputed that invoice included actual flood damage repairs; RSA's own telephone call with contractor revealed flood had caused a 'decent sized hole' which contractor made good when laying the drain",
        "Leak Source": "Groundwater — rising water table following heavy rain flooded basement flat; RSA accepted flood peril; dispute concerns whether contractor's invoice covered actual flood damage repairs or only preventative work",
        "Property Type": "Residential home (basement flat)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "RSA: claim form and contractor's invoice showed only preventative work (drain laying) and making good damage the contractor caused while laying the drain; no flood repair work identifiable in invoice; RSA confirmed with contractor by telephone who said only preventative work was done",
        "Evidence Dispute": "RSA: contractor confirmed by phone only preventative work done and making good contractor's own damage; invoice described drain laying. Consumer: provided photographs, video evidence and contractor's further letter asserting invoice included actual flood damage repairs; flood damaged walls and floor. FOS: in the same telephone call with RSA, contractor also mentioned flood had caused 'a decent sized hole' in the corner of the room where water entered; contractor made good all damage in that area (original flood damage + contractor's damage) when laying the drain; RSA failed to probe this disclosure.",
        "Outcome Category": "Upheld in Part",
        "Outcome": "RSA must take further steps to determine the amount payable to Mr and Mrs L for flood damage — whether by assessing further photographic/video evidence they have available, speaking again to the contractor to probe the disclosed flood damage, and/or carrying out a site visit",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Handling Dispute",
        "Key Policy Clause": "Policy covers flood damage but not preventative flood work (drain laying, item removal/storage to prevent damage or enable preventative work); insurer must adequately probe all material disclosures made during investigative conversations with contractors — including disclosures made incidentally during phone calls focused on other matters; contractor making good original flood damage while completing preventative works does not eliminate insurer's liability for that flood damage",
        "Missing Evidence": "Specific quantification of flood damage in corner of room (not separately identified in contractor's invoice); photographic/video evidence of original flood damage before contractor commenced work",
        "Ombudsman Reasoning": "RSA correctly declined preventative drain-laying costs and item removal/storage. But in the same call confirming preventative-only work, contractor mentioned flood had caused 'a decent sized hole' in the corner where water entered, and that he had made good all damage in that area when laying the drain. RSA had the opportunity to probe this disclosure in the call but did not. This confirms there was payable flood damage. Amount is likely modest (as conceded by contractor) but RSA must now quantify it. RSA was not excused from investigating this disclosure simply because repairs had already been completed.",
        "Workflow Insight": "Insurer handling investigative calls with contractors must probe all disclosures — including incidental remarks about original damage made during calls focused on other matters. Contractor's admission of original flood damage made good during preventative works is material and requires follow-up. Distinguishing flood damage repair from flood prevention in a single contractor invoice requires thorough investigation. The fact that repairs are already complete does not mean a site visit would be valueless — photographic/video evidence and further contractor conversations remain investigative options. Failure to probe a material contractor disclosure is a handling failure justifying FOS intervention.",
        "AI Rule Candidate": "IF contractor_confirms_preventative_work_in_phone_call AND contractor_also_mentions_original_flood_damage_made_good_in_same_call THEN insurer_must_probe_and_quantify_flood_damage_element; preventative_flood_work_not_covered_BUT_original_flood_damage_made_good_by_contractor_IS_covered; IF claim_decline_based_solely_on_invoice_description AND consumer_disputes_description THEN insurer_must_speak_to_contractor_and_probe_all_disclosures; repairs_already_completed_does_not_prevent_further_investigation",
        "Source PDF": "DRN7939869.pdf",
    },
    {
        "Case ID": "FLOOD-054",
        "FOS Decision ID": "DRN8469463",
        "Insurer Name": "TBO Services Limited",
        "FOS Decision Date": "May 2016",
        "Claim Type": "Commercial property insurance — broker (TBO Services Limited) failed to disclose P's prior flood history when completing 2015/2016 renewal statement of facts; insurer (overseas regulated, not FOS member) cancelled policy from start date for non-disclosure and declined 2015 flood claim; P seeks compensation from broker for loss of insurance proceeds",
        "Leak Source": "Surface water — heavy rain causing water to enter commercial premises through doors (2015 flood event); property had prior flood claim from blocked drain in road outside property",
        "Property Type": "Commercial (business premises)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "TBO's renewal process: at 2015 renewal TBO did not ask P about any questions except one (piece of equipment); TBO inserted 'no history of flooding' in the statement of facts using 2014 data, despite knowing P had prior flood from a blocked road drain; insurer cancelled policy from start date for non-disclosure of flood susceptibility; TBO said P answered 'no' to flood history questions in 2014 and P should have corrected the statements of fact",
        "Evidence Dispute": "TBO: in 2014 P was only asked about weather-caused flooding, and a blocked drain flood is not weather-caused; P should have noticed the incorrect answer in the 2015 statement of facts. FOS: TBO's own 2014 statement of fact recorded 'floor flooded from blocked drain on road' — it knew of prior flood; flooding from a blocked drain during heavy rain is reasonably attributed to weather conditions; at 2015 renewal TBO only asked one question — it inserted all other answers, including 'no history of flooding', without asking P; inserting an incorrect answer without asking the client is careless; P cannot be blamed for not noticing TBO's own insertion on page four of a form. Loss adjusters for second insurer: P's director confirmed several flash floods in the area prior to 2015. TBO: cannot produce recording of 2014 conversation.",
        "Outcome Category": "Upheld",
        "Outcome": "TBO Services Limited must jointly with P instruct a loss adjuster to assess what (if anything) the insurer would likely have paid in respect of the 2015 flood claim if the policy had not been cancelled; TBO must then compensate P by the loss adjuster's assessed sum less the 2015/2016 premium P paid; TBO must also pay the loss adjuster's fee",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": "Broker's duty to accurately complete insurance application and renewal statements of fact; where broker inserts answers on behalf of client without asking questions, broker assumes full responsibility for the accuracy of those answers; flooding from a blocked drain during heavy rain is reasonably characterised as weather-related and relevant to a 'history of flooding' question; broker cannot rely on client not having noticed broker's own incorrect insertion; where insurer is not FOS-regulated, FOS can order broker to compensate equivalent to what the insurer would have paid",
        "Missing Evidence": "Recording of original 2014 conversation in which P provided flood history information (TBO could not produce this)",
        "Ombudsman Reasoning": "At 2015 renewal TBO asked P only one question; TBO inserted all other answers including 'no history of flooding' from 2014 data without discussing them with P. TBO knew from its own 2014 statement of fact that P's property had flooded from a blocked road drain. Flooding from a blocked drain during heavy rain is reasonably caused by weather conditions. TBO had a duty to discuss the 'history of flooding' question with P before inserting 'no' — especially knowing of the prior flood and especially given its own staff guidance to clarify questions where clients may be confused. TBO cannot shift responsibility to P for not spotting its incorrect insertion on page four. Underlying insurer not FOS member: FOS orders TBO to compensate equivalent to likely insurer settlement.",
        "Workflow Insight": "Brokers who insert renewal answers without re-asking questions assume responsibility for accuracy — especially for material flood history questions where a prior flood claim is known. A prior flood from a blocked drain during heavy rain counts as flood history relevant to flood risk questions. Where the underlying insurer is not FOS-regulated, FOS treats the broker as the responsible party and orders compensatory payment equivalent to lost insurer proceeds. Broker's own internal standards (staff to clarify confused answers) are used against it when it fails to follow them.",
        "AI Rule Candidate": "IF broker_inserts_answers_in_renewal_without_asking_client AND known_prior_flood_claim_exists AND history_of_flooding_answered_no THEN broker_liable_for_resulting_non_disclosure; flood_from_blocked_road_drain_during_heavy_rain_counts_as_flood_history; IF underlying_insurer_not_FOS_member AND broker_at_fault THEN FOS_orders_broker_to_compensate_equivalent_to_insurer_settlement; broker_cannot_blame_client_for_not_spotting_brokers_own_incorrect_insertion",
        "Source PDF": "DRN8469463.pdf",
    },
    {
        "Case ID": "FLOOD-055",
        "FOS Decision ID": "DRN9152389",
        "Insurer Name": "WR Berkley Insurance (Europe) Limited",
        "FOS Decision Date": "Sep 2016",
        "Claim Type": "Residential buildings insurance — 2015 flood damage; standard insurance schedule listed storm, tempest and flood as covered perils; policy endorsement specifically excluded 'damage caused by flooding'; consumer argued the schedule showed they had flood cover; insurer declined under the endorsement",
        "Leak Source": "External flood (2015 flood event; source not specified in decision; property known to be in a high flood risk area at inception)",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Berkley: endorsement on policy schedule specifically excluded 'damage caused by flooding'; endorsement overrides the standard schedule peril list; consumer signed a statement of fact at inception identifying property as high flood risk; initially property was unoccupied and all three perils (storm, tempest, flood) were excluded; when property became occupied, storm and tempest were included but flood was excluded by endorsement; endorsement clearly and expressly removed flood cover",
        "Evidence Dispute": "Consumer: standard schedule listed storm, tempest and flood — they believed they had flood cover under the policy. Berkley: endorsement is a key and integral part of the policy; it clearly excluded flood while retaining storm and tempest; Berkley has since updated its standard schedule to treat storm and flood as separate perils, reflecting the original intent. FOS: endorsements are integral to any insurance policy; the flood exclusion endorsement clearly and expressly removed flood cover; not unfair or unreasonable to rely on it.",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld; Berkley acted fairly and reasonably in declining the flood claim under the clearly expressed flood exclusion endorsement",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Endorsements form a key and integral part of an insurance policy and override the standard peril list in the policy schedule; a clear flood exclusion endorsement takes precedence over 'storm, tempest and flood' wording in the body of the schedule; consumer signed statement of fact identifying high flood risk at inception — consistent with expectation that flood cover might be excluded or restricted; standard policy schedule listing 'flood' as a peril does not guarantee flood cover where that cover is removed by endorsement",
        "Missing Evidence": "N/A — endorsement was unambiguous; consumer's statement of fact acknowledged high flood risk at inception",
        "Ombudsman Reasoning": "Consumer signed a statement of fact at inception confirming the property was in a high flood risk area. Initially the property was unoccupied and storm, tempest and flood perils were all excluded. When the property became occupied, storm and tempest were restored to the standard peril list but flood was specifically excluded by endorsement. The endorsement was an integral part of the policy, clearly drafted, and expressly excluded flood damage. Berkley's subsequent update to its standard schedule (separating storm and flood as distinct perils) confirms the original intent was always to retain storm/tempest cover while excluding flood. Relying on the endorsement to decline was not unfair or unreasonable.",
        "Workflow Insight": "Consumers should read policy endorsements as an integral part of their cover — not as ancillary documents. An endorsement expressly removing flood cover takes precedence over the standard schedule's peril list. Where property is in a known high flood risk area and consumer signed a statement of fact to that effect at inception, a flood exclusion endorsement is expected and cannot be overridden by argument that the standard schedule listed flood. Insurer subsequently updating its standard schedule to reflect its original intent reinforces that the endorsement was a deliberate coverage limitation.",
        "AI Rule Candidate": "IF policy_endorsement_explicitly_excludes_flood AND standard_schedule_lists_flood_as_covered_peril THEN endorsement_overrides_schedule; IF property_known_high_flood_risk_at_inception AND flood_exclusion_endorsement_present THEN endorsement_expected_and_valid; endorsement_is_integral_part_of_policy_not_optional_addendum; standard_peril_list_does_not_guarantee_cover_where_endorsement_removes_it",
        "Source PDF": "DRN9152389.pdf",
    },
    {
        "Case ID": "FLOOD-056",
        "FOS Decision ID": "DRN9771710",
        "Insurer Name": "Society of Lloyd's",
        "FOS Decision Date": "Mar 2017",
        "Claim Type": "Holiday caravan insurance — Storm Frank caused river to burst banks and flood caravan site; caravan floated 120 metres and suffered extensive damage; policy covered storm and accidental damage but specifically excluded 'all claims due to flood'; insurer declined under flood exclusion; consumer argued storm was the true (proximate) cause",
        "Leak Source": "Flood water from river bursting banks — Storm Frank conditions (sustained winds of 75mph+, extreme and unusual rainfall intensity) caused river to overflow and flood the caravan site; caravan floated 120 metres in flood water; damage caused by flood water immersion, not wind",
        "Property Type": "Holiday caravan",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Society of Lloyd's: policy expressly excluded 'all claims due to flood'; flood water caused the direct damage to the caravan; consumer himself described 'Storm Frank floated caravan 120 metres approx.' — confirming flood water as damaging agent, not wind; consumer was informed of flood exclusion approximately three years before the event and continued the policy",
        "Evidence Dispute": "Consumer: storm conditions (which are covered) were the true cause of everything — storm caused flood which caused damage; storm is the proximate cause and should override the flood exclusion. Lloyd's: caravan was damaged by flood water not wind; consumer's own description confirms this; storm and flood are distinguishable events; policy clearly excluded all flood claims. FOS: weather records confirmed storm conditions (75mph gusts, extreme rainfall intensity) — qualifies as storm. But consumer's own statement ('floated caravan 120 metres') confirms flood water was the direct damaging cause; if caravan had been beyond reach of flood water, storm winds would not have caused the damage; proximate cause of loss = flood, not storm.",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld; Lloyd's reasonably declined claim under the express 'all claims due to flood' exclusion; flood water was the proximate cause of loss; storm conditions were the antecedent cause not the direct cause of damage",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Explicit 'all claims due to flood' exclusion covers damage caused by flood water even when storm conditions caused the flood event; proximate cause rule — the immediate/direct cause of damage (flood water) determines which exclusion applies, not the antecedent cause (storm); consumer's own incident description is an important evidential admission about the actual mechanism of damage; consumer awareness of flood exclusion for approximately three years and continuation of policy = informed acceptance of exclusion terms; storm-triggered floods fall within a broad flood exclusion",
        "Missing Evidence": "N/A — consumer's own description of the incident confirmed flood water as the direct cause; weather records verified storm conditions; damage extent confirmed to exceed sum insured",
        "Ombudsman Reasoning": "Weather records confirmed storm conditions (75mph+ gusts, extreme rainfall). Consumer himself reported 'Storm Frank floated caravan 120 metres approx.' — flood water was the direct damaging agent. FOS: storm caused river to burst banks (accepted), but it was the flood water that caused the loss. If caravan had been outside the flood's reach, storm winds alone would not have caused the damage. Proximate/immediate cause = flood. Policy excluded 'all claims due to flood' — broad exclusion covering storm-triggered flood events. Consumer knew about exclusion ~3 years before and continued the policy. Additional damage claim when caravan was moved irrelevant as total repair cost exceeded sum insured.",
        "Workflow Insight": "Where storm causes a flood which causes the damage, the immediate cause of damage (flood water) is the proximate cause — not the antecedent storm. A broad flood exclusion ('all claims due to flood') will encompass storm-triggered flood events and defeat arguments that storm cover should apply. Consumer's own description of the incident is a powerful admission: 'Storm Frank floated my caravan' confirms flood water was the damaging mechanism. Long-standing consumer awareness of an exclusion and continued policy renewal = acceptance of those exclusion terms. Additional damage claimed during recovery/removal is not separately recoverable if total repair cost exceeds sum insured.",
        "AI Rule Candidate": "IF storm_causes_flood_which_causes_damage AND policy_has_broad_flood_exclusion THEN flood_is_proximate_cause_and_exclusion_applies; storm_cover_does_not_override_express_flood_exclusion_where_flood_is_direct_damaging_mechanism; consumer_own_description_of_flood_water_as_damaging_agent_is_decisive_evidential_admission; consumer_awareness_of_flood_exclusion_for_multiple_years_and_policy_continuation_is_acceptance_of_exclusion_terms",
        "Source PDF": "DRN9771710.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Flood_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
