"""
Standard append script for Flood Case Database — Schema v1 (21 columns).
Batch 2: FLOOD-011 to FLOOD-020

Usage
-----
Run from the repo root:
    py scripts/append_flood_v2.py

Appends NEW_CASES rows to:
    knowledge/case-databases/Flood_Case_Database.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "Case ID",
    "FOS Decision ID",
    "Insurer Name",
    "FOS Decision Date",
    "Claim Type",
    "Leak Source",
    "Property Type",
    "Dispute Type",
    "Coverage Decision",
    "Rejection Reason",
    "Evidence Dispute",
    "Outcome Category",
    "Outcome",
    "Compensation Awarded (£)",
    "Is Core Case",
    "Key Policy Clause",
    "Missing Evidence",
    "Ombudsman Reasoning",
    "Workflow Insight",
    "AI Rule Candidate",
    "Source PDF",
]

CONTROLLED_VOCAB = {
    "Dispute Type": {
        "Coverage Dispute",
        "Handling / Reinstatement Dispute",
        "Endorsement / Exclusion Challenge",
        "Pre-Inception Damage Dispute",
        "Peril Classification Dispute",
        "Claim Recording / Administrative Dispute",
        "Broker Conduct Dispute",
    },
    "Coverage Decision": {
        "Declined — Full",
        "Declined — Partial",
        "Accepted",
        "Accepted — Disputed Settlement",
        "Not Applicable",
    },
    "Outcome Category": {
        "Upheld",
        "Upheld in Part",
        "Not Upheld",
        "Compensation Only",
    },
    "Is Core Case": {
        "Yes",
        "No — Administrative",
        "No — Handling Dispute",
        "No — Commercial",
        "No — Broker Dispute",
    },
}

CENTRED_COLS = {
    "Case ID", "FOS Decision ID", "Insurer Name", "FOS Decision Date",
    "Property Type", "Dispute Type", "Coverage Decision",
    "Outcome Category", "Compensation Awarded (£)", "Is Core Case",
    "Source PDF",
}

# ---------------------------------------------------------------------------
# NEW CASES — Batch 2: FLOOD-011 to FLOOD-020
# ---------------------------------------------------------------------------
NEW_CASES: list[dict] = [
    {
        "Case ID": "FLOOD-011",
        "FOS Decision ID": "DRN2341674",
        "Insurer Name": "Society of Lloyd's",
        "FOS Decision Date": "5 Jul 2017",
        "Claim Type": "Static caravan destroyed by flooding at caravan park following major storm; flood exclusion endorsement applied at inception; claimant argued storm was proximate cause",
        "Leak Source": "River burst banks due to cumulative heavy rainfall over 2 months prior to storm; storm occurred on already-saturated ground; ground-level floodwater from river inundated caravan park",
        "Property Type": "Other (static caravan / holiday park)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Specific flood exclusion endorsement applied at inception due to high flood risk at caravan park; flood was proximate cause of damage, not storm — cumulative rainfall caused river to rise, storm was not the active and efficient cause",
        "Evidence Dispute": "Claimant argued storm was proximate cause and but-for storm flood would not have occurred; insurer cited 2 months of heavy rainfall saturating ground and Met Office data; damage caused solely by floodwater not by storm action",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — flood exclusion endorsement at inception valid and clearly communicated; proximate cause of damage was flood (cumulative rainfall), not storm as a single event; insurer not required to cover claim",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood exclusion endorsement specific to caravan park (high-risk location); proximate cause test — heavy rainfall over preceding months was the active and efficient cause of the river flooding, not the storm; 'but for' test insufficient where independent cause (cumulative rainfall) would likely have produced same result",
        "Missing Evidence": "No independent hydrological report distinguishing storm contribution from pre-existing rainfall accumulation; no evidence storm alone would have caused flood absent prior saturation",
        "Ombudsman Reasoning": "Heavy rainfall 2 months before storm saturated ground; river likely to have burst regardless of storm; damage caused solely by floodwater — if barriers had been in place, storm alone would not have caused damage; claimant agreed to flood endorsement terms at inception; 'but for' storm argument fails because prior rainfall was itself an independent operative cause",
        "Workflow Insight": "Specific flood endorsements at inception are enforceable even when storm immediately precedes flood; proximate cause test requires identifying the efficient cause of the flood itself (not just the triggering weather event); cumulative prior rainfall qualifying as an independent cause defeats the 'but for storm' argument",
        "AI Rule Candidate": "IF flood_exclusion_endorsed_at_inception = TRUE AND proximate_cause = 'flood' AND cumulative_prior_rainfall_was_independent_cause THEN decline; storm as contributing factor insufficient to override explicit endorsement",
        "Source PDF": "DRN2341674.pdf",
    },
    {
        "Case ID": "FLOOD-012",
        "FOS Decision ID": "DRN2449131",
        "Insurer Name": "Woodland Insurance Services Ltd",
        "FOS Decision Date": "11 Dec 2015",
        "Claim Type": "Broker conduct dispute — broker failed to properly disclose prior flood history to insurer causing insurer to remove flood cover and decline commercial property flood claim (2014 heavy rain, drain overflow)",
        "Leak Source": "Surface water / drain overflow — drains unable to cope with heavy rainfall; water entered ground floor of commercial buy-to-let property; culvert blocked causing water to overflow and enter premises",
        "Property Type": "Commercial (buy-to-let rental)",
        "Dispute Type": "Broker Conduct Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Insurer removed flood cover and declined claim after discovering undisclosed flood history (two prior floods in 2007 and 2012); broker had described prior events as 'blocked drains' rather than flooding, understating the risk",
        "Evidence Dispute": "Broker claimed Mr R only mentioned 'blocked drains'; Mr R said he disclosed full flood history and showed conveyancer documents; call recording confirmed broker knew dates and water company involvement — sufficient to trigger duty to investigate further; broker's own post-decline letter referenced culvert flooding inconsistent with 'only just learned' defence",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld against broker — broker should have disclosed or further investigated flood history before submitting to insurer; insurer separately directed to consider claim; broker to pay Mr R £100 for trouble and upset caused by delay",
        "Compensation Awarded (£)": 100,
        "Is Core Case": "No — Broker Dispute",
        "Key Policy Clause": "Broker's duty to accurately communicate material prior flood history to insurer; duty to probe ambiguous prior claim descriptions; client entitled to rely on broker's competence in handling disclosure",
        "Missing Evidence": "No written record of what Mr R disclosed during face-to-face broker visit; call recording available but confirmed 'blocked drains' framing despite broker knowing dates and water company context",
        "Ombudsman Reasoning": "Broker knew the dates and water company's remedial works — this information came from Mr R and was sufficient to trigger further enquiry; 'blocked drains' framing understated risk; Mr R had no incentive to downplay severity; broker's post-decline letter referencing culvert flooding undermined defence that full facts were not known at outset; Mr R entitled to assume questions were answered correctly having signed summary on broker's advice",
        "Workflow Insight": "Brokers must probe ambiguous prior claim descriptions; 'blocked drains' terminology can conceal material flood history — any description involving internal water damage or prior claims should prompt specific flood disclosure questions; signing a statement of fact prepared by the broker does not shift responsibility to the client",
        "AI Rule Candidate": "IF prior_claim_described_as_blocked_drains AND internal_flooding_known_or_inferable THEN broker_must_verify_flood_classification_before_submission_to_insurer; IF broker_has_dates_and_remediation_detail THEN duty_to_investigate_triggered",
        "Source PDF": "DRN2449131.pdf",
    },
    {
        "Case ID": "FLOOD-013",
        "FOS Decision ID": "DRN-2482936",
        "Insurer Name": "AXA Insurance UK Plc",
        "FOS Decision Date": "5 Apr 2021",
        "Claim Type": "Peril classification dispute — AXA classified claim as flood after flash flood caused drain backup and internal water ingress; policyholders argued classification should be escape of water from blocked drain, fearing adverse impact on future premiums and resale value",
        "Leak Source": "External surface water / flash flood — exceptionally heavy rainfall caused drains to back up unable to cope with volume; rainwater unable to enter drain accumulated outside and entered property; no standing water observed at time of loss adjuster visit",
        "Property Type": "Residential home",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "Not applicable — claim accepted and repaired under flood peril; policyholders disputed classification only (concerned about premium and resale impact)",
        "Evidence Dispute": "Policyholders argued drain was blocked by building debris before rain (escape of water); AXA cited flash flood evidence including regional claims data and very heavy rainfall in preceding days; policyholders provided river level data showing levels had previously been higher without water ingress; no emergency plumber's report retained",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — AXA's flood classification upheld; proximate cause was external rainwater flooding the ground outside; water never entered the drain so cannot be escape of water; river level data irrelevant to local circumstances; absence of plumber's report immaterial",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Flood defined as 'invasion of property by large volume of water caused by rapid build-up or sudden release of water from outside the buildings'; escape of water requires water to have entered and then escaped an apparatus — water that never entered the drain cannot escape from it; proximate cause is the external rainwater flooding the ground, not a blockage within the drainage apparatus",
        "Missing Evidence": "Emergency plumber's report not retained by AXA (AXA arranged the plumber); CCTV drain survey to confirm pre-existing blockage vs. flood-induced backup; plumber's report would not have changed outcome given undisputed that water could not enter drain",
        "Ombudsman Reasoning": "Flash flood confirmed by regional claims data and heavy rainfall evidence; drain backed up because it could not cope with rainfall volume — flood overwhelmed drainage; water that cannot enter the drain has not 'escaped' from it; escape of water requires release from apparatus containing water, not failure of drainage to absorb external water; local river level data does not determine whether local circumstances constituted a flood; AXA acted reasonably in its classification",
        "Workflow Insight": "Flood vs EOW classification turns on whether water entered the drain/pipe (EOW) or was unable to enter at all (flood); absence of plumber's report does not change outcome where cause is undisputed; consumer concern about premium and resale impact does not alter correct peril classification",
        "AI Rule Candidate": "IF external_surface_water_confirmed = TRUE AND water_unable_to_enter_drain = TRUE THEN classify_as_flood NOT escape_of_water; escape_of_water requires prior containment within apparatus",
        "Source PDF": "DRN-2482936.pdf",
    },
    {
        "Case ID": "FLOOD-014",
        "FOS Decision ID": "DRN2512341",
        "Insurer Name": "Zurich Insurance PLC",
        "FOS Decision Date": "28 Feb 2019",
        "Claim Type": "Latent flood damage claim — wet rot on joists discovered after purchase; previous owner had a 'serious flood' approximately 2 years before; claimant sought to invoke secondary flooding protocol; Zurich declined as damage pre-dated policy inception",
        "Leak Source": "Unknown — original cause unconfirmed; possibly escape of water (pipe/appliance) or external flood; property not near river; no evidence of damage to neighbouring properties consistent with natural flood",
        "Property Type": "Residential home",
        "Dispute Type": "Pre-Inception Damage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Damage occurred before policy inception; secondary flooding protocol requires confirmed flood as original cause — cannot establish on balance of probabilities that original damage was caused by flood rather than escape of water; escape of water is the more likely cause given property location; secondary escape of water not covered by secondary flooding protocol",
        "Evidence Dispute": "Only hearsay from neighbour that previous owner had 'serious flood'; no access to original claim file; property not near river or body of water; no evidence of damage to neighbouring properties; claimant later noted property in medium-high surface water flood risk area and found water in floor void during renovation — FOS found this did not add to case",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — Zurich's decline upheld; secondary flooding protocol does not apply where original cause unconfirmed; escape of water more likely given location and absence of neighbouring flood damage; hearsay from neighbour insufficient to establish balance of probabilities; water found during renovation not linked to original damage",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Coverage limited to damage occurring during policy period; secondary flooding protocol (industry practice) applies only to confirmed flood damage — does not extend to secondary escape of water; even if escape of water causes a flood, predominant cause remains escape of water; policyholder must show on balance of probabilities that original damage was caused by flood",
        "Missing Evidence": "Original claim file from previous owner's insurer; professional confirmation of how original damage was classified (flood or EOW); hydrological flood risk mapping for the specific property; evidence of flood damage to neighbouring properties at the relevant time",
        "Ombudsman Reasoning": "Neighbour hearsay insufficient — 'serious flood' ambiguous as to cause, extent, and mechanism; property not near natural water body; no evidence neighbouring properties were damaged suggesting natural flood; escape of water scenario more consistent with property location; subsequent water found in floor void during renovation not conclusively linked to original damage or to flood; secondary flooding protocol narrow and confirmed-flood-only",
        "Workflow Insight": "To invoke secondary flooding protocol, claimant must establish on balance of probabilities that original damage was caused by flood (not merely that a flood happened); property location, absence of neighbouring damage, and classification of original claim are key factors; hearsay from a neighbour is insufficient; the distinction between flood and escape of water matters even when the practical outcome is similar",
        "AI Rule Candidate": "IF latent_damage_pre_policy = TRUE AND original_cause_unconfirmed THEN decline_secondary_flooding_protocol UNLESS flood_established_on_balance_of_probabilities_by_evidence_beyond_hearsay",
        "Source PDF": "DRN2512341.pdf",
    },
    {
        "Case ID": "FLOOD-015",
        "FOS Decision ID": "DRN-2541699",
        "Insurer Name": "The National Farmers' Union Mutual Insurance Society Limited",
        "FOS Decision Date": "23 Feb 2021",
        "Claim Type": "Peril classification dispute — NFUM classified water-soaked carpet as flood leading to Flood Re cession and near-doubled renewal premium; claimant argued storm was main cause as rainfall was of exceptional intensity; inadequate loss adjuster investigation",
        "Leak Source": "Surface water ingress — carpet soaked in corner of sitting room after exceptionally heavy rain (15.4mm/hour); entry point not inspected by loss adjuster; possibly gutter overflow or drain backup; no standing water accumulation observed at loss adjuster visit 9 days later",
        "Property Type": "Residential home",
        "Dispute Type": "Peril Classification Dispute",
        "Coverage Decision": "Accepted",
        "Rejection Reason": "Not applicable — claim accepted and carpet paid (approx £1,100); dispute is classification: NFUM classified as flood leading to Flood Re cession and large premium increase; claimant sought reclassification as storm",
        "Evidence Dispute": "Loss adjuster concluded flood (drains backed up) based on inspection 9 days post-incident with no drain or gutter inspection; weather records showed 15.4mm/hour rainfall (exceptional intensity) on the day; no evidence of standing water accumulation required for NFUM's flood definition; claimant argued storm caused damage via gutter overflow",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld — storm reclassified as main cause; NFUM to reclassify claim as storm, update records including external databases, reassess and rewrite 2020/21 policy cover, refund excess premium with 8% simple interest per annum, and withdraw policy from Flood Re scheme",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Storm (no policy definition — FOS applies: violent winds or exceptional intensity rainfall; 15.4mm/hour qualifies); NFUM flood definition requires substantial/abnormal body of water entering at ground level — no evidence of standing accumulation; Flood Re cession criteria require a flood claim; where both storm and flood could apply, FOS assesses which was main cause",
        "Missing Evidence": "Loss adjuster did not inspect drain or gutter at time of claim; no record of whether standing water accumulation was present at time of damage (visited 9 days later); loss adjuster report conclusory and did not explain how flood determination was reached; no gutter or drain inspection conducted",
        "Ombudsman Reasoning": "15.4mm/hour rainfall = exceptional intensity constituting a storm; NFUM's own flood definition requires substantial/abnormal body of water — no evidence of standing accumulation at time of damage; loss adjuster's report inadequate — visited 9 days post-incident with no drain or gutter inspection; given drastic premium consequences of flood classification and probability that damage was caused by storm, fair to reclassify; NFUM's underwriting technician's reliance on 'drains backed up' finding in inadequate report insufficient to maintain flood classification",
        "Workflow Insight": "When flood vs storm classification triggers Flood Re cession and large premium increase, FOS applies heightened scrutiny to adequacy of loss adjuster investigation; a conclusory report produced after 9-day delay without drain or gutter inspection cannot support flood classification where storm evidence (weather records) is strong; insurers must conduct contemporaneous thorough investigations before classifying perils with significant premium consequences",
        "AI Rule Candidate": "IF rainfall_intensity >= 15mm_per_hour AND no_standing_water_evidenced AND loss_adjuster_inspection_inadequate_or_delayed THEN consider_storm_reclassification; IF flood_classification_triggers_flood_re_cession THEN require_adequate_contemporaneous_investigation_before_flood_classification_stands",
        "Source PDF": "DRN-2541699.pdf",
    },
    {
        "Case ID": "FLOOD-016",
        "FOS Decision ID": "DRN-2632381",
        "Insurer Name": "Royal & Sun Alliance Insurance Plc",
        "FOS Decision Date": "29 Mar 2021",
        "Claim Type": "Commercial basement flood claim declined — RSA applied rising water table exclusion; claimant argued heavy rain was proximate cause and exclusion should not apply",
        "Leak Source": "Rising water table / groundwater — prolonged heavy rainfall raised water table to basement level; basement (recently tanked) inundated; contractor confirmed cause was 'rising groundwater associated with rainfall and water table changes'",
        "Property Type": "Commercial (rental property with basement)",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Rising water table exclusion in policy — flooding caused by change in groundwater/water table level; contractor's own evidence confirmed rising groundwater as cause; prior tanking consistent with known groundwater vulnerability; heavy rain as contributing cause does not override explicit exclusion for water table mechanism",
        "Evidence Dispute": "Claimant argued heavy rain was proximate cause and causal link between rainfall and water table rise should engage flood cover; contractor's evidence attributed flooding to rainfall and water table changes; RSA relied on exclusion wording and contractor's own description of the mechanism",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — RSA's application of rising water table exclusion upheld; causal link between heavy rain and water table rise acknowledged but does not override the explicit exclusion; contractor's evidence itself confirmed the exclusion mechanism; insurer met burden of showing exclusion applies on balance of probabilities",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Explicit exclusion for loss or damage caused by rising water table levels; onus on insurer to show exclusion applies on balance of probabilities; rising water table exclusion not unusual in commercial property policies; even if heavy rain caused the water table to rise, the mechanism of water entry (rising groundwater) falls within the exclusion",
        "Missing Evidence": "No independent hydrological report distinguishing groundwater flood from surface water flood; contractor's report was the only professional evidence — and it supported the exclusion mechanism; no evidence of surface water entry independent of the water table rise",
        "Ombudsman Reasoning": "Contractor explicitly described cause as rising groundwater associated with rainfall — this is the exclusion mechanism; tanking already installed indicates known groundwater vulnerability; while heavy rain contributed, this does not change the mechanism of entry; RSA's exclusion is clear and not unusual for commercial policies; insurer met burden on balance of probabilities via contractor's own evidence; causal link between rain and water table rise is direct but does not negate the exclusion",
        "Workflow Insight": "Rising water table exclusion is enforceable even where heavy rain ultimately caused the water table to rise; policyholders with basements in groundwater-susceptible areas should be alerted to this exclusion at point of sale or renewal; prior installation of tanking is consistent with known groundwater risk and supports application of exclusion",
        "AI Rule Candidate": "IF flood_mechanism = 'groundwater_rise' AND rising_water_table_exclusion_in_policy = TRUE THEN decline EVEN IF prolonged_rainfall_was_contributing_cause; onus_on_insurer_to_show_exclusion_applies_on_balance_of_probabilities",
        "Source PDF": "DRN-2632381.pdf",
    },
    {
        "Case ID": "FLOOD-017",
        "FOS Decision ID": "DRN-2785967",
        "Insurer Name": "Ageas Insurance Limited",
        "FOS Decision Date": "9 Jul 2021",
        "Claim Type": "Flood claim declined — Ageas argued damage was limited to garden/riverbank (uncovered) after August 2020 flood caused riverbank erosion; claimant contended a garden wall (covered building element) was destroyed",
        "Leak Source": "River flooding / bank erosion — property at confluence of two rivers; heavy rain in August 2020 caused riverbank at end of garden to collapse into river; county council drainage engineer confirmed flooding event; structural engineer confirmed flood causation",
        "Property Type": "Residential home",
        "Dispute Type": "Coverage Dispute",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Damage limited to garden/riverbank — not covered under buildings definition; insufficient evidence that a garden wall (a covered building element per policy definition) was present and destroyed by the flood; policyholder bears burden of showing covered property was damaged",
        "Evidence Dispute": "Claimant asserted permanent stone garden wall on stone foundations existed at riverbank (partially vegetation-covered; lower sections/foundations not visible); none of initial expert reports (council engineer, structural engineer, loss adjuster) mentioned a wall; repair engineer later referenced 'remains of existing damaged block stone foundation and stone wall'; insurance risk manager used uncertain language ('able to ascertain remains of a stone wall'); initial call to Ageas described riverbank collapse not wall; daughter mentioned gabions and boulders but not a wall",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — burden of proof not met; insufficient evidence that a covered building structure (garden wall) was damaged by flood as opposed to uncovered garden/riverbank; no contemporaneous expert confirmed wall existed; policyholder must show on balance that covered property was damaged by insured event; no award",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Policy buildings definition includes garden walls but not garden items (flower beds, hedges, lawns, shrubs, trees); burden on policyholder to show on balance of probabilities that a covered property element was damaged by an insured event; Ageas not required to prove absence of a wall — policyholder must prove presence of covered structure",
        "Missing Evidence": "Pre-flood photographs or structural survey confirming existence and classification of a garden wall; contemporaneous inspection report from council or structural engineer mentioning wall; CCTV or survey of riverbank prior to flood; no expert who attended the site before or immediately after the flood mentioned a wall",
        "Ombudsman Reasoning": "No contemporaneous expert confirmed a garden wall; multiple early expert reports (council engineer, structural engineer, loss adjuster) silent on any wall despite detailed inspections; Mr J's own initial report to Ageas described riverbank collapse only; daughter mentioned gabions and boulders; insurance risk manager's language implied uncertainty; repair engineer only saw damaged remains post-flood making pre-flood description unreliable; structure only partially visible (vegetation-covered) before flood creating unrecoverable evidential gap",
        "Workflow Insight": "Claims for flood damage to garden structures require pre-loss documentary evidence of the structure's existence and its eligibility as a covered building element; structures partially obscured by vegetation or only partially visible present an unrecoverable evidential gap if not pre-documented; multiple contemporaneous expert visits without mention of a wall is strong evidence against its existence as a covered structure",
        "AI Rule Candidate": "IF claimed_damaged_structure_not_mentioned_in_contemporaneous_expert_reports AND no_pre_loss_documentation_of_covered_structure THEN burden_of_proof_not_met AND decline; policyholder_must_prove_covered_property_was_damaged_on_balance_of_probabilities",
        "Source PDF": "DRN-2785967.pdf",
    },
    {
        "Case ID": "FLOOD-018",
        "FOS Decision ID": "DRN-2787998",
        "Insurer Name": "Ocaso SA, Compania de Seguros y Reaseguros",
        "FOS Decision Date": "11 Jun 2021",
        "Claim Type": "Flood claim declined — Ocaso retrospectively excluded flood cover after Ms J disclosed 1998 area flood history; flood exclusion endorsement added December 2019; property flooded February 2020 following named storm",
        "Leak Source": "External flooding following named storm (February 2020); property confirmed high flood risk by government data; area had flooded in 1998 when nearby river burst its banks",
        "Property Type": "Residential home",
        "Dispute Type": "Endorsement / Exclusion Challenge",
        "Coverage Decision": "Declined — Full",
        "Rejection Reason": "Flood cover excluded by endorsement added December 2019 (retrospective to inception August 2018) after Ms J disclosed 1998 flood history; insurer's underwriting guidelines demonstrate it would never have offered flood cover if prior flood history had been known; property confirmed as high flood risk by government website; answers at inception were inaccurate even if given in good faith",
        "Evidence Dispute": "Ms J argued she was unaware of 1998 flood history when taking out policy (lived there after); answered flood risk questions to best of knowledge at time; Ocaso provided underwriting guidelines confirming it would never have offered flood cover for this risk; government website confirmed property at high flood risk; Ms J argued neighbour with Ocaso policy was covered for flood",
        "Outcome Category": "Not Upheld",
        "Outcome": "Complaint not upheld — flood exclusion endorsement fairly applied retrospectively; inaccurate answers at inception (even in good faith) are material; Ocaso demonstrated via underwriting guidelines it would not have offered flood cover; broker instructed to notify Ms J of change; claim fairly declined; Flood Re not available via price comparison site route; neighbour's policy circumstances not comparable",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Material change in risk — prior flood history is material to flood cover underwriting; inaccurate answers at inception permit retrospective endorsement even if made in good faith; Flood Re scheme accessible only via direct broker/insurer route, not price comparison sites; insurer must demonstrate via underwriting evidence it would not have offered cover if true facts known",
        "Missing Evidence": "Not applicable — insurer had underwriting guidelines and government flood risk data; outcome did not depend on missing evidence",
        "Ombudsman Reasoning": "Ms J answered questions to best of knowledge — accepted in good faith; but inaccuracy was material regardless of intent; Ocaso provided underwriting guidelines confirming flood cover would never have been offered for this property; property classified as high flood risk by government; Ocaso instructed broker to notify Ms J promptly; Ms J confirmed she wanted policy to continue without flood cover; Flood Re unavailable via PCW channel; neighbour's circumstances unknown and irrelevant",
        "Workflow Insight": "Properties acquired via price comparison sites cannot access Flood Re underwriting — Flood Re eligibility requires direct broker/insurer engagement at point of sale; undisclosed prior flood history (even where policyholder is unaware) permits retrospective flood exclusion if insurer can demonstrate it would not have offered cover; good faith at inception does not prevent retrospective endorsement where misrepresentation is established",
        "AI Rule Candidate": "IF undisclosed_prior_flood_history_discovered AND insurer_underwriting_evidence_shows_cover_would_not_have_been_offered THEN flood_exclusion_endorsement_valid_retrospective_to_inception; Flood_Re NOT available IF policy_obtained_via_price_comparison_site",
        "Source PDF": "DRN-2787998.pdf",
    },
    {
        "Case ID": "FLOOD-019",
        "FOS Decision ID": "DRN-2800821",
        "Insurer Name": "Aviva Insurance Limited",
        "FOS Decision Date": "3 Jun 2021",
        "Claim Type": "Premium pricing dispute — Aviva rated property as high flood risk from 2012 causing 73-47% annual premium increases; dispute whether 2012-2015 and 2016-2018 premiums were fair; Flood Re launched 2016 but premiums not proportionally reduced; also: 2006 claim incorrectly included causing overcharge 2014-2019 (already refunded £1,161)",
        "Leak Source": "Not applicable — no flood damage claim; complaint concerns premium pricing linked to high flood risk rating",
        "Property Type": "Residential home",
        "Dispute Type": "Claim Recording / Administrative Dispute",
        "Coverage Decision": "Not Applicable",
        "Rejection Reason": "Not applicable — no flood claim was made; dispute concerns premium pricing and administration",
        "Evidence Dispute": "Aviva cited flood mapping model rating property at significant flood risk from 2012; ABI Statement of Principles 2008 as basis for continuing cover at elevated premium; Mr D disputed the flood risk rating and ability to find alternative quotes at ~£400/year; Aviva could not provide pre-2014 premium breakdown (bank administered policy until 2014); Flood Re Scheme cost was £330-335 per year for property",
        "Outcome Category": "Upheld in Part",
        "Outcome": "Upheld in part — 2012-2015 premiums fair (flood risk rating justified per ABI Statement of Principles); 2016-2018 premiums excessive post-Flood Re cession; Aviva to recalculate 2016-2018 premiums as (2011 base premium less insurance premium tax less 10% flood loading) plus Flood Re cost plus applicable insurance premium tax; refund difference with 8% simple interest per annum from date of each instalment; 2006 claim overcharge already remedied (£1,161 refunded). Aviva must also cancel and reissue uncashed cheques if required.",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "No — Administrative",
        "Key Policy Clause": "ABI Statement of Principles 2008 — insurer obligation to continue cover for existing high flood risk policyholders; Flood Re scheme (launched 2016) — flood risk cession mechanism; duty not to exploit policyholder non-engagement; insurer deemed to know policyholder is not engaging by fifth renewal (2015 for Mrs D); premium after Flood Re cession must reflect actual retained risk only",
        "Missing Evidence": "Pre-2014 premium breakdown unavailable (bank administered policy until 2014); proportion of 2011 premium attributable to flood risk unknown — FOS applied 10% proxy deduction in absence of true figure",
        "Ombudsman Reasoning": "Insurer entitled to change flood risk assessment and increase premium; ABI Statement of Principles justified continued cover at elevated premiums 2012-2015 without sudden increases (phased increases acknowledged); once Flood Re cession in 2016 transferred flood risk, retained premium should have returned to pre-risk-rating base plus Flood Re charges; residual premium still 234% above pre-2012 level despite cession; insurer exploited non-engagement after 5th renewal; 10% proxy deduction applied as equitable substitute for unknown 2011 flood loading",
        "Workflow Insight": "After Flood Re cession, insurers must recalibrate retained premium to exclude the flood risk element now covered by Flood Re; inflated pre-cession flood premiums cannot simply continue post-cession; policy administration handovers should audit prior claim loadings for accuracy; FOS considers an insurer ought to know a policyholder is not engaging by their fifth renewal",
        "AI Rule Candidate": "IF flood_re_cession = TRUE AND premium_not_reduced_to_reflect_ceded_flood_risk THEN investigate_post_cession_overcharge; IF policy_administration_transferred THEN audit_prior_claim_loadings_for_accuracy; IF policyholder_has_not_engaged_by_5th_renewal THEN insurer_must_not_exploit_inertia_by_above_market_increases",
        "Source PDF": "DRN-2800821.pdf",
    },
    {
        "Case ID": "FLOOD-020",
        "FOS Decision ID": "DRN-2882609",
        "Insurer Name": "Aviva Insurance Limited",
        "FOS Decision Date": "3 Aug 2021",
        "Claim Type": "Flood claim accepted but settlement inadequate — Aviva paid £4,629.19 for drying and redecoration only; refused full structural waterproofing repairs citing betterment and wear and tear; 200-year-old stone property flooded following named storm February 2020",
        "Leak Source": "External flood water via ground floor — 200-year-old solid stone walls; external ground levels higher than internal floor in places; no damp proof course or internal tanking; water entered ground floor via wall/floor junction; property had no prior flood claim",
        "Property Type": "Residential home (200-year-old stone property with 45-year extension)",
        "Dispute Type": "Handling / Reinstatement Dispute",
        "Coverage Decision": "Accepted — Disputed Settlement",
        "Rejection Reason": "Aviva argued full structural waterproofing and DPC works would constitute betterment — structural deficiencies (no DPC, no tanking, high external ground levels) attributed to age, wear and tear, and poor construction pre-dating the flood; settlement limited to drying out and redecoration (£4,629.19) as restoring pre-loss condition",
        "Evidence Dispute": "Aviva surveyor attributed structural deficiencies to age/wear/tear not flood — concluded drying and redecoration restored property to pre-loss condition; claimant argued property was watertight before flood (no prior claim); water re-entered after Aviva's settlement confirming inadequacy; claimant self-funded approximately £16,000 in repairs; surveyor's own comments confirmed drying/redecoration would not prevent future water ingress",
        "Outcome Category": "Upheld",
        "Outcome": "Complaint upheld — Aviva must arrange effective and long-lasting repairs to prevent future flood damage, or reimburse Mr O's repair costs with 8% simple interest per annum from date of payment; drying and redecoration insufficient as confirmed by surveyor's own evidence; once claim accepted insurer cannot rely on wear and tear exclusion to avoid paying for effective repairs; betterment principle does not apply where repairs restore (not improve) pre-loss condition",
        "Compensation Awarded (£)": 0,
        "Is Core Case": "Yes",
        "Key Policy Clause": "Indemnity principle — insurer must provide effective and long-lasting repairs, not merely cosmetic restoration; betterment does not apply where works are necessary to restore pre-loss watertight condition (not to improve upon it); once claim accepted, insurer cannot rely on wear and tear exclusion to deny effective repairs for the same claim; property built to building regulations at time of construction is assumed adequate unless pre-existing deficiency is evidenced",
        "Missing Evidence": "Pre-loss structural survey or DPC testing confirming property was or was not watertight before flood; independent assessment distinguishing flood-caused damage from pre-existing deficiencies; itemised repair costs for flood-attributable works vs structural improvement; no prior flood claim or water ingress evidence to support pre-existing deficiency argument",
        "Ombudsman Reasoning": "No prior flood claim and no evidence property was not watertight before February 2020 event; property and extension assumed built to building regulations at time; once claim accepted, insurer bound to provide effective and lasting repair; surveyor's own letter confirmed drying/redecoration would not prevent recurrence — this undermines Aviva's position; water re-entered confirming inadequacy of settlement; betterment only applies where repairs improve property beyond pre-loss state — here they merely restore it; wear and tear exclusion cannot be invoked after accepting claim to avoid liability for effective repairs",
        "Workflow Insight": "Betterment argument fails when property had no prior water ingress and structural works are necessary to restore pre-flood watertight condition; insurers should conduct thorough pre-settlement structural surveys to separate pre-existing deficiencies from flood-caused vulnerabilities before accepting; accepting a claim forecloses subsequent reliance on wear and tear to deny effective reinstatement; surveyor comments acknowledging inadequacy of proposed repairs are binding admissions",
        "AI Rule Candidate": "IF claim_accepted = TRUE AND no_prior_water_ingress_history AND surveyor_confirms_proposed_repairs_inadequate THEN insurer_must_fund_effective_lasting_repairs; betterment_exclusion NOT applicable WHERE repairs_restore_not_improve_pre_loss_condition",
        "Source PDF": "DRN-2882609.pdf",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _row_fill(even: bool) -> PatternFill:
    color = "D6E4F0" if even else "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def _validate(case: dict) -> None:
    case_id = case.get("Case ID", "?")
    for field, allowed in CONTROLLED_VOCAB.items():
        val = case.get(field, "")
        if val not in allowed:
            raise ValueError(
                f"{case_id} — '{field}' value '{val}' not in controlled vocab.\n"
                f"  Allowed: {sorted(allowed)}"
            )
    if not isinstance(case.get("Compensation Awarded (£)", 0), (int, float)):
        raise TypeError(
            f"{case_id} — 'Compensation Awarded (£)' must be a number."
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not NEW_CASES:
        print("NEW_CASES is empty — nothing to append.")
        return

    for case in NEW_CASES:
        _validate(case)

    repo_root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.join(
        repo_root, "knowledge", "case-databases", "Flood_Case_Database.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    live_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if live_headers != COLUMNS:
        mismatch = [
            (i + 1, live_headers[i] if i < len(live_headers) else "—", COLUMNS[i])
            for i in range(max(len(live_headers), len(COLUMNS)))
            if i >= len(live_headers) or i >= len(COLUMNS) or live_headers[i] != COLUMNS[i]
        ]
        raise RuntimeError(
            "Live workbook columns do not match COLUMNS definition.\n"
            "Mismatches (col, live, expected):\n"
            + "\n".join(f"  {c}: '{l}' vs '{e}'" for c, l, e in mismatch)
        )

    cell_font = Font(name="Calibri", size=10)
    border    = _border()
    wrap      = Alignment(wrap_text=True, vertical="top")
    centre    = Alignment(horizontal="center", vertical="top")

    first_new_row = ws.max_row + 1

    for i, case in enumerate(NEW_CASES):
        row_idx = first_new_row + i
        fill = _row_fill(row_idx % 2 == 0)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = centre if col_name in CENTRED_COLS else wrap
        ws.row_dimensions[row_idx].height = 120

    wb.save(xlsx_path)

    last_row  = ws.max_row
    last_case = ws.cell(row=last_row, column=1).value
    total_data = last_row - 1

    print(f"Appended {len(NEW_CASES)} case(s) to {xlsx_path}")
    print(f"Total data rows : {total_data}")
    print(f"Last Case ID    : {last_case}")


if __name__ == "__main__":
    main()
